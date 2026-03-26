from __future__ import annotations

import argparse
import math
from pathlib import Path

from openpyxl import load_workbook


def _load(path: Path):
    return load_workbook(path, data_only=True)


def _normalize(value):
    if value == "":
        return None
    return value


def _values_equal(left, right, tolerance: float) -> bool:
    left = _normalize(left)
    right = _normalize(right)

    if left is None and right is None:
        return True

    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=0.0, abs_tol=tolerance)

    return left == right


def compare_workbooks(
    baseline_path: Path,
    candidate_path: Path,
    tolerance: float,
    max_diffs: int,
    include_sheets: list[str] | None = None,
    ignore_extra_sheets: bool = False,
):
    baseline_wb = _load(baseline_path)
    candidate_wb = _load(candidate_path)

    diffs: list[str] = []

    baseline_sheets = include_sheets or list(baseline_wb.sheetnames)
    candidate_sheets = list(candidate_wb.sheetnames)

    for sheet_name in baseline_sheets:
        if sheet_name not in candidate_sheets:
            diffs.append(f"Missing sheet in candidate: {sheet_name}")
            if len(diffs) >= max_diffs:
                return diffs
            continue

        baseline_ws = baseline_wb[sheet_name]
        candidate_ws = candidate_wb[sheet_name]

        max_row = max(baseline_ws.max_row, candidate_ws.max_row)
        max_col = max(baseline_ws.max_column, candidate_ws.max_column)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                baseline_value = baseline_ws.cell(row, col).value
                candidate_value = candidate_ws.cell(row, col).value
                if _values_equal(baseline_value, candidate_value, tolerance):
                    continue

                diffs.append(
                    f"{sheet_name}!R{row}C{col}: baseline={baseline_value!r} candidate={candidate_value!r}"
                )
                if len(diffs) >= max_diffs:
                    return diffs

    if not ignore_extra_sheets:
        for sheet_name in candidate_sheets:
            if sheet_name not in baseline_sheets:
                diffs.append(f"Unexpected sheet in candidate: {sheet_name}")
                if len(diffs) >= max_diffs:
                    return diffs

    return diffs


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare two Excel workbooks cell by cell.")
    parser.add_argument("baseline", type=Path, help="Path to the approved baseline workbook")
    parser.add_argument("candidate", type=Path, help="Path to the candidate workbook")
    parser.add_argument("--tolerance", type=float, default=1e-9, help="Absolute tolerance for numeric comparisons")
    parser.add_argument("--max-diffs", type=int, default=200, help="Maximum number of diffs to print")
    parser.add_argument("--sheet", action="append", dest="sheets", help="Restrict comparison to a specific sheet. Can be passed multiple times.")
    parser.add_argument("--ignore-extra-sheets", action="store_true", help="Do not fail on sheets present only in the candidate workbook.")
    args = parser.parse_args()

    diffs = compare_workbooks(
        args.baseline,
        args.candidate,
        args.tolerance,
        args.max_diffs,
        include_sheets=args.sheets,
        ignore_extra_sheets=args.ignore_extra_sheets,
    )
    if diffs:
        print("=== WORKBOOK DIFF ===")
        for diff in diffs:
            print(diff)
        return 1

    print("=== WORKBOOKS MATCH ===")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())