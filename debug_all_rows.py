#!/usr/bin/env python
"""Check ALL acciones exterior and their ratios + verify output v3."""
import sys
sys.path.insert(0, '.')
from pathlib import Path
from openpyxl import load_workbook

wb = load_workbook('TEST_output_v3_values.xlsx')

# Check PrecioTenenciasIniciales ratio column
ws_pt = wb['PrecioTenenciasIniciales']
print("=== PrecioTenenciasIniciales - ALL rows ===")
headers = [ws_pt.cell(1, c).value for c in range(1, ws_pt.max_column + 1)]
print(f"Headers: {headers}")
print()
for row in range(2, ws_pt.max_row + 1):
    vals = [ws_pt.cell(row, c).value for c in range(1, ws_pt.max_column + 1)]
    ratio_val = vals[7] if len(vals) > 7 else 'N/A'  # col 8 = Ratio CEDEAR
    if ratio_val and ratio_val != '':
        print(f"  Row {row}: {vals[0]:>8} {str(vals[1]):<12} ratio={ratio_val}")

# Check Posicion Inicial Gallo
ws_pi = wb['Posicion Inicial Gallo']
print("\n=== Posicion Inicial Gallo - Precio a Utilizar vs Precio Nominal ===")
for row in range(2, ws_pi.max_row + 1):
    ticker = ws_pi.cell(row, 2).value
    pu = ws_pi.cell(row, 16).value
    pn = ws_pi.cell(row, 22).value
    same = "SAME" if pu == pn else f"DIFF (pu={pu}, pn={pn})"
    print(f"  {str(ticker):<15} P.Util={str(pu):>18}  P.Nom={str(pn):>18}  {same}")

# Check RatiosCedearsAcciones sheet
print(f"\n=== RatiosCedearsAcciones sheet exists: {'RatiosCedearsAcciones' in wb.sheetnames} ===")
if 'RatiosCedearsAcciones' in wb.sheetnames:
    ws_r = wb['RatiosCedearsAcciones']
    print(f"  Rows: {ws_r.max_row - 1}, Cols: {ws_r.max_column}")
    print(f"  Headers: {[ws_r.cell(1, c).value for c in range(1, ws_r.max_column + 1)]}")
    # Show first 3 rows
    for r in range(2, min(5, ws_r.max_row + 1)):
        print(f"  Row {r}: {[ws_r.cell(r, c).value for c in range(1, ws_r.max_column + 1)]}")
