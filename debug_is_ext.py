#!/usr/bin/env python
"""Check which tickers are acciones exterior."""
import sys
sys.path.insert(0, '.')
from pathlib import Path
from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger
from openpyxl import load_workbook

m = GalloVisualMerger(
    '12128_LANDRO_VERONICA_INES_Gallo_Generado_OK.xlsx',
    '12128_LANDRO_VERONICA_INES_Visual_Generado_OK.xlsx',
    str(Path('pdf_converter/datalab/aux_data')),
    precio_tenencias_path='precio_estructurado.xlsx'
)

# Check tickers from Posicion Inicial
tickers_to_check = ['KO','XOM','JNJ','GOOGL','MELI','AAPL','AMZN','ALUA','PAMP','TXAR','NVDA-US','TSLA-US']
for t in tickers_to_check:
    code = m._get_codigo_from_ticker(t)
    code_str = str(code) if code else ""
    is_ext = m._is_accion_exterior(code_str) if code_str else False
    ratio = m._get_ratio_for_especie(t, '')
    # Also check EspeciesVisual data
    data = m._especies_visual_cache.get(m._clean_codigo(code_str), {}) if code_str else {}
    moneda = data.get('moneda_emision', '')
    tipo = data.get('tipo_especie', '')
    print(f"{t:>10}: code={str(code):>6}, is_ext={is_ext}, ratio={ratio}, moneda='{moneda}', tipo='{tipo}'")

# Now check ALL PrecioTenenciasIniciales rows
print("\n=== PrecioTenenciasIniciales codigos vs EspeciesVisual ===")
wb = load_workbook('precio_estructurado.xlsx')
ws = wb['PrecioTenenciasIniciales']
for row in range(2, ws.max_row + 1):
    cod = ws.cell(row, 1).value
    ticker = ws.cell(row, 2).value
    if cod:
        cod_clean = m._clean_codigo(str(cod))
        is_ext = m._is_accion_exterior(cod_clean)
        ratio = m._get_ratio_for_especie(str(ticker) if ticker else '', '')
        if is_ext or ratio:
            print(f"  Row {row}: cod={cod} ({cod_clean}), ticker={ticker}, is_ext={is_ext}, ratio={ratio}")
