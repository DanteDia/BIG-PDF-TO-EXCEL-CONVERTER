#!/usr/bin/env python
"""Debug ratio column in PrecioTenenciasIniciales."""
import sys
sys.path.insert(0, '.')
from pathlib import Path
from openpyxl import load_workbook

# Load the precio_estructurado to see what _add_precio_tenencias_sheet sees
wb = load_workbook('precio_estructurado.xlsx')
ws = wb['PrecioTenenciasIniciales'] if 'PrecioTenenciasIniciales' in wb.sheetnames else wb.active

max_col = ws.max_column
headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, max_col + 1)]
headers_lower = [h.lower() for h in headers]
print(f"Headers: {headers}")
print(f"Max col: {max_col}")

def find_col(keyword):
    for idx, h in enumerate(headers_lower, 1):
        if keyword in h:
            return idx
    return None

print(f"find_col('cod') = {find_col('cod')}")
print(f"find_col('ticker') = {find_col('ticker')}")
print(f"find_col('cantidad') = {find_col('cantidad')}")
print(f"find_col('importe') = {find_col('importe')}")
print(f"find_col('especie') = {find_col('especie')}")  # This might match 'cod.especie' first!

# Check NVIDIA and TESLA rows
for row in range(2, ws.max_row + 1):
    ticker = ws.cell(row, find_col('ticker')).value
    if ticker and str(ticker).upper() in ['NVIDIA', 'TESLA']:
        codigo = ws.cell(row, find_col('cod')).value
        especie_col = find_col('especie')
        especie_val = ws.cell(row, especie_col).value if especie_col else 'N/A'
        print(f"\nRow {row}: ticker={ticker}, codigo={codigo}")
        print(f"  especie (col {especie_col}): {especie_val}")
        for c in range(1, max_col + 1):
            print(f"  col {c} ({headers[c-1]}): {ws.cell(row, c).value}")

# Now test with the merger
from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger

merger = GalloVisualMerger(
    '12128_LANDRO_VERONICA_INES_Gallo_Generado_OK.xlsx',
    '12128_LANDRO_VERONICA_INES_Visual_Generado_OK.xlsx',
    str(Path('pdf_converter/datalab/aux_data')),
    precio_tenencias_path='precio_estructurado.xlsx'
)

# Check which codigos are acciones exterior
print("\n=== Acciones exterior check ===")
for row in range(2, ws.max_row + 1):
    codigo = ws.cell(row, find_col('cod')).value
    if codigo:
        cod_clean = merger._clean_codigo(str(codigo))
        is_ext = merger._is_accion_exterior(cod_clean)
        if is_ext:
            ticker = ws.cell(row, find_col('ticker')).value
            print(f"  codigo={codigo} ({cod_clean}), ticker={ticker}, is_accion_exterior=True")
            ratio = merger._get_ratio_for_especie(str(ticker) if ticker else '', '')
            print(f"    ratio for ticker '{ticker}': {ratio}")
