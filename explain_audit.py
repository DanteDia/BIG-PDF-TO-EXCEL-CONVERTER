from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from pdf_converter.datalab.economic_sanity import validate_workbook


ROOT = Path(__file__).resolve().parent


def load_registry(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"cases": {}}
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {"cases": {}}


def normalize_path(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except ValueError:
        return path.as_posix()


def find_case_for_workbook(workbook_path: Path, registry: dict[str, Any]) -> tuple[str | None, dict[str, Any] | None]:
    normalized = normalize_path(workbook_path)
    for case_id, case in (registry.get("cases") or {}).items():
        outputs = case.get("local_outputs") or {}
        audit = case.get("latest_audit") or {}
        candidates = [outputs.get("values"), audit.get("target")]
        if normalized in {str(candidate).replace("\\", "/") for candidate in candidates if candidate}:
            return case_id, case
    return None, None


def header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value:
            headers[str(value).strip().lower()] = col
    return headers


def find_col(headers: dict[str, int], *needles: str) -> int | None:
    normalized = [needle.lower() for needle in needles]
    for needle in normalized:
        if needle in headers:
            return headers[needle]
    for header, col in headers.items():
        if any(needle in header for needle in normalized):
            return col
    return None


def cell(ws, row: int, col: int | None) -> Any:
    if not col:
        return None
    return ws.cell(row, col).value


def row_context(wb, sheet_name: str, row: int) -> dict[str, Any]:
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers = header_map(ws)
    context = {
        "instrument": cell(ws, row, find_col(headers, "instrumento", "especie")),
        "code": cell(ws, row, find_col(headers, "cod.instrum", "cod instrum", "codigo")),
        "operation": cell(ws, row, find_col(headers, "tipo operacion", "operacion")),
        "quantity": cell(ws, row, find_col(headers, "cantidad")),
        "bruto": cell(ws, row, find_col(headers, "bruto")),
        "cost": cell(ws, row, find_col(headers, "costo por venta", "costo")),
        "result": cell(ws, row, find_col(headers, "resultado calculado", "resultado")),
        "audit": cell(ws, row, find_col(headers, "chequeado", "auditoria")),
    }
    return {key: value for key, value in context.items() if value not in (None, "")}


def explain_workbook(path: Path, registry_path: Path) -> int:
    workbook_path = path if path.is_absolute() else ROOT / path
    registry = load_registry(registry_path)
    case_id, case = find_case_for_workbook(workbook_path, registry)
    wb = load_workbook(workbook_path, data_only=True)
    report = validate_workbook(wb)

    print(f"Audit explanation: {normalize_path(workbook_path)}")
    if case_id and case:
        print(f"Case: {case_id} ({case.get('status', 'unknown')})")
        handoff = case.get("handoff")
        if handoff:
            print(f"Handoff: {handoff}")
    print(f"Issues: {report.issue_count} severity={report.counts_by_severity()}")

    grouped: dict[str, list] = {}
    for issue in report.issues:
        grouped.setdefault(issue.rule_id, []).append(issue)

    for rule_id, issues in sorted(grouped.items()):
        print(f"\n{rule_id}: {len(issues)}")
        for issue in issues[:20]:
            context = row_context(wb, issue.sheet, issue.row)
            context_text = ", ".join(f"{key}={value!r}" for key, value in context.items())
            print(f"  {issue.sheet} row {issue.row}: {issue.message} value={issue.value}")
            if context_text:
                print(f"    {context_text}")
        if len(issues) > 20:
            print(f"  ... {len(issues) - 20} more")

    if case and case.get("remaining_reviews"):
        print("\nRegistry remaining reviews:")
        for item in case["remaining_reviews"]:
            print(f"  - {item}")
    return 0


def explain_audit_json(path: Path) -> int:
    audit_path = path if path.is_absolute() else ROOT / path
    data = json.loads(audit_path.read_text(encoding="utf-8"))
    print(f"Audit JSON: {normalize_path(audit_path)}")
    print(f"Summary: {data.get('summary', {})}")
    for target in data.get("targets", []):
        print(
            f"[{target.get('bucket')}] {target.get('status')}: {target.get('label')} "
            f"issues={target.get('issue_count')} severity={target.get('counts_by_severity')} rules={target.get('counts_by_rule')}"
        )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Explain economic sanity audit triggers for a workbook or audit JSON.")
    parser.add_argument("path", type=Path, help="Values workbook (.xlsx) or release audit JSON.")
    parser.add_argument("--registry", type=Path, default=ROOT / "CASE_REGISTRY.yml")
    args = parser.parse_args()

    registry_path = args.registry if args.registry.is_absolute() else ROOT / args.registry
    if args.path.suffix.lower() == ".json":
        return explain_audit_json(args.path)
    return explain_workbook(args.path, registry_path)


if __name__ == "__main__":
    raise SystemExit(main())