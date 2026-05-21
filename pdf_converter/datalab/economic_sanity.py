from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Iterable
import unicodedata

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


VALIDATION_SHEET_NAME = "Validacion"


@dataclass(frozen=True)
class EconomicSanityConfig:
    usd_large_boleto: float = 3_000_000.0
    ars_large_boleto: float = 3_000_000_000.0
    usd_tiny_boleto: float = 5.0
    ars_tiny_boleto: float = 5_000.0
    result_ratio_review: float = 0.80
    result_ratio_high: float = 1.00
    max_issues_per_rule: int = 200


@dataclass(frozen=True)
class ValidationIssue:
    severity: str
    rule_id: str
    sheet: str
    row: int
    metric: str
    value: float | str | None
    message: str
    suggested_review: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "rule_id": self.rule_id,
            "sheet": self.sheet,
            "row": self.row,
            "metric": self.metric,
            "value": self.value,
            "message": self.message,
            "suggested_review": self.suggested_review,
        }


@dataclass(frozen=True)
class ValidationReport:
    generated_at: str
    issues: tuple[ValidationIssue, ...] = field(default_factory=tuple)

    @property
    def issue_count(self) -> int:
        return len(self.issues)

    def counts_by_severity(self) -> dict[str, int]:
        counts: dict[str, int] = {}
        for issue in self.issues:
            counts[issue.severity] = counts.get(issue.severity, 0) + 1
        return counts

    def to_dict(self) -> dict[str, Any]:
        return {
            "generated_at": self.generated_at,
            "issue_count": self.issue_count,
            "counts_by_severity": self.counts_by_severity(),
            "issues": [issue.to_dict() for issue in self.issues],
        }


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.lower().replace(".", " ").replace("_", " ").split())


def _to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]
    if text.endswith("-"):
        negative = True
        text = text[:-1]
    text = text.replace("$", "").replace("USD", "").replace("U$S", "")
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if negative else number


def _header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        normalized = _normalize_text(ws.cell(1, col).value)
        if normalized:
            headers[normalized] = col
    return headers


def _find_col(headers: dict[str, int], candidates: Iterable[str]) -> int | None:
    normalized_candidates = [_normalize_text(candidate) for candidate in candidates]
    for candidate in normalized_candidates:
        if candidate in headers:
            return headers[candidate]
    for header, col in headers.items():
        if any(candidate and candidate in header for candidate in normalized_candidates):
            return col
    return None


def _clean_code(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = "".join(char for char in text if char.isdigit())
    if digits:
        return str(int(digits))
    return _normalize_text(text)


def _recovered_cost_codes(wb) -> set[str]:
    if "Posicion Inicial Gallo" not in wb.sheetnames:
        return set()

    ws = wb["Posicion Inicial Gallo"]
    headers = _header_map(ws)
    code_col = _find_col(headers, ["Codigo especie", "Cod.Especie", "Codigo"])
    origin_col = _find_col(headers, ["Origen precio costo"])
    if not code_col or not origin_col:
        return set()

    codes: set[str] = set()
    for row in range(2, ws.max_row + 1):
        origin = _normalize_text(_cell(ws, row, origin_col))
        if "preciotenenciascostorecuperado" in origin.replace(" ", ""):
            code = _clean_code(_cell(ws, row, code_col))
            if code:
                codes.add(code)
    return codes


def _cell(ws, row: int, col: int | None) -> Any:
    if not col:
        return None
    return ws.cell(row, col).value


def _is_usd_currency(value: Any) -> bool:
    text = _normalize_text(value)
    return "dolar" in text or "usd" in text or "u$s" in text


def _is_future_row(*values: Any) -> bool:
    combined = " ".join(_normalize_text(value) for value in values)
    return "futuro" in combined


def _append_limited(issues: list[ValidationIssue], issue: ValidationIssue, per_rule_counts: dict[str, int], limit: int) -> None:
    count = per_rule_counts.get(issue.rule_id, 0)
    if count >= limit:
        return
    per_rule_counts[issue.rule_id] = count + 1
    issues.append(issue)


def _validate_boletos(wb, config: EconomicSanityConfig, issues: list[ValidationIssue], per_rule_counts: dict[str, int]) -> None:
    if "Boletos" not in wb.sheetnames:
        return
    ws = wb["Boletos"]
    headers = _header_map(ws)
    moneda_col = _find_col(headers, ["Moneda"])
    bruto_col = _find_col(headers, ["Bruto"])
    neto_col = _find_col(headers, ["Neto", "Neto Calculado"])
    cantidad_col = _find_col(headers, ["Cantidad"])
    precio_col = _find_col(headers, ["Precio Nominal", "Precio", "Precio Unitario"])
    operacion_col = _find_col(headers, ["Tipo Operación", "Operacion"])
    instrumento_col = _find_col(headers, ["Tipo de Instrumento", "Instrumento"])

    if not bruto_col:
        return

    for row in range(2, ws.max_row + 1):
        moneda = _cell(ws, row, moneda_col)
        bruto = _to_float(_cell(ws, row, bruto_col))
        neto = _to_float(_cell(ws, row, neto_col))
        abs_bruto = abs(bruto)
        abs_neto = abs(neto)
        is_usd = _is_usd_currency(moneda)
        is_future = _is_future_row(_cell(ws, row, instrumento_col), _cell(ws, row, operacion_col))

        if abs_bruto > 0:
            if is_usd and abs_bruto > config.usd_large_boleto:
                _append_limited(
                    issues,
                    ValidationIssue(
                        "review", "BOL-USD-LARGE-001", "Boletos", row, "abs_bruto_usd", round(abs_bruto, 4),
                        f"Boleto USD con bruto mayor a {config.usd_large_boleto:,.0f}.",
                        "Revisar escala de cantidad/precio, moneda y tipo de cambio.",
                    ),
                    per_rule_counts,
                    config.max_issues_per_rule,
                )
            if not is_usd and abs_bruto > config.ars_large_boleto:
                _append_limited(
                    issues,
                    ValidationIssue(
                        "review", "BOL-ARS-LARGE-001", "Boletos", row, "abs_bruto_ars", round(abs_bruto, 4),
                        f"Boleto ARS con bruto mayor a {config.ars_large_boleto:,.0f}.",
                        "Revisar si la escala del cliente justifica el monto o si hay multiplicacion OCR.",
                    ),
                    per_rule_counts,
                    config.max_issues_per_rule,
                )
            tiny_threshold = config.usd_tiny_boleto if is_usd else config.ars_tiny_boleto
            if 0 < abs_bruto <= tiny_threshold and not is_future:
                _append_limited(
                    issues,
                    ValidationIssue(
                        "review", "BOL-TINY-001", "Boletos", row, "abs_bruto", round(abs_bruto, 4),
                        f"Boleto con bruto muy chico para revision ({abs_bruto:,.4f}).",
                        "Revisar si es una operacion real o un corrimiento/parseo de centavos.",
                    ),
                    per_rule_counts,
                    config.max_issues_per_rule,
                )

        if abs_bruto > 0 and abs_neto > 0 and not is_future:
            ratio = abs_neto / abs_bruto
            if ratio > 1.5 or ratio < 0.5:
                _append_limited(
                    issues,
                    ValidationIssue(
                        "review", "BOL-NETO-BRUTO-001", "Boletos", row, "abs_neto_abs_bruto_ratio", round(ratio, 6),
                        "Neto y bruto difieren en una magnitud inusual.",
                        "Revisar gastos, signos, moneda y si bruto/neto fueron rescatados correctamente.",
                    ),
                    per_rule_counts,
                    config.max_issues_per_rule,
                )

        cantidad = abs(_to_float(_cell(ws, row, cantidad_col)))
        precio = abs(_to_float(_cell(ws, row, precio_col)))
        if cantidad > 0 and precio > 0 and abs_bruto > 0 and not is_future:
            expected = cantidad * precio
            ratio = expected / abs_bruto if abs_bruto else 0
            if ratio > 1_000 or ratio < 0.001:
                _append_limited(
                    issues,
                    ValidationIssue(
                        "review", "BOL-QTY-PRICE-BRUTO-001", "Boletos", row, "qty_price_vs_bruto", round(ratio, 6),
                        "Cantidad x precio queda extremadamente lejos del bruto.",
                        "Revisar precio cada 100, tipo de cambio, cantidad OCR o columnas corridas.",
                    ),
                    per_rule_counts,
                    config.max_issues_per_rule,
                )


def _validate_resultado_sheet(
    wb,
    sheet_name: str,
    moneda_tipo: str,
    config: EconomicSanityConfig,
    issues: list[ValidationIssue],
    per_rule_counts: dict[str, int],
) -> None:
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    headers = _header_map(ws)
    bruto_col = _find_col(headers, ["Bruto", "Bruto en USD"])
    costo_col = _find_col(headers, ["Costo", "Costo Computable", "Costo Calculado"])
    resultado_col = _find_col(headers, ["Resultado Calculado(final)", "Resultado Calculado", "Resultado"])
    operacion_col = _find_col(headers, ["Tipo Operación", "Operacion"])
    stock_qty_col = _find_col(headers, ["Cantidad Stock Inicial"])
    code_col = _find_col(headers, ["Cod.Instrum", "Cod Instrum", "Codigo instrumento", "Codigo"])
    recovered_codes = _recovered_cost_codes(wb) if moneda_tipo == "ARS" else set()

    if not resultado_col:
        return

    for row in range(2, ws.max_row + 1):
        operacion_text = _normalize_text(_cell(ws, row, operacion_col))
        if "compra" in operacion_text and "venta" not in operacion_text:
            continue
        resultado = _to_float(_cell(ws, row, resultado_col))
        bruto = _to_float(_cell(ws, row, bruto_col))
        costo = _to_float(_cell(ws, row, costo_col))
        if resultado == 0:
            continue

        denominator_name = "bruto"
        denominator = abs(bruto)
        if moneda_tipo == "USD" and abs(costo) > 0:
            denominator_name = "costo"
            denominator = abs(costo)
        if denominator <= 0:
            continue

        ratio = abs(resultado) / denominator
        code = _clean_code(_cell(ws, row, code_col))
        recovered_cost_stock = moneda_tipo == "ARS" and code in recovered_codes and _to_float(_cell(ws, row, stock_qty_col)) > 0
        if recovered_cost_stock and ratio <= 1.000001:
            continue
        if ratio > config.result_ratio_high:
            severity = "high"
        elif ratio > config.result_ratio_review:
            severity = "review"
        else:
            severity = ""
        if severity:
            _append_limited(
                issues,
                ValidationIssue(
                    severity,
                    f"RES-{moneda_tipo}-RATIO-001",
                    sheet_name,
                    row,
                    f"abs_resultado_abs_{denominator_name}",
                    round(ratio, 6),
                    f"Resultado representa {ratio:.2%} del {denominator_name}.",
                    "Revisar stock inicial, costo PPP, cantidad, precio y routing ARS/USD.",
                ),
                per_rule_counts,
                config.max_issues_per_rule,
            )


def validate_workbook(wb, config: EconomicSanityConfig | None = None) -> ValidationReport:
    config = config or EconomicSanityConfig()
    issues: list[ValidationIssue] = []
    per_rule_counts: dict[str, int] = {}
    _validate_boletos(wb, config, issues, per_rule_counts)
    _validate_resultado_sheet(wb, "Resultado Ventas ARS", "ARS", config, issues, per_rule_counts)
    _validate_resultado_sheet(wb, "Resultado Ventas USD", "USD", config, issues, per_rule_counts)
    return ValidationReport(datetime.now().isoformat(timespec="seconds"), tuple(issues))


def add_validation_sheet(wb, report: ValidationReport) -> None:
    if VALIDATION_SHEET_NAME in wb.sheetnames:
        del wb[VALIDATION_SHEET_NAME]
    ws = wb.create_sheet(VALIDATION_SHEET_NAME, 0)
    ws.append(["Validacion economica", report.generated_at])
    ws.append(["Total issues", report.issue_count])
    for severity, count in sorted(report.counts_by_severity().items()):
        ws.append([f"Issues {severity}", count])
    ws.append([])

    headers = ["Severidad", "Regla", "Hoja", "Fila", "Metrica", "Valor", "Mensaje", "Revision sugerida"]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")

    severity_fills = {
        "high": PatternFill("solid", fgColor="F4CCCC"),
        "review": PatternFill("solid", fgColor="FFF2CC"),
    }
    for issue in report.issues:
        ws.append([
            issue.severity,
            issue.rule_id,
            issue.sheet,
            issue.row,
            issue.metric,
            issue.value,
            issue.message,
            issue.suggested_review,
        ])
        fill = severity_fills.get(issue.severity)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    if not report.issues:
        ws.append(["ok", "SIN-TRIGGERS", "", "", "", "", "No se detectaron triggers economicos.", ""])

    for column in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row, column).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(column)].width = min(max_len + 2, 55)
    ws.freeze_panes = "A6"
