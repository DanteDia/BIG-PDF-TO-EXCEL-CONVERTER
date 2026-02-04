"""
Módulo para exportar el Excel consolidado a PDF con formato Visual.

Genera un PDF con todas las secciones del reporte financiero:
- Boletos (por tipo de instrumento)
- Resultado Ventas ARS
- Resultado Ventas USD
- Rentas y Dividendos ARS
- Rentas y Dividendos USD
- Cauciones Tomadoras
- Cauciones Colocadoras
- Resumen
- Posición de Títulos

Usa Datalab API para leer valores de fórmulas Excel (cross-platform).
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, 
    PageBreak, KeepTogether, Image
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import io

# Version para debugging en Streamlit Cloud
__version__ = "2.0.0-datalab"


class ExcelToPdfExporter:
    """
    Exporta un Excel consolidado (merge Gallo+Visual) a PDF con formato Visual.
    
    Versión 2.0: Usa Datalab markdown para leer valores de fórmulas.
    """
    
    # Colores corporativos
    HEADER_BG = colors.Color(0.2, 0.3, 0.5)  # Azul oscuro
    HEADER_TEXT = colors.white
    ROW_ALT_BG = colors.Color(0.95, 0.95, 0.95)  # Gris claro
    SECTION_BG = colors.Color(0.85, 0.85, 0.9)  # Gris azulado
    SUBSECTION_BG = colors.Color(0.9, 0.9, 0.95)  # Gris más claro
    
    def __init__(self, excel_path: str, cliente_info: Dict[str, str] = None, 
                 datalab_api_key: str = None, datalab_markdown: str = None):
        """
        Inicializa el exportador.
        
        Args:
            excel_path: Ruta al Excel consolidado
            cliente_info: Diccionario con info del cliente (numero, nombre)
            datalab_api_key: API key de Datalab para leer valores de fórmulas
            datalab_markdown: Markdown ya convertido por Datalab (recomendado para evitar re-conversión)
        """
        self.excel_path = Path(excel_path)
        self.wb = load_workbook(excel_path, data_only=True)
        
        # Inicializar atributos (COM ya no se usa, pero mantener para compatibilidad)
        self._com_data = None
        self._datalab_reader = None
        
        # 1. Usar markdown pre-convertido si está disponible (más eficiente)
        if datalab_markdown:
            try:
                from .datalab_excel_reader import DatalabExcelReader
                print("[INFO] Usando markdown Datalab pre-convertido")
                self._datalab_reader = DatalabExcelReader(datalab_api_key or "")
                self._datalab_reader.parse_all_sections(datalab_markdown)
                print(f"[INFO] Secciones parseadas: {list(self._datalab_reader._parsed_data.keys())}")
            except Exception as e:
                print(f"[WARNING] No se pudo parsear Datalab markdown: {e}")
        
        # 2. Si no hay markdown pero hay API key, convertir Excel ahora
        elif datalab_api_key:
            try:
                from .datalab_excel_reader import DatalabExcelReader
                print("[INFO] Convirtiendo Excel con Datalab API...")
                self._datalab_reader = DatalabExcelReader(datalab_api_key)
                markdown = self._datalab_reader.convert_to_markdown(str(excel_path))
                if markdown:
                    self._datalab_reader.parse_all_sections(markdown)
                    print(f"[INFO] Secciones parseadas: {list(self._datalab_reader._parsed_data.keys())}")
            except Exception as e:
                print(f"[WARNING] No se pudo usar Datalab API: {e}")
        
        # 3. Fallback: openpyxl (valores de fórmulas serán None)
        if not self._datalab_reader:
            print("[WARNING] Sin Datalab - valores de fórmulas pueden estar vacíos")
        
        # Info del cliente
        self.cliente_info = cliente_info or {
            'numero': 'XXXXX',
            'nombre': 'CLIENTE'
        }
        
        # Período del reporte (detectar de datos o usar default)
        self.periodo_inicio = "Enero 1"
        self.periodo_fin = "Diciembre 31"
        self.anio = datetime.now().year
        
        # Estilos
        self.styles = getSampleStyleSheet()
        self._setup_styles()
    
    def _get_cell_value(self, sheet_name: str, row: int, col: int) -> Any:
        """
        Obtiene el valor de una celda, usando COM si está disponible.
        row y col son 1-indexed (como en Excel/openpyxl).
        """
        if self._com_data and sheet_name in self._com_data:
            data = self._com_data[sheet_name]
            if row <= len(data) and col <= len(data[row-1]):
                return data[row-1][col-1]
            return None
        else:
            # Fallback a openpyxl
            if sheet_name in self.wb.sheetnames:
                return self.wb[sheet_name].cell(row, col).value
            return None
    
    def _get_sheet_data(self, sheet_name: str) -> List[List[Any]]:
        """
        Obtiene todos los datos de una hoja como lista de listas.
        """
        if self._com_data and sheet_name in self._com_data:
            return self._com_data[sheet_name]
        else:
            # Fallback a openpyxl
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                data = []
                for row in range(1, ws.max_row + 1):
                    row_data = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
                    data.append(row_data)
                return data
            return []
    
    def _setup_styles(self):
        """Configura estilos personalizados."""
        # Título de sección
        self.styles.add(ParagraphStyle(
            'SectionTitle',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceAfter=6,
            spaceBefore=12,
            textColor=colors.Color(0.2, 0.2, 0.4),
            fontName='Helvetica-Bold'
        ))
        
        # Subtítulo (tipo de instrumento)
        self.styles.add(ParagraphStyle(
            'SubsectionTitle',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceBefore=6,
            spaceAfter=3,
            textColor=colors.Color(0.3, 0.3, 0.5),
            fontName='Helvetica-Bold'
        ))
        
        # Normal pequeño para tablas
        self.styles.add(ParagraphStyle(
            'TableCell',
            parent=self.styles['Normal'],
            fontSize=7,
            leading=9
        ))
        
        # Header de tabla
        self.styles.add(ParagraphStyle(
            'TableHeader',
            parent=self.styles['Normal'],
            fontSize=7,
            fontName='Helvetica-Bold',
            textColor=colors.white
        ))
    
    def _format_number(self, value: Any, decimals: int = 2) -> str:
        """Formatea un número al estilo argentino (punto miles, coma decimales)."""
        if value is None:
            return ""
        try:
            num = float(value)
            # Manejar negativos con paréntesis
            is_negative = num < 0
            num = abs(num)
            
            # Formatear con separadores
            if decimals == 0:
                formatted = f"{num:,.0f}"
            else:
                formatted = f"{num:,.{decimals}f}"
            
            # Convertir a formato argentino
            # Primero reemplazar comas por placeholder, luego puntos por comas, finalmente placeholder por puntos
            formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
            
            if is_negative:
                return f"({formatted})"
            return formatted
        except (ValueError, TypeError):
            return str(value) if value else ""
    
    def _format_date(self, value: Any) -> str:
        """Formatea una fecha como D/M/AAAA."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        return str(value)
    
    def _get_header_footer(self, page_num: int, total_pages: int) -> str:
        """Genera el encabezado del reporte."""
        return f"REPORTE DE GANANCIAS / Período {self.periodo_inicio} - {self.periodo_fin}, {self.anio}   Página {page_num} de {total_pages}       {self.cliente_info['numero']} - {self.cliente_info['nombre']}"
    
    def _read_sheet_data(self, sheet_name: str) -> Tuple[List[str], List[List[Any]]]:
        """
        Lee datos de una hoja Excel.
        Usa datos de Datalab parser si están disponibles, sino openpyxl (fallback).
        
        Returns:
            Tuple de (headers, rows)
        """
        # 1. Usar datos de Datalab parser (recomendado - tiene valores de fórmulas)
        if self._datalab_reader:
            return self._read_from_datalab(sheet_name)
        
        # 2. Fallback a openpyxl (puede tener valores None para fórmulas)
        return self._read_from_openpyxl(sheet_name)
    
    def _read_from_datalab(self, sheet_name: str) -> Tuple[List[str], List[List[Any]]]:
        """Lee datos desde el parser Datalab."""
        # Mapeo de nombre de hoja a sección del parser
        section_map = {
            'Boletos': 'boletos',
            'Resultado Ventas ARS': 'ventas_ars',
            'Resultado Ventas USD': 'ventas_usd',
            'Rentas Dividendos ARS': 'rentas_dividendos_ars',
            'Rentas Dividendos USD': 'rentas_dividendos_usd',
            'Cauciones': 'cauciones',
            'Cauciones Tomadoras': 'cauciones',
            'Cauciones Colocadoras': 'cauciones',
        }
        
        section_key = section_map.get(sheet_name)
        if not section_key:
            # No tenemos datos Datalab para esta hoja, usar openpyxl
            print(f"[DEBUG] Hoja '{sheet_name}' no mapeada a Datalab, usando openpyxl")
            return self._read_from_openpyxl(sheet_name)
        
        # Obtener datos del parser
        parsed = self._datalab_reader._parsed_data
        if not parsed:
            return self._read_from_openpyxl(sheet_name)
        
        data = parsed.get(section_key, [])
        if not data:
            return [], []
        
        # Construir headers y rows desde los dicts
        headers = list(data[0].keys())
        rows = []
        for item in data:
            row = [item.get(h) for h in headers]
            rows.append(row)
        
        return headers, rows
    
    def _read_from_openpyxl(self, sheet_name: str) -> Tuple[List[str], List[List[Any]]]:
        """Lee datos desde openpyxl (fallback)."""
        if sheet_name not in self.wb.sheetnames:
            return [], []
        
        ws = self.wb[sheet_name]
        
        # Headers
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(1, col).value
            headers.append(str(val) if val else "")
        
        # Data rows
        rows = []
        for row_num in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row_num, col).value
                row_data.append(val)
            rows.append(row_data)
        
        return headers, rows
    
    def _create_table(self, headers: List[str], rows: List[List[Any]], 
                      col_widths: List[float] = None,
                      col_formatters: Dict[int, str] = None,
                      font_size: int = 6) -> Table:
        """
        Crea una tabla formateada.
        
        Args:
            headers: Lista de encabezados
            rows: Lista de filas de datos
            col_widths: Anchos de columnas en mm
            col_formatters: Diccionario {col_index: 'date'|'number'|'integer'|'text'}
            font_size: Tamaño de fuente para el cuerpo (default 6)
        """
        if not headers:
            return None
        
        col_formatters = col_formatters or {}
        
        # Formatear datos
        formatted_rows = [headers]
        for row in rows:
            formatted_row = []
            for i, val in enumerate(row):
                fmt_type = col_formatters.get(i, 'text')
                if fmt_type == 'date':
                    formatted_row.append(self._format_date(val))
                elif fmt_type == 'number':
                    formatted_row.append(self._format_number(val, 2))
                elif fmt_type == 'integer':
                    formatted_row.append(self._format_number(val, 0))
                elif fmt_type == 'text_truncate':
                    # Truncar texto largo a 25 caracteres
                    text = str(val) if val is not None else ""
                    formatted_row.append(text[:28] + ".." if len(text) > 30 else text)
                else:
                    formatted_row.append(str(val) if val is not None else "")
            formatted_rows.append(formatted_row)
        
        # Crear tabla
        if col_widths:
            table = Table(formatted_rows, colWidths=[w * mm for w in col_widths])
        else:
            table = Table(formatted_rows)
        
        # Estilo
        style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), self.HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), self.HEADER_TEXT),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 6),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            
            # Números alineados a la derecha
            *[(('ALIGN', (i, 1), (i, -1), 'RIGHT')) for i in range(len(headers)) 
              if col_formatters.get(i) in ('number', 'integer')],
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ])
        
        # Filas alternadas
        for i in range(1, len(formatted_rows)):
            if i % 2 == 0:
                style.add('BACKGROUND', (0, i), (-1, i), self.ROW_ALT_BG)
        
        table.setStyle(style)
        return table
    
    def _get_col_index(self, headers: List[str], col_name: str, alt_names: List[str] = None) -> int:
        """
        Obtiene el índice de una columna por nombre.
        Busca primero coincidencia exacta, luego substring.
        """
        all_names = [col_name] + (alt_names or [])
        
        # Primera pasada: buscar coincidencia exacta (case-insensitive)
        for name in all_names:
            for i, h in enumerate(headers):
                if h and h.lower().strip() == name.lower().strip():
                    return i
        
        # Segunda pasada: buscar substring (case-insensitive)
        for name in all_names:
            for i, h in enumerate(headers):
                if h and name.lower() in h.lower():
                    return i
        return -1
    
    def _get_value(self, row: List, headers: List[str], col_name: str, alt_names: List[str] = None) -> Any:
        """Obtiene un valor de una fila por nombre de columna."""
        idx = self._get_col_index(headers, col_name, alt_names)
        if idx >= 0 and idx < len(row):
            return row[idx]
        return None

    def _build_boletos_section(self) -> List:
        """Construye la sección de Boletos."""
        elements = []
        elements.append(Paragraph("Boletos", self.styles['SectionTitle']))
        
        headers, rows = self._read_sheet_data('Boletos')
        if not rows:
            elements.append(Paragraph("Sin operaciones en el período", self.styles['Normal']))
            elements.append(Spacer(1, 10*mm))
            return elements
        
        # Mapeo de columnas a mostrar: (nombre_excel, nombre_display, formato)
        col_map = [
            ('Concertación', 'Fecha', 'date'),
            ('Liquidación', 'Liquid.', 'date'),
            ('Nro. Boleto', 'Boleto', 'integer'),
            ('Moneda', 'Mon.', 'text'),
            ('Tipo Operación', 'Operación', 'text'),
            ('Cod.Instrum', 'Cód.', 'integer'),
            ('Instrumento Crudo', 'Instrumento', 'text_truncate'),
            ('Cantidad', 'Cantidad', 'number'),
            ('Precio Nominal', 'Precio Nom.', 'number'),
            ('Tipo Cambio', 'T.C.', 'number'),
            ('Bruto', 'Bruto', 'number'),
            ('Interés', 'Interés', 'number'),
            ('Gastos', 'Gastos', 'number'),
            ('Neto Calculado', 'Neto', 'number'),
        ]
        
        # Obtener índice de cada columna
        col_indices = []
        for col_name, display_name, fmt in col_map:
            idx = self._get_col_index(headers, col_name)
            col_indices.append((idx, display_name, fmt))
        
        # Índice de Tipo de Instrumento para agrupar
        tipo_idx = self._get_col_index(headers, 'Tipo de Instrumento')
        
        # Agrupar por tipo de instrumento
        by_tipo = {}
        for row in rows:
            tipo = row[tipo_idx] if tipo_idx >= 0 and tipo_idx < len(row) else "Otros"
            tipo = tipo if tipo else "Otros"
            if tipo not in by_tipo:
                by_tipo[tipo] = []
            by_tipo[tipo].append(row)
        
        # Ordenar tipos
        tipos_order = ['Acciones', 'Títulos Públicos', 'Obligaciones Negociables', 
                       'Letras del Tesoro', 'CEDEAR', 'Cedears', 'FCI', 'Otros']
        sorted_tipos = sorted(by_tipo.keys(), 
                            key=lambda x: next((i for i, t in enumerate(tipos_order) if t.lower() in x.lower()), 999))
        
        for tipo in sorted_tipos:
            tipo_rows = by_tipo[tipo]
            elements.append(Paragraph(tipo, self.styles['SubsectionTitle']))
            
            # Preparar datos para tabla
            table_headers = [c[1] for c in col_indices]
            table_rows = []
            col_formatters = {i: c[2] for i, c in enumerate(col_indices)}
            
            for row in tipo_rows:
                table_row = []
                for idx, _, _ in col_indices:
                    val = row[idx] if idx >= 0 and idx < len(row) else None
                    table_row.append(val)
                table_rows.append(table_row)
            
            # Anchos de columnas (total ~270mm para landscape A4)
            # Fecha, Liquid, Boleto, Mon, Operación, Cód, Instrumento, Cantidad, Precio, TC, Bruto, Interés, Gastos, Neto
            col_widths = [16, 16, 12, 16, 18, 12, 48, 22, 18, 14, 26, 18, 16, 26]
            
            table = self._create_table(table_headers, table_rows, col_widths, col_formatters, font_size=5)
            if table:
                elements.append(table)
            elements.append(Spacer(1, 3*mm))
        
        elements.append(Spacer(1, 10*mm))
        return elements
    
    def _build_resultado_ventas_section(self, moneda: str) -> List:
        """Construye la sección de Resultado Ventas (ARS o USD).
        
        Agrupa por Tipo de Instrumento (ej: Acciones, Obligaciones Negociables, etc.)
        y muestra una tabla con todas las operaciones del tipo, incluyendo
        Instrumento y Código en las primeras columnas.
        """
        elements = []
        sheet_name = f"Resultado Ventas {moneda}"
        
        elements.append(Paragraph(f"Resultado Ventas", self.styles['SectionTitle']))
        elements.append(Paragraph(moneda, self.styles['SubsectionTitle']))
        
        headers, rows = self._read_sheet_data(sheet_name)
        if not rows:
            elements.append(Paragraph("Sin operaciones en el período", self.styles['Normal']))
            elements.append(Spacer(1, 10*mm))
            return elements
        
        # Columnas a mostrar por moneda usando nombres
        if moneda == 'ARS':
            col_map = [
                ('Instrumento', 'Instrumento', 'text_truncate'),
                ('Cod.Instrum', 'Cód.', 'integer'),
                ('Concertación', 'Fecha', 'date'),
                ('Tipo Operación', 'Tipo Op.', 'text'),
                ('Cantidad', 'Cantidad', 'number'),
                ('Precio Nominal', 'Precio Nom.', 'number'),
                ('Bruto', 'Bruto', 'number'),
                ('Gastos', 'Gastos', 'number'),
                ('IVA', 'IVA', 'number'),
                ('Resultado Calculado(final)', 'Resultado', 'number'),
            ]
        else:  # USD
            col_map = [
                ('Instrumento', 'Instrumento', 'text_truncate'),
                ('Cod.Instrum', 'Cód.', 'integer'),
                ('Concertación', 'Fecha', 'date'),
                ('Tipo Operación', 'Tipo Op.', 'text'),
                ('Cantidad', 'Cantidad', 'number'),
                ('Precio Nominal', 'Precio Nom.', 'number'),
                ('Bruto en USD', 'Bruto USD', 'number'),
                ('Gastos', 'Gastos', 'number'),
                ('IVA', 'IVA', 'number'),
                ('Resultado Calculado(final)', 'Resultado', 'number'),
            ]
        
        # Obtener índices de columnas
        col_indices = []
        for col_name, display_name, fmt in col_map:
            idx = self._get_col_index(headers, col_name)
            col_indices.append((idx, display_name, fmt))
        
        # Índice para agrupar por tipo de instrumento
        tipo_idx = self._get_col_index(headers, 'Tipo de Instrumento')
        
        # Agrupar por tipo de instrumento solamente
        by_tipo = {}
        for row in rows:
            tipo = row[tipo_idx] if tipo_idx >= 0 and tipo_idx < len(row) else "Otros"
            tipo = tipo if tipo else "Otros"
            
            if tipo not in by_tipo:
                by_tipo[tipo] = []
            by_tipo[tipo].append(row)
        
        for tipo in sorted(by_tipo.keys()):
            # Solo mostrar el tipo de instrumento como encabezado (ej: "Acciones")
            elements.append(Paragraph(tipo, self.styles['SubsectionTitle']))
            
            tipo_rows = by_tipo[tipo]
            
            # Preparar tabla con todas las filas del tipo
            # La tabla incluye Instrumento y Código como primeras columnas
            table_headers = [c[1] for c in col_indices]
            table_rows = []
            col_formatters = {i: c[2] for i, c in enumerate(col_indices)}
            
            for row in tipo_rows:
                table_row = []
                for idx, _, _ in col_indices:
                    val = row[idx] if idx >= 0 and idx < len(row) else None
                    table_row.append(val)
                table_rows.append(table_row)
            
            # Anchos de columna ajustados
            # Instr, Cód, Fecha, TipoOp, Cantidad, Precio, Bruto, Gastos, IVA, Resultado
            col_widths = [42, 14, 16, 22, 22, 22, 26, 20, 18, 28]
            
            table = self._create_table(table_headers, table_rows, col_widths, col_formatters, font_size=5)
            if table:
                elements.append(table)
            elements.append(Spacer(1, 4*mm))
        
        elements.append(Spacer(1, 10*mm))
        return elements
    
    def _build_rentas_dividendos_section(self, moneda: str) -> List:
        """Construye la sección de Rentas y Dividendos (ARS o USD)."""
        elements = []
        sheet_name = f"Rentas Dividendos {moneda}"
        
        elements.append(Paragraph("Rentas y Dividendos", self.styles['SectionTitle']))
        elements.append(Paragraph(moneda, self.styles['SubsectionTitle']))
        
        headers, rows = self._read_sheet_data(sheet_name)
        if not rows:
            elements.append(Paragraph("Sin operaciones en el período", self.styles['Normal']))
            elements.append(Spacer(1, 10*mm))
            return elements
        
        # Columnas a mostrar por nombre
        col_map = [
            ('Concertación', 'Concertación', 'date'),
            ('Liquidación', 'Liquidación', 'date'),
            ('Nro. NDC', 'Nro. NDC', 'integer'),
            ('Tipo Operación', 'Tipo Operación', 'text'),
            ('Cantidad', 'Cantidad', 'number'),
            ('Moneda', 'Moneda', 'text'),
            ('Tipo de Cambio', 'T.C.', 'number'),
            ('Gastos', 'Gastos', 'number'),
            ('Importe', 'Importe', 'number'),
        ]
        
        # Obtener índices de columnas
        col_indices = []
        for col_name, display_name, fmt in col_map:
            idx = self._get_col_index(headers, col_name)
            col_indices.append((idx, display_name, fmt))
        
        # Índices para agrupar
        cat_idx = self._get_col_index(headers, 'Categoría')
        tipo_instr_idx = self._get_col_index(headers, 'tipo_instrumento', ['Tipo de Instrumento'])
        instr_idx = self._get_col_index(headers, 'Instrumento')
        
        # Agrupar por categoría y tipo_instrumento
        by_cat = {}
        for row in rows:
            cat = row[cat_idx] if cat_idx >= 0 and cat_idx < len(row) else "Otros"
            cat = cat if cat else "Otros"
            tipo_instr = row[tipo_instr_idx] if tipo_instr_idx >= 0 and tipo_instr_idx < len(row) else "Sin tipo"
            tipo_instr = tipo_instr if tipo_instr else "Sin tipo"
            instr = row[instr_idx] if instr_idx >= 0 and instr_idx < len(row) else "Sin nombre"
            instr = instr if instr else "Sin nombre"
            
            if cat not in by_cat:
                by_cat[cat] = {}
            if tipo_instr not in by_cat[cat]:
                by_cat[cat][tipo_instr] = {}
            if instr not in by_cat[cat][tipo_instr]:
                by_cat[cat][tipo_instr][instr] = []
            by_cat[cat][tipo_instr][instr].append(row)
        
        # Ordenar: Rentas primero, luego Dividendos
        cat_order = ['Rentas', 'Dividendos', 'Otros']
        sorted_cats = sorted(by_cat.keys(), 
                           key=lambda x: cat_order.index(x) if x in cat_order else 999)
        
        for cat in sorted_cats:
            elements.append(Paragraph(cat, self.styles['SubsectionTitle']))
            
            for tipo_instr in sorted(by_cat[cat].keys()):
                elements.append(Paragraph(f"  {tipo_instr}", self.styles['Normal']))
                
                for instr in sorted(by_cat[cat][tipo_instr].keys()):
                    instr_rows = by_cat[cat][tipo_instr][instr]
                    
                    # Nombre del instrumento
                    elements.append(Paragraph(f"    {instr}", 
                                            ParagraphStyle('InstrName', 
                                                          parent=self.styles['Normal'],
                                                          fontSize=8,
                                                          textColor=colors.grey)))
                    
                    table_headers = [c[1] for c in col_indices]
                    table_rows = []
                    col_formatters = {i: c[2] for i, c in enumerate(col_indices)}
                    
                    for row in instr_rows:
                        table_row = []
                        for idx, _, _ in col_indices:
                            val = row[idx] if idx >= 0 and idx < len(row) else None
                            table_row.append(val)
                        table_rows.append(table_row)
                    
                    col_widths = [18, 18, 15, 28, 18, 22, 18, 18, 25]
                    
                    table = self._create_table(table_headers, table_rows, col_widths, col_formatters)
                    if table:
                        elements.append(table)
                    elements.append(Spacer(1, 2*mm))
        
        elements.append(Spacer(1, 10*mm))
        return elements
    
    def _build_cauciones_section(self, tipo: str = "tomadoras") -> List:
        """
        Construye la sección de Cauciones.
        
        Args:
            tipo: 'tomadoras' o 'colocadoras'
        """
        elements = []
        titulo = f"Cauciones {tipo.capitalize()}"
        elements.append(Paragraph(titulo, self.styles['SectionTitle']))
        
        # Buscar hoja específica según tipo (nuevo formato con hojas separadas)
        sheet_name = f"Cauciones {tipo.capitalize()}"
        headers, rows = self._read_sheet_data(sheet_name)
        
        # Fallback a hoja única "Cauciones" si no existe la hoja específica
        if not rows:
            headers, rows = self._read_sheet_data('Cauciones')
            
            if rows:
                # Filtrar por tipo de operación usando nombre de columna
                tipo_op_idx = self._get_col_index(headers, 'Operación', ['Tipo Operación'])
                filtered_rows = []
                for row in rows:
                    operacion = str(row[tipo_op_idx]).upper() if tipo_op_idx >= 0 and tipo_op_idx < len(row) and row[tipo_op_idx] else ""
                    if tipo == "tomadoras" and "TOM" in operacion:
                        filtered_rows.append(row)
                    elif tipo == "colocadoras" and "COL" in operacion:
                        filtered_rows.append(row)
                rows = filtered_rows
        
        if not rows:
            elements.append(Paragraph("Sin operaciones en el período", self.styles['Normal']))
            elements.append(Spacer(1, 10*mm))
            return elements
        
        # Columnas a mostrar por nombre
        col_map = [
            ('Concertación', 'Concertación', 'date'),
            ('Plazo', 'Plazo', 'integer'),
            ('Liquidación', 'Liquidación', 'date'),
            ('Operación', 'Operación', 'text'),
            ('# Boleto', '# Boleto', 'integer'),
            ('Contado', 'Contado', 'number'),
            ('Futuro', 'Futuro', 'number'),
            ('Tipo de cambio', 'T.C.', 'number'),
            ('Tasa (%)', 'Tasa (%)', 'number'),
            ('Interés Bruto', 'Int. Bruto', 'number'),
            ('Interés Devengad', 'Int. Dev.', 'number'),
            ('Aranceles', 'Aranceles', 'number'),
            ('Derechos', 'Derechos', 'number'),
            ('Costo financiero', 'Costo Fin.', 'number'),
        ]
        
        # Obtener índices de columnas
        col_indices = []
        for col_name, display_name, fmt in col_map:
            idx = self._get_col_index(headers, col_name)
            col_indices.append((idx, display_name, fmt))
        
        # Índice de moneda para agrupar
        moneda_idx = self._get_col_index(headers, 'Moneda')
        # Índice de costo financiero para totales
        costo_idx = self._get_col_index(headers, 'Costo financiero')
        
        # Agrupar por moneda
        by_moneda = {}
        for row in rows:
            moneda = row[moneda_idx] if moneda_idx >= 0 and moneda_idx < len(row) and row[moneda_idx] else "Pesos"
            if moneda not in by_moneda:
                by_moneda[moneda] = []
            by_moneda[moneda].append(row)
        
        for moneda in ['Pesos', 'Dólares', 'Dolar MEP', 'Dolar Cable']:
            if moneda not in by_moneda:
                continue
            
            elements.append(Paragraph(f"  {moneda}", self.styles['SubsectionTitle']))
            
            table_headers = [c[1] for c in col_indices]
            table_rows = []
            col_formatters = {i: c[2] for i, c in enumerate(col_indices)}
            
            for row in by_moneda[moneda]:
                table_row = []
                for idx, _, _ in col_indices:
                    val = row[idx] if idx >= 0 and idx < len(row) else None
                    table_row.append(val)
                table_rows.append(table_row)
            
            col_widths = [18, 10, 18, 25, 15, 22, 22, 16, 14, 18, 18, 15, 14, 18]
            
            table = self._create_table(table_headers, table_rows, col_widths, col_formatters)
            if table:
                elements.append(table)
            
            # Total usando índice de columna
            total_cf = 0.0
            for r in by_moneda[moneda]:
                val = r[costo_idx] if costo_idx >= 0 and costo_idx < len(r) else 0
                try:
                    total_cf += float(val or 0)
                except (ValueError, TypeError):
                    pass
            elements.append(Paragraph(f"Totales: {self._format_number(total_cf)}", 
                                     self.styles['Normal']))
            elements.append(Spacer(1, 3*mm))
        
        elements.append(Spacer(1, 10*mm))
        return elements
    
    def _build_resumen_section(self) -> List:
        """Construye la sección de Resumen.
        
        Calcula los totales directamente de las hojas de datos,
        ya que las fórmulas de Excel no se evalúan al guardar con openpyxl.
        Si tenemos valores de Datalab o COM, los usamos directamente.
        """
        elements = []
        elements.append(Paragraph("Resumen", self.styles['SectionTitle']))
        
        # Usar valores de Datalab si están disponibles (para Streamlit Cloud)
        datalab_resumen = None
        if self._datalab_reader and self._datalab_reader._parsed_data:
            datalab_resumen = self._datalab_reader.get_resumen()
        
        if datalab_resumen:
            ventas_ars = datalab_resumen.get('ventas_ars', 0.0)
            ventas_usd = datalab_resumen.get('ventas_usd', 0.0)
            print(f"[INFO] Usando valores Datalab: ARS={ventas_ars:,.2f}, USD={ventas_usd:,.2f}")
        else:
            # Calcular totales directamente de las hojas de datos
            ventas_ars = self._calculate_ventas_total('Resultado Ventas ARS')
            ventas_usd = self._calculate_ventas_total('Resultado Ventas USD')
        
        rentas_ars = self._calculate_rentas_dividendos('Rentas Dividendos ARS', ['Rentas', 'AMORTIZACION'])
        dividendos_ars = self._calculate_rentas_dividendos('Rentas Dividendos ARS', ['Dividendos'])
        
        rentas_usd = self._calculate_rentas_dividendos('Rentas Dividendos USD', ['Rentas', 'AMORTIZACION'])
        dividendos_usd = self._calculate_rentas_dividendos('Rentas Dividendos USD', ['Dividendos'])
        
        # Cau(Int) = suma de Interés Devengado (col K=11) de ambas hojas de cauciones
        cau_int_ars = (self._calculate_cauciones('Cauciones Tomadoras', 'ARS', 'interes') +
                      self._calculate_cauciones('Cauciones Colocadoras', 'ARS', 'interes'))
        # Cau(CF) = suma de Costo Financiero (col N=14) de ambas hojas de cauciones
        cau_cf_ars = (self._calculate_cauciones('Cauciones Tomadoras', 'ARS', 'costo') +
                     self._calculate_cauciones('Cauciones Colocadoras', 'ARS', 'costo'))
        
        # Cau(Int) = suma de Interés Devengado (col K=11) de ambas hojas de cauciones
        cau_int_usd = (self._calculate_cauciones('Cauciones Tomadoras', 'USD', 'interes') +
                      self._calculate_cauciones('Cauciones Colocadoras', 'USD', 'interes'))
        # Cau(CF) = suma de Costo Financiero (col N=14) de ambas hojas de cauciones
        cau_cf_usd = (self._calculate_cauciones('Cauciones Tomadoras', 'USD', 'costo') +
                     self._calculate_cauciones('Cauciones Colocadoras', 'USD', 'costo'))
        
        total_ars = ventas_ars + rentas_ars + dividendos_ars + cau_int_ars + cau_cf_ars
        total_usd = ventas_usd + rentas_usd + dividendos_usd + cau_int_usd + cau_cf_usd
        
        # Headers
        table_headers = ['Moneda', 'Resultados', '', '', '', '', '', '', '', '', '', 'Total']
        sub_headers = ['', 'Ventas', 'FCI', 'Opciones', 'Rentas', 'Dividendos', 
                      'Ef. CPD', 'Pagarés', 'Futuros', 'Cau (int)', 'Cau (CF)', '']
        
        table_data = [table_headers, sub_headers]
        
        # Fila ARS
        table_data.append([
            'ARS',
            self._format_number(ventas_ars),
            self._format_number(0),  # FCI
            self._format_number(0),  # Opciones
            self._format_number(rentas_ars),
            self._format_number(dividendos_ars),
            self._format_number(0),  # Ef. CPD
            self._format_number(0),  # Pagarés
            self._format_number(0),  # Futuros
            self._format_number(cau_int_ars),
            self._format_number(cau_cf_ars),
            self._format_number(total_ars),
        ])
        
        # Fila USD
        table_data.append([
            'USD',
            self._format_number(ventas_usd),
            self._format_number(0),
            self._format_number(0),
            self._format_number(rentas_usd),
            self._format_number(dividendos_usd),
            self._format_number(0),
            self._format_number(0),
            self._format_number(0),
            self._format_number(cau_int_usd),
            self._format_number(cau_cf_usd),
            self._format_number(total_usd),
        ])
        
        col_widths = [20, 28, 20, 20, 22, 24, 20, 20, 20, 22, 22, 32]
        
        table = Table(table_data, colWidths=[w * mm for w in col_widths])
        
        style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 1), self.HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 1), self.HEADER_TEXT),
            ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 1), 8),
            ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
            
            # Merge "Resultados" header
            ('SPAN', (1, 0), (10, 0)),
            
            # Body
            ('FONTNAME', (0, 2), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 2), (-1, -1), 8),
            ('ALIGN', (1, 2), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 2), (0, -1), 'LEFT'),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 1), (-1, 1), 1, colors.black),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ])
        
        table.setStyle(style)
        elements.append(table)
        elements.append(Spacer(1, 15*mm))
        
        return elements
    
    def _calculate_ventas_total(self, sheet_name: str) -> float:
        """Calcula el resultado total de ventas sumando la columna Resultado Calculado(final).
        
        ARS: Col U (21) = Resultado Calculado(final)
        USD: Col X (24) = Resultado Calculado(final)
        """
        if sheet_name not in self.wb.sheetnames:
            return 0
        
        ws = self.wb[sheet_name]
        
        # Determinar columna de resultado según el tipo
        # ARS: Col U (21) = Resultado Calculado(final)
        # USD: Col X (24) = Resultado Calculado(final)
        resultado_col = 21 if 'ARS' in sheet_name else 24
        
        total = 0
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, resultado_col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        
        return total
    
    def _calculate_rentas_dividendos(self, sheet_name: str, tipos: List[str]) -> float:
        """Calcula el total de rentas o dividendos de una hoja."""
        if sheet_name not in self.wb.sheetnames:
            return 0
        
        ws = self.wb[sheet_name]
        total = 0
        
        # Columnas: C=Tipo(3), M=Importe Neto(13)
        for row in range(2, ws.max_row + 1):
            tipo = str(ws.cell(row, 3).value or '').upper()
            if any(t.upper() in tipo for t in tipos):
                importe = ws.cell(row, 13).value
                if importe and isinstance(importe, (int, float)):
                    total += importe
        
        return total
    
    def _calculate_cauciones(self, sheet_name: str, moneda: str, campo: str) -> float:
        """Calcula el total de cauciones (interés devengado o costo financiero)."""
        if sheet_name not in self.wb.sheetnames:
            return 0
        
        ws = self.wb[sheet_name]
        total = 0
        
        # Columnas: 11=Interés Devengado (K), 14=Costo Financiero (N), 15=Moneda
        col = 11 if campo == 'interes' else 14
        
        for row in range(2, ws.max_row + 1):
            moneda_val = str(ws.cell(row, 15).value or '').upper()
            if moneda == 'ARS' and 'PESO' in moneda_val:
                val = ws.cell(row, col).value
                if val and isinstance(val, (int, float)):
                    total += val
            elif moneda == 'USD' and ('DOLAR' in moneda_val or 'USD' in moneda_val):
                val = ws.cell(row, col).value
                if val and isinstance(val, (int, float)):
                    total += val
        
        return total
    
    def _build_posicion_titulos_section(self) -> List:
        """Construye la sección de Posición de Títulos."""
        elements = []
        elements.append(Paragraph("Posición de Títulos", self.styles['SectionTitle']))
        
        headers, rows = self._read_sheet_data('Posicion Titulos')
        if not rows:
            elements.append(Paragraph("Sin posiciones", self.styles['Normal']))
            return elements
        
        # Agregar subtítulo "Es Disponible Sí"
        elements.append(Paragraph("Es Disponible Sí", self.styles['Normal']))
        elements.append(Spacer(1, 3*mm))
        
        # Headers: Instrumento(0), Código(1), Ticker(2), Cantidad(3), Importe(4), Moneda(5)
        
        col_map = [
            (0, 'Instrumento', 'text'),
            (1, 'Código', 'integer'),
            (2, 'Ticker', 'text'),
            (3, 'Cantidad', 'number'),
            (4, 'Importe', 'number'),
            (5, 'Moneda', 'text'),
        ]
        
        table_headers = [c[1] for c in col_map]
        table_rows = []
        col_formatters = {i: c[2] for i, c in enumerate(col_map)}
        
        for row in rows:
            table_row = [row[c[0]] if c[0] < len(row) else None for c in col_map]
            table_rows.append(table_row)
        
        col_widths = [90, 20, 20, 30, 35, 25]
        
        table = self._create_table(table_headers, table_rows, col_widths, col_formatters)
        if table:
            elements.append(table)
        
        return elements
    
    def export_to_pdf(self, output_path: str = None) -> bytes:
        """
        Exporta el Excel a PDF.
        
        Args:
            output_path: Ruta opcional para guardar el archivo
            
        Returns:
            bytes del PDF generado
        """
        # Buffer para el PDF
        buffer = io.BytesIO()
        
        # Crear documento
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=10*mm,
            leftMargin=10*mm,
            topMargin=20*mm,
            bottomMargin=15*mm
        )
        
        # Construir contenido
        elements = []
        
        # Header principal (se repetirá en cada página via onPage)
        
        # Boletos
        elements.extend(self._build_boletos_section())
        elements.append(PageBreak())
        
        # Resultado Ventas ARS
        elements.extend(self._build_resultado_ventas_section('ARS'))
        elements.append(PageBreak())
        
        # Resultado Ventas USD
        elements.extend(self._build_resultado_ventas_section('USD'))
        elements.append(PageBreak())
        
        # Rentas y Dividendos ARS
        elements.extend(self._build_rentas_dividendos_section('ARS'))
        elements.append(PageBreak())
        
        # Rentas y Dividendos USD
        elements.extend(self._build_rentas_dividendos_section('USD'))
        elements.append(PageBreak())
        
        # Cauciones Tomadoras
        elements.extend(self._build_cauciones_section('tomadoras'))
        
        # Cauciones Colocadoras
        elements.extend(self._build_cauciones_section('colocadoras'))
        elements.append(PageBreak())
        
        # Resumen
        elements.extend(self._build_resumen_section())
        
        # Posición de Títulos
        elements.extend(self._build_posicion_titulos_section())
        
        # Generar PDF
        def add_header_footer(canvas, doc):
            canvas.saveState()
            page_width = landscape(A4)[0]
            page_height = landscape(A4)[1]
            
            # Fondo del header
            canvas.setFillColor(colors.Color(0.1, 0.2, 0.4))  # Azul oscuro
            canvas.rect(0, page_height - 18*mm, page_width, 18*mm, fill=True, stroke=False)
            
            # Logo (si existe)
            logo_path = Path(__file__).parent.parent / 'assets' / 'logo.png'
            if logo_path.exists():
                try:
                    canvas.drawImage(
                        str(logo_path), 
                        8*mm, 
                        page_height - 15*mm, 
                        width=12*mm, 
                        height=12*mm,
                        preserveAspectRatio=True,
                        mask='auto'
                    )
                    text_x = 22*mm
                except:
                    text_x = 10*mm
            else:
                text_x = 10*mm
            
            # Título del reporte
            canvas.setFillColor(colors.white)
            canvas.setFont('Helvetica-Bold', 10)
            canvas.drawString(text_x, page_height - 8*mm, "REPORTE DE GANANCIAS")
            
            # Info del período y cliente
            canvas.setFont('Helvetica', 8)
            info_text = f"Período: {self.periodo_inicio} - {self.periodo_fin}, {self.anio}   |   {self.cliente_info['numero']} - {self.cliente_info['nombre']}"
            canvas.drawString(text_x, page_height - 14*mm, info_text)
            
            # Page number (derecha)
            canvas.setFont('Helvetica-Bold', 9)
            page_num = f"Pág. {doc.page}"
            canvas.drawRightString(page_width - 10*mm, page_height - 11*mm, page_num)
            
            # Línea decorativa debajo del header
            canvas.setStrokeColor(colors.Color(0.9, 0.7, 0.1))  # Dorado/amarillo
            canvas.setLineWidth(2)
            canvas.line(0, page_height - 18*mm, page_width, page_height - 18*mm)
            
            # Footer
            canvas.setFillColor(colors.Color(0.5, 0.5, 0.5))
            canvas.setFont('Helvetica', 6)
            canvas.drawCentredString(page_width/2, 8*mm, f"Generado automáticamente - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            
            canvas.restoreState()
        
        doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
        
        # Obtener bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        # Guardar si se especificó ruta
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(pdf_bytes)
        
        return pdf_bytes


def export_excel_to_pdf(excel_path: str, output_path: str = None,
                        cliente_numero: str = None, cliente_nombre: str = None,
                        periodo_inicio: str = None, periodo_fin: str = None,
                        anio: int = None) -> bytes:
    """
    Función de conveniencia para exportar un Excel consolidado a PDF.
    
    Args:
        excel_path: Ruta al archivo Excel consolidado
        output_path: Ruta opcional para guardar el PDF
        cliente_numero: Número de comitente
        cliente_nombre: Nombre del cliente
        periodo_inicio: Fecha inicio del período (ej: "Junio 1")
        periodo_fin: Fecha fin del período (ej: "Diciembre 12")
        anio: Año del reporte
        
    Returns:
        bytes del PDF generado
    """
    cliente_info = {}
    if cliente_numero:
        cliente_info['numero'] = cliente_numero
    if cliente_nombre:
        cliente_info['nombre'] = cliente_nombre
    
    exporter = ExcelToPdfExporter(excel_path, cliente_info or None)
    
    if periodo_inicio:
        exporter.periodo_inicio = periodo_inicio
    if periodo_fin:
        exporter.periodo_fin = periodo_fin
    if anio:
        exporter.anio = anio
    
    return exporter.export_to_pdf(output_path)


if __name__ == "__main__":
    # Test
    import sys
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = "TEST_Merge_v8.xlsx"
    
    output_path = excel_path.replace('.xlsx', '.pdf')
    
    pdf_bytes = export_excel_to_pdf(
        excel_path,
        output_path,
        cliente_numero="12345",
        cliente_nombre="TEST CLIENT",
        periodo_inicio="Junio 1",
        periodo_fin="Diciembre 12",
        anio=2025
    )
    
    print(f"PDF generado: {output_path} ({len(pdf_bytes)} bytes)")
