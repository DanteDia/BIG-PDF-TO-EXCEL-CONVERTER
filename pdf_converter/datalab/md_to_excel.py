#!/usr/bin/env python
"""
Markdown to Excel converter for Datalab-extracted financial reports.
Parses the Markdown tables and creates a structured Excel file.
"""

import re
import sys
from pathlib import Path
from typing import Optional, Tuple
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rich.console import Console

console = Console()


def extract_comitente_info(markdown_content: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Extract comitente number and name from markdown content.
    
    Returns:
        Tuple of (comitente_number, comitente_name) or (None, None) if not found
    
    Examples:
        - Visual format: "12128 - LANDRO, VERONICA INES"
        - Gallo format: "Comitente: 12128 LANDRO, VERONICA INES"
    """
    # Try Visual format first: "12128 - LANDRO, VERONICA INES"
    # Use [^\n] instead of \s to avoid capturing newlines
    visual_pattern = r'^\*?\*?(\d{4,6})\s*-\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ ,\.]+)'
    match = re.search(visual_pattern, markdown_content, re.MULTILINE)
    if match:
        # Clean name - remove newlines and extra spaces
        name = re.sub(r'\s+', ' ', match.group(2)).strip()
        return match.group(1), name
    
    # Try Gallo format: "Comitente: 12128 LANDRO, VERONICA INES"
    gallo_pattern = r'Comitente:\s*(\d{4,6})\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ ,\.]+)'
    match = re.search(gallo_pattern, markdown_content, re.MULTILINE | re.IGNORECASE)
    if match:
        # Clean name - remove newlines and extra spaces
        name = re.sub(r'\s+', ' ', match.group(2)).strip()
        return match.group(1), name
    
    return None, None


@dataclass
class TableData:
    """Represents an extracted table."""
    section: str
    headers: list[str]
    rows: list[list[str]]
    metadata: dict = None  # Optional metadata (e.g., fecha for Posicion sheets)
    
    def __post_init__(self):
        if self.metadata is None:
            self.metadata = {}


class MarkdownTableParser:
    """Parse markdown tables into structured data."""
    
    # Section mappings for sheet names - Gallo format
    GALLO_SECTIONS = {
        "RESULTADOS TOTALES": "Resultado Totales",
        "TIT.PRIVADOS EXENTOS": "Tit.Privados Exentos",
        "TIT.PRIVADOS DEL EXTERIOR": "Tit.Privados Exterior",
        "RENTA FIJA EN PESOS": "Renta Fija Pesos",
        "RENTA FIJA EN DOLARES": "Renta Fija Dolares",
        "CAUCIONES EN PESOS": "Cauciones Pesos",
        "CAUCIONES EN DOLARES": "Cauciones Dolares",
        "POSICION AL 01/01": "Posicion Inicial",
        "POSICION AL 31/12": "Posicion Final",
    }
    
    # Section mappings for sheet names - Visual format
    VISUAL_SECTIONS = {
        "BOLETOS": "Boletos",
        "RESULTADO VENTAS": "Resultado Ventas",
        "RENTAS Y DIVIDENDOS": "Rentas Dividendos",
        "CAUCIONES TOMADORAS": "Cauciones",
        "CAUCIONES": "Cauciones",
        "RESUMEN": "Resumen",
        "POSICIÓN DE TÍTULOS": "Posicion Titulos",
        "POSICION DE TITULOS": "Posicion Titulos",
    }
    
    def __init__(self, markdown_content: str):
        self.content = markdown_content
        self.tables: dict[str, TableData] = {}
        self.format_type = self._detect_format()
    
    def _detect_format(self) -> str:
        """Detect if this is Gallo or Visual format."""
        content_upper = self.content.upper()
        if "BOLETOS" in content_upper or "RESULTADO VENTAS" in content_upper:
            return "visual"
        return "gallo"
    
    def parse(self) -> dict[str, TableData]:
        """Parse all tables from the markdown content."""
        if self.format_type == "visual":
            return self._parse_visual()
        return self._parse_gallo()
    
    def _parse_gallo(self) -> dict[str, TableData]:
        """Parse Gallo format tables."""
        lines = self.content.split('\n')
        current_section = None
        current_headers = None
        current_rows = []
        current_metadata = {}  # Store metadata like fecha
        in_table = False
        skip_section = False  # Flag to skip INCREMENTOS/DECREMENTOS section
        posicion_count = 0  # Track order of POSICION AL sections (1st=Inicial, 2nd=Final)
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            # Skip INCREMENTOS/DECREMENTOS section entirely
            if 'INCREMENTOS' in line.upper() or 'DECREMENTOS' in line.upper():
                skip_section = True
                continue
            
            # Detect section headers (## or ### followed by section name)
            if line.startswith('#'):
                section_name = line.lstrip('#').strip()
                
                # Check if this is INCREMENTOS/DECREMENTOS to skip
                if 'INCREMENTOS' in section_name.upper() or 'DECREMENTOS' in section_name.upper():
                    skip_section = True
                    continue
                
                # Check if this is a POSICION AL section and increment counter
                # Extract fecha FIRST before anything else
                temp_metadata = {}
                if 'POSICION AL' in section_name.upper():
                    posicion_count += 1
                    # Extract fecha from "POSICION AL DD/MM/YY"
                    fecha_match = re.search(r'POSICION AL (\d{2}/\d{2}/\d{2,4})', section_name.upper())
                    if fecha_match:
                        temp_metadata = {'fecha': fecha_match.group(1)}
                
                matched_section = self._match_section(section_name, self.GALLO_SECTIONS, posicion_count)
                
                if matched_section:
                    # Exit skip mode when we find a valid section
                    skip_section = False
                    # Save previous section data with OLD metadata
                    if current_section and current_headers:
                        self._save_table(current_section, current_headers, current_rows, current_metadata)
                    
                    current_section = matched_section
                    current_headers = None
                    current_rows = []
                    # Set NEW metadata for this section
                    current_metadata = temp_metadata
                    in_table = False
                continue
            
            # Skip lines while in skip mode
            if skip_section:
                continue
            
            # Detect table start (line starting with |)
            if line.startswith('|') and current_section:
                cells = self._parse_table_row(line)
                
                if not in_table:
                    # This is potentially a header row
                    # Check if next line is separator (|---|)
                    if i + 1 < len(lines) and '---' in lines[i + 1]:
                        current_headers = cells
                        in_table = True
                    continue
                
                # Skip separator rows
                if '---' in line:
                    continue
                
                # This is a data row
                if cells and any(c.strip() for c in cells):
                    # Skip rows that are just periods or empty
                    if not all(c.strip() in ['', '.'] for c in cells):
                        current_rows.append(cells)
        
        # Save last section
        if current_section and current_headers:
            self._save_table(current_section, current_headers, current_rows, current_metadata)
        
        return self.tables
    
    def _extract_instrument_info(self, text: str) -> tuple[Optional[str], Optional[str]]:
        """
        Extract instrument name and code from a bold line like:
        '<b>PAMPA ENERGIA S.A. ESCRIT. 1 VOTO - Pesos / 457</b>'
        Returns: (instrument_name, cod_instrum)
        """
        import re
        # Remove bold tags
        text = re.sub(r'</?b>', '', text).strip()
        
        # Try to match pattern: "INSTRUMENT NAME - Currency / CODE"
        match = re.match(r'^(.+?)\s*-\s*[^/]+/\s*(\d+(?:\.\d+)?)\s*$', text)
        if match:
            return (match.group(1).strip(), match.group(2).replace('.', ''))
        
        return (None, None)
    
    def _extract_tipo_instrumento(self, text: str) -> Optional[str]:
        """
        Extract tipo de instrumento from lines like:
        '<b>1 / Acciones</b>' or '<b>Acciones</b>'
        """
        import re
        # Remove bold tags
        text = re.sub(r'</?b>', '', text).strip()
        
        # Pattern: "NUMBER / TYPE" or just "TYPE"
        match = re.match(r'^(?:\d+\s*/\s*)?(.+)$', text)
        if match:
            return match.group(1).strip()
        return text
    
    def _parse_visual(self) -> dict[str, TableData]:
        """Parse Visual format tables - splits by currency (ARS/USD) for some sections."""
        import re
        lines = self.content.split('\n')
        current_section = None
        current_currency = None
        current_headers = None
        current_rows = []
        in_table = False
        
        # Context tracking for metadata columns
        current_tipo_instrumento = None
        current_instrumento = None
        current_cod_instrum = None
        current_categoria = None  # For Rentas/Dividendos
        
        # Sections that should NOT be split by currency
        no_currency_sections = {"Resumen", "Posicion Titulos", "Boletos"}
        
        # Sections that need extra columns
        boletos_section = "Boletos"
        resultado_ventas_section = "Resultado Ventas"
        rentas_dividendos_section = "Rentas Dividendos"
        cauciones_section = "Cauciones"
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            # Detect section headers
            if line.startswith('#'):
                section_name = line.lstrip('#').strip()
                matched_section = self._match_section(section_name, self.VISUAL_SECTIONS)
                
                if matched_section:
                    # Save previous section data
                    if current_section and current_headers and current_rows:
                        sheet_name = current_section
                        if current_currency and current_section not in no_currency_sections:
                            sheet_name = f"{current_section} {current_currency}"
                        self._save_table(sheet_name, current_headers, current_rows)
                    
                    current_section = matched_section
                    current_currency = None
                    current_headers = None
                    current_rows = []
                    in_table = False
                    # Reset metadata
                    current_tipo_instrumento = None
                    current_instrumento = None
                    current_cod_instrum = None
                    current_categoria = None
                continue
            
            # Detect currency markers in Visual format (only for sections that use currency)
            if current_section and current_section not in no_currency_sections:
                # Standard currency markers
                if '<b>ARS</b>' in line or line == 'ARS':
                    if current_section and current_headers and current_rows:
                        sheet_name = f"{current_section} {current_currency}" if current_currency else current_section
                        self._save_table(sheet_name, current_headers, current_rows)
                        current_rows = []
                    current_currency = "ARS"
                    current_tipo_instrumento = None
                    current_instrumento = None
                    current_cod_instrum = None
                    current_categoria = None
                    continue
                elif '<b>USD</b>' in line or line == 'USD':
                    if current_section and current_headers and current_rows:
                        sheet_name = f"{current_section} {current_currency}" if current_currency else current_section
                        self._save_table(sheet_name, current_headers, current_rows)
                        current_rows = []
                    current_currency = "USD"
                    current_tipo_instrumento = None
                    current_instrumento = None
                    current_cod_instrum = None
                    current_categoria = None
                    continue
                
                # Cauciones-specific currency markers: "<b>1 / Pesos</b>" = ARS, "<b>2 / Dólares</b>" = USD
                if current_section == cauciones_section:
                    if '<b>1 / Pesos</b>' in line or '<b>1 /Pesos</b>' in line:
                        if not current_currency:  # Only set if not already set
                            current_currency = "ARS"
                        continue
                    elif '<b>2 / D' in line and ('lares</b>' in line or 'LARES</b>' in line):
                        if current_headers and current_rows:
                            sheet_name = f"{current_section} {current_currency}" if current_currency else current_section
                            self._save_table(sheet_name, current_headers, current_rows)
                            current_rows = []
                        current_currency = "USD"
                        continue
            
            # Detect table start
            if line.startswith('|') and current_section:
                cells = self._parse_table_row(line)
                
                if not in_table:
                    if i + 1 < len(lines) and '---' in lines[i + 1]:
                        # Modify headers based on section
                        if current_section == boletos_section:
                            # Add Tipo de Instrumento as first column
                            current_headers = ['Tipo de Instrumento'] + cells
                        elif current_section == resultado_ventas_section:
                            # Add Tipo de Instrumento, Instrumento, Cod.Instrum at the start
                            current_headers = ['Tipo de Instrumento', 'Instrumento', 'Cod.Instrum'] + cells
                        elif current_section == rentas_dividendos_section:
                            # Add Instrumento, Cod.Instrum, Categoría, tipo_instrumento at the start
                            current_headers = ['Instrumento', 'Cod.Instrum', 'Categoría', 'tipo_instrumento'] + cells
                        elif current_section == "Resumen":
                            # For Resumen, use hardcoded headers (OCR often splits them incorrectly)
                            current_headers = ['Moneda', 'Ventas', 'FCI', 'Opciones', 'Rentas', 'Dividendos Ef.', 
                                             'CPD', 'Pagarés', 'Futuros', 'Cau (int)', 'Cau (CF)', 'Total']
                        else:
                            current_headers = cells
                        in_table = True
                    continue
                
                if '---' in line:
                    continue
                
                # Special handling for Resumen section
                if current_section == "Resumen":
                    # Skip the sub-header row (Ventas, FCI, etc.) - we use hardcoded headers
                    if not cells[0].strip() and any('Ventas' in c or 'FCI' in c for c in cells):
                        continue
                    # Data rows (ARS/USD)
                    if cells[0].strip() in ('ARS', 'USD'):
                        # Get all non-empty values
                        data_cells = [cells[0].strip()]
                        for c in cells[1:]:
                            if c.strip():
                                data_cells.append(c.strip())
                        # Ensure row has same length as headers
                        while len(data_cells) < len(current_headers):
                            data_cells.append('')
                        data_cells = data_cells[:len(current_headers)]
                        current_rows.append(data_cells)
                    continue
                
                # Process rows based on content
                if cells and any(c.strip() for c in cells):
                    first_cell = cells[0] if cells else ""
                    
                    # Check if this is a category/type row (bold text only in first cell, rest empty)
                    is_category_row = first_cell.startswith('<b>') and all(
                        not c.strip() or c.startswith('<b>') for c in cells[1:5]
                    )
                    
                    if is_category_row:
                        clean_text = re.sub(r'</?b>', '', first_cell).strip()
                        
                        # For Boletos section: detect tipo instrumento
                        if current_section == boletos_section:
                            # Types like "Acciones", "Títulos Públicos", "Obligaciones Negociables", "Cedears", "Letras del Tesoro nac"
                            current_tipo_instrumento = clean_text
                            continue
                        
                        # For Resultado Ventas: detect tipo instrumento (like "1 / Acciones")
                        # Or instrument details (like "PAMPA ENERGIA S.A. ... - Pesos / 457")
                        elif current_section == resultado_ventas_section:
                            # Check if it's a tipo like "1 / Acciones" (starts with number, then /, then type)
                            tipo_match = re.match(r'^(\d+)\s*/\s*(.+)$', clean_text)
                            if tipo_match:
                                current_tipo_instrumento = tipo_match.group(2).strip()
                            elif ' - ' in clean_text and clean_text.endswith(')') == False:
                                # This is instrument line like "PAMPA ENERGIA S.A. ... - Pesos / 457"
                                instr, cod = self._extract_instrument_info(first_cell)
                                if instr:
                                    current_instrumento = instr
                                    current_cod_instrum = cod
                            continue
                        
                        # For Rentas y Dividendos: detect categoria and tipo instrumento
                        elif current_section == rentas_dividendos_section:
                            # Categoria markers: "Rentas" or "Dividendos"
                            if clean_text in ['Rentas', 'Dividendos']:
                                current_categoria = clean_text
                            # Check if it's a tipo like "59 / Letras del Tesoro nac"
                            elif re.match(r'^\d+\s*/\s*.+$', clean_text):
                                tipo_match = re.match(r'^\d+\s*/\s*(.+)$', clean_text)
                                if tipo_match:
                                    current_tipo_instrumento = tipo_match.group(1).strip()
                            elif ' - ' not in clean_text:
                                # Simple tipo like "Cedears", "Acciones", "Títulos Públicos", "Obligaciones Negociables"
                                current_tipo_instrumento = clean_text
                            else:
                                # This is instrument line like "CEDEAR APPLE INC. - Pesos / 8.445"
                                instr, cod = self._extract_instrument_info(first_cell)
                                if instr:
                                    current_instrumento = instr
                                    current_cod_instrum = cod
                            continue
                        else:
                            continue
                    
                    # Skip "Saldo Anterior" rows
                    if 'Saldo Anterior' in first_cell:
                        continue
                    
                    # Skip "Total" rows
                    if first_cell.strip().startswith('Total') or first_cell.strip().startswith('<b>Total'):
                        continue
                    
                    # This is a data row - add metadata columns based on section
                    if current_section == boletos_section:
                        new_row = [current_tipo_instrumento] + cells
                        current_rows.append(new_row)
                    elif current_section == resultado_ventas_section:
                        new_row = [current_tipo_instrumento, current_instrumento, current_cod_instrum] + cells
                        current_rows.append(new_row)
                    elif current_section == rentas_dividendos_section:
                        # Order: Instrumento, Cod.Instrum, Categoría, tipo_instrumento
                        new_row = [current_instrumento, current_cod_instrum, current_categoria, current_tipo_instrumento] + cells
                        current_rows.append(new_row)
                    else:
                        current_rows.append(cells)
        
        # Save last section
        if current_section and current_headers and current_rows:
            sheet_name = current_section
            if current_currency and current_section not in no_currency_sections:
                sheet_name = f"{current_section} {current_currency}"
            self._save_table(sheet_name, current_headers, current_rows)
        
        return self.tables
    
    def _match_section(self, text: str, sections_dict: dict, posicion_count: int = 0) -> Optional[str]:
        """Match text to a known section name."""
        text_upper = text.upper()
        
        # Special handling for Gallo POSICION sections
        # First POSICION AL -> Posicion Inicial, Second -> Posicion Final
        if "POSICION AL" in text_upper and sections_dict == self.GALLO_SECTIONS:
            if posicion_count == 1:
                return "Posicion Inicial"
            else:
                return "Posicion Final"
        
        for pattern, sheet_name in sections_dict.items():
            if pattern in text_upper:
                return sheet_name
        return None
    
    def _parse_table_row(self, line: str) -> list[str]:
        """Parse a markdown table row into cells."""
        # Remove leading/trailing |
        line = line.strip('|')
        cells = [cell.strip() for cell in line.split('|')]
        return cells
    
    def _save_table(self, section: str, headers: list[str], rows: list[list[str]], metadata: dict = None):
        """Save or merge table data for a section."""
        if metadata is None:
            metadata = {}
            
        if section in self.tables:
            # Merge with existing data
            existing = self.tables[section]
            # Add rows (assuming same headers)
            for row in rows:
                # Skip if row is already present (duplicate from page break)
                if row not in existing.rows:
                    existing.rows.append(row)
            # Merge metadata
            if metadata:
                existing.metadata.update(metadata)
        else:
            self.tables[section] = TableData(
                section=section,
                headers=headers,
                rows=rows.copy(),
                metadata=metadata.copy() if metadata else {}
            )


class ExcelExporter:
    """Export parsed tables to Excel."""
    
    # Column widths by header keywords
    COLUMN_WIDTHS = {
        "especie": 35,
        "detalle": 30,
        "tipo": 35,
        "operacion": 12,
        "fecha": 12,
        "numero": 10,
        "cantidad": 12,
        "precio": 12,
        "importe": 15,
        "costo": 12,
        "resultado": 18,
        "gastos": 15,
        "custodia": 15,
        "cartera": 12,
    }
    
    def __init__(self):
        self.wb = Workbook()
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def add_table(self, table: TableData):
        """Add a table as a new worksheet."""
        ws = self.wb.create_sheet(title=table.section[:31])  # Excel sheet name limit
        
        # Write headers
        for col, header in enumerate(table.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Write data rows
        for row_idx, row_data in enumerate(table.rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._clean_value(value))
                cell.border = self.border
                
                # Right-align numeric values
                if self._is_numeric(value):
                    cell.alignment = Alignment(horizontal="right")
        
        # Auto-width columns
        self._auto_width(ws, table.headers)
    
    def _clean_value(self, value: str) -> str:
        """Clean cell value."""
        if not value:
            return ""
        
        # Remove HTML tags like <b>, </b>
        value = re.sub(r'<[^>]+>', '', value)
        
        # Clean up escaped characters
        value = value.replace('\\$', '$')
        value = value.replace('<br>', ' ')
        
        return value.strip()
    
    def _is_numeric(self, value: str) -> bool:
        """Check if value looks numeric."""
        if not value:
            return False
        # Remove common number formatting
        clean = value.replace(',', '').replace('-', '').replace(' ', '').replace('.', '')
        return clean.isdigit()
    
    def _auto_width(self, ws, headers: list[str]):
        """Set column widths based on header content."""
        for col_idx, header in enumerate(headers, 1):
            header_lower = header.lower()
            width = 12  # default
            
            for keyword, w in self.COLUMN_WIDTHS.items():
                if keyword in header_lower:
                    width = w
                    break
            
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width
    
    def save(self, output_path: str):
        """Save workbook to file."""
        self.wb.save(output_path)


def convert_markdown_to_excel(
    markdown_path: str,
    output_path: Optional[str] = None,
    apply_postprocess: bool = True
) -> str:
    """
    Convert Datalab markdown output to structured Excel.
    
    Args:
        markdown_path: Path to the .datalab.md file
        output_path: Optional output Excel path
        apply_postprocess: Whether to apply format-specific post-processing
    
    Returns:
        Path to the generated Excel file
    """
    md_path = Path(markdown_path)
    
    if not md_path.exists():
        raise FileNotFoundError(f"Markdown file not found: {md_path}")
    
    # Default output path
    if not output_path:
        output_path = str(md_path.with_suffix('.xlsx'))
    
    console.print(f"[cyan]📊 Parsing markdown tables...[/cyan]")
    
    # Read markdown
    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Parse tables
    parser = MarkdownTableParser(content)
    tables = parser.parse()
    format_type = parser.format_type
    
    if not tables:
        console.print("[yellow]⚠️ No tables found in markdown[/yellow]")
        return output_path
    
    console.print(f"[green]✓ Found {len(tables)} sections ({format_type} format)[/green]")
    for section, data in tables.items():
        console.print(f"  • {section}: {len(data.rows)} rows")
    
    # Export to Excel
    console.print(f"\n[cyan]📝 Creating Excel file...[/cyan]")
    
    exporter = ExcelExporter()
    for table in tables.values():
        if table.rows:  # Only add non-empty tables
            exporter.add_table(table)
    
    # Apply post-processing if enabled
    if apply_postprocess:
        from .postprocess import postprocess_gallo_workbook, postprocess_visual_workbook
        
        if format_type == "gallo":
            postprocess_gallo_workbook(exporter.wb, tables)
        else:
            postprocess_visual_workbook(exporter.wb)
    
    exporter.save(output_path)
    console.print(f"[green]✓ Saved to: {output_path}[/green]")
    
    return output_path
    
    return output_path


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Convert Datalab markdown to Excel")
    parser.add_argument("markdown", help="Path to the .datalab.md file")
    parser.add_argument("-o", "--output", help="Output Excel path")
    
    args = parser.parse_args()
    
    try:
        output = convert_markdown_to_excel(args.markdown, args.output)
        console.print(f"\n[bold green]✓ Conversion complete: {output}[/bold green]")
        return 0
    except Exception as e:
        console.print(f"[red]❌ Error: {e}[/red]")
        return 1


if __name__ == "__main__":
    sys.exit(main())
