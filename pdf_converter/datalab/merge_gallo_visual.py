"""
Módulo para unificar archivos Excel de Gallo y Visual en un resumen impositivo consolidado.
Traduce la estructura de Gallo al esquema de Visual y genera hojas de resultados.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple
import re


class GalloVisualMerger:
    """
    Clase principal para unificar Excel de Gallo y Visual.
    
    Genera un Excel consolidado con:
    - Posicion Inicial Gallo
    - Posicion Final Gallo  
    - Boletos (merge de transacciones)
    - Rentas y Dividendos Gallo
    - Resultado Ventas ARS
    - Resultado Ventas USD
    - Rentas Dividendos ARS
    - Rentas Dividendos USD
    - Resumen
    - Hojas auxiliares (EspeciesVisual, EspeciesGallo, etc.)
    """
    
    # Operaciones de compra/venta para Boletos
    OPERACIONES_BOLETOS = ['compra', 'venta', 'cpra', 'canje', 'licitacion', 'licitaciones', 
                           'compra usd', 'venta usd', 'cpra cable', 'venta cable',
                           'trf titulos']
    
    # Operaciones de rentas/dividendos
    OPERACIONES_RENTAS = ['renta', 'dividendo', 'dividendos', 'amortizacion', 'amortizaciones']
    
    # Hojas de transacciones en Gallo
    HOJAS_GALLO_TRANSACCIONES = ['Tit Privados Exentos', 'Renta Fija Dolares', 'Tit Privados Exterior',
                                  'Cauciones', 'Titulos Publicos', 'Cedears']
    
    # Tipos de instrumento que expresan precio cada 100 unidades (dividir por 100 para nominal)
    # NOTA: NO incluir 'on' porque matchea con 'Acciones' usando substring matching
    TIPOS_PRECIO_CADA_100 = ['obligaciones negociables', 'obligacion negociable', 
                             'títulos públicos', 'titulos publicos', 'titulo publico',
                             'letras del tesoro', 'letra del tesoro', 'letras']

    # Para Excel 2007+ (incluye 2013), las fórmulas en XLSX deben almacenarse en inglés (invariante)
    USE_INVARIANT_FORMULAS = True
    
    def _es_tipo_precio_cada_100(self, tipo_instrumento: str) -> bool:
        """Verifica si el tipo de instrumento expresa precio cada 100 unidades."""
        if not tipo_instrumento:
            return False
        tipo_lower = tipo_instrumento.lower().strip()
        # Usar matching parcial para los términos de la lista
        return any(t in tipo_lower for t in self.TIPOS_PRECIO_CADA_100)

    def _resolve_option_underlying(self, option_ticker: str) -> Optional[Tuple[str, str]]:
        """Map an option ticker to the underlying stock (code, ticker).

        Argentine option tickers follow the pattern:
            ESPECIE + C|V + STRIKE + MES_LETTER
        e.g. YPFC49000D → prefix=YPF, underlying YPFD (code 710).

        Returns (code_clean, ticker) or None if not resolvable.
        """
        if not option_ticker:
            return None
        m = re.match(r'^([A-Z]+?)([CV])(\d+)([A-X])$', option_ticker.strip().upper())
        if not m:
            return None
        prefix = m.group(1)

        candidates = []
        for code, data in self._especies_visual_cache.items():
            if data.get('tipo_especie') == 'Acciones':
                ticker = data.get('ticker') or ''
                if ticker.upper().startswith(prefix):
                    candidates.append((code, ticker))

        if not candidates:
            return None

        # Prefer exact prefix match, then shortest ticker
        for code, ticker in candidates:
            if ticker.upper() == prefix:
                return (code, ticker)
        candidates.sort(key=lambda x: len(x[1]))
        return candidates[0]

    def _is_dollar_related(self, *values) -> bool:
        """Determina si los textos refieren a operatoria/activos dolarizados."""
        text = " ".join(str(v or "") for v in values).lower()
        return any(token in text for token in ['dolar', 'dólar', 'usd', 'cable', 'mep'])

    def _resultado_bucket_hint_from_origen(self, origen: str) -> Optional[str]:
        """Da prioridad al bucket implícito del origen cuando la hoja fuente lo deja explícito."""
        origen_text = str(origen or '').strip().lower()
        if 'gallo-renta fija pesos' in origen_text:
            return 'ARS'
        if 'gallo-renta fija dolares' in origen_text or 'gallo-tit privados exterior' in origen_text:
            return 'USD'
        return None

    def _classify_rentas_currency(self, moneda, moneda_emision=None, origen: str = "") -> str:
        """Clasifica una renta/dividendo en ARS o USD priorizando la moneda efectiva de la fila."""
        origen_text = str(origen or '').strip().lower()
        if 'visual-rentas dividendos ars' in origen_text:
            return 'ARS'
        if 'visual-rentas dividendos usd' in origen_text:
            return 'USD'
        if 'gallo' in origen_text and self._is_dollar_related(moneda_emision, origen):
            return 'USD'

        moneda_text = str(moneda or '').strip().lower()
        if 'peso' in moneda_text or moneda_text == 'ars':
            return 'ARS'
        if any(token in moneda_text for token in ['dolar', 'dólar', 'usd', 'mep', 'cable']):
            return 'USD'

        moneda_emision_text = str(moneda_emision or '').strip().lower()
        if 'peso' in moneda_emision_text or moneda_emision_text == 'ars':
            return 'ARS'
        if any(token in moneda_emision_text for token in ['dolar', 'dólar', 'usd', 'mep', 'cable']):
            return 'USD'

        return 'ARS'

    def _normalize_visual_rentas_currency(self, visual_sheet_name: str, moneda, moneda_default: str) -> str:
        """Preserva la moneda efectiva para rentas/dividendos importados desde Visual."""
        sheet_name = str(visual_sheet_name or '').strip().lower()
        if sheet_name.endswith('ars'):
            return 'Pesos'

        moneda_text = str(moneda or '').strip().lower()
        if 'mep' in moneda_text:
            return 'Dolar MEP'
        if 'cable' in moneda_text:
            return 'Dolar Cable'
        if any(token in moneda_text for token in ['dolar', 'dólar', 'usd']):
            return 'Dolar'
        return moneda_default

    def _is_visual_origin(self, origen: str) -> bool:
        """Indica si la fila proviene de Visual."""
        return 'visual' in str(origen or '').lower()

    def _uses_visual_ars_raw_nominal(self, price, tipo_instrumento: str, origen: str = "", moneda_tipo: str = "") -> bool:
        """Determina si una fila Visual ARS ya viene con precio nominal y no debe dividirse por 100."""
        moneda_text = str(moneda_tipo or "").strip().lower()
        if moneda_text not in ("ars", "pesos") or not self._is_visual_origin(origen):
            return False
        if not self._es_tipo_precio_cada_100(tipo_instrumento):
            return False

        price_num = abs(self._to_float(price))
        # En Visual ARS los títulos/letras problemáticos vienen ya nominales (~1.xx).
        # Parma mostró que el corte seguro debe ser más estricto para no capturar precios
        # de pantalla intermedios que todavía están expresados cada 100.
        return 0 < price_num < 2

    def _is_visual_usd_micro_price(self, price, tipo_instrumento: str, origen: str = "", moneda: str = "") -> bool:
        """Detecta micro-precios Visual USD que no deben seguir la heurística normal raw<2."""
        if not self._is_visual_origin(origen):
            return False
        if not self._es_tipo_precio_cada_100(tipo_instrumento):
            return False

        moneda_text = str(moneda or "").strip().lower()
        if not any(token in moneda_text for token in ['dolar', 'dólar', 'usd', 'mep', 'cabl']):
            return False

        price_num = abs(self._to_float(price))
        return 0 < price_num < 0.01

    def _normalize_nominal_price(self, price, tipo_instrumento: str, origen: str = "", moneda_tipo: str = "") -> float:
        """Normaliza un precio a base nominal consistente para cálculos de stock/resultados."""
        price_num = self._to_float(price)
        if self._uses_visual_ars_raw_nominal(price_num, tipo_instrumento, origen, moneda_tipo):
            return price_num
        if self._es_tipo_precio_cada_100(tipo_instrumento):
            return price_num / 100
        return price_num

    def _uses_visual_raw_trade_price(self, price, tipo_instrumento: str, origen: str = "", moneda: str = "") -> bool:
        """Detecta filas Visual cuyo precio de boleto ya viene en precio operado correcto.

        Parma mostró que en Boletos/Resultados de Visual para Títulos Públicos y
        Obligaciones Negociables hay combinaciones donde el precio no debe volver a
        dividirse por 100: pesos con precios ~1.200/1.500 y dólares con precios ~0.63/0.95
        que representan cotización operada correcta del trade.
        """
        if not self._is_visual_origin(origen):
            return False
        if not self._es_tipo_precio_cada_100(tipo_instrumento):
            return False

        price_num = abs(self._to_float(price))
        moneda_text = str(moneda or "").strip().lower()

        if 'peso' in moneda_text or moneda_text == 'ars':
            return price_num >= 100

        if any(token in moneda_text for token in ['dolar', 'dólar', 'usd', 'mep', 'cabl']):
            if self._is_visual_usd_micro_price(price_num, tipo_instrumento, origen, moneda):
                return False
            return 0 < price_num < 2

        return False

    def _normalize_trade_price(self, price, tipo_instrumento: str, origen: str = "", moneda: str = "", moneda_tipo: str = "") -> float:
        """Normaliza precios de trade respetando excepciones validadas por origen/capa."""
        price_num = self._to_float(price)
        if moneda_tipo == "USD" and self._is_visual_origin(origen) and self._es_tipo_precio_cada_100(tipo_instrumento):
            if 0 < abs(price_num) < 0.01:
                return price_num
        if self._is_visual_usd_micro_price(price_num, tipo_instrumento, origen, moneda):
            return price_num
        if self._uses_visual_raw_trade_price(price_num, tipo_instrumento, origen, moneda):
            return price_num
        moneda_context = moneda_tipo or moneda
        return self._normalize_nominal_price(price_num, tipo_instrumento, origen, moneda_context)

    def _should_standardize_visual_usd_price(self, price, tipo_instrumento: str, origen: str = "", moneda: str = "") -> bool:
        """Decide si una fila Visual debe elevarse a pantalla x100 al construir Resultado Ventas USD."""
        if not self._is_visual_origin(origen):
            return False
        if not self._es_tipo_precio_cada_100(tipo_instrumento):
            return False
        if self._uses_visual_ars_raw_nominal(price, tipo_instrumento, origen, moneda):
            return False
        if self._is_visual_usd_micro_price(price, tipo_instrumento, origen, moneda):
            return False
        return True

    def _should_preserve_visual_source_money(
        self,
        origen: str,
        bruto_fuente: float,
        neto_fuente: float,
        bruto_calc: float,
        neto_calc: float,
        moneda: str = "",
    ) -> bool:
        """Decide si conviene preservar los monetarios fuente de Visual frente a una recomputación dañada."""
        if not self._is_visual_origin(origen):
            return False
        if abs(bruto_fuente) <= 0 and abs(neto_fuente) <= 0:
            return False

        calc_ref = max(abs(bruto_calc), abs(neto_calc))
        source_ref = max(abs(bruto_fuente), abs(neto_fuente))
        source_material = source_ref >= 100
        calc_collapsed = calc_ref < 1
        calc_far_below_source = source_ref > 0 and calc_ref <= source_ref * 0.1
        if source_material and (calc_collapsed or calc_far_below_source):
            return True

        # For dollar legs in Visual Boletos, the source bruto/neto from the PDF is often
        # more reliable than recomputing from a rounded micro-price after quantity rescue.
        if self._is_dollar_related(moneda) and source_material:
            same_bruto_sign = bruto_fuente == 0 or bruto_calc == 0 or (bruto_fuente * bruto_calc) > 0
            same_neto_sign = neto_fuente == 0 or neto_calc == 0 or (neto_fuente * neto_calc) > 0
            if same_bruto_sign and same_neto_sign:
                bruto_rel_error = abs(bruto_calc - bruto_fuente) / abs(bruto_fuente) if bruto_fuente else 0
                neto_rel_error = abs(neto_calc - neto_fuente) / abs(neto_fuente) if neto_fuente else 0
                if bruto_rel_error <= 0.10 or neto_rel_error <= 0.10:
                    return True

        return False

    def _should_preserve_visual_usd_micro_source_money(self, origen: str, moneda: str, tipo_instrumento: str,
                                                       price, bruto_fuente: float, bruto_calc: float) -> bool:
        """Preserva el bruto fuente en micro-precios Visual dolarizados cuando el cálculo diverge."""
        if not self._is_visual_origin(origen):
            return False
        if not self._is_dollar_related(moneda):
            return False
        if not self._es_tipo_precio_cada_100(tipo_instrumento):
            return False

        price_num = abs(self._to_float(price))
        if not (0 < price_num < 0.01):
            return False

        source_abs = abs(self._to_float(bruto_fuente))
        calc_abs = abs(self._to_float(bruto_calc))
        if source_abs <= 0 or calc_abs <= 0:
            return False

        relative_error = abs(calc_abs - source_abs) / source_abs
        return relative_error > 0.02

    def _effective_unit_price_from_bruto(self, cantidad, bruto) -> float:
        """Deriva el precio unitario económico desde el bruto fuente cuando éste es más confiable."""
        qty = abs(self._to_float(cantidad))
        gross = abs(self._to_float(bruto))
        if qty <= 0 or gross <= 0:
            return 0.0
        return gross / qty

    def _date_key(self, value) -> str:
        """Normaliza fechas para poder vincular legs del mismo boleto entre hojas."""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if value is None:
            return ""
        return str(value).strip()

    def _meaningful_fx_rate(self, tipo_cambio) -> float:
        """Filtra tipos de cambio fuente que realmente representan una conversión de monedas."""
        tipo_cambio_num = abs(self._to_float(tipo_cambio))
        if tipo_cambio_num > 1:
            return tipo_cambio_num
        return 0.0

    def _build_visual_source_fx_map(self, boletos_ws) -> Dict[int, float]:
        """Busca el TC bruto utilizable por fila Visual, incluso heredándolo del leg apareado."""
        exact_groups: Dict[Tuple[str, str, str, str], List[float]] = {}
        loose_groups: Dict[Tuple[str, str, str], List[float]] = {}
        row_meta = []

        for row in range(2, boletos_ws.max_row + 1):
            origen = boletos_ws.cell(row, 17).value
            if not self._is_visual_origin(origen):
                continue

            cod_instrum = boletos_ws.cell(row, 7).value
            cod_clean = self._clean_codigo(str(cod_instrum)) if cod_instrum else None
            if not cod_clean:
                continue

            concertacion_key = self._date_key(boletos_ws.cell(row, 2).value)
            liquidacion_key = self._date_key(boletos_ws.cell(row, 3).value)
            cantidad_abs = abs(self._to_float(boletos_ws.cell(row, 10).value))
            cantidad_key = f"{cantidad_abs:.10f}"
            tipo_cambio = self._meaningful_fx_rate(boletos_ws.cell(row, 12).value)

            exact_key = (cod_clean, concertacion_key, liquidacion_key, cantidad_key)
            loose_key = (cod_clean, concertacion_key, liquidacion_key)
            if tipo_cambio > 0:
                exact_groups.setdefault(exact_key, []).append(tipo_cambio)
                loose_groups.setdefault(loose_key, []).append(tipo_cambio)

            row_meta.append((row, tipo_cambio, exact_key, loose_key))

        row_fx_map: Dict[int, float] = {}
        for row, own_tipo_cambio, exact_key, loose_key in row_meta:
            if own_tipo_cambio > 0:
                row_fx_map[row] = own_tipo_cambio
                continue

            exact_candidates = exact_groups.get(exact_key, [])
            if exact_candidates:
                row_fx_map[row] = exact_candidates[0]
                continue

            loose_candidates = list(dict.fromkeys(loose_groups.get(loose_key, [])))
            if len(loose_candidates) == 1:
                row_fx_map[row] = loose_candidates[0]

        return row_fx_map

    def _normalize_ars_result_nominal_price(self, price, tipo_instrumento: str, origen: str = "",
                                            moneda: str = "", tipo_cambio=1) -> float:
        """Convierte a precio nominal ARS usando el TC fuente de Visual cuando existe."""
        price_num = self._to_float(price)
        if not price_num:
            return 0.0

        if self._is_visual_origin(origen) and self._is_dollar_related(moneda):
            tipo_cambio_num = abs(self._to_float(tipo_cambio))
            if tipo_cambio_num <= 0:
                tipo_cambio_num = 1.0
            converted_price = price_num * tipo_cambio_num
            if self._is_visual_usd_micro_price(price_num, tipo_instrumento, origen, moneda):
                return converted_price
            if self._uses_visual_raw_trade_price(price_num, tipo_instrumento, origen, moneda):
                return converted_price
            return self._normalize_nominal_price(converted_price, tipo_instrumento, origen, "ARS")

        return self._normalize_trade_price(price_num, tipo_instrumento, origen, moneda, "ARS")

    def _convert_usd_sheet_gastos(self, gastos, tipo_cambio) -> float:
        """Convierte gastos de la hoja USD a USD económico usando valor absoluto."""
        gastos_num = self._to_float(gastos)
        tipo_cambio_num = self._to_float(tipo_cambio)
        if tipo_cambio_num <= 0:
            tipo_cambio_num = 1.0
        return abs(gastos_num * tipo_cambio_num)

    def _get_usd_conversion_factor(self, moneda: str, valor_usd_dia) -> float:
        """Devuelve el factor para llevar monetarios de la hoja USD a USD."""
        moneda_text = str(moneda or "").lower()
        if 'dolar' in moneda_text:
            return 1.0

        valor_usd = self._to_float(valor_usd_dia)
        if valor_usd > 0:
            return 1.0 / valor_usd
        return 1.0

    def _apply_signed_expense(self, bruto, gastos_abs, cantidad) -> float:
        """Empuja el neto más lejos de cero según el signo económico de la operación."""
        bruto_num = self._to_float(bruto)
        gastos_num = abs(self._to_float(gastos_abs))
        cantidad_num = self._to_float(cantidad)
        is_sale = cantidad_num < 0 or (cantidad_num == 0 and bruto_num < 0)
        if is_sale:
            return bruto_num - gastos_num
        return bruto_num + gastos_num

    def _normalize_usd_result_nominal_price(self, nominal_price, tipo_instrumento: str, origen: str = "", moneda: str = "") -> float:
        """Convierte a precio económico unitario para Resultado Ventas USD.

        En BYMA, bonos y ON suelen operarse con pantalla "cada 100". En resultado USD
        necesitamos comparar venta y costo en la misma base económica por unidad.
        Si una fila TP/ON/Letras queda con nominal ~100, se lleva a ~1 dividiendo por 100.
        """
        price_num = self._to_float(nominal_price)
        if not self._es_tipo_precio_cada_100(tipo_instrumento):
            return price_num
        if price_num <= 0:
            return price_num

        if abs(price_num) >= 10:
            return price_num / 100
        return price_num

    def _should_guardrail_stock_price(self, stock_price: float, nominal_price: float, tipo_instrumento: str, moneda_tipo: str) -> bool:
        """Detecta precios de stock absurdos frente al precio nominal de la operación.

        Se usa como protocolo de emergencia para no reportar resultados claramente inválidos
        cuando el dato fuente o el fallback de costo inicial entra en escala incorrecta.
        """
        if moneda_tipo != "USD":
            return False
        if not self._es_tipo_precio_cada_100(tipo_instrumento):
            return False

        stock_val = abs(self._to_float(stock_price))
        nominal_val = abs(self._to_float(nominal_price))
        if stock_val <= 0 or nominal_val <= 0:
            return False

        ratio = stock_val / nominal_val
        return ratio < 0.01 or ratio > 100

    def _resolve_usd_stock_price(self, raw_price, nominal_price, tipo_instrumento: str, cod_instrum: str = "") -> float:
        """Normaliza el precio de stock inicial USD usando reglas y fallbacks de seguridad."""
        stock_price = self._to_float(raw_price)
        nominal_val = abs(self._to_float(nominal_price))

        if stock_price <= 0:
            return stock_price

        if self._es_tipo_precio_cada_100(tipo_instrumento) and nominal_val > 0:
            if stock_price > 2 and nominal_val < 5:
                stock_price = stock_price / 100

        return stock_price

    def _normalize_initial_cost_price(self, price, tipo_instrumento: str, origen_precio: str = "") -> float:
        """Normaliza precios de costo inicial sin volver a dividir fuentes ya nominales.

        `PrecioTenenciasIniciales` ya entra como precio por unidad ajustado; volver a
        aplicar la regla de instrumentos cotizados cada 100 provoca costos casi nulos.
        """
        price_num = self._to_float(price)
        if not price_num:
            return 0.0

        origen_precio_text = str(origen_precio or "").strip().lower()
        if origen_precio_text == 'preciotenenciasiniciales':
            return price_num

        return self._normalize_nominal_price(price_num, tipo_instrumento)

    def _build_ars_nominal_formula(self, row_out: int) -> str:
        """Fórmula de Precio Nominal para Resultado Ventas ARS usando TC fuente de Visual cuando aplica."""
        tipo_checks = (
            f'O(ESNUMERO(HALLAR("OBLIGACION";MAYUSC(B{row_out})));'
            f'ESNUMERO(HALLAR("TITULO";MAYUSC(B{row_out})));'
            f'ESNUMERO(HALLAR("TÍTULO";MAYUSC(B{row_out})));'
            f'ESNUMERO(HALLAR("LETRA";MAYUSC(B{row_out}))))'
        )
        is_visual = f'ESNUMERO(HALLAR("VISUAL";MAYUSC(A{row_out})))'
        is_dollar_source = (
            f'O(ESNUMERO(HALLAR("DOLAR";MAYUSC(G{row_out})));'
            f'ESNUMERO(HALLAR("DÓLAR";MAYUSC(G{row_out})));'
            f'ESNUMERO(HALLAR("USD";MAYUSC(G{row_out})));'
            f'ESNUMERO(HALLAR("MEP";MAYUSC(G{row_out})));'
            f'ESNUMERO(HALLAR("CABLE";MAYUSC(G{row_out}))))'
        )
        base_price = f'SI(Y({is_visual};{is_dollar_source};M{row_out}>0);J{row_out}*M{row_out};J{row_out})'
        visual_raw_nominal = (
            f'Y({is_visual};{tipo_checks};'
            f'O('
            f'Y(NO({is_dollar_source});ABS(J{row_out})<20);'
            f'Y(NO({is_dollar_source});ABS(J{row_out})>=100);'
            f'Y({is_dollar_source};ABS(J{row_out})<2)'
            f'))'
        )
        return f'=SI({visual_raw_nominal};{base_price};SI({tipo_checks};{base_price}/100;{base_price}))'

    def _build_resultado_currency_overrides(self, boletos_ws) -> Dict[str, str]:
        """Asigna una única hoja ARS/USD por código para mantener íntegro el running stock."""
        grouped: Dict[str, Dict[str, bool]] = {}

        for row in range(2, boletos_ws.max_row + 1):
            cod_instrum = boletos_ws.cell(row, 7).value
            cod_clean = self._clean_codigo(str(cod_instrum)) if cod_instrum else None
            if not cod_clean:
                continue

            especie_data = self._especies_visual_cache.get(cod_clean, {})
            moneda = boletos_ws.cell(row, 5).value
            tipo_operacion = boletos_ws.cell(row, 6).value
            origen = boletos_ws.cell(row, 17).value
            moneda_emision = especie_data.get('moneda_emision')

            state = grouped.setdefault(cod_clean, {'usd': False, 'ars': False})
            explicit_bucket = self._resultado_bucket_hint_from_origen(origen)
            if explicit_bucket == 'USD':
                state['usd'] = True
                continue
            if explicit_bucket == 'ARS':
                state['ars'] = True
                continue

            if self._is_dollar_related(moneda, tipo_operacion, moneda_emision):
                state['usd'] = True

            moneda_text = str(moneda or '').lower()
            tipo_text = str(tipo_operacion or '').lower()
            if 'peso' in moneda_text or moneda_text == 'ars' or 'peso' in tipo_text or str(moneda_emision or '').strip() == 'Pesos':
                state['ars'] = True

        overrides: Dict[str, str] = {}
        for cod_clean, state in grouped.items():
            if state['usd']:
                overrides[cod_clean] = 'USD'
            elif state['ars']:
                overrides[cod_clean] = 'ARS'
        return overrides
    
    def __init__(self, gallo_path: str = None, visual_path: str = None, aux_data_dir: str = None, precio_tenencias_path: str = None):
        """
        Inicializa el merger con las rutas a los archivos.
        
        Args:
            gallo_path: Ruta al Excel generado de Gallo (opcional para casos Visual-only)
            visual_path: Ruta al Excel generado de Visual
            aux_data_dir: Directorio con hojas auxiliares (default: pdf_converter/datalab/aux_data)
            precio_tenencias_path: Ruta al Excel generado desde el PDF de Precio Tenencias (opcional)
        """
        if not visual_path:
            raise ValueError("visual_path es obligatorio")

        self.gallo_path = Path(gallo_path) if gallo_path else None
        self.visual_path = Path(visual_path)
        self.precio_tenencias_path = Path(precio_tenencias_path) if precio_tenencias_path else None
        
        if aux_data_dir is None:
            aux_data_dir = Path(__file__).parent / 'aux_data'
        self.aux_data_dir = Path(aux_data_dir)
        
        # Cargar workbooks
        self.gallo_wb = load_workbook(gallo_path) if gallo_path else self._create_empty_gallo_workbook()
        self.visual_wb = load_workbook(visual_path)
        self.precio_tenencias_wb = load_workbook(precio_tenencias_path) if precio_tenencias_path else None
        
        # Cargar hojas auxiliares
        self.especies_visual = self._load_aux('EspeciesVisual.xlsx')
        self.especies_gallo = self._load_aux('EspeciesGallo.xlsx')
        self.cotizacion_dolar = self._load_aux('Cotizacion_Dolar_Historica.xlsx')
        self.precios_iniciales = self._load_aux('PreciosInicialesEspecies.xlsx')
        
        # Cache de mapeos
        self._especies_visual_cache = {}
        self._especies_gallo_cache = {}
        self._cotizacion_cache = {}
        self._precios_iniciales_cache = {}
        self._precios_iniciales_by_codigo = {}  # codigo -> {ticker, precio}
        self._precio_tenencias_by_codigo = {}
        self._precio_tenencias_by_ticker = {}
        self._ratios_cedears_cache = {}
        
        # Load ratio cache first (needed by _build_precio_tenencias_cache)
        self._ratios_cedears_cache = self._load_ratio_cache()
        
        # Construir caches
        self._build_caches()

    def _create_empty_gallo_workbook(self) -> Workbook:
        """Crea un workbook Gallo vacío para flujos Visual-only."""
        wb = Workbook()
        wb.active.title = 'EMPTY_GALLO'
        return wb
    
    def _load_aux(self, filename: str) -> Workbook:
        """Carga un archivo auxiliar."""
        path = self.aux_data_dir / filename
        if not path.exists():
            raise FileNotFoundError(f"Archivo auxiliar no encontrado: {path}")
        return load_workbook(path)
    
    def _build_caches(self):
        """Construye caches para búsquedas rápidas."""
        # Cache EspeciesVisual: codigo -> {nombre, moneda_emision, tipo_especie, ...}
        ws = self.especies_visual.active
        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row, 3).value  # Columna C = codigo
            if codigo:
                codigo_clean = self._clean_codigo(codigo)
                self._especies_visual_cache[codigo_clean] = {
                    'codigo': codigo,
                    'moneda_emision': ws.cell(row, 7).value,  # Col G
                    'ticker': ws.cell(row, 8).value,  # Col H
                    'nombre_con_moneda': ws.cell(row, 17).value,  # Col Q
                    'tipo_especie': ws.cell(row, 18).value,  # Col R
                }
        
        # Cache EspeciesGallo: codigo -> {nombre, ticker, moneda_emision}
        ws = self.especies_gallo.active
        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row, 1).value  # Columna A
            if codigo:
                codigo_clean = self._clean_codigo(codigo)
                self._especies_gallo_cache[codigo_clean] = {
                    'codigo': codigo,
                    'nombre': ws.cell(row, 2).value,  # Col B
                    'ticker': ws.cell(row, 10).value,  # Col J
                    'moneda_emision': ws.cell(row, 14).value,  # Col N
                }
        
        # Cache Cotización Dólar: (fecha, tipo_dolar) -> cotizacion
        ws = self.cotizacion_dolar.active
        for row in range(2, ws.max_row + 1):
            fecha = ws.cell(row, 1).value
            cotizacion = ws.cell(row, 2).value
            tipo_dolar = ws.cell(row, 3).value
            if fecha and cotizacion:
                # Normalizar fecha
                if isinstance(fecha, datetime):
                    fecha_key = fecha.date()
                else:
                    fecha_key = fecha
                self._cotizacion_cache[(fecha_key, tipo_dolar)] = cotizacion
        
        # Cache Precios Iniciales: ticker -> {codigo, precio}
        # Col A = codigo, Col B = nombre, Col C = ticker/ORDEN, Col G = precio
        ws = self.precios_iniciales.active
        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row, 1).value  # Col A = codigo especie
            ticker = ws.cell(row, 3).value  # Col C = ORDEN/ticker
            precio = ws.cell(row, 7).value  # Col G = precio
            if ticker:
                ticker_key = str(ticker).upper().strip()
                self._precios_iniciales_cache[ticker_key] = {
                    'codigo': int(codigo) if codigo else None,
                    'precio': precio if precio else 0
                }
            # Cache adicional por código
            if codigo:
                codigo_clean = self._clean_codigo(codigo)
                self._precios_iniciales_by_codigo[codigo_clean] = {
                    'ticker': ticker_key if ticker else None,
                    'precio': precio if precio else 0
                }
        # Cache PrecioTenencias (si existe)
        if self.precio_tenencias_wb:
            if 'PrecioTenenciasIniciales' in self.precio_tenencias_wb.sheetnames:
                ws_precio = self.precio_tenencias_wb['PrecioTenenciasIniciales']
            else:
                ws_precio = self.precio_tenencias_wb.active
            self._build_precio_tenencias_cache(ws_precio)

    def _build_precio_tenencias_cache(self, ws):
        """Construye cache de PrecioTenenciasIniciales por código y ticker.
        
        Stores the adjusted price: raw / ratio for acciones del exterior,
        raw / 1 for everything else.
        """
        headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]

        def find_col(keyword: str):
            for idx, h in enumerate(headers, start=1):
                if keyword in h:
                    return idx
            return None

        col_codigo = find_col('cod')
        col_ticker = find_col('ticker')
        col_especie = find_col('especie')
        col_precio = find_col('precio tenencia')
        col_cantidad = find_col('cantidad')
        col_importe = find_col('importe')
        col_resultado = find_col('resultado')

        if not col_precio:
            return

        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row, col_codigo).value if col_codigo else None
            ticker = ws.cell(row, col_ticker).value if col_ticker else None
            especie_name = ws.cell(row, col_especie).value if col_especie else None
            precio = ws.cell(row, col_precio).value
            cantidad = ws.cell(row, col_cantidad).value if col_cantidad else None
            importe = ws.cell(row, col_importe).value if col_importe else None
            resultado = ws.cell(row, col_resultado).value if col_resultado else None

            # Compute raw price = importe / cantidad
            raw_price = 0
            if cantidad is not None and importe is not None:
                cantidad_num = self._to_float(cantidad)
                importe_num = self._to_float(importe)
                resultado_num = self._to_float(resultado) if resultado is not None else 0
                # Fix invalid rows: cantidad > 0 but importe <= 0
                if cantidad_num > 0 and importe_num <= 0:
                    if importe_num == 0:
                        raw_price = abs(resultado_num / cantidad_num) if cantidad_num else 0
                    else:
                        raw_price = abs(importe_num) / cantidad_num
                elif cantidad_num:
                    raw_price = importe_num / cantidad_num
            if not raw_price:
                try:
                    raw_price = float(precio) if precio is not None else 0
                except Exception:
                    raw_price = 0

            # Ratio: real CEDEAR ratio for acciones del exterior, 1 for everything else
            ratio = 1
            cod_clean = self._clean_codigo(str(codigo)) if codigo else ''
            if cod_clean and self._is_accion_exterior(cod_clean):
                r = self._get_ratio_for_especie(
                    str(ticker) if ticker else '',
                    str(especie_name) if especie_name else '',
                )
                if r:
                    ratio = r

            adjusted_price = raw_price / ratio

            if codigo:
                codigo_clean = self._clean_codigo(str(codigo))
                self._precio_tenencias_by_codigo[codigo_clean] = adjusted_price
            if ticker:
                self._precio_tenencias_by_ticker[str(ticker).strip().upper()] = adjusted_price

    def _normalize_ratio_key(self, val: str) -> str:
        if not val:
            return ""
        return re.sub(r"[^A-Z0-9]", "", str(val).strip().upper())

    def _load_ratio_cache(self) -> dict:
        try:
            aux_path = self.aux_data_dir / 'RatiosCedearsAcciones.xlsx'
            if not aux_path.exists():
                return {}
            wb_ratios = load_workbook(aux_path)
            ws_ratios = wb_ratios.active
            cache = {}
            for r in range(2, ws_ratios.max_row + 1):
                nombre = ws_ratios.cell(r, 1).value
                ratio_val = ws_ratios.cell(r, 2).value
                key = ws_ratios.cell(r, 3).value
                if ratio_val is None:
                    continue
                try:
                    ratio_num = float(ratio_val)
                except Exception:
                    continue
                if key:
                    normalized_key = self._normalize_ratio_key(key)
                    if normalized_key:
                        cache[normalized_key] = ratio_num
                if nombre:
                    nombre_str = str(nombre).strip()
                    nombre_key = self._normalize_ratio_key(nombre_str.split()[0])
                    if nombre_key:
                        cache.setdefault(nombre_key, ratio_num)
                    # Extract stock ticker from Nombre (format: "Company Name TICKER EXCHANGE")
                    tokens = nombre_str.split()
                    if len(tokens) >= 2:
                        # Second-to-last token is usually the ticker symbol
                        ticker_candidate = tokens[-2]
                        ticker_key = self._normalize_ratio_key(ticker_candidate)
                        if ticker_key and len(ticker_key) <= 6:
                            cache.setdefault(ticker_key, ratio_num)
            return cache
        except Exception:
            return {}

    def _get_ratio_for_especie(self, ticker: str, especie: str) -> float:
        if not self._ratios_cedears_cache:
            return 0.0
        # Try full ticker first (e.g. "NVDAUS")
        search_text = f"{ticker} {especie}".strip().upper()
        key = self._normalize_ratio_key(search_text.split()[0]) if search_text else ""
        ratio = float(self._ratios_cedears_cache.get(key, 0) or 0)
        if ratio:
            return ratio
        # Try ticker without -US/-D suffix (e.g. "NVDA-US" -> "NVDA")
        base_ticker = str(ticker).strip().upper().split('-')[0] if ticker else ""
        if base_ticker:
            key2 = self._normalize_ratio_key(base_ticker)
            ratio = float(self._ratios_cedears_cache.get(key2, 0) or 0)
            if ratio:
                return ratio
        return 0.0

    def _is_accion_exterior(self, codigo: str) -> bool:
        if not codigo:
            return False
        cod_clean = self._clean_codigo(codigo)
        data = self._especies_visual_cache.get(cod_clean, {})
        return (str(data.get('moneda_emision', '')).strip() == "Dolar Cable (exterior)" and
                str(data.get('tipo_especie', '')).strip() == "Acciones")
    
    def _clean_codigo(self, codigo) -> str:
        """Limpia código de especie: quita puntos, ceros a izquierda, etc."""
        if codigo is None:
            return ""
        codigo_str = str(codigo).strip()
        # Quitar puntos
        codigo_str = codigo_str.replace('.', '').replace(',', '')
        # Quitar ceros a la izquierda
        try:
            return str(int(float(codigo_str)))
        except:
            return codigo_str

    def _resolve_gallo_coupon_alias(self, cod_especie, especie: str) -> Optional[Dict[str, str]]:
        """Resuelve cupones Gallo aliasados como `1 + codigo subyacente`.

        Caso validado: Gallo puede emitir un flujo de cupon/dividendo con especie
        textual tipo `CUPON GGAL` y código `01534`, donde el activo subyacente real
        es la acción `00534`.
        """
        especie_text = str(especie or '').strip()
        if 'cupon' not in especie_text.lower():
            return None

        coupon_code = self._clean_codigo(cod_especie)
        if not coupon_code or not coupon_code.startswith('1') or len(coupon_code) <= 1:
            return None

        underlying_code = self._clean_codigo(coupon_code[1:])
        if not underlying_code:
            return None

        underlying_data = self._especies_visual_cache.get(underlying_code, {})
        if not underlying_data:
            return None

        underlying_tipo = str(underlying_data.get('tipo_especie') or '').strip()
        cashflow_operacion = 'RENTA'
        if any(token in underlying_tipo.lower() for token in ['accion', 'cedear', 'fci']):
            cashflow_operacion = 'DIVIDENDO'

        return {
            'coupon_code': coupon_code,
            'underlying_code': underlying_code,
            'underlying_name': underlying_data.get('nombre') or especie_text,
            'underlying_tipo': underlying_tipo,
            'cashflow_operacion': cashflow_operacion,
        }
    
    def _split_especie(self, especie: str) -> Tuple[str, str]:
        """Divide especie en Ticker y resto del nombre."""
        if not especie:
            return "", ""
        parts = str(especie).strip().split(' ', 1)
        ticker = parts[0] if parts else ""
        resto = parts[1] if len(parts) > 1 else ""
        return ticker, resto
    
    def _get_moneda(self, resultado_pesos, resultado_usd, gastos_pesos, gastos_usd, hoja_origen: str, operacion: str = "") -> str:
        """Determina la moneda basándose en los valores, la hoja de origen y la operación."""
        operacion_lower = str(operacion).lower() if operacion else ""
        
        # Si la operación menciona USD/EXT/CABLE
        if 'usd' in operacion_lower or 'ext' in operacion_lower or 'cable' in operacion_lower:
            if 'exterior' in hoja_origen.lower() or 'cable' in operacion_lower:
                return "Dolar Cable"
            else:
                return "Dolar MEP"
        
        # Si hay valores en pesos
        if resultado_pesos and float(resultado_pesos) != 0:
            return "Pesos"
        if gastos_pesos and float(gastos_pesos) != 0:
            return "Pesos"
        
        # Si hay valores en USD, determinar tipo
        if resultado_usd or gastos_usd:
            if 'exterior' in hoja_origen.lower():
                return "Dolar Cable"
            else:
                return "Dolar MEP"
        
        # Por defecto usar la hoja de origen
        if 'dolar' in hoja_origen.lower():
            return "Dolar MEP"
        
        return "Pesos"  # Default
    
    def _parse_fecha(self, fecha_value) -> Tuple[datetime, int]:
        """Parsea fecha de varios formatos. Retorna (datetime, año)."""
        if fecha_value is None:
            return None, 0
        
        if isinstance(fecha_value, datetime):
            return fecha_value, fecha_value.year
        
        if isinstance(fecha_value, str):
            fecha_str = fecha_value.strip()
            # Formato dd/mm/yy o dd/mm/yyyy
            try:
                parts = fecha_str.split('/')
                if len(parts) == 3:
                    day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                    # Ajustar año de 2 dígitos
                    if year < 100:
                        year = 2000 + year if year < 50 else 1900 + year
                    return datetime(year, month, day), year
            except:
                pass
        
        return None, 0
    
    def _is_year_2025(self, fecha_value) -> bool:
        """Verifica si una fecha corresponde a 2025."""
        _, year = self._parse_fecha(fecha_value)
        return year == 2025
    
    def _buscar_codigo_especie(self, especie: str, tipo_especie: str = None) -> Tuple[str, str]:
        """
        Busca el código de especie en las hojas de transacciones de Gallo.
        Retorna (codigo, origen) donde origen es 'Gallo' o 'EspeciesGallo'.
        """
        especie_upper = str(especie).upper().strip()
        
        # Primero buscar en transacciones de Gallo
        for sheet_name in self.gallo_wb.sheetnames:
            if sheet_name in ['Posicion Inicial', 'Posicion Final', 'Resultados']:
                continue
            try:
                ws = self.gallo_wb[sheet_name]
                for row in range(2, min(ws.max_row + 1, 500)):  # Limitar búsqueda
                    cod = ws.cell(row, 2).value  # Col B = cod_especie
                    esp = ws.cell(row, 3).value  # Col C = especie
                    if esp and self._match_especie(especie_upper, str(esp).upper()):
                        if cod:
                            return self._clean_codigo(cod), "Gallo"
            except:
                continue
        
        # Fallback: buscar en EspeciesGallo por nombre similar
        for codigo, data in self._especies_gallo_cache.items():
            nombre = data.get('nombre', '')
            if nombre and self._match_especie(especie_upper, str(nombre).upper()):
                return codigo, "EspeciesGallo"
        
        return "", "NoEncontrado"
    
    def _match_especie(self, especie1: str, especie2: str) -> bool:
        """Verifica si dos especies hacen match (fuzzy)."""
        # Limpieza básica
        e1 = especie1.replace('-', ' ').replace('/', ' ').strip()
        e2 = especie2.replace('-', ' ').replace('/', ' ').strip()
        
        # Match exacto
        if e1 == e2:
            return True
        
        # Match sin ticker (primera palabra)
        words1 = e1.split()
        words2 = e2.split()
        
        # Si uno contiene al otro
        if e1 in e2 or e2 in e1:
            return True
        
        # Match por palabras clave (excluyendo ticker)
        if len(words1) > 1 and len(words2) > 1:
            rest1 = ' '.join(words1[1:])
            rest2 = ' '.join(words2[1:])
            if rest1 == rest2 or rest1 in rest2 or rest2 in rest1:
                return True
        
        return False
    
    def _get_cotizacion(self, fecha, tipo_moneda: str) -> float:
        """Obtiene cotización del dólar para una fecha y tipo."""
        if tipo_moneda == "Pesos":
            return 1.0
        
        # Normalizar fecha
        if isinstance(fecha, datetime):
            fecha_key = fecha.date()
        else:
            fecha_key = fecha
        
        # Buscar en cache con diferentes variantes del tipo
        for tipo_key in ["Dolar MEP (local)", "Dolar MEP", "Dolar Cable", tipo_moneda]:
            if (fecha_key, tipo_key) in self._cotizacion_cache:
                return self._cotizacion_cache[(fecha_key, tipo_key)]
        
        return 1.0
    
    def _generate_ticker_variations(self, ticker: str) -> List[str]:
        """
        Genera variaciones de ticker cambiando 0↔O para manejar errores de OCR.
        Ej: TLC10 -> [TLC10, TLC1O], TL0C0 -> [TL0C0, TLOCO, TL0CO, TLOC0]
        """
        ticker_upper = str(ticker).upper().strip()
        variations = [ticker_upper]
        
        # Encontrar posiciones de 0 y O
        positions_0 = [i for i, c in enumerate(ticker_upper) if c == '0']
        positions_O = [i for i, c in enumerate(ticker_upper) if c == 'O']
        
        # Si hay 0, generar versión con O
        for pos in positions_0:
            new_ticker = ticker_upper[:pos] + 'O' + ticker_upper[pos+1:]
            if new_ticker not in variations:
                variations.append(new_ticker)
        
        # Si hay O, generar versión con 0
        for pos in positions_O:
            new_ticker = ticker_upper[:pos] + '0' + ticker_upper[pos+1:]
            if new_ticker not in variations:
                variations.append(new_ticker)
        
        return variations
    
    def _get_precio_inicial(self, ticker: str) -> float:
        """Obtiene precio inicial de una especie por ticker."""
        ticker_upper = str(ticker).upper().strip()
        
        # Valores fijos para monedas
        if ticker_upper in ['PESOS', '$']:
            return 1.0
        if ticker_upper in ['DOLARES', 'USD', 'U$S', 'DOLAR']:
            return 1167.806
        if 'CABLE' in ticker_upper:
            return 1148.93
        
        # Primero probar ticker exacto
        data = self._precios_iniciales_cache.get(ticker_upper, {})
        if isinstance(data, dict) and data.get('precio'):
            return data.get('precio', 0)
        
        # Si no encuentra, probar variaciones de ticker (OCR 0↔O)
        for ticker_var in self._generate_ticker_variations(ticker_upper):
            data = self._precios_iniciales_cache.get(ticker_var, {})
            if isinstance(data, dict) and data.get('precio'):
                return data.get('precio', 0)
        
        return 0

    def _get_precio_tenencia_inicial(self, codigo: Optional[str], ticker: str) -> float:
        """Obtiene precio tenencia inicial desde PrecioTenenciasIniciales (por código o ticker)."""
        if codigo:
            codigo_clean = self._clean_codigo(str(codigo))
            precio = self._precio_tenencias_by_codigo.get(codigo_clean, 0)
            if precio:
                return precio

        ticker_upper = str(ticker).upper().strip()
        if ticker_upper:
            precio = self._precio_tenencias_by_ticker.get(ticker_upper, 0)
            if precio:
                return precio
            # Probar variaciones OCR 0↔O
            for ticker_var in self._generate_ticker_variations(ticker_upper):
                precio = self._precio_tenencias_by_ticker.get(ticker_var, 0)
                if precio:
                    return precio

        return 0

    def _get_codigo_from_ticker(self, ticker: str) -> Optional[int]:
        """Obtiene código de especie desde el ticker usando PreciosInicialesEspecies."""
        ticker_upper = str(ticker).upper().strip()
        
        # Las monedas no tienen código
        if ticker_upper in ['PESOS', '$', 'DOLARES', 'USD', 'U$S', 'DOLAR', 'DOLAR CABLE']:
            return None
        
        # Primero probar ticker exacto
        data = self._precios_iniciales_cache.get(ticker_upper, {})
        if isinstance(data, dict) and data.get('codigo'):
            return data.get('codigo')
        
        # Si no encuentra, probar variaciones de ticker (OCR 0↔O)
        for ticker_var in self._generate_ticker_variations(ticker_upper):
            data = self._precios_iniciales_cache.get(ticker_var, {})
            if isinstance(data, dict) and data.get('codigo'):
                return data.get('codigo')
        
        return None
    
    def _is_moneda(self, ticker: str) -> bool:
        """Verifica si el ticker corresponde a una moneda (PESOS, DOLARES, DOLAR CABLE)."""
        ticker_upper = str(ticker).upper().strip()
        return ticker_upper in ['PESOS', '$', 'DOLARES', 'USD', 'U$S', 'DOLAR', 'DOLAR CABLE', 'CABLE']
    
    def _vlookup_especies_visual(self, codigo, columna: int):
        """Simula VLOOKUP en EspeciesVisual."""
        codigo_clean = self._clean_codigo(codigo)
        data = self._especies_visual_cache.get(codigo_clean, {})
        
        if columna == 5:  # Moneda emisión
            return data.get('moneda_emision', '')
        elif columna == 15:  # Nombre con moneda
            return data.get('nombre_con_moneda', '')
        elif columna == 16:  # Tipo especie
            return data.get('tipo_especie', '')
        
        return ''
    
    def merge(self, output_mode: str = "both") -> Tuple[Workbook, Workbook]:
        """
        Ejecuta el merge completo y retorna el/los workbook(s) consolidado(s).
        
        Args:
            output_mode: "formulas" (solo fórmulas), "values" (solo valores), 
                        o "both" (ambas versiones, default)
        
        Returns:
            Tuple (wb_formulas, wb_values). Si output_mode != "both", 
            una de las dos será None.
        """
        wb = Workbook()
        # Eliminar hoja default
        wb.remove(wb.active)
        
        # Crear hojas en orden
        self._create_posicion_inicial(wb)
        self._create_posicion_final(wb)
        self._create_boletos(wb)
        self._create_cauciones_tomadoras(wb)  # Cauciones Tomadoras
        self._create_cauciones_colocadoras(wb)  # Cauciones Colocadoras
        self._create_rentas_dividendos_gallo(wb)
        self._create_resultado_ventas_ars(wb)
        self._create_resultado_ventas_usd(wb)
        self._create_rentas_dividendos_ars(wb)
        self._create_rentas_dividendos_usd(wb)
        self._create_fci(wb)
        self._create_opciones(wb)
        self._create_futuros(wb)
        self._create_pagare_cpd(wb)
        self._create_resumen(wb)
        self._create_posicion_titulos(wb)  # Copia directa de Visual
        
        # Agregar hojas auxiliares
        self._add_aux_sheets(wb)
        self._add_precio_tenencias_sheet(wb)
        self._add_ratios_cedears_sheet(wb)

        if self.USE_INVARIANT_FORMULAS:
            self._normalize_formulas_to_english(wb)
        
        if output_mode == "formulas":
            return (wb, None)
        
        # Crear copia para materializar valores
        wb_values = self._deep_copy_workbook(wb)
        
        # Materializar todas las fórmulas en la copia
        self._materialize_formulas(wb_values)
        
        if output_mode == "values":
            return (None, wb_values)
        
        # Modo "both": retornar ambas versiones
        return (wb, wb_values)

    def _normalize_formulas_to_english(self, wb: Workbook):
        """
        Convierte fórmulas a formato invariante (inglés + separador coma + punto decimal)
        para compatibilidad con Excel 2007+.
        """
        func_map = {
            "SI(": "IF(",
            "ESERROR(": "ISERROR(",
            "BUSCARV(": "VLOOKUP(",
            "HALLAR(": "SEARCH(",
            "ESNUMERO(": "ISNUMBER(",
            "O(": "OR(",
            "MINUSC(": "LOWER(",
            "IZQUIERDA(": "LEFT(",
        }

        def _replace_functions(segment: str) -> str:
            for es, en in func_map.items():
                segment = re.sub(rf"\b{re.escape(es)}", en, segment, flags=re.IGNORECASE)
            return segment

        def _normalize_formula(formula: str) -> str:
            # Separar por comillas para no tocar strings literales
            parts = formula.split('"')
            for i in range(0, len(parts), 2):
                part = parts[i]
                part = _replace_functions(part)
                # Booleanos ES -> EN
                part = re.sub(r"\bFALSO\b", "FALSE", part, flags=re.IGNORECASE)
                part = re.sub(r"\bVERDADERO\b", "TRUE", part, flags=re.IGNORECASE)
                # Decimal coma -> punto (solo entre dígitos)
                part = re.sub(r"(?<=\d),(?=\d)", ".", part)
                # Separador de argumentos -> coma
                part = part.replace(";", ",")
                parts[i] = part
            return '"'.join(parts)

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith('='):
                        cell.value = _normalize_formula(val)
    
    def _materialize_formulas(self, wb: Workbook):
        """
        Convierte todas las fórmulas de Excel a valores calculados en Python.
        Esto es necesario porque openpyxl no evalúa fórmulas y el PDF mostraría celdas vacías.
        """
        # 0. Materializar fórmulas en Posicion Inicial y Final (PRIMERO, porque Resultado Ventas las usa)
        if 'Posicion Inicial Gallo' in wb.sheetnames:
            self._materialize_posicion(wb['Posicion Inicial Gallo'])
        if 'Posicion Final Gallo' in wb.sheetnames:
            self._materialize_posicion(wb['Posicion Final Gallo'])
        
        # 1. Materializar fórmulas en Boletos
        if 'Boletos' in wb.sheetnames:
            self._materialize_boletos(wb['Boletos'])
        
        # 2. Materializar fórmulas en Rentas y Dividendos Gallo
        if 'Rentas y Dividendos Gallo' in wb.sheetnames:
            self._materialize_rentas_dividendos_gallo(wb['Rentas y Dividendos Gallo'])
        
        # 3. Materializar fórmulas en Resultado Ventas ARS
        if 'Resultado Ventas ARS' in wb.sheetnames:
            self._materialize_resultado_ventas(wb['Resultado Ventas ARS'], "ARS")
        
        # 4. Materializar fórmulas en Resultado Ventas USD
        if 'Resultado Ventas USD' in wb.sheetnames:
            self._materialize_resultado_ventas(wb['Resultado Ventas USD'], "USD")

        # 5. Materializar Resumen (usa valores ya calculados)
        if 'Resumen' in wb.sheetnames:
            self._materialize_resumen(wb)

    def _materialize_resumen(self, wb: Workbook):
        """Calcula valores del Resumen a partir de hojas ya materializadas."""
        ws = wb['Resumen']

        # Ventas (sumas directas de columnas de resultado calculado)
        ventas_ars = self._sum_column(wb, 'Resultado Ventas ARS', 21)  # U
        ventas_usd = self._sum_column(wb, 'Resultado Ventas USD', 24)  # X

        # Rentas/Dividendos (col C = categoría, col M = Importe)
        rentas_ars = self._sum_by_tipo(wb, 'Rentas Dividendos ARS', 3, 13, ['Rentas'], moneda_filter='ARS')
        dividendos_ars = self._sum_by_tipo(wb, 'Rentas Dividendos ARS', 3, 13, ['Dividendos'], moneda_filter='ARS')
        rentas_usd = self._sum_by_tipo(wb, 'Rentas Dividendos USD', 3, 13, ['Rentas'], moneda_filter='USD')
        dividendos_usd = self._sum_by_tipo(wb, 'Rentas Dividendos USD', 3, 13, ['Dividendos'], moneda_filter='USD')
        fci_ars = self._sum_sheet_result_by_moneda(wb, 'FCI', 'ARS')
        fci_usd = self._sum_sheet_result_by_moneda(wb, 'FCI', 'USD')
        opciones_ars = self._sum_sheet_result_by_moneda(wb, 'Opciones', 'ARS')
        opciones_usd = self._sum_sheet_result_by_moneda(wb, 'Opciones', 'USD')
        pagare_cpd_ars = self._sum_sheet_result_by_moneda(wb, 'Pagare_CPD', 'ARS')
        pagare_cpd_usd = self._sum_sheet_result_by_moneda(wb, 'Pagare_CPD', 'USD')
        futuros_ars = self._sum_sheet_result_by_moneda(wb, 'Futuros', 'ARS')
        futuros_usd = self._sum_sheet_result_by_moneda(wb, 'Futuros', 'USD')

        # Cauciones por moneda (nuevo criterio):
        # J = Cau (Tom): suma costo financiero de Tomadoras
        # K = Cau (Col): suma costo financiero de Colocadoras
        cauciones_tom_ars = self._sum_column(wb, 'Cauciones Tomadoras', 14, moneda_filter='Pesos')
        cauciones_col_ars = self._sum_column(wb, 'Cauciones Colocadoras', 14, moneda_filter='Pesos')
        cauciones_tom_usd = self._sum_column(wb, 'Cauciones Tomadoras', 14, moneda_filter='Dolar')
        cauciones_col_usd = self._sum_column(wb, 'Cauciones Colocadoras', 14, moneda_filter='Dolar')

        # Fila ARS (row 2)
        ws.cell(2, 2, ventas_ars)
        ws.cell(2, 3, fci_ars)
        ws.cell(2, 4, opciones_ars)
        ws.cell(2, 5, rentas_ars)
        ws.cell(2, 6, dividendos_ars)
        ws.cell(2, 7, pagare_cpd_ars)
        ws.cell(2, 8, futuros_ars)
        ws.cell(2, 9, cauciones_tom_ars)
        ws.cell(2, 10, cauciones_col_ars)
        ws.cell(2, 11, (ventas_ars or 0) + (fci_ars or 0) + (opciones_ars or 0) + (rentas_ars or 0) + (dividendos_ars or 0) + (pagare_cpd_ars or 0) + (futuros_ars or 0) + (cauciones_tom_ars or 0) + (cauciones_col_ars or 0))

        # Fila USD (row 3)
        ws.cell(3, 2, ventas_usd)
        ws.cell(3, 3, fci_usd)
        ws.cell(3, 4, opciones_usd)
        ws.cell(3, 5, rentas_usd)
        ws.cell(3, 6, dividendos_usd)
        ws.cell(3, 7, pagare_cpd_usd)
        ws.cell(3, 8, futuros_usd)
        ws.cell(3, 9, cauciones_tom_usd)
        ws.cell(3, 10, cauciones_col_usd)
        ws.cell(3, 11, (ventas_usd or 0) + (fci_usd or 0) + (opciones_usd or 0) + (rentas_usd or 0) + (dividendos_usd or 0) + (pagare_cpd_usd or 0) + (futuros_usd or 0) + (cauciones_tom_usd or 0) + (cauciones_col_usd or 0))

    def _find_header_column(self, ws, aliases: List[str]) -> Optional[int]:
        alias_list = [a.lower() for a in aliases]
        for c in range(1, ws.max_column + 1):
            header = str(ws.cell(1, c).value or '').strip().lower()
            if any(alias == header or alias in header for alias in alias_list):
                return c
        return None

    def _sum_sheet_result_by_moneda(self, wb: Workbook, sheet_name: str, moneda: str) -> float:
        if sheet_name not in wb.sheetnames:
            return 0

        ws = wb[sheet_name]
        moneda_col = self._find_header_column(ws, ['moneda'])
        value_col = self._find_header_column(ws, ['resultado', 'total', 'neto'])

        if not value_col:
            return 0

        total = 0
        for row in range(2, ws.max_row + 1):
            if moneda_col:
                moneda_val = str(ws.cell(row, moneda_col).value or '').upper()
                if moneda == 'ARS' and 'PESO' not in moneda_val and moneda_val != 'ARS':
                    continue
                if moneda == 'USD' and not any(token in moneda_val for token in ['DOLAR', 'DÓLAR', 'USD']):
                    continue

            val = ws.cell(row, value_col).value
            if val is not None and isinstance(val, (int, float)):
                total += float(val)

        return total
    
    def _materialize_posicion(self, ws):
        """
        Materializa fórmulas en hojas de Posición (Inicial y Final).
        
        Col U (21) = Tipo Instrumento = VLOOKUP a EspeciesVisual
        Col V (22) = Precio Nominal = Precio a Utilizar (col P)
        """
        for row in range(2, ws.max_row + 1):
            cod_especie = ws.cell(row, 4).value  # Col D = Codigo especie
            precio_a_utilizar = self._to_float(ws.cell(row, 16).value)  # Col P = Precio a Utilizar
            origen_precio = ws.cell(row, 14).value  # Col N = Origen precio costo
            
            # Obtener tipo de instrumento desde cache de EspeciesVisual
            cod_clean = self._clean_codigo(str(cod_especie)) if cod_especie else None
            especie_data = self._especies_visual_cache.get(cod_clean, {}) if cod_clean else {}
            tipo_instrumento = especie_data.get('tipo_especie', '')
            
            # Guardar Tipo Instrumento en Col U (21)
            ws.cell(row, 21, tipo_instrumento)
            
            # Precio Nominal = Precio a Utilizar normalizado para tipos cotizados cada 100
            ws.cell(row, 22, self._normalize_initial_cost_price(precio_a_utilizar, tipo_instrumento, origen_precio))
    
    def _materialize_boletos(self, ws):
        """
        Materializa fórmulas en la hoja Boletos.
        
        Agrega columna 'Precio Nominal' (Col 20, después de las 19 originales) para tipos que expresan precio cada 100.
        El Bruto y Neto se calculan usando el Precio Nominal.
        """
        # Agregar header para Precio Nominal en Col 20 (después de las 19 originales)
        col_precio_nominal = 20
        if ws.cell(1, col_precio_nominal).value != 'Precio Nominal':
            ws.cell(1, col_precio_nominal, 'Precio Nominal')
            ws.cell(1, col_precio_nominal).font = Font(bold=True)
        
        for row in range(2, ws.max_row + 1):
            # Col G = Cod.Instrum (valor directo)
            cod_instrum = ws.cell(row, 7).value
            cod_clean = self._clean_codigo(str(cod_instrum)) if cod_instrum else None
            especie_data = self._especies_visual_cache.get(cod_clean, {}) if cod_clean else {}
            
            # Col A (1): Tipo de Instrumento - Si es fórmula, buscar en cache
            cell_val = ws.cell(row, 1).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                ws.cell(row, 1, especie_data.get('tipo_especie', ''))
            
            # Obtener tipo de instrumento (ya materializado o valor directo)
            tipo_instrumento = ws.cell(row, 1).value or especie_data.get('tipo_especie', '')

            # Fallback para códigos que solo existen en Gallo (no en Visual/EspeciesVisual):
            # derival el tipo de instrumento del nombre de la hoja Gallo de origen.
            if not tipo_instrumento:
                origen = ws.cell(row, 17).value or ''
                origen_lower = str(origen).lower()
                if 'gallo' in origen_lower:
                    sheet_part = origen_lower.replace('gallo-', '')
                    if any(k in sheet_part for k in ['renta fija', 'titulos publicos', 'tit privados']):
                        tipo_instrumento = 'Títulos Públicos'
                    elif 'cedear' in sheet_part:
                        tipo_instrumento = 'Cedears'
                    if tipo_instrumento:
                        ws.cell(row, 1, tipo_instrumento)
            
            # Col I (9): InstrumentoConMoneda - Si es fórmula, buscar en cache
            cell_val = ws.cell(row, 9).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                ws.cell(row, 9, especie_data.get('nombre_con_moneda', ''))
            
            # Col L (12): Tipo Cambio - Si es fórmula, calcular
            cell_val = ws.cell(row, 12).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                moneda = ws.cell(row, 5).value  # Col E = Moneda
                fecha = ws.cell(row, 2).value   # Col B = Fecha Concertación
                if moneda == "Pesos":
                    tc = 1.0
                else:
                    tc = self._get_cotizacion(fecha, str(moneda) if moneda else "Dolar MEP")
                ws.cell(row, 12, tc)
            
            # Obtener precio original (Col K, 11)
            precio_original = ws.cell(row, 11).value
            try:
                precio_num = float(precio_original or 0)
            except:
                precio_num = 0
            
            origen = ws.cell(row, 17).value  # Col Q = Origen
            moneda = ws.cell(row, 5).value   # Col E = Moneda

            # Calcular Precio Nominal con reglas por origen/capa
            precio_nominal = self._normalize_trade_price(
                precio_num,
                tipo_instrumento,
                origen,
                moneda,
            )
            
            # Guardar Precio Nominal en Col 20 (nueva columna después de las 19 originales)
            ws.cell(row, col_precio_nominal, precio_nominal)

            # Futuros: bruto is meaningless (not qty × precio); preserve OCR gastos/neto.
            if tipo_instrumento and 'futuro' in str(tipo_instrumento).lower():
                ws.cell(row, col_precio_nominal, precio_num)  # keep original price
                ws.cell(row, 13, 0)                           # Bruto = 0
                gastos_ocr = self._to_float(ws.cell(row, 15).value)
                neto_ocr = self._to_float(ws.cell(row, 16).value)
                value = gastos_ocr if gastos_ocr != 0 else neto_ocr
                ws.cell(row, 15, value)
                ws.cell(row, 16, value)
                continue
            
            # Capturar monetarios fuente de Visual antes de recomputar.
            bruto_fuente = self._to_float(ws.cell(row, 13).value)
            neto_fuente = self._to_float(ws.cell(row, 16).value)

            # Col M (13): Bruto = Cantidad * Precio Nominal
            cantidad = ws.cell(row, 10).value  # Col J
            try:
                cantidad_num = float(cantidad or 0)
            except:
                cantidad_num = 0

            # Visual USD micro-price: derive effective price from OCR bruto.
            # When qty was rescued from Resultado Ventas, the rounded OCR price
            # (e.g. 0.0009) × large qty overshoots.  OCR bruto is more accurate;
            # derive price = bruto / qty so downstream calculations stay consistent.
            if (self._is_visual_origin(origen)
                    and self._is_visual_usd_micro_price(precio_num, tipo_instrumento, origen, moneda)
                    and abs(bruto_fuente) > 0
                    and abs(cantidad_num) > 0):
                effective_price = abs(bruto_fuente) / abs(cantidad_num)
                if 0 < effective_price < 0.01:
                    precio_nominal = effective_price
                    ws.cell(row, col_precio_nominal, precio_nominal)

            bruto = cantidad_num * precio_nominal
            ws.cell(row, 13, bruto)
            
            # Col P (16): Neto = SI(J>0, J*PrecioNominal+O, J*PrecioNominal-O)
            gastos = ws.cell(row, 15).value    # Col O
            try:
                gastos_num = float(gastos or 0)
            except:
                gastos_num = 0

            # Zero out gastos when it is nearly equal to bruto (column-width OCR artifact).
            if bruto != 0 and gastos_num != 0 and abs(gastos_num / bruto) > 0.9:
                gastos_num = 0
                ws.cell(row, 15, 0)

            if cantidad_num > 0:
                neto = cantidad_num * precio_nominal + gastos_num
            else:
                neto = cantidad_num * precio_nominal - gastos_num

            if self._should_preserve_visual_source_money(origen, bruto_fuente, neto_fuente, bruto, neto, moneda):
                bruto = bruto_fuente
                neto = neto_fuente
                ws.cell(row, 13, bruto)
                neto = neto_fuente

            # Hard guardrail: no single USD trade should have |bruto| > 1 B.
            # If it does the row is definitely garbage (user rule: > 1000 M USD ⇒ wrong).
            if self._is_dollar_related(moneda) and abs(bruto) > 1_000_000_000:
                if 0 < abs(bruto_fuente) < 1_000_000_000 and abs(cantidad_num) > 0:
                    bruto = bruto_fuente
                    precio_nominal = self._effective_unit_price_from_bruto(cantidad_num, bruto_fuente)
                    ws.cell(row, col_precio_nominal, precio_nominal)
                else:
                    bruto = 0
                    precio_nominal = 0
                    ws.cell(row, col_precio_nominal, 0)
                ws.cell(row, 13, bruto)
                neto = bruto  # best-effort approximation
            ws.cell(row, 16, neto)
            
            # Col R (18): Moneda Emisión - Si es fórmula, buscar en cache
            cell_val = ws.cell(row, 18).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                ws.cell(row, 18, especie_data.get('moneda_emision', ''))
    
    def _materialize_rentas_dividendos_gallo(self, ws):
        """
        Materializa fórmulas en la hoja Rentas y Dividendos Gallo.
        
        Columnas con fórmulas:
        - A (1): Tipo de Instrumento = VLOOKUP a EspeciesVisual
        - I (9): InstrumentoConMoneda = VLOOKUP a EspeciesVisual
        - L (12): Tipo Cambio = 1 si Pesos, else cotización
        - Q (17): Neto Calculado = Para amortización: M*(-1), para otros: J-O+P
        - S (19): Moneda Emisión = VLOOKUP a EspeciesVisual
        """
        for row in range(2, ws.max_row + 1):
            # Col G (7) = Cod.Instrum
            cod_instrum = ws.cell(row, 7).value
            cod_clean = self._clean_codigo(str(cod_instrum)) if cod_instrum else None
            especie_data = self._especies_visual_cache.get(cod_clean, {}) if cod_clean else {}
            
            # Col A (1): Tipo de Instrumento
            cell_val = ws.cell(row, 1).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                ws.cell(row, 1, especie_data.get('tipo_especie', ''))
            
            # Col I (9): InstrumentoConMoneda
            cell_val = ws.cell(row, 9).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                ws.cell(row, 9, especie_data.get('nombre_con_moneda', ''))
            
            # Col L (12): Tipo Cambio
            cell_val = ws.cell(row, 12).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                moneda = ws.cell(row, 5).value  # Col E = Moneda
                fecha = ws.cell(row, 2).value   # Col B = Fecha
                if moneda == "Pesos":
                    tc = 1.0
                else:
                    tc = self._get_cotizacion(fecha, str(moneda) if moneda else "Dolar MEP")
                ws.cell(row, 12, tc)
            
            # Col Q (17): Neto Calculado
            cell_val = ws.cell(row, 17).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                tipo_op = ws.cell(row, 4).value  # Col D = Tipo Operación
                bruto = self._to_float(ws.cell(row, 10).value)   # Col J = Bruto
                gastos = self._to_float(ws.cell(row, 13).value)  # Col M = Gastos/Amortización
                interes = self._to_float(ws.cell(row, 14).value) # Col N = Interés
                iva = self._to_float(ws.cell(row, 15).value)     # Col O = IVA
                iibb = self._to_float(ws.cell(row, 16).value)    # Col P = IIBB
                
                tipo_op_lower = str(tipo_op).lower() if tipo_op else ""
                if 'amortizacion' in tipo_op_lower or 'amortización' in tipo_op_lower:
                    # Para amortización: Neto = Monto (col M) * -1
                    neto = gastos * -1  # gastos contiene el monto de amortización
                else:
                    # Para rentas/dividendos: Neto = Bruto - Gastos + Interés - IVA - IIBB
                    neto = bruto - abs(gastos) + interes - iva - iibb
                ws.cell(row, 17, neto)
            
            # Col S (19): Moneda Emisión
            cell_val = ws.cell(row, 19).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                ws.cell(row, 19, especie_data.get('moneda_emision', ''))
    
    def _materialize_resultado_ventas(self, ws, moneda_tipo: str):
        """
        Materializa fórmulas en hojas de Resultado Ventas.
        
        Algoritmo: Itera secuencialmente fila por fila, comparando D{row} vs D{row-1}
        para detectar cambio de instrumento. Esto replica exactamente el comportamiento
        de las fórmulas Excel.
        
        IMPORTANTE: Para ON, Títulos Públicos, Letras del Tesoro, el precio viene
        expresado cada 100 unidades. Se crea columna "Precio Nominal" = Precio/100.
        
        ARS: 26 columnas originales, Precio Nominal en col 27
        USD: 28 columnas originales, Precio Nominal en col 29
        """
        wb = ws.parent
        
        # Agregar header para Precio Nominal al final
        if moneda_tipo == "ARS":
            col_precio_nominal = 27  # Después de las 26 columnas originales
        else:
            col_precio_nominal = 29  # Después de las 28 columnas originales
        
        if ws.cell(1, col_precio_nominal).value != 'Precio Nominal':
            ws.cell(1, col_precio_nominal, 'Precio Nominal')
            ws.cell(1, col_precio_nominal).font = Font(bold=True)
        
        # Variables de running stock (persisten entre filas del mismo instrumento)
        stock_cantidad = 0.0
        stock_precio = 0.0  # Este será el precio nominal promedio
        prev_cod_instrum = None
        
        for row in range(2, ws.max_row + 1):
            # Leer valores de la fila actual
            cod_instrum_raw = ws.cell(row, 4).value  # Col D = Cod.Instrum
            if not cod_instrum_raw:
                continue
            cod_instrum = self._clean_codigo(str(cod_instrum_raw))
            
            # Obtener tipo de instrumento desde cache
            especie_data = self._especies_visual_cache.get(cod_instrum, {})
            tipo_instrumento = ws.cell(row, 2).value or especie_data.get('tipo_especie', '')
            es_precio_cada_100 = self._es_tipo_precio_cada_100(tipo_instrumento)
            
            origen = ws.cell(row, 1).value or ""  # Col A = Origen
            is_gallo = origen.upper().startswith("GALLO")
            
            cantidad = self._to_float(ws.cell(row, 9).value)  # Col I = Cantidad
            
            # Columnas varían entre ARS y USD
            if moneda_tipo == "ARS":
                precio_original = self._to_float(ws.cell(row, 10).value)   # Col J = Precio
                interes = self._to_float(ws.cell(row, 12).value)  # Col L = Interés
                tipo_cambio = self._to_float(ws.cell(row, 13).value)  # Col M = Tipo de Cambio
                gastos = self._to_float(ws.cell(row, 14).value)   # Col N = Gastos
                
                # Calcular Precio Nominal
                moneda_val = ws.cell(row, 7).value  # Col G = Moneda
                precio_nominal = self._normalize_ars_result_nominal_price(
                    precio_original,
                    tipo_instrumento,
                    origen,
                    moneda_val,
                    tipo_cambio,
                )
                current_nominal_price = precio_nominal
                ws.cell(row, col_precio_nominal, precio_nominal)
                
                # Dividir gastos e intereses por 100 para ON/TP/Letras
                if es_precio_cada_100:
                    gastos = gastos / 100
                    interes = interes / 100
                    # Actualizar celdas con valores divididos
                    ws.cell(row, 12, interes)  # Col L = Interés
                    ws.cell(row, 14, gastos)   # Col N = Gastos
                
                # Recalcular Bruto con precio nominal
                bruto = cantidad * precio_nominal
                ws.cell(row, 11, bruto)  # Col K = Bruto (sobrescribir)
                
                # Columnas de running stock: Q(17)-W(23)
                col_stock_ini_qty = 17   # Q
                col_stock_ini_price = 18 # R
                col_costo = 19           # S
                col_neto = 20            # T
                col_resultado = 21       # U
                col_stock_fin_qty = 22   # V
                col_stock_fin_price = 23 # W
            else:  # USD
                precio_original = self._to_float(ws.cell(row, 10).value)   # Col J = Precio base
                precio_std_original = self._to_float(ws.cell(row, 11).value)  # Col K = Precio Standarizado
                interes = self._to_float(ws.cell(row, 14).value)  # Col N = Interés
                gastos = self._to_float(ws.cell(row, 17).value)   # Col Q = Gastos (ya es valor)
                bruto_fuente = self._to_float(ws.cell(row, 13).value)
                
                # Materializar P (Valor USD Día) - si es fórmula, calcular VLOOKUP
                valor_usd_dia_cell = ws.cell(row, 16).value
                if isinstance(valor_usd_dia_cell, str) and valor_usd_dia_cell.startswith('='):
                    fecha = ws.cell(row, 5).value  # Col E = Concertación
                    valor_usd_dia = self._get_cotizacion(fecha, "Dolar MEP")
                    ws.cell(row, 16, valor_usd_dia)
                else:
                    valor_usd_dia = self._to_float(valor_usd_dia_cell)
                if valor_usd_dia == 0:
                    fecha = ws.cell(row, 5).value
                    valor_usd_dia = self._get_cotizacion(fecha, "Dolar MEP")
                    ws.cell(row, 16, valor_usd_dia)
                
                # Materializar O (Tipo Cambio) - 1 para dolar, sino 1/P
                tipo_cambio_cell = ws.cell(row, 15).value
                moneda_val = ws.cell(row, 7).value or ""  # Col G = Moneda
                if isinstance(tipo_cambio_cell, str) and tipo_cambio_cell.startswith('='):
                    if 'dolar' in str(moneda_val).lower():
                        tipo_cambio = 1.0
                    else:
                        tipo_cambio = 1.0 / valor_usd_dia if valor_usd_dia > 0 else 1.0
                    ws.cell(row, 15, tipo_cambio)
                else:
                    tipo_cambio = self._to_float(tipo_cambio_cell)
                    if tipo_cambio == 0:
                        tipo_cambio = 1.0
                
                # Calcular Precio Nominal USD:
                # 1. Primero convertir precio_std_original a USD: precio_std_original * tipo_cambio
                # 2. Si es ON/TP/Letras, dividir por 100
                precio_std_usd_raw = precio_std_original * tipo_cambio
                precio_std_usd = self._normalize_trade_price(precio_std_usd_raw, tipo_instrumento, origen, moneda_val, moneda_tipo)
                precio_resultado_usd = self._normalize_usd_result_nominal_price(precio_std_usd, tipo_instrumento, origen, moneda_val)
                current_nominal_price = precio_resultado_usd
                
                # Materializar L (Precio Std USD) - Este es el precio por 100VN en USD
                ws.cell(row, 12, precio_std_usd_raw)
                
                # Guardar Precio Nominal (en USD, dividido por 100 si corresponde)
                ws.cell(row, col_precio_nominal, precio_resultado_usd)

                # Reconciliar stock inicial/fallback con la misma escala nominal de la venta.
                if cod_instrum != prev_cod_instrum:
                    pass
                
                # Materializar M (Bruto USD) = I * Precio Nominal (ya en USD y ajustado)
                bruto_usd = cantidad * precio_resultado_usd
                preserve_micro_source = self._should_preserve_visual_usd_micro_source_money(
                    origen,
                    moneda_val,
                    tipo_instrumento,
                    precio_std_usd_raw,
                    bruto_fuente,
                    bruto_usd,
                )
                if preserve_micro_source:
                    bruto_usd = bruto_fuente
                    effective_unit_price = self._effective_unit_price_from_bruto(cantidad, bruto_usd)
                    if effective_unit_price > 0:
                        precio_resultado_usd = effective_unit_price
                        current_nominal_price = effective_unit_price
                ws.cell(row, 13, bruto_usd)
                
                # Los gastos en la hoja USD son monetarios fuente; no deben desescalarse por precio cada 100.
                if es_precio_cada_100:
                    interes = interes / 100
                    # Actualizar celdas con valores divididos
                    ws.cell(row, 14, interes)  # Col N = Interés
                gastos_usd = self._convert_usd_sheet_gastos(gastos, tipo_cambio)
                ws.cell(row, 17, gastos_usd)  # Col Q = Gastos USD visibles
                
                # Columnas de running stock: T(20)-Z(26)
                col_stock_ini_qty = 20   # T
                col_stock_ini_price = 21 # U
                col_costo = 22           # V
                col_neto = 23            # W
                col_resultado = 24       # X
                col_stock_fin_qty = 25   # Y
                col_stock_fin_price = 26 # Z
            
            # ========== LÓGICA DE RUNNING STOCK ==========
            # Si es nuevo instrumento (D{row} != D{row-1}), buscar posición inicial
            if cod_instrum != prev_cod_instrum:
                # Buscar en hoja de posición correspondiente
                # NOTA: _get_posicion_inicial ahora devuelve Precio Nominal (col U=21)
                # que ya está dividido por 100 para ON/TP/Letras
                # Para USD, pasamos for_usd=True para que el fallback ya venga en USD
                is_usd_sheet = (moneda_tipo == "USD")
                stock_cantidad, stock_precio = self._get_posicion_inicial(wb, cod_instrum, is_gallo, for_usd=is_usd_sheet)
                
                # Para USD: convertir precio de posición a USD 
                # SOLO si vino de Posicion (stock_cantidad > 0), porque el fallback ya viene en USD
                if is_usd_sheet and stock_cantidad > 0 and valor_usd_dia > 0:
                    stock_precio = stock_precio / valor_usd_dia
                # Nota: si stock_cantidad == 0, el precio ya viene en USD (del fallback)

                if is_usd_sheet:
                    stock_precio = self._resolve_usd_stock_price(stock_precio, precio_resultado_usd, tipo_instrumento, cod_instrum)
            # else: usar valores de stock_cantidad y stock_precio de la fila anterior
            
            # Guardar stock inicial para esta fila
            cantidad_stock_inicial = stock_cantidad
            precio_stock_inicial = stock_precio  # Ya es nominal y en USD para hojas USD

            guardrail_stock_price = self._should_guardrail_stock_price(
                precio_stock_inicial,
                current_nominal_price,
                tipo_instrumento,
                moneda_tipo,
            )
            
            # Calcular costo, neto, resultado según fórmulas Excel
            if cantidad < 0:  # VENTA
                if moneda_tipo == "ARS":
                    costo = cantidad * precio_stock_inicial  # negativo (cantidad < 0)
                    neto = bruto + interes  # Bruto (con precio nominal) + Interés
                else:  # USD
                    costo = cantidad * precio_stock_inicial  # negativo, precio ya nominal y USD
                    neto = self._apply_signed_expense(bruto_usd, gastos_usd, cantidad)
                # Resultado = IF(Costo<>0, ABS(Neto)-ABS(Costo), 0)
                resultado = abs(neto) - abs(costo) if costo != 0 else 0
            else:  # COMPRA
                costo = 0
                if moneda_tipo == "ARS":
                    neto = bruto + interes
                else:
                    neto = self._apply_signed_expense(bruto_usd, gastos_usd, cantidad)
                resultado = 0  # No hay resultado en compras
            
            # Actualizar stock para la próxima fila
            if cantidad > 0:  # COMPRA - promedio ponderado
                valor_anterior = stock_cantidad * stock_precio
                if moneda_tipo == "USD":
                    # Para USD: usar precio nominal en USD
                    valor_nuevo = cantidad * precio_resultado_usd
                else:
                    valor_nuevo = cantidad * precio_nominal
                stock_cantidad += cantidad
                if stock_cantidad > 0:
                    stock_precio = (valor_anterior + valor_nuevo) / stock_cantidad
            elif cantidad < 0:  # VENTA - solo reduce cantidad
                stock_cantidad += cantidad  # cantidad es negativo
                # Precio promedio no cambia en ventas
            
            cantidad_stock_final = stock_cantidad
            precio_stock_final = stock_precio
            
            # ========== MATERIALIZAR VALORES ==========
            if moneda_tipo == "ARS":
                # Col O (15): IVA = IF(N>0, N*0.1736, N*-0.1736)
                if gastos > 0:
                    iva = gastos * 0.1736
                else:
                    iva = gastos * -0.1736
                ws.cell(row, 15, iva)
            else:  # USD
                # Col R (18): IVA = ABS(Gastos USD) * 0.1736
                iva = gastos_usd * 0.1736 if gastos_usd > 0 else 0
                ws.cell(row, 18, iva)
            
            # Running stock columns
            ws.cell(row, col_stock_ini_qty, cantidad_stock_inicial)
            ws.cell(row, col_stock_ini_price, precio_stock_inicial)  # Ya es nominal y en USD para USD
            
            ws.cell(row, col_costo, costo)
            ws.cell(row, col_neto, neto)
            if guardrail_stock_price:
                ws.cell(row, col_resultado, '|')
            else:
                ws.cell(row, col_resultado, resultado)
            ws.cell(row, col_stock_fin_qty, cantidad_stock_final)
            
            if cantidad_stock_final != 0:
                ws.cell(row, col_stock_fin_price, precio_stock_final)
            else:
                ws.cell(row, col_stock_fin_price, 0)

            bruto_ref = bruto if moneda_tipo == "ARS" else bruto_usd
            audit_col = 26 if moneda_tipo == "ARS" else 28
            warnings = []
            if bruto_ref and abs(resultado) > abs(bruto_ref):
                warnings.append('RESULTADO>BRUTO')
            if current_nominal_price and precio_stock_inicial:
                current_abs = abs(current_nominal_price)
                stock_abs = abs(precio_stock_inicial)
                min_price = min(current_abs, stock_abs)
                max_price = max(current_abs, stock_abs)
                if min_price > 0 and (max_price / min_price) > 20:
                    warnings.append('ESCALA_PRECIO_STOCK')
            if guardrail_stock_price:
                warnings.append('STOCK_PRICE_GUARDRAIL')
            if warnings:
                prev_audit = str(ws.cell(row, audit_col).value or '').strip()
                suffix = ' | ALERTA: ' + ', '.join(warnings)
                ws.cell(row, audit_col, f'{prev_audit}{suffix}' if prev_audit else suffix.lstrip(' |'))
            
            # Actualizar código previo para siguiente iteración
            prev_cod_instrum = cod_instrum
    
    def _to_float(self, value) -> float:
        """Convierte un valor a float de forma segura."""
        if value is None:
            return 0.0
        if isinstance(value, str):
            # Si es fórmula, retornar 0
            if value.startswith('='):
                return 0.0
            try:
                return float(value.replace(',', '.').replace(' ', ''))
            except:
                return 0.0
        try:
            return float(value)
        except:
            return 0.0

    def _fmt_num_es(self, value: float) -> str:
        """Formatea un número para fórmulas en Excel ES (coma decimal)."""
        try:
            num = float(value)
        except:
            return "0"
        s = f"{num}".replace(".", ",")
        return s
    
    def _deep_copy_workbook(self, wb: Workbook) -> Workbook:
        """Crea una copia profunda del workbook guardando a BytesIO y recargando."""
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return load_workbook(buffer)
    
    # Cotización del dólar MEP al inicio del período (31/12/2024)
    COTIZACION_INICIO_PERIODO = 1167.806
    
    def _get_posicion_inicial(self, wb: Workbook, cod_instrum: str, is_gallo: bool, for_usd: bool = False) -> Tuple[float, float]:
        """
        Obtiene (cantidad, precio_nominal) de la hoja de posición correspondiente.
        
        Args:
            wb: Workbook consolidado
            cod_instrum: Código del instrumento (limpio)
            is_gallo: True si origen es Gallo, False si es Visual
            for_usd: True si el resultado se usará para hoja USD (para fallback en USD)
            
        Returns:
            Tuple (cantidad, precio_nominal). Si no encuentra en Posicion, 
            busca en PreciosInicialesEspecies y retorna (0, precio_nominal_fallback).
            Para fallback USD: ya retorna el precio en USD (dividido por cotización inicio período).
            
        NOTA: Ahora devuelve Precio Nominal (Col V=22) que ya está dividido por 100
        para ON, Títulos Públicos y Letras del Tesoro.
        Col U=21 es Tipo Instrumento (VLOOKUP a EspeciesVisual).
        """
        # Para cálculos de Resultado Ventas, siempre usar Posicion Inicial Gallo si existe
        pos_sheet = 'Posicion Inicial Gallo' if 'Posicion Inicial Gallo' in wb.sheetnames else (
            'Posicion Final Gallo' if 'Posicion Final Gallo' in wb.sheetnames else None
        )

        if not pos_sheet:
            return (0.0, 0.0)
        
        if pos_sheet in wb.sheetnames:
            pos_ws = wb[pos_sheet]
            for r in range(2, pos_ws.max_row + 1):
                pos_cod = pos_ws.cell(r, 4).value  # Col D = Codigo especie
                if pos_cod and self._clean_codigo(str(pos_cod)) == cod_instrum:
                    cantidad = self._to_float(pos_ws.cell(r, 9).value)   # Col I = cantidad
                    precio_nominal = self._to_float(pos_ws.cell(r, 22).value)  # Col V = Precio Nominal
                    return (cantidad, precio_nominal)
        
        # No encontrado en Posicion - usar PrecioTenenciasIniciales si está disponible
        precio_tenencia = self._precio_tenencias_by_codigo.get(cod_instrum, 0)
        if precio_tenencia:
            precio_nominal = self._to_float(precio_tenencia)
            if for_usd:
                precio_nominal = precio_nominal / self.COTIZACION_INICIO_PERIODO
            return (0.0, precio_nominal)

        # Si no está, buscar en PreciosInicialesEspecies como fallback
        # Esto cubre el caso de instrumentos que se venden sin haber estado en posicion
        precio_data = self._precios_iniciales_by_codigo.get(cod_instrum, {})
        if precio_data.get('precio'):
            precio_bruto = precio_data['precio']
            # Obtener tipo de instrumento para saber si dividir por 100
            tipo_instrumento = self._vlookup_especies_visual(cod_instrum, 16)
            if self._es_tipo_precio_cada_100(tipo_instrumento):
                precio_nominal = precio_bruto / 100
            else:
                precio_nominal = precio_bruto
            # Para USD: convertir a USD usando cotización del inicio del período
            # (PreciosInicialesEspecies tiene precios al 31/12/2024)
            if for_usd:
                precio_nominal = precio_nominal / self.COTIZACION_INICIO_PERIODO
            # Retornar cantidad=0 (no hay stock previo) pero con precio de referencia
            return (0.0, precio_nominal)
        
        # No encontrado en ningún lado - retornar 0, 0
        return (0.0, 0.0)

    def _compute_synthetic_initial_positions(self) -> dict:
        """Compute initial positions from ALL pre-2025 historical operations in Gallo.

        For securities that have transactions in Gallo but no row in
        Posicion Inicial, we build a running weighted-average stock from
        every pre-2025 operation (COMPRA, VENTA, TRF TITULOS, CANJE,
        LICITACION, AMORTIZACION, etc.) to derive the correct cost basis
        at 01/01/2025.

        Returns:
            dict keyed by cleaned cod_instrum -> {
                'cantidad': float,
                'precio_per100': float,   # weighted-average price in per-100 terms
                'especie': str,
                'tipo_especie': str,      # sheet name (e.g. 'Renta Fija Dolares')
            }
        """
        from collections import defaultdict

        # Collect all pre-2025 operations grouped by security
        ops_by_cod: dict = defaultdict(list)
        seen = set()

        for sheet_name in self.gallo_wb.sheetnames:
            if any(skip in sheet_name for skip in ['Posicion', 'Posición', 'Resultado', 'Cauciones', 'Totales']):
                continue
            gallo_ws = self.gallo_wb[sheet_name]
            for row in range(2, gallo_ws.max_row + 1):
                tipo_fila = gallo_ws.cell(row, 1).value
                if not tipo_fila or str(tipo_fila).lower().strip() != 'transaccion':
                    continue

                operacion = gallo_ws.cell(row, 5).value
                fecha = gallo_ws.cell(row, 4).value
                if not operacion or not fecha:
                    continue

                # Only pre-2025 operations
                if self._is_year_2025(fecha):
                    continue

                operacion_lower = str(operacion).lower().strip()
                # Accept buy/sell/trf/canje/amort — anything that moves stock
                ops_validas = ['compra', 'venta', 'cpra', 'canje', 'licitacion', 'trf', 'amortizacion']
                if not any(op in operacion_lower for op in ops_validas):
                    continue

                cod_especie = gallo_ws.cell(row, 2).value
                especie = gallo_ws.cell(row, 3).value
                numero = gallo_ws.cell(row, 6).value
                cantidad = self._to_float(gallo_ws.cell(row, 7).value)
                precio = self._to_float(gallo_ws.cell(row, 8).value)

                cod_clean = self._clean_codigo(cod_especie)
                if not cod_clean:
                    continue

                # Deduplicate (OCR often produces duplicate rows)
                dedup_key = (sheet_name, str(fecha), str(numero), cod_clean,
                             operacion_lower, cantidad, precio)
                if dedup_key in seen:
                    continue
                seen.add(dedup_key)

                fecha_dt, _ = self._parse_fecha(fecha)
                ops_by_cod[cod_clean].append({
                    'fecha': fecha_dt,
                    'operacion': operacion_lower,
                    'cantidad': cantidad,
                    'precio': precio,
                    'especie': especie,
                    'tipo_especie': sheet_name,
                })

        # For each security, compute running stock through chronological operations
        result = {}
        for cod, ops in ops_by_cod.items():
            ops.sort(key=lambda o: (o['fecha'] or datetime(1900, 1, 1), 0 if o['cantidad'] >= 0 else 1))
            stock_qty = 0.0
            stock_price = 0.0  # weighted average in per-100 terms

            for op in ops:
                qty = op['cantidad']
                px = op['precio']

                if qty > 0:  # Buy-like: COMPRA, TRF, CANJE, LICITACION
                    valor_anterior = stock_qty * stock_price
                    valor_nuevo = qty * px
                    stock_qty += qty
                    if stock_qty > 0:
                        stock_price = (valor_anterior + valor_nuevo) / stock_qty
                elif qty < 0:  # Sell-like: VENTA, AMORTIZACION
                    stock_qty += qty  # qty is negative
                    # Price doesn't change on sales

            if stock_qty > 0 and stock_price > 0:
                # Determine if this security's sheet uses USD pricing
                is_usd_sheet = any(tok in ops[-1]['tipo_especie'].lower()
                                   for tok in ['dolar', 'exterior'])
                result[cod] = {
                    'cantidad': stock_qty,
                    'precio_per100': stock_price,
                    'especie': ops[-1]['especie'],
                    'tipo_especie': ops[-1]['tipo_especie'],
                    'is_usd': is_usd_sheet,
                }

        return result

    def _create_posicion_inicial(self, wb: Workbook):
        """Crea hoja Posicion Inicial Gallo con las mismas columnas que Posicion Final."""
        ws = wb.create_sheet("Posicion Inicial Gallo")
        
        # Headers (22 columnas) - misma estructura que Posicion Final pero con nombres "Inicial"
        headers = ['tipo_especie', 'Ticker', 'especie', 'Codigo especie',
                   'Codigo Especie Origen', 'comentario especies', 'detalle', 'custodia', 'cantidad',
                   'precio Tenencia Inicial Pesos', 'precio Tenencia Inicial USD', 'Precio de PreciosIniciales',
                   'precio costo(en proceso)', 'Origen precio costo', 'comentarios precio costo',
                   'Precio a Utilizar', 'importe_pesos', 'porc_cartera_pesos', 'importe_dolares', 
                   'porc_cartera_dolares', 'Tipo Instrumento', 'Precio Nominal']
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Copiar datos de Gallo
        try:
            gallo_ws = self.gallo_wb['Posicion Inicial']
        except KeyError:
            return
        
        row_out = 2
        for row in range(2, gallo_ws.max_row + 1):
            tipo_especie = gallo_ws.cell(row, 1).value
            especie_full = gallo_ws.cell(row, 2).value
            
            if not especie_full:
                continue
            
            ticker, especie = self._split_especie(especie_full)
            
            # Para monedas (PESOS, DOLARES, DOLAR CABLE), especie = ticker (no vacía)
            is_moneda = self._is_moneda(ticker)
            if is_moneda:
                especie = ticker  # Col C muestra PESOS, DOLARES, DOLAR CABLE
            
            # Buscar código de especie usando ticker en PreciosInicialesEspecies
            codigo = None
            codigo_origen = ""
            if not is_moneda:
                codigo = self._get_codigo_from_ticker(ticker)
                if codigo:
                    codigo_origen = "PreciosInicialesEspecies"
                else:
                    # Fallback: buscar en transacciones de Gallo
                    codigo_str, codigo_origen = self._buscar_codigo_especie(especie_full, tipo_especie)
                    if codigo_str:
                        try:
                            codigo = int(codigo_str)
                        except:
                            codigo = codigo_str
            
            # Datos originales
            detalle = gallo_ws.cell(row, 3).value
            custodia = gallo_ws.cell(row, 4).value
            cantidad = gallo_ws.cell(row, 5).value
            precio_orig = gallo_ws.cell(row, 6).value
            importe_pesos = gallo_ws.cell(row, 7).value
            porc_pesos = gallo_ws.cell(row, 8).value
            importe_usd = gallo_ws.cell(row, 9).value
            porc_usd = gallo_ws.cell(row, 10).value
            
            # Determinar si es renta fija dólares (precio dividido x100) o TIT.PRIVADOS EXTERIOR (precio en USD)
            tipo_lower = str(tipo_especie).lower() if tipo_especie else ""
            es_renta_fija_usd = 'renta fija' in tipo_lower and ('dolar' in tipo_lower or 'usd' in tipo_lower)
            es_tit_privados_ext = 'privados' in tipo_lower and 'exterior' in tipo_lower
            
            # Calcular precios tenencia inicial
            precio_pesos = 0
            precio_usd = 0
            if cantidad and float(cantidad) != 0:
                if importe_pesos:
                    try:
                        precio_pesos = float(importe_pesos) / float(cantidad)
                        # Para renta fija dólares, el precio viene dividido x100
                        if es_renta_fija_usd:
                            precio_pesos = precio_pesos * 100
                    except:
                        pass
                if importe_usd:
                    try:
                        precio_usd = float(importe_usd) / float(cantidad)
                    except:
                        pass
            elif precio_orig:
                precio_pesos = precio_orig
            
            # Precio de PreciosInicialesEspecies (via ticker)
            precio_inicial = self._get_precio_inicial(ticker)
            
            # Para TIT.PRIVADOS EXTERIOR, precio viene en USD - convertir a ARS
            # Usamos cotización del dólar cable al 31/12/2024 (inicio del año)
            if es_tit_privados_ext and precio_inicial > 0:
                # Cotización dólar cable al inicio del período (usamos el valor fijo)
                cotizacion_usd = 1148.93  # Dólar Cable 31/12/2024
                precio_inicial = precio_inicial * cotizacion_usd
            
            # Precio a utilizar = PrecioTenenciasIniciales (adjusted) si existe, sino PreciosInicialesEspecies
            precio_tenencia = self._get_precio_tenencia_inicial(codigo, ticker)
            if precio_tenencia > 0:
                precio_a_utilizar = precio_tenencia
                origen_precio = "PrecioTenenciasIniciales"
            else:
                precio_a_utilizar = precio_inicial
                origen_precio = "PreciosInicialesEspecies"
            
            # Escribir fila
            ws.cell(row_out, 1, tipo_especie)
            ws.cell(row_out, 2, ticker)
            ws.cell(row_out, 3, especie)
            ws.cell(row_out, 4, codigo)  # Forzar número
            ws.cell(row_out, 5, codigo_origen)
            ws.cell(row_out, 6, "")  # comentario especies
            ws.cell(row_out, 7, detalle)
            ws.cell(row_out, 8, custodia)
            ws.cell(row_out, 9, cantidad)
            ws.cell(row_out, 10, precio_pesos)
            ws.cell(row_out, 11, precio_usd)
            ws.cell(row_out, 12, precio_inicial)
            ws.cell(row_out, 13, "")  # precio costo (en proceso)
            ws.cell(row_out, 14, origen_precio)  # origen precio costo
            ws.cell(row_out, 15, "")  # comentarios precio costo
            ws.cell(row_out, 16, precio_a_utilizar)
            ws.cell(row_out, 17, importe_pesos)
            ws.cell(row_out, 18, porc_pesos)
            ws.cell(row_out, 19, importe_usd)
            ws.cell(row_out, 20, porc_usd)
            # Col U (21): Tipo Instrumento = VLOOKUP desde EspeciesVisual usando Codigo especie (col D)
            tipo_instrumento = f'=SI(ESERROR(BUSCARV(D{row_out};EspeciesVisual!C:R;16;FALSO));"";BUSCARV(D{row_out};EspeciesVisual!C:R;16;FALSO))'
            ws.cell(row_out, 21, tipo_instrumento)
            # Col V (22): Precio Nominal = Precio a Utilizar normalizado
            ws.cell(row_out, 22, self._normalize_initial_cost_price(
                precio_a_utilizar,
                self._vlookup_especies_visual(codigo, 16) if codigo else '',
                origen_precio,
            ))
            
            row_out += 1

        # --- Synthetic initial positions from pre-2025 historical operations ---
        # For securities that have transactions in Gallo but no Posicion Inicial
        # entry, compute cost basis from all pre-2025 ops (TRF TITULOS, COMPRA, etc.)
        existing_codes = set()
        for r in range(2, row_out):
            cod_val = ws.cell(r, 4).value
            if cod_val:
                existing_codes.add(self._clean_codigo(str(cod_val)))

        synthetic = self._compute_synthetic_initial_positions()
        for cod, data in sorted(synthetic.items()):
            if cod in existing_codes:
                continue  # Already in Posicion Inicial from Gallo source

            try:
                cod_num = int(cod)
            except (ValueError, TypeError):
                cod_num = cod

            tipo_instrumento_val = self._vlookup_especies_visual(cod, 16) if cod else ''
            # The PPP is in the Gallo sheet's native units (per-100).
            # For USD sheets (Renta Fija Dolares, Tit Privados Exterior),
            # prices are in USD per-100. Col V must be in ARS (nominal)
            # because the stock init code divides by valor_usd_dia for
            # Posicion entries with stock_cantidad > 0.
            precio_per100 = data['precio_per100']
            if data.get('is_usd'):
                # Convert USD per-100 → ARS per-100
                precio_per100 = precio_per100 * self.COTIZACION_INICIO_PERIODO
            precio_nominal = self._normalize_nominal_price(precio_per100, tipo_instrumento_val)

            ws.cell(row_out, 1, data['tipo_especie'])  # tipo_especie (sheet name)
            ws.cell(row_out, 2, "")  # Ticker (unknown for synthetic)
            ws.cell(row_out, 3, data['especie'])
            ws.cell(row_out, 4, cod_num)
            ws.cell(row_out, 5, "Synthetic-from-history")
            ws.cell(row_out, 6, "Computed from pre-2025 ops (TRF/COMPRA/CANJE)")
            ws.cell(row_out, 7, "")  # detalle
            ws.cell(row_out, 8, "")  # custodia
            ws.cell(row_out, 9, data['cantidad'])
            ws.cell(row_out, 10, 0)  # precio pesos - unknown
            ws.cell(row_out, 11, 0)  # precio usd - unknown
            ws.cell(row_out, 12, 0)  # PreciosIniciales - not used
            ws.cell(row_out, 13, "")  # precio costo
            ws.cell(row_out, 14, "Synthetic-from-history")  # origen
            ws.cell(row_out, 15, f"PPP from {len(synthetic)} pre-2025 ops")
            ws.cell(row_out, 16, precio_per100)  # Precio a Utilizar — ARS per-100 for USD sheets
            ws.cell(row_out, 17, 0)  # importe_pesos
            ws.cell(row_out, 18, 0)  # porc_cartera_pesos
            ws.cell(row_out, 19, 0)  # importe_dolares
            ws.cell(row_out, 20, 0)  # porc_cartera_dolares
            tipo_inst_formula = f'=SI(ESERROR(BUSCARV(D{row_out};EspeciesVisual!C:R;16;FALSO));"";BUSCARV(D{row_out};EspeciesVisual!C:R;16;FALSO))'
            ws.cell(row_out, 21, tipo_inst_formula)
            ws.cell(row_out, 22, precio_nominal)  # Precio Nominal (already normalized)

            row_out += 1

    def _create_posicion_final(self, wb: Workbook):
        """Crea hoja Posicion Final Gallo con columnas adicionales."""
        ws = wb.create_sheet("Posicion Final Gallo")
        
        # Headers (22 columnas)
        headers = ['tipo_especie', 'Ticker', 'especie', 'Codigo especie',
                   'Codigo Especie Origen', 'comentario especies', 'detalle', 'custodia', 'cantidad',
                   'precio Tenencia Final Pesos', 'precio Tenencia Final USD', 'Precio Tenencia Inicial',
                   'precio costo(en proceso)', 'Origen precio costo', 'comentarios precio costo',
                   'Precio a Utilizar', 'importe_pesos', 'porc_cartera_pesos', 'importe_dolares', 
                   'porc_cartera_dolares', 'Tipo Instrumento', 'Precio Nominal']
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Copiar datos de Gallo
        try:
            gallo_ws = self.gallo_wb['Posicion Final']
        except KeyError:
            return
        
        row_out = 2
        for row in range(2, gallo_ws.max_row + 1):
            tipo_especie = gallo_ws.cell(row, 1).value
            especie_full = gallo_ws.cell(row, 2).value
            
            if not especie_full:
                continue
            
            ticker, especie = self._split_especie(especie_full)
            
            # Para monedas (PESOS, DOLARES, DOLAR CABLE), especie = ticker
            is_moneda = self._is_moneda(ticker)
            if is_moneda:
                especie = ticker  # Col C muestra PESOS, DOLARES, DOLAR CABLE
            
            # Buscar código de especie usando ticker en PreciosInicialesEspecies
            codigo = None
            codigo_origen = ""
            if not is_moneda:
                codigo = self._get_codigo_from_ticker(ticker)
                if codigo:
                    codigo_origen = "PreciosInicialesEspecies"
                else:
                    # Fallback: buscar en transacciones de Gallo
                    codigo_str, codigo_origen = self._buscar_codigo_especie(especie_full, tipo_especie)
                    if codigo_str:
                        try:
                            codigo = int(codigo_str)
                        except:
                            codigo = codigo_str
            
            # Datos originales
            detalle = gallo_ws.cell(row, 3).value
            custodia = gallo_ws.cell(row, 4).value
            cantidad = gallo_ws.cell(row, 5).value
            importe_pesos = gallo_ws.cell(row, 7).value
            porc_pesos = gallo_ws.cell(row, 8).value
            importe_usd = gallo_ws.cell(row, 9).value
            porc_usd = gallo_ws.cell(row, 10).value
            
            # Determinar si es renta fija dólares (precio dividido x100) o TIT.PRIVADOS EXTERIOR (precio en USD)
            tipo_lower = str(tipo_especie).lower() if tipo_especie else ""
            es_tit_privados_ext = 'privados' in tipo_lower and 'exterior' in tipo_lower
            
            # Calcular precios tenencia final
            precio_pesos = 0
            precio_usd = 0
            if cantidad and float(cantidad) != 0:
                if importe_pesos:
                    try:
                        precio_pesos = float(importe_pesos) / float(cantidad)
                    except:
                        pass
                if importe_usd:
                    try:
                        precio_usd = float(importe_usd) / float(cantidad)
                    except:
                        pass
            
            # Precio tenencia inicial (de PreciosInicialesEspecies)
            precio_inicial = self._get_precio_inicial(ticker)
            
            # Para TIT.PRIVADOS EXTERIOR, precio viene en USD - convertir a ARS
            if es_tit_privados_ext and precio_inicial > 0:
                cotizacion_usd = 1148.93  # Dólar Cable 31/12/2024
                precio_inicial = precio_inicial * cotizacion_usd
            
            # Precio a utilizar = precio tenencia inicial
            precio_a_utilizar = precio_inicial
            
            # Escribir fila
            ws.cell(row_out, 1, tipo_especie)
            ws.cell(row_out, 2, ticker)
            ws.cell(row_out, 3, especie)
            ws.cell(row_out, 4, codigo)  # Forzar número (int)
            ws.cell(row_out, 5, codigo_origen)
            ws.cell(row_out, 6, "")  # comentario especies
            ws.cell(row_out, 7, detalle)
            ws.cell(row_out, 8, custodia)
            ws.cell(row_out, 9, cantidad)
            ws.cell(row_out, 10, precio_pesos)
            ws.cell(row_out, 11, precio_usd)
            ws.cell(row_out, 12, precio_inicial)
            ws.cell(row_out, 13, "")  # precio costo (en proceso)
            ws.cell(row_out, 14, "PreciosInicialesEspecies")  # origen precio costo
            ws.cell(row_out, 15, "")  # comentarios precio costo
            ws.cell(row_out, 16, precio_a_utilizar)
            ws.cell(row_out, 17, importe_pesos)
            ws.cell(row_out, 18, porc_pesos)
            ws.cell(row_out, 19, importe_usd)
            ws.cell(row_out, 20, porc_usd)
            # Col U (21): Tipo Instrumento = VLOOKUP desde EspeciesVisual usando Codigo especie (col D)
            tipo_instrumento = f'=SI(ESERROR(BUSCARV(D{row_out};EspeciesVisual!C:R;16;FALSO));"";BUSCARV(D{row_out};EspeciesVisual!C:R;16;FALSO))'
            ws.cell(row_out, 21, tipo_instrumento)
            # Col V (22): Precio Nominal = Precio a Utilizar normalizado
            ws.cell(row_out, 22, self._normalize_initial_cost_price(
                precio_a_utilizar,
                self._vlookup_especies_visual(codigo, 16) if codigo else '',
                "PreciosInicialesEspecies",
            ))
            
            row_out += 1

    def _create_boletos(self, wb: Workbook):
        """Crea hoja Boletos con transacciones de Gallo y Visual, ordenadas por Cod.Instrum y fecha."""
        ws = wb.create_sheet("Boletos")
        
        # Headers (20 columnas - agregamos Precio Nominal)
        headers = ['Tipo de Instrumento', 'Concertación', 'Liquidación', 'Nro. Boleto',
                   'Moneda', 'Tipo Operación', 'Cod.Instrum', 'Instrumento Crudo',
                   'InstrumentoConMoneda', 'Cantidad', 'Precio', 'Tipo Cambio',
                   'Bruto', 'Interés', 'Gastos', 'Neto Calculado', 'Origen', 
                   'moneda emision', 'Auditoría', 'Precio Nominal']
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Recolectar todas las transacciones para ordenar
        all_transactions = []
        seen_gallo_transactions = set()
        
        # Procesar hojas de Gallo
        for sheet_name in self.gallo_wb.sheetnames:
            # Saltear hojas de posición y resultados
            if any(skip in sheet_name for skip in ['Posicion', 'Resultado', 'Posición']):
                continue
            
            # SKIP cauciones - van en hoja separada
            if 'caucion' in sheet_name.lower():
                continue
            
            try:
                gallo_ws = self.gallo_wb[sheet_name]
            except:
                continue
            
            # Pre-scan: in "Pesos" sheets, detect instrument codes that have
            # a "VENTA USD" operation.  Gallo glitches and puts all rows for
            # that instrument in USD despite being in the Pesos section.
            # ALL rows for such an instrument must use "Dolar MEP".
            _usd_override_codes = set()
            if 'pesos' in sheet_name.lower():
                for _r in range(2, gallo_ws.max_row + 1):
                    _op = gallo_ws.cell(_r, 5).value
                    if _op and 'venta' in str(_op).lower() and 'usd' in str(_op).lower():
                        _cod = gallo_ws.cell(_r, 2).value
                        if _cod:
                            _usd_override_codes.add(str(_cod).strip())
            
            for row in range(2, gallo_ws.max_row + 1):
                operacion = gallo_ws.cell(row, 5).value  # Col E
                fecha = gallo_ws.cell(row, 4).value
                numero = gallo_ws.cell(row, 6).value
                
                if not operacion:
                    continue
                
                operacion_lower = str(operacion).lower().strip()
                
                # SKIP operaciones de cauciones (COL CAU TER con instrumento VARIAS)
                especie = gallo_ws.cell(row, 3).value
                if especie and 'varias' in str(especie).lower():
                    continue
                if 'col cau' in operacion_lower:
                    continue

                coupon_alias = self._resolve_gallo_coupon_alias(gallo_ws.cell(row, 2).value, especie)
                if coupon_alias:
                    continue
                
                # Solo operaciones de compra/venta/trf para Boletos
                operaciones_validas = ['compra', 'venta', 'cpra', 'canje', 'licitacion', 'trf']
                if not any(op in operacion_lower for op in operaciones_validas):
                    continue
                
                # Filtrar solo 2025
                if not self._is_year_2025(fecha):
                    continue
                
                # Extraer datos
                cod_especie = gallo_ws.cell(row, 2).value
                cantidad = gallo_ws.cell(row, 7).value
                precio = gallo_ws.cell(row, 8).value
                resultado_pesos = gallo_ws.cell(row, 11).value
                resultado_usd = gallo_ws.cell(row, 12).value
                gastos_pesos = gallo_ws.cell(row, 13).value
                gastos_usd = gallo_ws.cell(row, 14).value
                
                # Determinar moneda PRIMERO basándose en el nombre de la hoja
                # Si la hoja dice "Pesos", es Pesos.
                # EXCEPCION: si alguna op del instrumento es "VENTA USD" en esta
                # sección Pesos, Gallo glitchea y TODAS las filas de ese
                # instrumento tienen valores en USD → override a "Dolar MEP".
                # Si la hoja dice "Dolares", es Dolar MEP
                # Si la hoja dice "Exterior", es Dolar Cable
                sheet_lower = sheet_name.lower()
                cod_str = str(cod_especie).strip() if cod_especie else ''
                if 'pesos' in sheet_lower:
                    if cod_str in _usd_override_codes:
                        moneda = "Dolar MEP"
                    else:
                        moneda = "Pesos"
                elif 'exterior' in sheet_lower:
                    moneda = "Dolar Cable"
                elif 'dolar' in sheet_lower:
                    moneda = "Dolar MEP"
                else:
                    # Fallback: usar la lógica original
                    moneda = self._get_moneda(resultado_pesos, resultado_usd, gastos_pesos, gastos_usd, sheet_name, operacion)
                
                # Gastos según moneda
                gastos = gastos_pesos if moneda == "Pesos" else gastos_usd
                if gastos is None:
                    gastos = 0
                
                # Código limpio y forzar a número
                cod_clean = self._clean_codigo(cod_especie)
                try:
                    cod_num = int(cod_clean) if cod_clean else None
                except:
                    cod_num = cod_clean
                
                # Convertir fecha a datetime para Excel
                fecha_dt, _ = self._parse_fecha(fecha)

                dedupe_key = (
                    sheet_name,
                    fecha_dt if fecha_dt else str(fecha),
                    str(numero or '').strip(),
                    str(cod_num or '').strip(),
                    str(moneda or '').strip().lower(),
                    operacion_lower,
                    self._to_float(cantidad),
                    self._to_float(precio),
                    self._to_float(gastos),
                    str(especie or '').strip().upper(),
                )
                if dedupe_key in seen_gallo_transactions:
                    continue
                seen_gallo_transactions.add(dedupe_key)
                
                # Auditoría
                auditoria = f"Origen: Gallo-{sheet_name} | Fecha: {fecha} | Cod: {cod_especie} | Op: {operacion}"
                
                # Guardar transacción (sin fórmulas, se generan al escribir)
                all_transactions.append({
                    'cod_instrum': cod_num,  # Forzado a número
                    'fecha': fecha_dt if fecha_dt else fecha,
                    'fecha_raw': fecha,
                    'liquidacion': "",
                    'numero': numero,
                    'moneda': moneda,
                    'operacion': operacion,
                    'especie': especie,
                    'cantidad': cantidad,
                    'precio': precio,
                    'interes': 0,
                    'gastos': gastos,
                    'bruto_fuente': None,
                    'neto_fuente': None,
                    'origen': f"gallo-{sheet_name}",
                    'auditoria': auditoria,
                    'tipo_instrumento_val': None,  # Se usará fórmula
                })
        
        # Agregar transacciones de Visual
        try:
            visual_boletos = self.visual_wb['Boletos']

            # --- Ejercicio dedup pre-pass ---
            # Exercise boletos can appear duplicated in Visual with the same boleto
            # number: one row carries the underlying nominal (larger |qty|, real)
            # and the other carries only the contract count (smaller |qty|, noise).
            # Also covers the Sturman 2797 case where one leg is negative (closure).
            # Rule: for each ejercicio boleto number, keep the row whose |qty| is
            # largest; when a positive row exists alongside a negative one, prefer
            # the positive one (shares received). Rows thus marked are skipped
            # inside the main iterator below.
            ejercicio_groups: Dict = {}
            for _row in range(2, visual_boletos.max_row + 1):
                _oper = str(visual_boletos.cell(_row, 6).value or '').lower()
                if 'ejercicio' not in _oper:
                    continue
                _bol = visual_boletos.cell(_row, 4).value
                if _bol in (None, ''):
                    continue
                _bol_key = str(_bol).strip()
                _qty_val = self._to_float(visual_boletos.cell(_row, 9).value)
                ejercicio_groups.setdefault(_bol_key, []).append((_row, _qty_val))

            _ejercicio_skip_rows: set = set()
            for _bol_key, _rows in ejercicio_groups.items():
                if len(_rows) < 2:
                    continue
                # Prefer positive-qty rows when any positive exists (covers 2797).
                _positives = [r for r in _rows if r[1] > 0]
                _candidates = _positives if _positives else _rows
                # Winner: row with the largest |qty| among candidates (covers Salvo).
                _winner = max(_candidates, key=lambda r: abs(r[1]))
                for _r, _q in _rows:
                    if _r != _winner[0]:
                        _ejercicio_skip_rows.add(_r)

            for row in range(2, visual_boletos.max_row + 1):
                if row in _ejercicio_skip_rows:
                    continue
                tipo_instrumento = visual_boletos.cell(row, 1).value
                concertacion = visual_boletos.cell(row, 2).value
                liquidacion = visual_boletos.cell(row, 3).value
                numero = visual_boletos.cell(row, 4).value
                moneda = visual_boletos.cell(row, 5).value
                operacion = visual_boletos.cell(row, 6).value
                cod_instrum = visual_boletos.cell(row, 7).value
                instrumento = visual_boletos.cell(row, 8).value
                cantidad = visual_boletos.cell(row, 9).value
                precio = visual_boletos.cell(row, 10).value
                tipo_cambio_fuente = visual_boletos.cell(row, 11).value
                bruto_fuente = visual_boletos.cell(row, 12).value
                interes = visual_boletos.cell(row, 13).value
                gastos = visual_boletos.cell(row, 14).value
                neto_fuente = visual_boletos.cell(row, 15).value
                
                if not operacion:
                    continue
                
                # Parsear fecha
                fecha_dt, year = self._parse_fecha(concertacion)
                if year != 2025:
                    continue
                
                # Código limpio y forzar a número
                cod_clean = self._clean_codigo(cod_instrum)
                try:
                    cod_num = int(cod_clean) if cod_clean else None
                except:
                    cod_num = cod_clean
                
                auditoria = f"Origen: Visual | Fecha: {concertacion} | Cod: {cod_instrum} | Op: {operacion}"
                
                # --- Column-shift detection for USD cada-100 instruments ---
                # When OCR drops a cell (e.g. Precio), all subsequent columns shift
                # left: the TC value lands in the Precio slot, Bruto in the TC slot, etc.
                # Detect this by checking if the "Precio" looks like a TC (>200 for USD).
                _pr = self._to_float(precio)
                _tc = self._to_float(tipo_cambio_fuente)
                _br = self._to_float(bruto_fuente)
                _qt = self._to_float(cantidad)
                if (self._is_dollar_related(moneda)
                        and self._es_tipo_precio_cada_100(tipo_instrumento)
                        and abs(_pr) > 200
                        and abs(_qt) > 0
                        and abs(_tc) > 0):
                    derived_price = abs(_tc) / abs(_qt)
                    if 0 < derived_price < 200:
                        # Shift confirmed: precio=TC, TC=bruto → correct them.
                        tipo_cambio_fuente = _pr
                        bruto_fuente = _tc
                        precio = derived_price
                        gastos = 0
                        interes = 0
                        neto_fuente = _tc   # approx neto ≈ bruto (gastos zeroed)
                        auditoria += " | SHIFT_CORRECTED"

                # --- Ejercicio handling ---
                # Duplicate/closure rows were already filtered by the pre-pass
                # above (see `_ejercicio_skip_rows`). The surviving row is the
                # one with the largest |qty| (underlying nominal) or, when
                # available, the positive-qty row (shares received).
                oper_lower = str(operacion).lower()
                if 'ejercicio' in oper_lower:
                    # Shares-received row — resolve underlying stock
                    option_ticker_orig = str(instrumento or '')
                    resolved = self._resolve_option_underlying(option_ticker_orig)
                    if resolved:
                        underlying_code, underlying_ticker = resolved
                        try:
                            cod_num = int(underlying_code)
                        except (ValueError, TypeError):
                            cod_num = underlying_code
                        instrumento = underlying_ticker
                        tipo_instrumento = 'Acciones'
                        auditoria += f" | EJERCICIO_MAPPED:{option_ticker_orig}->{underlying_ticker}(cod={underlying_code})"

                all_transactions.append({
                    'cod_instrum': cod_num,  # Forzado a número
                    'fecha': fecha_dt if fecha_dt else concertacion,
                    'fecha_raw': concertacion,
                    'liquidacion': liquidacion,
                    'numero': numero,
                    'moneda': moneda,
                    'operacion': operacion,
                    'especie': instrumento,
                    'cantidad': cantidad,
                    'precio': precio,
                    'tipo_cambio_fuente': tipo_cambio_fuente,
                    'bruto_fuente': bruto_fuente,
                    'interes': interes if interes else 0,
                    'gastos': gastos if gastos else 0,
                    'neto_fuente': neto_fuente,
                    'origen': "Visual",
                    'auditoria': auditoria,
                    'tipo_instrumento_val': tipo_instrumento,
                })
        except KeyError:
            pass  # Visual no tiene hoja Boletos
        
        # Ordenar por fecha de concertación
        def sort_key(t):
            fecha = t.get('fecha')
            if isinstance(fecha, datetime):
                return fecha
            else:
                return datetime.min
        
        all_transactions.sort(key=sort_key)
        
        # Escribir transacciones ordenadas
        for row_out, trans in enumerate(all_transactions, start=2):
            # Fórmulas con row_out correcto
            # Tipo de Instrumento: usa VLOOKUP si no viene de Visual
            tipo_instrumento = f'=SI(ESERROR(BUSCARV(G{row_out};EspeciesVisual!C:R;16;FALSO));"";BUSCARV(G{row_out};EspeciesVisual!C:R;16;FALSO))' if not trans['tipo_instrumento_val'] else trans['tipo_instrumento_val']
            
            # InstrumentoConMoneda
            instrumento_con_moneda = f'=SI(ESERROR(BUSCARV(G{row_out};EspeciesVisual!C:Q;15;FALSO));"";BUSCARV(G{row_out};EspeciesVisual!C:Q;15;FALSO))'
            
            # Tipo Cambio: fórmula simplificada compatible con Excel 2013 español
            # Usa el TC crudo de Visual cuando existe; si no, cae al histórico.
            tipo_cambio_fuente = self._to_float(trans.get('tipo_cambio_fuente'))
            if self._is_visual_origin(trans.get('origen')) and tipo_cambio_fuente > 0:
                tipo_cambio = tipo_cambio_fuente
            else:
                tipo_cambio = f'=SI(E{row_out}="Pesos";1;SI(ESERROR(BUSCARV(B{row_out};\'Cotizacion Dolar Historica\'!A:B;2;FALSO));0;BUSCARV(B{row_out};\'Cotizacion Dolar Historica\'!A:B;2;FALSO)))'
            
            # Precio Nominal: dividir por 100 si es ON, Títulos Públicos o Letras del Tesoro
            # Busca en el Tipo de Instrumento (col A) si contiene Obligacion, Titulo/Título o Letra
            precio_nominal = f'=SI(O(ESNUMERO(HALLAR("Obligacion";A{row_out}));ESNUMERO(HALLAR("Titulo";A{row_out}));ESNUMERO(HALLAR("Título";A{row_out}));ESNUMERO(HALLAR("Letra";A{row_out})));K{row_out}/100;K{row_out})'
            
            # Bruto y Neto usan Precio Nominal (col T) en lugar de Precio (col K)
            bruto = f'=J{row_out}*T{row_out}'
            neto = f'=SI(J{row_out}>0;J{row_out}*T{row_out}+O{row_out};J{row_out}*T{row_out}-O{row_out})'
            moneda_emision = f'=SI(ESERROR(BUSCARV(G{row_out};EspeciesVisual!C:Q;5;FALSO));"";BUSCARV(G{row_out};EspeciesVisual!C:Q;5;FALSO))'
            
            ws.cell(row_out, 1, tipo_instrumento)
            ws.cell(row_out, 2, trans['fecha'])
            ws.cell(row_out, 3, trans['liquidacion'])
            ws.cell(row_out, 4, trans['numero'])
            ws.cell(row_out, 5, trans['moneda'])
            ws.cell(row_out, 6, trans['operacion'])
            ws.cell(row_out, 7, trans['cod_instrum'])  # Ya es número
            ws.cell(row_out, 8, trans['especie'])
            ws.cell(row_out, 9, instrumento_con_moneda)
            ws.cell(row_out, 10, trans['cantidad'])
            ws.cell(row_out, 11, trans['precio'])
            ws.cell(row_out, 12, tipo_cambio)
            if trans.get('bruto_fuente') is not None and 'visual' in str(trans.get('origen') or '').lower():
                ws.cell(row_out, 13, trans.get('bruto_fuente'))
            else:
                ws.cell(row_out, 13, bruto)
            ws.cell(row_out, 14, trans['interes'])
            ws.cell(row_out, 15, trans['gastos'])
            if trans.get('neto_fuente') is not None and 'visual' in str(trans.get('origen') or '').lower():
                ws.cell(row_out, 16, trans.get('neto_fuente'))
            else:
                ws.cell(row_out, 16, neto)
            ws.cell(row_out, 17, trans['origen'])
            ws.cell(row_out, 18, moneda_emision)
            ws.cell(row_out, 19, trans['auditoria'])
            ws.cell(row_out, 20, precio_nominal)  # Precio Nominal
    
    def _create_cauciones_tomadoras(self, wb: Workbook):
        """
        Crea hoja Cauciones Tomadoras con operaciones donde el comitente TOMA prestado.
        Incluye datos de Gallo (operación contiene 'TOM') y Visual (hoja 'Cauciones Tomadoras').
        """
        self._create_cauciones_by_type(wb, "Cauciones Tomadoras", "TOM")
    
    def _create_cauciones_colocadoras(self, wb: Workbook):
        """
        Crea hoja Cauciones Colocadoras con operaciones donde el comitente COLOCA fondos.
        Incluye datos de Gallo (operación contiene 'COL') y Visual (hoja 'Cauciones Colocadoras').
        """
        self._create_cauciones_by_type(wb, "Cauciones Colocadoras", "COL")
    
    def _create_cauciones_by_type(self, wb: Workbook, sheet_name: str, tipo_filtro: str):
        """
        Crea hoja Cauciones con operaciones de caución filtradas por tipo.
        
        Args:
            sheet_name: Nombre de la hoja a crear
            tipo_filtro: "TOM" para tomadoras, "COL" para colocadoras
        
        Columnas según estructura Visual:
        - Concertación, Plazo, Liquidación, Operación, Boleto
        - Contado, Futuro, Tipo de Cambio, Tasa (%)
        - Interés Bruto, Interés Devengado, Aranceles, Derechos, Costo Financiero
        """
        ws = wb.create_sheet(sheet_name)
        
        # Headers según estructura Visual
        headers = ['Concertación', 'Plazo', 'Liquidación', 'Operación', 'Boleto',
                   'Contado', 'Futuro', 'Tipo de Cambio', 'Tasa (%)', 
                   'Interés Bruto', 'Interés Devengado', 'Aranceles', 'Derechos',
                   'Costo Financiero', 'Moneda', 'Origen', 'Auditoría']
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Recolectar todas las cauciones para ordenar
        all_cauciones = []
        
        # Procesar hojas de Cauciones de Gallo
        for gallo_sheet_name in self.gallo_wb.sheetnames:
            if 'caucion' not in gallo_sheet_name.lower():
                continue
            
            # Determinar moneda del nombre de la hoja de Gallo
            if 'pesos' in gallo_sheet_name.lower():
                moneda = "Pesos"
                tipo_cambio = 1
            elif 'dolar' in gallo_sheet_name.lower():
                moneda = "Dolar MEP"
                tipo_cambio = 1167.806  # Cotización dólar al 31/12/2024
            else:
                moneda = "Pesos"
                tipo_cambio = 1
            
            try:
                gallo_ws = self.gallo_wb[gallo_sheet_name]
            except:
                continue
            
            # Estructura Cauciones Gallo según OneShotSpec:
            # ["tipo_fila", "cod_especie", "especie", "fecha", "vencimiento", "operacion", 
            #  "numero", "colocado", "al_vencimiento", "interes_pesos", "interes_usd", 
            #  "gastos_pesos", "gastos_usd"]
            for row in range(2, gallo_ws.max_row + 1):
                tipo_fila = gallo_ws.cell(row, 1).value
                fecha = gallo_ws.cell(row, 4).value
                vencimiento = gallo_ws.cell(row, 5).value
                operacion = gallo_ws.cell(row, 6).value
                numero = gallo_ws.cell(row, 7).value
                colocado = gallo_ws.cell(row, 8).value
                al_vencimiento = gallo_ws.cell(row, 9).value
                interes_pesos = gallo_ws.cell(row, 10).value
                interes_usd = gallo_ws.cell(row, 11).value
                gastos_pesos = gallo_ws.cell(row, 12).value
                gastos_usd = gallo_ws.cell(row, 13).value
                
                # Saltear filas de total
                if tipo_fila and 'total' in str(tipo_fila).lower():
                    continue
                
                if not operacion:
                    continue
                
                # Filtrar por tipo de caución (TOM o COL)
                operacion_upper = str(operacion).upper()
                if tipo_filtro not in operacion_upper:
                    continue
                
                # Filtrar solo 2025 usando vencimiento (col E del origen visual del usuario)
                if not self._is_year_2025(vencimiento):
                    continue
                
                # Parsear fechas
                fecha_dt, _ = self._parse_fecha(fecha)
                venc_dt, _ = self._parse_fecha(vencimiento)
                
                # Calcular plazo (diferencia en días)
                plazo = 0
                if fecha_dt and venc_dt:
                    plazo = (venc_dt - fecha_dt).days
                
                # Interés según moneda (para interes_devengado)
                interes = interes_pesos if moneda == "Pesos" else interes_usd
                interes = float(interes) if interes else 0
                interes = abs(interes)
                
                # Gastos: tomar el que tenga valor (cualquiera de los dos)
                gastos = float(gastos_pesos) if gastos_pesos else (float(gastos_usd) if gastos_usd else 0)
                gastos = abs(gastos)
                
                # Interes Bruto = Futuro - Contado
                try:
                    contado_val = float(colocado) if colocado else 0
                    futuro_val = float(al_vencimiento) if al_vencimiento else 0
                    interes_bruto = abs(futuro_val - contado_val)
                except:
                    interes_bruto = 0
                
                # Reglas por tipo de caución:
                # - TOM: interés/aranceles/derechos negativos (costo),
                #        CF = Interés Bruto + Aranceles + Derechos (más negativo)
                # - COL: interés positivo, aranceles/derechos restan,
                #        CF = Interés Bruto - Aranceles - Derechos
                if tipo_filtro == "TOM":
                    interes_bruto = -abs(interes_bruto)
                    interes_devengado = -abs(interes)
                    aranceles = -abs(gastos)
                    derechos = 0
                    costo_financiero = interes_bruto + aranceles + derechos
                else:
                    interes_bruto = abs(interes_bruto)
                    interes_devengado = abs(interes)
                    aranceles = abs(gastos)
                    derechos = 0
                    costo_financiero = interes_bruto - aranceles - derechos
                
                auditoria = f"Origen: Gallo-{gallo_sheet_name}"
                
                all_cauciones.append({
                    'fecha': fecha_dt if fecha_dt else fecha,
                    'plazo': plazo,
                    'liquidacion': venc_dt if venc_dt else vencimiento,
                    'operacion': operacion,
                    'boleto': numero,
                    'contado': colocado,
                    'futuro': al_vencimiento,
                    'tipo_cambio': tipo_cambio,
                    'tasa': None,  # No disponible en Gallo
                    'interes_bruto': interes_bruto,
                    'interes_devengado': interes_devengado,
                    'aranceles': aranceles,
                    'derechos': derechos,
                    'costo_financiero': costo_financiero,
                    'moneda': moneda,
                    'origen': f"Gallo-{gallo_sheet_name}",
                    'auditoria': auditoria,
                })
        
        # Agregar cauciones de Visual (si existen hojas correspondientes)
        visual_sheet_name = sheet_name  # "Cauciones Tomadoras" o "Cauciones Colocadoras"
        if visual_sheet_name in self.visual_wb.sheetnames:
            visual_ws = self.visual_wb[visual_sheet_name]

            # Mapeo robusto por encabezados para evitar corrimientos OCR
            v_headers = [str(visual_ws.cell(1, c).value or '').strip().lower() for c in range(1, visual_ws.max_column + 1)]
            def _find_col(options, default_idx):
                for i, h in enumerate(v_headers, start=1):
                    if any(opt in h for opt in options):
                        return i
                return default_idx

            col_fecha = _find_col(['concert'], 1)
            col_plazo = _find_col(['plaz'], 2)
            col_liq = _find_col(['liquid'], 3)
            col_op = _find_col(['operaci'], 4)
            col_bol = _find_col(['# boleto', 'nro. boleto', 'boleto'], 5)
            col_contado = _find_col(['contado'], 6)
            col_futuro = _find_col(['futuro'], 7)
            col_tc = _find_col(['tipo de cambio', 'tipo cambio'], 8)
            col_tasa = _find_col(['tasa'], 9)
            col_ib = _find_col(['interés bruto', 'interes bruto'], 10)
            col_id = _find_col(['interés deveng', 'interes deveng'], 11)
            col_ara = _find_col(['arancel'], 12)
            col_der = _find_col(['derech'], 13)
            
            for row in range(2, visual_ws.max_row + 1):
                # Estructura esperada de Visual cauciones:
                # Concertación, Plazo, Liquidación, Operación, Boleto,
                # Contado, Futuro, Tipo de Cambio, Tasa (%),
                # Interés Bruto, Interés Devengado, Aranceles, Derechos, Costo Financiero
                fecha = visual_ws.cell(row, col_fecha).value
                plazo = visual_ws.cell(row, col_plazo).value
                liquidacion = visual_ws.cell(row, col_liq).value
                operacion = visual_ws.cell(row, col_op).value
                boleto = visual_ws.cell(row, col_bol).value
                contado = visual_ws.cell(row, col_contado).value
                futuro = visual_ws.cell(row, col_futuro).value
                tipo_cambio = visual_ws.cell(row, col_tc).value
                tasa = visual_ws.cell(row, col_tasa).value
                interes_bruto_raw = visual_ws.cell(row, col_ib).value
                interes_devengado_raw = visual_ws.cell(row, col_id).value
                aranceles_raw = visual_ws.cell(row, col_ara).value
                derechos_raw = visual_ws.cell(row, col_der).value
                
                if not operacion:
                    continue
                
                interes_bruto_raw = float(interes_bruto_raw) if isinstance(interes_bruto_raw, (int, float)) else 0
                interes_devengado_raw = float(interes_devengado_raw) if isinstance(interes_devengado_raw, (int, float)) else 0
                aranceles_raw = float(aranceles_raw) if isinstance(aranceles_raw, (int, float)) else 0
                derechos_raw = float(derechos_raw) if isinstance(derechos_raw, (int, float)) else 0

                if tipo_filtro == "TOM":
                    interes_bruto = -abs(interes_bruto_raw)
                    interes_devengado = -abs(interes_devengado_raw)
                    aranceles = -abs(aranceles_raw)
                    derechos = -abs(derechos_raw)
                    costo_financiero = interes_bruto + aranceles + derechos
                else:
                    interes_bruto = abs(interes_bruto_raw)
                    interes_devengado = abs(interes_devengado_raw)
                    aranceles = abs(aranceles_raw)
                    derechos = abs(derechos_raw)
                    costo_financiero = interes_bruto - aranceles - derechos
                
                # Determinar moneda (asumimos Pesos por default, o buscar en columna si existe)
                moneda = "Pesos"
                if tipo_cambio and float(tipo_cambio) > 1:
                    moneda = "Dolar MEP"
                
                fecha_dt, _ = self._parse_fecha(fecha)
                liq_dt, _ = self._parse_fecha(liquidacion)
                
                auditoria = f"Origen: Visual-{visual_sheet_name}"
                
                all_cauciones.append({
                    'fecha': fecha_dt if fecha_dt else fecha,
                    'plazo': plazo,
                    'liquidacion': liq_dt if liq_dt else liquidacion,
                    'operacion': operacion,
                    'boleto': boleto,
                    'contado': contado,
                    'futuro': futuro,
                    'tipo_cambio': tipo_cambio or 1,
                    'tasa': tasa,
                    'interes_bruto': interes_bruto,
                    'interes_devengado': interes_devengado,
                    'aranceles': aranceles,
                    'derechos': derechos,
                    'costo_financiero': costo_financiero,
                    'moneda': moneda,
                    'origen': f"Visual-{visual_sheet_name}",
                    'auditoria': auditoria,
                })
        
        # Ordenar por fecha de concertación
        def sort_key(t):
            fecha = t.get('fecha')
            if isinstance(fecha, datetime):
                return fecha
            else:
                return datetime.min
        
        all_cauciones.sort(key=sort_key)
        
        # Escribir cauciones ordenadas
        for row_out, cau in enumerate(all_cauciones, start=2):
            ws.cell(row_out, 1, cau['fecha'])
            ws.cell(row_out, 2, cau['plazo'])
            ws.cell(row_out, 3, cau['liquidacion'])
            ws.cell(row_out, 4, cau['operacion'])
            ws.cell(row_out, 5, cau['boleto'])
            ws.cell(row_out, 6, cau['contado'])
            ws.cell(row_out, 7, cau['futuro'])
            ws.cell(row_out, 8, cau['tipo_cambio'])
            ws.cell(row_out, 9, cau['tasa'])
            ws.cell(row_out, 10, cau['interes_bruto'])
            ws.cell(row_out, 11, cau['interes_devengado'])
            ws.cell(row_out, 12, cau['aranceles'])
            ws.cell(row_out, 13, cau['derechos'])
            ws.cell(row_out, 14, cau['costo_financiero'])
            ws.cell(row_out, 15, cau['moneda'])
            ws.cell(row_out, 16, cau['origen'])
            ws.cell(row_out, 17, cau['auditoria'])
    
    def _create_rentas_dividendos_gallo(self, wb: Workbook):
        """Crea hoja Rentas y Dividendos Gallo, ordenada por fecha de concertación."""
        ws = wb.create_sheet("Rentas y Dividendos Gallo")
        
        # Headers (20 columnas)
        headers = ['Tipo de Instrumento', 'Concertación', 'Liquidación', 'Nro. Boleto',
                   'Moneda', 'Tipo Operación', 'Cod.Instrum', 'Instrumento Crudo',
                   'InstrumentoConMoneda', 'Cantidad', 'Precio', 'Tipo Cambio',
                   'Bruto', 'Interés', 'Gastos', 'Costo', 'Neto Calculado', 
                   'Origen', 'moneda emision', 'Auditoría']
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Recolectar todas las transacciones para ordenar
        all_rentas = []
        
        # Procesar hojas de Gallo
        for sheet_name in self.gallo_wb.sheetnames:
            # Saltear hojas de posición y resultados
            if any(skip in sheet_name for skip in ['Posicion', 'Resultado', 'Posición']):
                continue
            
            try:
                gallo_ws = self.gallo_wb[sheet_name]
            except:
                continue
            
            for row in range(2, gallo_ws.max_row + 1):
                operacion = gallo_ws.cell(row, 5).value  # Col E
                if not operacion:
                    continue
                
                operacion_lower = str(operacion).lower().strip()
                cod_especie = gallo_ws.cell(row, 2).value
                especie = gallo_ws.cell(row, 3).value
                coupon_alias = self._resolve_gallo_coupon_alias(cod_especie, especie)
                is_coupon_cashflow = coupon_alias is not None
                
                # Solo operaciones de rentas/dividendos/amortización.
                # Excepción validada: algunos cupones Gallo vienen como `VENTA`
                # con código aliasado `1 + codigo subyacente`; deben rutearse como
                # flujo de renta/dividendo del activo base, no como boleto.
                if not is_coupon_cashflow and not any(op in operacion_lower for op in self.OPERACIONES_RENTAS):
                    continue

                fecha = gallo_ws.cell(row, 4).value
                numero = gallo_ws.cell(row, 6).value
                cantidad = gallo_ws.cell(row, 7).value
                precio = gallo_ws.cell(row, 8).value
                importe = gallo_ws.cell(row, 9).value
                costo = gallo_ws.cell(row, 10).value
                resultado_pesos = gallo_ws.cell(row, 11).value
                resultado_usd = gallo_ws.cell(row, 12).value
                gastos_pesos = gallo_ws.cell(row, 13).value
                gastos_usd = gallo_ws.cell(row, 14).value
                
                # Filtrar solo 2025
                if not self._is_year_2025(fecha):
                    continue

                if coupon_alias:
                    cod_especie = coupon_alias['underlying_code']
                    especie = coupon_alias['underlying_name']
                    operacion = coupon_alias['cashflow_operacion']
                    operacion_lower = operacion.lower()
                
                # Determinar moneda basándose en nombre de hoja
                sheet_lower = sheet_name.lower()
                if 'pesos' in sheet_lower:
                    moneda = "Pesos"
                elif 'exterior' in sheet_lower:
                    moneda = "Dolar Cable"
                elif 'dolar' in sheet_lower:
                    moneda = "Dolar MEP"
                else:
                    moneda = self._get_moneda(resultado_pesos, resultado_usd, gastos_pesos, gastos_usd, sheet_name, operacion)
                
                # Gastos según moneda - siempre positivos
                gastos = gastos_pesos if moneda == "Pesos" else gastos_usd
                if gastos is None:
                    gastos = 0
                gastos = abs(float(gastos)) if gastos else 0
                
                # Cantidad siempre positiva
                if cantidad:
                    cantidad = abs(float(cantidad))
                
                # Para amortizaciones: costo /100 y precio ajustado
                is_amortizacion = 'amortizacion' in operacion_lower or 'amortización' in operacion_lower
                if is_amortizacion:
                    # Costo dividido por 100
                    if costo:
                        costo = abs(float(costo)) / 100
                    # Ajustar precio para amortizaciones (100 -> 1)
                    if precio and float(precio) == 100:
                        precio = 1
                
                # Código limpio y forzar a número
                cod_clean = self._clean_codigo(cod_especie)
                try:
                    cod_num = int(cod_clean) if cod_clean else None
                except:
                    cod_num = cod_clean
                
                # Bruto - siempre positivo
                bruto = importe if importe else (cantidad * precio if cantidad and precio else 0)
                bruto = abs(float(bruto)) if bruto else 0
                
                # Parsear fecha
                fecha_dt, _ = self._parse_fecha(fecha)
                
                auditoria = f"Origen: Gallo-{sheet_name} | Operación: {operacion}"
                if coupon_alias:
                    auditoria += (
                        f" | CUPON_ALIAS:{coupon_alias['coupon_code']}->"
                        f"{coupon_alias['underlying_code']}"
                    )
                
                all_rentas.append({
                    'tipo_instrumento_val': None,  # Usará fórmula
                    'fecha': fecha_dt if fecha_dt else fecha,
                    'liquidacion': "",
                    'numero': numero,
                    'moneda': moneda,
                    'operacion': operacion.upper(),
                    'cod_num': cod_num,
                    'especie': especie,
                    'cantidad': cantidad if cantidad else 0,
                    'precio': precio if precio else 0,
                    'bruto': bruto,
                    'interes': 0,
                    'gastos': gastos,
                    'costo': costo if costo else 0,
                    'origen': f"Gallo-{sheet_name}",
                    'is_amortizacion': is_amortizacion,
                    'auditoria': auditoria,
                })
        
        # Agregar Rentas/Dividendos de Visual
        visual_sheets = [('Rentas Dividendos ARS', 'Pesos'), ('Rentas Dividendos USD', 'Dolar')]
        for visual_sheet_name, moneda_default in visual_sheets:
            try:
                visual_ws = self.visual_wb[visual_sheet_name]
            except KeyError:
                continue
            
            for row in range(2, visual_ws.max_row + 1):
                instrumento = visual_ws.cell(row, 1).value
                cod_instrum = visual_ws.cell(row, 2).value
                categoria = visual_ws.cell(row, 3).value
                tipo_instrum = visual_ws.cell(row, 4).value
                concertacion = visual_ws.cell(row, 5).value
                liquidacion = visual_ws.cell(row, 6).value
                nro_ndc = visual_ws.cell(row, 7).value
                tipo_operacion = visual_ws.cell(row, 8).value
                cantidad = visual_ws.cell(row, 9).value
                moneda = visual_ws.cell(row, 10).value
                tipo_cambio_val = visual_ws.cell(row, 11).value
                gastos = visual_ws.cell(row, 12).value
                importe = visual_ws.cell(row, 13).value
                
                if not tipo_operacion:
                    continue
                
                # Parsear fecha y filtrar 2025
                fecha_dt, year = self._parse_fecha(concertacion)
                if year != 2025:
                    continue
                
                # Código limpio y convertir a número
                cod_clean = self._clean_codigo(cod_instrum)
                try:
                    cod_num = int(cod_clean) if cod_clean else None
                except (ValueError, TypeError):
                    cod_num = cod_clean
                
                # En rentas/dividendos de Visual la hoja de origen define el bucket ARS/USD.
                moneda_final = self._normalize_visual_rentas_currency(
                    visual_sheet_name,
                    moneda,
                    moneda_default,
                )
                
                bruto = abs(float(importe)) if importe else 0
                auditoria = f"Origen: Visual-{visual_sheet_name} | Cat: {categoria} | Op: {tipo_operacion}"
                
                # Cantidad siempre positiva
                cantidad_val = abs(float(cantidad)) if cantidad else 0
                
                all_rentas.append({
                    'tipo_instrumento_val': tipo_instrum,
                    'fecha': fecha_dt if fecha_dt else concertacion,
                    'liquidacion': liquidacion,
                    'numero': nro_ndc,
                    'moneda': moneda_final,
                    'operacion': str(tipo_operacion).upper() if tipo_operacion else "",
                    'cod_num': cod_num,
                    'especie': instrumento,
                    'cantidad': cantidad_val,
                    'precio': 1,  # Precio = 1 para rentas/dividendos de Visual
                    'bruto': bruto,
                    'interes': 0,
                    'gastos': abs(float(gastos)) if gastos else 0,
                    'costo': 0,
                    'origen': f"Visual-{visual_sheet_name}",
                    'is_amortizacion': False,
                    'auditoria': auditoria,
                })
        
        # Ordenar por fecha de concertación
        def sort_key(t):
            fecha = t.get('fecha')
            if isinstance(fecha, datetime):
                return fecha
            else:
                return datetime.min
        
        all_rentas.sort(key=sort_key)
        
        # Escribir transacciones ordenadas con fórmulas correctas
        for row_out, renta in enumerate(all_rentas, start=2):
            # Fórmulas con row_out correcto
            tipo_instrumento = renta['tipo_instrumento_val'] if renta['tipo_instrumento_val'] else f'=SI(ESERROR(BUSCARV(G{row_out};EspeciesVisual!C:R;16;FALSO));"";BUSCARV(G{row_out};EspeciesVisual!C:R;16;FALSO))'
            instrumento_con_moneda = f'=SI(ESERROR(BUSCARV(G{row_out};EspeciesVisual!C:Q;15;FALSO));"";BUSCARV(G{row_out};EspeciesVisual!C:Q;15;FALSO))'
            tipo_cambio = f'=SI(E{row_out}="Pesos";1;SI(ESERROR(BUSCARV(B{row_out};\'Cotizacion Dolar Historica\'!A:B;2;FALSO));0;BUSCARV(B{row_out};\'Cotizacion Dolar Historica\'!A:B;2;FALSO)))'
            moneda_emision = f'=SI(ESERROR(BUSCARV(G{row_out};EspeciesVisual!C:Q;5;FALSO));"";BUSCARV(G{row_out};EspeciesVisual!C:Q;5;FALSO))'
            
            # Neto calculado: M - N - O - P (siempre la misma fórmula)
            neto = f'=M{row_out}-N{row_out}-O{row_out}-P{row_out}'
            
            ws.cell(row_out, 1, tipo_instrumento)
            ws.cell(row_out, 2, renta['fecha'])
            ws.cell(row_out, 3, renta['liquidacion'])
            ws.cell(row_out, 4, renta['numero'])
            ws.cell(row_out, 5, renta['moneda'])
            ws.cell(row_out, 6, renta['operacion'])
            ws.cell(row_out, 7, renta['cod_num'])  # Forzado a número
            ws.cell(row_out, 8, renta['especie'])
            ws.cell(row_out, 9, instrumento_con_moneda)
            ws.cell(row_out, 10, renta['cantidad'])
            ws.cell(row_out, 11, renta['precio'])
            ws.cell(row_out, 12, tipo_cambio)
            ws.cell(row_out, 13, renta['bruto'])
            ws.cell(row_out, 14, renta['interes'])
            ws.cell(row_out, 15, renta['gastos'])
            ws.cell(row_out, 16, renta['costo'])
            ws.cell(row_out, 17, neto)
            ws.cell(row_out, 18, renta['origen'])
            ws.cell(row_out, 19, moneda_emision)
            ws.cell(row_out, 20, renta['auditoria'])
    
    def _create_resultado_ventas_ars(self, wb: Workbook):
        """Crea hoja Resultado Ventas ARS con transacciones de Boletos filtradas por Pesos."""
        ws = wb.create_sheet("Resultado Ventas ARS")
        use_precio_tenencias = bool(self.precio_tenencias_wb and 'PrecioTenenciasIniciales' in self.precio_tenencias_wb.sheetnames)
        
        # Headers (27 columnas - agregamos Precio Nominal)
        headers = ['Origen', 'Tipo de Instrumento', 'Instrumento', 'Cod.Instrum',
                   'Concertación', 'Liquidación', 'Moneda', 'Tipo Operación',
                   'Cantidad', 'Precio', 'Bruto', 'Interés', 'Tipo de Cambio',
                   'Gastos', 'IVA', 'Resultado', 'Cantidad Stock Inicial',
                   'Precio Stock Inicial', 'Costo por venta(gallo)', 'Neto Calculado(visual)',
                   'Resultado Calculado(final)', 'Cantidad de Stock Final', 
                   'Precio Stock Final', 'Explicación Q', 'Explicación R-U', 'chequeado', 'Precio Nominal']
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Recolectar transacciones de Boletos con moneda == "Pesos"
        boletos_ws = wb['Boletos']
        currency_overrides = self._build_resultado_currency_overrides(boletos_ws)
        visual_fx_map = self._build_visual_source_fx_map(boletos_ws)
        transactions = []
        
        for boletos_row in range(2, boletos_ws.max_row + 1):
            # Leer cod_instrum y buscar datos en cache
            cod_instrum = boletos_ws.cell(boletos_row, 7).value  # Col G
            cod_clean = self._clean_codigo(str(cod_instrum)) if cod_instrum else None
            
            # Obtener datos del cache
            especie_data = self._especies_visual_cache.get(cod_clean, {}) if cod_clean else {}
            moneda_emision = especie_data.get('moneda_emision')
            
            moneda = boletos_ws.cell(boletos_row, 5).value  # Col E
            tipo_operacion = boletos_ws.cell(boletos_row, 6).value  # Col F
            route_currency = currency_overrides.get(cod_clean)

            # Mantener todos los legs del mismo instrumento en una sola hoja
            if route_currency == 'USD':
                continue
            if route_currency is None and self._is_dollar_related(moneda, tipo_operacion, moneda_emision):
                continue
            if route_currency is None and moneda_emision != "Pesos":
                continue
            
            # Extraer valores REALES (no fórmulas)
            origen = boletos_ws.cell(boletos_row, 17).value  # Col Q - Origen (texto)
            
            # Tipo de instrumento: obtener del cache si es fórmula
            tipo_instrumento_cell = boletos_ws.cell(boletos_row, 1).value
            if isinstance(tipo_instrumento_cell, str) and tipo_instrumento_cell.startswith('='):
                tipo_instrumento = especie_data.get('tipo_especie', '')
            else:
                tipo_instrumento = tipo_instrumento_cell
            
            # Instrumento: obtener del cache si es fórmula
            instrumento_cell = boletos_ws.cell(boletos_row, 9).value
            if isinstance(instrumento_cell, str) and instrumento_cell.startswith('='):
                instrumento = especie_data.get('nombre_con_moneda', '')
            else:
                instrumento = instrumento_cell
            
            # Valores directos
            concertacion = boletos_ws.cell(boletos_row, 2).value  # Col B - fecha
            liquidacion = boletos_ws.cell(boletos_row, 3).value  # Col C
            cantidad = boletos_ws.cell(boletos_row, 10).value  # Col J
            precio = boletos_ws.cell(boletos_row, 11).value  # Col K
            interes = boletos_ws.cell(boletos_row, 14).value  # Col N
            gastos = boletos_ws.cell(boletos_row, 15).value  # Col O
            bruto_fuente = boletos_ws.cell(boletos_row, 13).value  # Col M
            neto_fuente = boletos_ws.cell(boletos_row, 16).value  # Col P
            tipo_cambio_boletos = boletos_ws.cell(boletos_row, 12).value  # Col L
            especie_raw = boletos_ws.cell(boletos_row, 8).value  # Col H - instrumento crudo
            
            # Calcular Bruto = Cantidad * Precio
            try:
                bruto = float(cantidad) * float(precio) if cantidad and precio else 0
            except:
                bruto = 0
            
            tipo_cambio = visual_fx_map.get(boletos_row)
            if tipo_cambio is None:
                tipo_cambio = self._to_float(tipo_cambio_boletos)
            if not tipo_cambio:
                tipo_cambio = 1
            
            transactions.append({
                'origen': origen,
                'tipo_instrumento': tipo_instrumento,
                'instrumento': instrumento if instrumento else especie_raw,
                'cod_instrum': cod_instrum,
                'concertacion': concertacion,
                'liquidacion': liquidacion,
                'moneda': moneda,
                'tipo_operacion': tipo_operacion,
                'cantidad': cantidad,
                'precio': precio,
                'bruto': bruto,
                'interes': interes if interes else 0,
                'tipo_cambio': tipo_cambio,
                'gastos': gastos if gastos else 0,
                '_idx': boletos_row,
            })

        def _to_sortable_date(value):
            if isinstance(value, datetime):
                return value
            if isinstance(value, date):
                return datetime.combine(value, datetime.min.time())
            if value is None:
                return datetime.min
            try:
                return datetime.fromisoformat(str(value))
            except Exception:
                return datetime.min

        def _sort_key(tx):
            cod = self._clean_codigo(str(tx.get('cod_instrum'))) if tx.get('cod_instrum') else ''
            concert = _to_sortable_date(tx.get('concertacion'))
            liquid = _to_sortable_date(tx.get('liquidacion'))
            # Buys (positive qty) before sells (negative qty) on same cod+date
            try:
                qty = float(tx.get('cantidad') or 0)
            except Exception:
                qty = 0
            buy_sell = 0 if qty > 0 else 1
            return (cod, concert, liquid, buy_sell, tx.get('_idx', 0))

        transactions.sort(key=_sort_key)
        
        def _si_error(expr: str, fallback: str) -> str:
            return f'SI(ESERROR({expr});{fallback};{expr})'

        def _si_error(expr: str, fallback: str) -> str:
            return f'SI(ESERROR({expr});{fallback};{expr})'

        # Escribir transacciones con VALORES (no fórmulas excepto para cálculos)
        for row_out, trans in enumerate(transactions, start=2):
            # Columnas A-N: Valores directos
            ws.cell(row_out, 1, trans['origen'])
            ws.cell(row_out, 2, trans['tipo_instrumento'])
            ws.cell(row_out, 3, trans['instrumento'])
            ws.cell(row_out, 4, trans['cod_instrum'])
            ws.cell(row_out, 5, trans['concertacion'])  # datetime
            ws.cell(row_out, 6, trans['liquidacion'])
            ws.cell(row_out, 7, trans['moneda'])
            ws.cell(row_out, 8, trans['tipo_operacion'])
            ws.cell(row_out, 9, trans['cantidad'])
            ws.cell(row_out, 10, trans['precio'])
            # Col K: Bruto = Cantidad * Precio Nominal (col AA) - FÓRMULA
            ws.cell(row_out, 11, f'=I{row_out}*AA{row_out}')
            ws.cell(row_out, 12, trans['interes'])
            ws.cell(row_out, 13, trans['tipo_cambio'])  # Valor 1, no fórmula
            ws.cell(row_out, 14, trans['gastos'])
            
            # Col O: IVA = SI(N>0, N*0.1736, N*-0.1736) basado en Gastos (col N)
            ws.cell(row_out, 15, f'=SI(N{row_out}>0;N{row_out}*0,1736;N{row_out}*-0,1736)')
            
            # Col P: Resultado (vacío por ahora)
            ws.cell(row_out, 16, "")
            
            # COLUMNAS Q-W: Fórmulas de Running Stock
            cod = trans['cod_instrum']
            origen_val = trans['origen'] or ""
            is_gallo = 'gallo' in origen_val.lower() if origen_val else False
            
            # Buscar cantidad y precio inicial desde cache de posición
            pos_inicial_cantidad = 0
            pos_inicial_precio = 0
            if cod:
                # Buscar en Posicion Inicial Gallo o Posicion Final según origen
                if is_gallo and 'Posicion Inicial Gallo' in wb.sheetnames:
                    pos_ws = wb['Posicion Inicial Gallo']
                    for r in range(2, pos_ws.max_row + 1):
                        if pos_ws.cell(r, 4).value == cod:  # Col D
                            pos_inicial_cantidad = pos_ws.cell(r, 9).value or 0  # Col I
                            pos_inicial_precio = pos_ws.cell(r, 16).value or 0  # Col P
                            break
                elif not is_gallo and 'Posicion Final Gallo' in wb.sheetnames:
                    pos_ws = wb['Posicion Final Gallo']
                    for r in range(2, pos_ws.max_row + 1):
                        if pos_ws.cell(r, 4).value == cod:  # Col D
                            pos_inicial_cantidad = pos_ws.cell(r, 9).value or 0  # Col I
                            pos_inicial_precio = pos_ws.cell(r, 16).value or 0  # Col P
                            break
            
            # Col Q: Cantidad Stock Inicial (siempre desde Posicion Inicial Gallo)
            if row_out == 2:
                ws.cell(row_out, 17, f'={_si_error(f"BUSCARV(D{row_out};\'Posicion Inicial Gallo\'!D:I;6;FALSO)", "0")}')
                # Col R: Precio Stock Inicial - Usa col V (19 desde D) = Precio Nominal
                # Con fallback a PrecioTenenciasIniciales y luego PreciosInicialesEspecies
                pos_lookup_gallo = _si_error(f"BUSCARV(D{row_out};'Posicion Inicial Gallo'!D:V;19;FALSO)", "0")
                if use_precio_tenencias:
                    fallback_precio = _si_error(
                        f"BUSCARV(D{row_out};PrecioTenenciasIniciales!A:G;7;FALSO)",
                        _si_error(f"BUSCARV(D{row_out};PreciosInicialesEspecies!A:I;9;FALSO)", "0")
                    )
                else:
                    fallback_precio = _si_error(f"BUSCARV(D{row_out};PreciosInicialesEspecies!A:I;9;FALSO)", "0")
                ws.cell(row_out, 18, f'=SI({pos_lookup_gallo}=0;{fallback_precio};{pos_lookup_gallo})')
                explicacion_q = f"BUSCARV(D{row_out}→Posicion Inicial Gallo col V=Precio Nominal, fallback PrecioTenenciasIniciales/PreciosInicialesEspecies)"
            else:
                prev = row_out - 1
                ws.cell(row_out, 17, f'=SI(D{row_out}=D{prev};V{prev};{_si_error(f"BUSCARV(D{row_out};\'Posicion Inicial Gallo\'!D:I;6;FALSO)", "0")})')
                # Col R: Con fallback
                pos_lookup_gallo = _si_error(f"BUSCARV(D{row_out};'Posicion Inicial Gallo'!D:V;19;FALSO)", "0")
                if use_precio_tenencias:
                    fallback_precio = _si_error(
                        f"BUSCARV(D{row_out};PrecioTenenciasIniciales!A:G;7;FALSO)",
                        _si_error(f"BUSCARV(D{row_out};PreciosInicialesEspecies!A:I;9;FALSO)", "0")
                    )
                else:
                    fallback_precio = _si_error(f"BUSCARV(D{row_out};PreciosInicialesEspecies!A:I;9;FALSO)", "0")
                ws.cell(row_out, 18, f'=SI(D{row_out}=D{prev};W{prev};SI({pos_lookup_gallo}=0;{fallback_precio};{pos_lookup_gallo}))')
                explicacion_q = f"SI D{row_out}=D{prev}: W{prev}, SINO: BUSCARV(D{row_out}→Posicion Inicial Gallo col V=Precio Nominal, fallback PrecioTenenciasIniciales/PreciosInicialesEspecies)"
            
            # Col S: Costo por venta = Cantidad * Precio Stock (si venta, cantidad < 0)
            ws.cell(row_out, 19, f'=SI(I{row_out}<0;I{row_out}*R{row_out};0)')
            
            # Col T: Neto Calculado = Bruto + Interés
            ws.cell(row_out, 20, f'=K{row_out}+L{row_out}')
            
            # Col U: Resultado Calculado = |Neto| - |Costo|
            ws.cell(row_out, 21, f'=SI(S{row_out}<>0;ABS(T{row_out})-ABS(S{row_out});0)')
            
            # Col V: Cantidad Stock Final = Cantidad + Stock Inicial
            ws.cell(row_out, 22, f'=I{row_out}+Q{row_out}')
            
            # Col W: Precio Stock Final (promedio ponderado si compra, mantiene si venta)
            # IMPORTANTE: Usar AA (Precio Nominal) en vez de J (Precio) para ON/TP/Letras
            ws.cell(row_out, 23, f'=SI(V{row_out}=0;0;SI(I{row_out}>0;SI((I{row_out}+Q{row_out})=0;0;(I{row_out}*AA{row_out}+Q{row_out}*R{row_out})/(I{row_out}+Q{row_out}));R{row_out}))')
            
            # Col X: Explicación Q (específica para esta fila)
            ws.cell(row_out, 24, explicacion_q)
            
            # Col Y: Explicación R-U (específica para esta fila)
            cantidad_val = trans['cantidad'] or 0
            explicacion_rstu = f"R=Precio stock previo | S={cantidad_val}*R{row_out}={cantidad_val}*[stock price] | T=K{row_out}+L{row_out}={trans['bruto']}+{trans['interes']} | U=|T{row_out}|-|S{row_out}|"
            ws.cell(row_out, 25, explicacion_rstu)
            
            # Col Z: Chequeado
            ws.cell(row_out, 26, f"Origen: {trans['origen']} | Cod: {cod}")
            
            # Col AA (27): Precio Nominal. En Visual ARS de renta fija/títulos ya nominales,
            # usar J directo; en el resto mantener /100 para instrumentos cotizados cada 100.
            ws.cell(row_out, 27, self._build_ars_nominal_formula(row_out))
    
    def _create_resultado_ventas_usd(self, wb: Workbook):
        """Crea hoja Resultado Ventas USD con transacciones de Boletos filtradas por Dolar."""
        ws = wb.create_sheet("Resultado Ventas USD")
        use_precio_tenencias = bool(self.precio_tenencias_wb and 'PrecioTenenciasIniciales' in self.precio_tenencias_wb.sheetnames)
        
        # Headers (29 columnas - agregamos Precio Nominal)
        headers = ['Origen', 'Tipo de Instrumento', 'Instrumento', 'Cod.Instrum',
                   'Concertación', 'Liquidación', 'Moneda', 'Tipo Operación',
                   'Cantidad', 'Precio', 'Precio Standarizado', 'Precio Standarizado en USD',
                   'Bruto en USD', 'Interés', 'Tipo de Cambio', 'Valor USD Dia',
               'Gastos USD', 'IVA USD', 'Resultado', 'Cantidad Stock Inicial',
                   'Precio Stock USD', 'Costo por venta(gallo)', 'Neto Calculado(visual)',
                   'Resultado Calculado(final)', 'Cantidad de Stock Final',
                   'Precio Stock Final', 'Explicación T-Z', 'Auditoría', 'Precio Nominal']
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        def _si_error(expr: str, fallback: str) -> str:
            return f'SI(ESERROR({expr});{fallback};{expr})'

        # Recolectar transacciones de Boletos con moneda contiene "Dolar"
        boletos_ws = wb['Boletos']
        currency_overrides = self._build_resultado_currency_overrides(boletos_ws)
        visual_fx_map = self._build_visual_source_fx_map(boletos_ws)
        transactions = []
        
        for boletos_row in range(2, boletos_ws.max_row + 1):
            # Leer cod_instrum y buscar datos en cache
            cod_instrum = boletos_ws.cell(boletos_row, 7).value  # Col G
            cod_clean = self._clean_codigo(str(cod_instrum)) if cod_instrum else None
            
            # Obtener datos del cache
            especie_data = self._especies_visual_cache.get(cod_clean, {}) if cod_clean else {}
            moneda_emision = especie_data.get('moneda_emision')
            
            moneda = boletos_ws.cell(boletos_row, 5).value  # Col E
            tipo_operacion = boletos_ws.cell(boletos_row, 6).value  # Col F
            route_currency = currency_overrides.get(cod_clean)

            # Filtrar operaciones dolarizadas manteniendo juntos todos los legs del código
            if route_currency == 'ARS':
                continue
            if route_currency is None and not self._is_dollar_related(moneda, tipo_operacion, moneda_emision):
                continue
            
            # Extraer valores REALES (no fórmulas)
            origen = boletos_ws.cell(boletos_row, 17).value  # Col Q - Origen (texto)
            
            # Tipo de instrumento: obtener del cache si es fórmula
            tipo_instrumento_cell = boletos_ws.cell(boletos_row, 1).value
            if isinstance(tipo_instrumento_cell, str) and tipo_instrumento_cell.startswith('='):
                tipo_instrumento = especie_data.get('tipo_especie', '')
            else:
                tipo_instrumento = tipo_instrumento_cell

            # Instrumento: obtener del cache si es fórmula
            instrumento_cell = boletos_ws.cell(boletos_row, 9).value
            if isinstance(instrumento_cell, str) and instrumento_cell.startswith('='):
                instrumento = especie_data.get('nombre_con_moneda', '')
            else:
                instrumento = instrumento_cell
            
            # Valores directos
            concertacion = boletos_ws.cell(boletos_row, 2).value  # Col B - fecha
            liquidacion = boletos_ws.cell(boletos_row, 3).value  # Col C
            cantidad = boletos_ws.cell(boletos_row, 10).value  # Col J
            precio = boletos_ws.cell(boletos_row, 11).value  # Col K
            interes = boletos_ws.cell(boletos_row, 14).value  # Col N
            gastos = boletos_ws.cell(boletos_row, 15).value  # Col O
            bruto_fuente = boletos_ws.cell(boletos_row, 13).value  # Col M
            neto_fuente = boletos_ws.cell(boletos_row, 16).value  # Col P
            tipo_cambio_boletos = boletos_ws.cell(boletos_row, 12).value  # Col L
            especie_raw = boletos_ws.cell(boletos_row, 8).value  # Col H - instrumento crudo
            
            tipo_cambio = visual_fx_map.get(boletos_row)
            if tipo_cambio is None:
                tipo_cambio = self._to_float(tipo_cambio_boletos)
            if not tipo_cambio and isinstance(concertacion, datetime):
                fecha_key = concertacion.strftime('%Y-%m-%d') if hasattr(concertacion, 'strftime') else str(concertacion)
                tipo_cambio = self._cotizacion_cache.get(fecha_key, 1) or 1
            if not tipo_cambio:
                tipo_cambio = 1
            
            transactions.append({
                'origen': origen,
                'tipo_instrumento': tipo_instrumento,
                'instrumento': instrumento if instrumento else especie_raw,
                'cod_instrum': cod_instrum,
                'concertacion': concertacion,
                'liquidacion': liquidacion,
                'moneda': moneda,
                'tipo_operacion': tipo_operacion,
                'cantidad': cantidad,
                'precio': precio,
                'interes': interes if interes else 0,
                'tipo_cambio': tipo_cambio,
                'gastos': gastos if gastos else 0,
                'bruto_fuente': bruto_fuente,
                'neto_fuente': neto_fuente,
                '_idx': boletos_row,
            })

        def _to_sortable_date(value):
            if isinstance(value, datetime):
                return value
            if isinstance(value, date):
                return datetime.combine(value, datetime.min.time())
            if value is None:
                return datetime.min
            try:
                return datetime.fromisoformat(str(value))
            except Exception:
                return datetime.min

        def _sort_key(tx):
            cod = self._clean_codigo(str(tx.get('cod_instrum'))) if tx.get('cod_instrum') else ''
            concert = _to_sortable_date(tx.get('concertacion'))
            liquid = _to_sortable_date(tx.get('liquidacion'))
            # Buys (positive qty) before sells (negative qty) on same cod+date
            try:
                qty = float(tx.get('cantidad') or 0)
            except Exception:
                qty = 0
            buy_sell = 0 if qty > 0 else 1
            return (cod, concert, liquid, buy_sell, tx.get('_idx', 0))

        transactions.sort(key=_sort_key)
        
        # Escribir transacciones con VALORES (no fórmulas excepto para cálculos)
        for row_out, trans in enumerate(transactions, start=2):
            origen_val = trans['origen'] or ""
            is_visual = 'visual' in origen_val.lower() if origen_val else False
            is_gallo = 'gallo' in origen_val.lower() if origen_val else False
            tipo_instrumento_val = str(trans['tipo_instrumento'] or '')
            precio_val = trans['precio'] or 0
            requires_standard_100 = self._should_standardize_visual_usd_price(
                precio_val,
                tipo_instrumento_val,
                origen_val,
                trans['moneda'],
            )
            
            # Columnas A-J: Valores directos
            ws.cell(row_out, 1, trans['origen'])
            ws.cell(row_out, 2, trans['tipo_instrumento'])
            ws.cell(row_out, 3, trans['instrumento'])
            ws.cell(row_out, 4, trans['cod_instrum'])
            ws.cell(row_out, 5, trans['concertacion'])  # datetime
            ws.cell(row_out, 6, trans['liquidacion'])
            ws.cell(row_out, 7, trans['moneda'])
            ws.cell(row_out, 8, trans['tipo_operacion'])
            ws.cell(row_out, 9, trans['cantidad'])
            ws.cell(row_out, 10, trans['precio'])
            
            # Col K: Precio Standarizado (x100 si Visual)
            try:
                precio_std = float(precio_val) * 100 if requires_standard_100 else float(precio_val)
            except:
                precio_std = 0
            ws.cell(row_out, 11, precio_std)  # Valor, no fórmula
            
            # Col L: Precio Standarizado en USD = K * O (Precio Std * Tipo Cambio)
            # O = 1 si moneda incluye "dolar", sino 1/P
            ws.cell(row_out, 12, f'=K{row_out}*O{row_out}')
            
            # Col M: Bruto en USD = Cantidad * Precio Nominal (col AC=29)
            bruto_fuente = self._to_float(trans.get('bruto_fuente'))
            bruto_calc = self._to_float(trans.get('cantidad')) * self._to_float(precio_val)
            preserve_micro_source = self._should_preserve_visual_usd_micro_source_money(
                origen_val,
                trans['moneda'],
                tipo_instrumento_val,
                precio_val,
                bruto_fuente,
                bruto_calc,
            )
            if preserve_micro_source:
                ws.cell(row_out, 13, bruto_fuente)
            else:
                ws.cell(row_out, 13, f'=I{row_out}*AC{row_out}')
            
            # Col N: Interés
            ws.cell(row_out, 14, trans['interes'])
            
            # Col O: Tipo de Cambio - Si moneda incluye "dolar" → 1, sino → 1/P (Valor USD Dia)
            moneda_val = trans['moneda'] or ""
            if 'dolar' in str(moneda_val).lower():
                ws.cell(row_out, 15, 1)  # Operaciones en dólares: tipo cambio = 1
            else:
                ws.cell(row_out, 15, f'=SI(P{row_out}=0;1;1/P{row_out})')  # Pesos: 1/ValorUSDDia
            
            # Col P: Valor USD Dia - preferir TC bruto de Visual cuando exista; si no, usar histórico.
            visual_tc = self._meaningful_fx_rate(trans.get('tipo_cambio'))
            if self._is_visual_origin(origen_val) and visual_tc > 0:
                ws.cell(row_out, 16, visual_tc)
            else:
                ws.cell(row_out, 16, f'={_si_error(f"BUSCARV(E{row_out};'Cotizacion Dolar Historica'!A:B;2;FALSO)", "0")}')
            
            gastos_fuente_formula = self._fmt_num_es(self._to_float(trans['gastos']))

            # Col Q: Gastos USD visibles, calculados desde el gasto fuente y el factor O.
            ws.cell(row_out, 17, f'=ABS({gastos_fuente_formula}*O{row_out})')

            # Col R: IVA USD = Q * 0.1736
            ws.cell(row_out, 18, f'=Q{row_out}*0,1736')
            
            # Col S: Resultado (vacío)
            ws.cell(row_out, 19, "")
            
            # COLUMNAS T-Z: Fórmulas de Running Stock
            cod = trans['cod_instrum']
            
            # Col T: Cantidad Stock Inicial - siempre desde Posicion Inicial Gallo
            if row_out == 2:
                ws.cell(row_out, 20, f'={_si_error(f"BUSCARV(D{row_out};\'Posicion Inicial Gallo\'!D:I;6;FALSO)", "0")}')
                # Col U: Precio Stock USD
                # Primero intenta VLOOKUP a Posicion / cotización día
                # Si es 0, usa fallback a PrecioTenenciasIniciales y luego PreciosInicialesEspecies
                pos_lookup_gallo = _si_error(f"BUSCARV(D{row_out};'Posicion Inicial Gallo'!D:V;19;FALSO)", "0")
                if use_precio_tenencias:
                    fallback_precio = _si_error(
                        f"BUSCARV(D{row_out};PrecioTenenciasIniciales!A:G;7;FALSO)",
                        _si_error(f"BUSCARV(D{row_out};PreciosInicialesEspecies!A:J;10;FALSO)", "0")
                    )
                else:
                    fallback_precio = _si_error(f"BUSCARV(D{row_out};PreciosInicialesEspecies!A:J;10;FALSO)", "0")
                # Fórmula: SI(P=0,0; usa precio fallback o posición ya expresado en USD nominal)
                ws.cell(row_out, 21, f'=SI(P{row_out}=0;0;SI({pos_lookup_gallo}=0;{fallback_precio};{pos_lookup_gallo}/P{row_out}))')
                explicacion_t = f"T=BUSCARV(D{row_out}→Posicion Inicial Gallo col V=Precio Nominal, fallback PrecioTenenciasIniciales/PreciosInicialesEspecies)"
            else:
                prev = row_out - 1
                ws.cell(row_out, 20, f'=SI(D{row_out}=D{prev};Y{prev};{_si_error(f"BUSCARV(D{row_out};\'Posicion Inicial Gallo\'!D:I;6;FALSO)", "0")})')
                # Col U: Con fallback
                pos_lookup_gallo = _si_error(f"BUSCARV(D{row_out};'Posicion Inicial Gallo'!D:V;19;FALSO)", "0")
                if use_precio_tenencias:
                    fallback_precio = _si_error(
                        f"BUSCARV(D{row_out};PrecioTenenciasIniciales!A:G;7;FALSO)",
                        _si_error(f"BUSCARV(D{row_out};PreciosInicialesEspecies!A:J;10;FALSO)", "0")
                    )
                else:
                    fallback_precio = _si_error(f"BUSCARV(D{row_out};PreciosInicialesEspecies!A:J;10;FALSO)", "0")
                ws.cell(row_out, 21, f'=SI(D{row_out}=D{prev};Z{prev};SI(P{row_out}=0;0;SI({pos_lookup_gallo}=0;{fallback_precio};{pos_lookup_gallo}/P{row_out})))')
                explicacion_t = f"SI D{row_out}=D{prev}: Z{prev}, SINO: BUSCARV(col V=Precio Nominal (Posicion Inicial Gallo), fallback PrecioTenenciasIniciales/PreciosInicialesEspecies)"
            
            # Col V: Costo por venta = Cantidad * Precio Stock USD (si venta)
            ws.cell(row_out, 22, f'=SI(I{row_out}<0;I{row_out}*U{row_out};0)')
            
            # Col W: Neto Calculado = Bruto USD +/- Gastos USD según signo económico
            ws.cell(row_out, 23, f'=SI(I{row_out}<0;M{row_out}-Q{row_out};M{row_out}+Q{row_out})')
            
            # Col X: Resultado Calculado = |Neto| - |Costo|
            ws.cell(row_out, 24, f'=SI(V{row_out}<>0;ABS(W{row_out})-ABS(V{row_out});0)')
            
            # Col Y: Cantidad Stock Final = Cantidad + Stock Inicial
            ws.cell(row_out, 25, f'=I{row_out}+T{row_out}')
            
            # Col Z: Precio Stock Final (promedio ponderado)
            # IMPORTANTE: Usar AC (Precio Nominal) en vez de L (Precio Std USD) para ON/TP/Letras
            ws.cell(row_out, 26, f'=SI(Y{row_out}=0;0;SI(I{row_out}>0;SI((I{row_out}+T{row_out})=0;0;(I{row_out}*AC{row_out}+T{row_out}*U{row_out})/(I{row_out}+T{row_out}));U{row_out}))')
            
            # Col AA: Explicación T-Z
            cantidad_val = trans['cantidad'] or 0
            explicacion_full = f"{explicacion_t} | U=PrecioPos/P{row_out} | V={cantidad_val}*U si venta | Q=GastosUSD | W=M +/- Q segun signo | X=|W|-|V| | Y=I+T | Z=Promedio"
            ws.cell(row_out, 27, explicacion_full)
            
            # Col AB: Auditoría
            ws.cell(row_out, 28, f"Origen: {trans['origen']} | Cod: {cod} | K(PrecioStd)={'x100' if requires_standard_100 else 'raw'} | L=K*O | GastosFuente={trans['gastos']}")
            
            # Col AC (29): Precio Nominal = Precio Standarizado en USD (L) /100 si es ON, Títulos Públicos o Letras
            if preserve_micro_source:
                ws.cell(row_out, 29, f'=SI(I{row_out}=0;0;ABS(M{row_out}/I{row_out}))')
            else:
                ws.cell(row_out, 29, f'=SI(O(ESNUMERO(HALLAR("Obligacion";B{row_out}));ESNUMERO(HALLAR("Titulo";B{row_out}));ESNUMERO(HALLAR("Título";B{row_out}));ESNUMERO(HALLAR("Letra";B{row_out})));SI(ABS(L{row_out})>=10;L{row_out}/100;L{row_out});L{row_out})')
    
    def _create_rentas_dividendos_ars(self, wb: Workbook):
        """Crea hoja Rentas Dividendos ARS con valores reales filtrados y ordenados.
        
        NOTA: Rentas y dividendos siempre son ganancias, por lo que:
        - Cantidad siempre positiva (especialmente amortizaciones)
        - Importe siempre positivo
        - Se guarda el importe original en columna O como referencia
        """
        ws = wb.create_sheet("Rentas Dividendos ARS")
        
        headers = ['Instrumento', 'Cod.Instrum', 'Categoría', 'tipo_instrumento',
                   'Concertación', 'Liquidación', 'Nro. NDC', 'Tipo Operación',
                   'Cantidad', 'Moneda', 'Tipo de Cambio', 'Gastos', 'Importe', 'Origen',
                   'Importe Original']  # Col O = referencia del importe original
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Recolectar datos de Rentas y Dividendos Gallo
        rentas_ws = wb['Rentas y Dividendos Gallo']
        transactions = []
        
        for rentas_row in range(2, rentas_ws.max_row + 1):
            # Obtener cod_instrum y buscar moneda_emision en cache
            cod_instrum = rentas_ws.cell(rentas_row, 7).value  # Col G
            cod_clean = self._clean_codigo(str(cod_instrum)) if cod_instrum else None
            
            # Obtener moneda_emision del cache
            especie_data = self._especies_visual_cache.get(cod_clean, {}) if cod_clean else {}
            moneda_emision = especie_data.get('moneda_emision')
            moneda = rentas_ws.cell(rentas_row, 5).value  # Col E
            
            # Filtrar solo ARS usando la moneda efectiva de la fila como verdad principal
            if self._classify_rentas_currency(moneda, moneda_emision, rentas_ws.cell(rentas_row, 18).value) != 'ARS':
                continue
            
            tipo_operacion = rentas_ws.cell(rentas_row, 6).value  # Col F
            instrumento = rentas_ws.cell(rentas_row, 8).value  # Col H = Instrumento Crudo
            
            # Determinar Categoría basado en tipo operación
            tipo_op_upper = str(tipo_operacion).upper() if tipo_operacion else ""
            # Incluir ambas variantes: con y sin tilde
            if tipo_op_upper in ["RENTA", "AMORTIZACION", "AMORTIZACIÓN"]:
                categoria = "Rentas"
            else:
                categoria = "Dividendos"
            
            tipo_instrumento = rentas_ws.cell(rentas_row, 1).value  # Col A (puede ser fórmula)
            if isinstance(tipo_instrumento, str) and tipo_instrumento.startswith('='):
                tipo_instrumento = especie_data.get('tipo_especie', '')
            
            concertacion = rentas_ws.cell(rentas_row, 2).value  # Col B
            liquidacion = rentas_ws.cell(rentas_row, 3).value  # Col C
            nro_ndc = rentas_ws.cell(rentas_row, 4).value  # Col D
            cantidad = rentas_ws.cell(rentas_row, 10).value  # Col J
            # Calcular gastos = Costo (P) + Gastos (O)
            costo = rentas_ws.cell(rentas_row, 16).value or 0  # Col P
            gastos_orig = rentas_ws.cell(rentas_row, 15).value or 0  # Col O
            gastos = (costo if isinstance(costo, (int, float)) else 0) + (gastos_orig if isinstance(gastos_orig, (int, float)) else 0)
            
            # Importe = Resultado (M) - Gastos (O) - Costo (P) como valor, no fórmula
            resultado = rentas_ws.cell(rentas_row, 13).value or 0  # Col M = Resultado
            if isinstance(resultado, (int, float)) and isinstance(gastos_orig, (int, float)):
                importe = resultado - gastos_orig - (costo if isinstance(costo, (int, float)) else 0)
            else:
                importe = 0
            # Amortizaciones no afectan el resumen anual
            if tipo_op_upper in ["AMORTIZACION", "AMORTIZACIÓN"]:
                importe = 0
            
            origen = rentas_ws.cell(rentas_row, 18).value  # Col R
            if moneda_emision and self._classify_rentas_currency(moneda, moneda_emision, origen) == 'ARS' and 'dolar' in str(moneda_emision).lower():
                origen = f"{origen} | ALERTA: MONEDA_EFECTIVA_ARS"
            
            # Tipo de Cambio ARS: 1 si Pesos, Valor USD del día si contiene dolar
            moneda_str = str(moneda).lower() if moneda else ""
            if moneda == "Pesos" or "peso" in moneda_str:
                tipo_cambio = 1
            elif "dolar" in moneda_str or "cable" in moneda_str:
                # Buscar valor USD del día
                fecha_conc = concertacion if isinstance(concertacion, datetime) else None
                tipo_cambio = self._get_cotizacion(fecha_conc, "Dolar MEP") if fecha_conc else 1
            else:
                tipo_cambio = 1
            
            # Solo agregar si tiene datos válidos
            if not instrumento and not cod_instrum:
                continue
            
            transactions.append({
                'instrumento': instrumento,
                'cod_instrum': cod_instrum,
                'categoria': categoria,
                'tipo_instrumento': tipo_instrumento,
                'concertacion': concertacion,
                'liquidacion': liquidacion,
                'nro_ndc': nro_ndc,
                'tipo_operacion': tipo_operacion,
                'cantidad': cantidad,
                'moneda': moneda,
                'tipo_cambio': tipo_cambio,
                'gastos': gastos,
                'importe': importe,
                'origen': origen,
            })
        
        # Ordenar por cod_instrum y luego por concertación
        def sort_key(t):
            cod = t.get('cod_instrum') or 0
            try:
                cod_num = int(cod)
            except:
                cod_num = 0
            fecha = t.get('concertacion')
            if isinstance(fecha, datetime):
                return (cod_num, fecha)
            else:
                return (cod_num, datetime.min)
        
        transactions.sort(key=sort_key)
        
        # Escribir transacciones
        for row_out, trans in enumerate(transactions, start=2):
            ws.cell(row_out, 1, trans['instrumento'])
            ws.cell(row_out, 2, trans['cod_instrum'])
            ws.cell(row_out, 3, trans['categoria'])
            ws.cell(row_out, 4, trans['tipo_instrumento'])
            ws.cell(row_out, 5, trans['concertacion'])  # datetime
            ws.cell(row_out, 6, trans['liquidacion'])
            ws.cell(row_out, 7, trans['nro_ndc'])
            ws.cell(row_out, 8, trans['tipo_operacion'])
            # Cantidad siempre positiva (especialmente para amortizaciones)
            cantidad_val = trans['cantidad']
            if isinstance(cantidad_val, (int, float)):
                cantidad_val = abs(cantidad_val)
            ws.cell(row_out, 9, cantidad_val)
            ws.cell(row_out, 10, trans['moneda'])
            ws.cell(row_out, 11, trans['tipo_cambio'])
            ws.cell(row_out, 12, trans['gastos'])
            # Importe siempre positivo (rentas/dividendos son ganancias)
            importe_val = trans['importe']
            importe_original = importe_val  # Guardar original
            if isinstance(importe_val, (int, float)):
                importe_val = abs(importe_val)
            ws.cell(row_out, 13, importe_val)
            ws.cell(row_out, 14, trans['origen'])
            ws.cell(row_out, 15, importe_original)  # Col O = Importe Original
    
    def _create_rentas_dividendos_usd(self, wb: Workbook):
        """Crea hoja Rentas Dividendos USD con valores reales filtrados y ordenados.
        
        NOTA: Rentas y dividendos siempre son ganancias, por lo que:
        - Cantidad siempre positiva (especialmente amortizaciones)
        - Importe siempre positivo
        - Se guarda el importe original en columna O como referencia
        """
        ws = wb.create_sheet("Rentas Dividendos USD")
        
        headers = ['Instrumento', 'Cod.Instrum', 'Categoría', 'tipo_instrumento',
                   'Concertación', 'Liquidación', 'Nro. NDC', 'Tipo Operación',
                   'Cantidad', 'Moneda', 'Tipo de Cambio', 'Gastos', 'Importe', 'Origen',
                   'Importe Original']  # Col O = referencia del importe original
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Recolectar datos de Rentas y Dividendos Gallo
        rentas_ws = wb['Rentas y Dividendos Gallo']
        transactions = []
        
        for rentas_row in range(2, rentas_ws.max_row + 1):
            # Obtener cod_instrum y buscar moneda_emision en cache
            cod_instrum = rentas_ws.cell(rentas_row, 7).value  # Col G
            cod_clean = self._clean_codigo(str(cod_instrum)) if cod_instrum else None
            
            # Obtener moneda_emision del cache
            especie_data = self._especies_visual_cache.get(cod_clean, {}) if cod_clean else {}
            moneda_emision = especie_data.get('moneda_emision')
            moneda = rentas_ws.cell(rentas_row, 5).value  # Col E
            
            # Filtrar solo USD usando la moneda efectiva de la fila como verdad principal
            if self._classify_rentas_currency(moneda, moneda_emision, rentas_ws.cell(rentas_row, 18).value) != 'USD':
                continue
            
            tipo_operacion = rentas_ws.cell(rentas_row, 6).value  # Col F
            instrumento = rentas_ws.cell(rentas_row, 8).value  # Col H = Instrumento Crudo
            
            # Determinar Categoría basado en tipo operación
            tipo_op_upper = str(tipo_operacion).upper() if tipo_operacion else ""
            # Incluir ambas variantes: con y sin tilde
            if tipo_op_upper in ["RENTA", "AMORTIZACION", "AMORTIZACIÓN"]:
                categoria = "Rentas"
            else:
                categoria = "Dividendos"
            
            tipo_instrumento = rentas_ws.cell(rentas_row, 1).value  # Col A (puede ser fórmula)
            if isinstance(tipo_instrumento, str) and tipo_instrumento.startswith('='):
                tipo_instrumento = especie_data.get('tipo_especie', '')
            
            concertacion = rentas_ws.cell(rentas_row, 2).value  # Col B
            liquidacion = rentas_ws.cell(rentas_row, 3).value  # Col C
            nro_ndc = rentas_ws.cell(rentas_row, 4).value  # Col D
            cantidad = rentas_ws.cell(rentas_row, 10).value  # Col J
            # Calcular gastos = Costo (P) + Gastos (O)
            costo = rentas_ws.cell(rentas_row, 16).value or 0  # Col P
            gastos_orig = rentas_ws.cell(rentas_row, 15).value or 0  # Col O
            gastos = (costo if isinstance(costo, (int, float)) else 0) + (gastos_orig if isinstance(gastos_orig, (int, float)) else 0)
            
            # Importe = Resultado (M) - Gastos (O) - Costo (P) como valor, no fórmula
            resultado = rentas_ws.cell(rentas_row, 13).value or 0  # Col M = Resultado
            if isinstance(resultado, (int, float)) and isinstance(gastos_orig, (int, float)):
                importe = resultado - gastos_orig - (costo if isinstance(costo, (int, float)) else 0)
            else:
                importe = 0
            # Amortizaciones no afectan el resumen anual
            if tipo_op_upper in ["AMORTIZACION", "AMORTIZACIÓN"]:
                importe = 0
            
            origen = rentas_ws.cell(rentas_row, 18).value  # Col R
            if moneda and ('peso' in str(moneda).lower()):
                origen = f"{origen} | ALERTA: MONEDA_EFECTIVA_PESOS_EN_USD"
            
            # Tipo de Cambio USD: 1 si contiene dolar/cable, Valor USD del día si Pesos
            moneda_str = str(moneda).lower() if moneda else ""
            if "dolar" in moneda_str or "cable" in moneda_str:
                tipo_cambio = 1
            elif moneda == "Pesos" or "peso" in moneda_str:
                # Buscar valor USD del día
                fecha_conc = concertacion if isinstance(concertacion, datetime) else None
                tipo_cambio = self._get_cotizacion(fecha_conc, "Dolar MEP") if fecha_conc else 1
            else:
                tipo_cambio = 1
            
            # Solo agregar si tiene datos válidos
            if not instrumento and not cod_instrum:
                continue
            
            transactions.append({
                'instrumento': instrumento,
                'cod_instrum': cod_instrum,
                'categoria': categoria,
                'tipo_instrumento': tipo_instrumento,
                'concertacion': concertacion,
                'liquidacion': liquidacion,
                'nro_ndc': nro_ndc,
                'tipo_operacion': tipo_operacion,
                'cantidad': cantidad,
                'moneda': moneda,
                'tipo_cambio': tipo_cambio,
                'gastos': gastos,
                'importe': importe,
                'origen': origen,
            })
        
        # Ordenar por cod_instrum y luego por concertación
        def sort_key(t):
            cod = t.get('cod_instrum') or 0
            try:
                cod_num = int(cod)
            except:
                cod_num = 0
            fecha = t.get('concertacion')
            if isinstance(fecha, datetime):
                return (cod_num, fecha)
            else:
                return (cod_num, datetime.min)
        
        transactions.sort(key=sort_key)
        
        # Escribir transacciones
        for row_out, trans in enumerate(transactions, start=2):
            ws.cell(row_out, 1, trans['instrumento'])
            ws.cell(row_out, 2, trans['cod_instrum'])
            ws.cell(row_out, 3, trans['categoria'])
            ws.cell(row_out, 4, trans['tipo_instrumento'])
            ws.cell(row_out, 5, trans['concertacion'])  # datetime
            ws.cell(row_out, 6, trans['liquidacion'])
            ws.cell(row_out, 7, trans['nro_ndc'])
            ws.cell(row_out, 8, trans['tipo_operacion'])
            # Cantidad siempre positiva (especialmente para amortizaciones)
            cantidad_val = trans['cantidad']
            if isinstance(cantidad_val, (int, float)):
                cantidad_val = abs(cantidad_val)
            ws.cell(row_out, 9, cantidad_val)
            ws.cell(row_out, 10, trans['moneda'])
            ws.cell(row_out, 11, trans['tipo_cambio'])
            ws.cell(row_out, 12, trans['gastos'])
            # Importe siempre positivo (rentas/dividendos son ganancias)
            importe_val = trans['importe']
            importe_original = importe_val  # Guardar original
            if isinstance(importe_val, (int, float)):
                importe_val = abs(importe_val)
            ws.cell(row_out, 13, importe_val)
            ws.cell(row_out, 14, trans['origen'])
            ws.cell(row_out, 15, importe_original)  # Col O = Importe Original
    
    def _create_resumen(self, wb: Workbook):
        """Crea hoja Resumen con totales por fórmula."""
        ws = wb.create_sheet("Resumen")
        
        headers = ['Moneda', 'Ventas', 'FCI', 'Opciones', 'Rentas', 'Dividendos',
               'Pagare/CPD', 'Futuros', 'Cau (Tom)', 'Cau (Col)', 'Total']
        
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)
        
        # Fila ARS
        ws.cell(2, 1, "ARS")
        ws.cell(2, 2, "=SUM('Resultado Ventas ARS'!U:U)")
        ws.cell(2, 3, '=SUMIF(FCI!C:C,"*Peso*",FCI!K:K)+SUMIF(FCI!C:C,"ARS",FCI!K:K)')
        ws.cell(2, 4, '=SUMIF(Opciones!C:C,"*Peso*",Opciones!K:K)+SUMIF(Opciones!C:C,"ARS",Opciones!K:K)')
        ws.cell(2, 5, "=SUMAR.SI.CONJUNTO('Rentas Dividendos ARS'!M:M;'Rentas Dividendos ARS'!C:C;\"Rentas\";'Rentas Dividendos ARS'!J:J;\"*Peso*\")+SUMAR.SI.CONJUNTO('Rentas Dividendos ARS'!M:M;'Rentas Dividendos ARS'!C:C;\"Rentas\";'Rentas Dividendos ARS'!J:J;\"ARS\")")
        ws.cell(2, 6, "=SUMAR.SI.CONJUNTO('Rentas Dividendos ARS'!M:M;'Rentas Dividendos ARS'!C:C;\"Dividendos\";'Rentas Dividendos ARS'!J:J;\"*Peso*\")+SUMAR.SI.CONJUNTO('Rentas Dividendos ARS'!M:M;'Rentas Dividendos ARS'!C:C;\"Dividendos\";'Rentas Dividendos ARS'!J:J;\"ARS\")")
        ws.cell(2, 7, '=SUMIF(Pagare_CPD!G:G,"*Peso*",Pagare_CPD!M:M)+SUMIF(Pagare_CPD!G:G,"ARS",Pagare_CPD!M:M)')
        ws.cell(2, 8, '=SUMIF(Futuros!A:A,"ARS",Futuros!D:D)+SUMIF(Futuros!A:A,"*Peso*",Futuros!D:D)')
        ws.cell(2, 9, "=SUMIF('Cauciones Tomadoras'!O:O,\"Pesos\",'Cauciones Tomadoras'!N:N)")
        ws.cell(2, 10, "=SUMIF('Cauciones Colocadoras'!O:O,\"Pesos\",'Cauciones Colocadoras'!N:N)")
        ws.cell(2, 11, "=B2+C2+D2+E2+F2+G2+H2+I2+J2")
        
        # Fila USD
        ws.cell(3, 1, "USD")
        ws.cell(3, 2, "=SUM('Resultado Ventas USD'!X:X)")
        ws.cell(3, 3, '=SUMIF(FCI!C:C,"*Dolar*",FCI!K:K)+SUMIF(FCI!C:C,"USD",FCI!K:K)')
        ws.cell(3, 4, '=SUMIF(Opciones!C:C,"*Dolar*",Opciones!K:K)+SUMIF(Opciones!C:C,"USD",Opciones!K:K)')
        ws.cell(3, 5, "=SUMAR.SI.CONJUNTO('Rentas Dividendos USD'!M:M;'Rentas Dividendos USD'!C:C;\"Rentas\";'Rentas Dividendos USD'!J:J;\"*Dolar*\")+SUMAR.SI.CONJUNTO('Rentas Dividendos USD'!M:M;'Rentas Dividendos USD'!C:C;\"Rentas\";'Rentas Dividendos USD'!J:J;\"USD\")")
        ws.cell(3, 6, "=SUMAR.SI.CONJUNTO('Rentas Dividendos USD'!M:M;'Rentas Dividendos USD'!C:C;\"Dividendos\";'Rentas Dividendos USD'!J:J;\"*Dolar*\")+SUMAR.SI.CONJUNTO('Rentas Dividendos USD'!M:M;'Rentas Dividendos USD'!C:C;\"Dividendos\";'Rentas Dividendos USD'!J:J;\"USD\")")
        ws.cell(3, 7, '=SUMIF(Pagare_CPD!G:G,"*Dolar*",Pagare_CPD!M:M)+SUMIF(Pagare_CPD!G:G,"USD",Pagare_CPD!M:M)')
        ws.cell(3, 8, '=SUMIF(Futuros!A:A,"USD",Futuros!D:D)+SUMIF(Futuros!A:A,"*Dolar*",Futuros!D:D)')
        ws.cell(3, 9, "=SUMIF('Cauciones Tomadoras'!O:O,\"*Dolar*\",'Cauciones Tomadoras'!N:N)")
        ws.cell(3, 10, "=SUMIF('Cauciones Colocadoras'!O:O,\"*Dolar*\",'Cauciones Colocadoras'!N:N)")
        ws.cell(3, 11, "=B3+C3+D3+E3+F3+G3+H3+I3+J3")
    
    def _sum_column(self, wb: Workbook, sheet_name: str, col: int, moneda_filter: str = None) -> float:
        """Suma una columna de una hoja, opcionalmente filtrando por moneda."""
        if sheet_name not in wb.sheetnames:
            return 0
        
        ws = wb[sheet_name]
        total = 0
        
        # Buscar columna de moneda (usualmente la 15 o la última)
        moneda_col = None
        if moneda_filter:
            for c in range(1, ws.max_column + 1):
                header = ws.cell(1, c).value
                if header and 'moneda' in str(header).lower():
                    moneda_col = c
                    break
        
        for row in range(2, ws.max_row + 1):
            # Filtrar por moneda si se especifica
            if moneda_filter and moneda_col:
                moneda_val = str(ws.cell(row, moneda_col).value or '').lower()
                if moneda_filter.lower() not in moneda_val:
                    continue
            
            val = ws.cell(row, col).value
            if val and isinstance(val, (int, float)):
                total += val
        
        return total
    
    def _sum_by_tipo(self, wb: Workbook, sheet_name: str, tipo_col: int, value_col: int, tipos: list, moneda_filter: str = None) -> float:
        """Suma valores de una columna filtrando por tipo."""
        if sheet_name not in wb.sheetnames:
            return 0
        
        ws = wb[sheet_name]
        total = 0
        moneda_col = None
        if moneda_filter:
            for c in range(1, ws.max_column + 1):
                header = ws.cell(1, c).value
                if header and 'moneda' in str(header).lower():
                    moneda_col = c
                    break
        
        for row in range(2, ws.max_row + 1):
            tipo = str(ws.cell(row, tipo_col).value or '').upper()
            if any(t.upper() in tipo for t in tipos):
                if moneda_filter and moneda_col:
                    moneda_val = str(ws.cell(row, moneda_col).value or '').lower()
                    if moneda_filter.lower() == 'ars':
                        if 'peso' not in moneda_val and moneda_val != 'ars':
                            continue
                    elif moneda_filter.lower() == 'usd':
                        if 'dolar' not in moneda_val and moneda_val != 'usd':
                            continue
                val = ws.cell(row, value_col).value
                if val and isinstance(val, (int, float)):
                    total += val
        
        return total
    
    def _calculate_ventas_real(self, wb: Workbook, sheet_name: str) -> float:
        """Calcula el resultado real de ventas usando running stock.
        
        Implementa el cálculo de running stock para obtener el precio promedio
        ponderado de compra y calcular el resultado real de cada venta:
        Resultado = Neto (Bruto + Interés) - Costo (Cantidad * Precio promedio)
        """
        from collections import defaultdict
        
        if sheet_name not in wb.sheetnames:
            return 0
        
        ws = wb[sheet_name]
        
        # Agrupar transacciones por código de instrumento
        transacciones_por_cod = defaultdict(list)
        
        for row in range(2, ws.max_row + 1):
            cod = ws.cell(row, 4).value  # Cod.Instrum (col D)
            cantidad_raw = ws.cell(row, 9).value  # Cantidad (col I)
            precio_raw = ws.cell(row, 10).value  # Precio (col J)
            bruto_raw = ws.cell(row, 11).value  # Bruto (col K)
            interes_raw = ws.cell(row, 12).value  # Interés (col L)
            
            # Convertir a números, ignorando fórmulas
            def to_float(val):
                if val is None:
                    return 0
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str) and val.startswith('='):
                    return 0  # Es una fórmula, no podemos evaluarla
                try:
                    return float(val)
                except:
                    return 0
            
            cantidad = to_float(cantidad_raw)
            precio = to_float(precio_raw)
            bruto = to_float(bruto_raw)
            interes = to_float(interes_raw)
            
            if cod:
                transacciones_por_cod[cod].append({
                    'cantidad': cantidad,
                    'precio': precio,
                    'bruto': bruto,
                    'interes': interes,
                    'neto': bruto + interes
                })
        
        resultado_total = 0
        
        for cod, transacciones in transacciones_por_cod.items():
            stock_cantidad = 0
            stock_precio_promedio = 0
            
            for t in transacciones:
                cantidad = t['cantidad']
                precio = t['precio']
                neto = t['neto']
                
                if cantidad > 0:  # COMPRA
                    # Actualizar precio promedio ponderado
                    valor_anterior = stock_cantidad * stock_precio_promedio
                    valor_nuevo = cantidad * precio
                    stock_cantidad += cantidad
                    if stock_cantidad > 0:
                        stock_precio_promedio = (valor_anterior + valor_nuevo) / stock_cantidad
                
                elif cantidad < 0:  # VENTA
                    # Calcular resultado = Neto - Costo
                    costo = abs(cantidad) * stock_precio_promedio
                    resultado = abs(neto) - costo
                    resultado_total += resultado
                    stock_cantidad += cantidad  # Resta porque cantidad es negativa
        
        return resultado_total

    def _create_visual_passthrough_sheet(self, wb: Workbook, output_name: str, source_candidates: List[str], default_headers: List[str]):
        """Copia una hoja de Visual al merge o crea una vacía si no existe."""
        ws = wb.create_sheet(output_name)

        src_name = next((name for name in source_candidates if name in self.visual_wb.sheetnames), None)
        if src_name:
            src_ws = self.visual_wb[src_name]
            max_col = src_ws.max_column
            for col in range(1, max_col + 1):
                header = src_ws.cell(1, col).value
                ws.cell(1, col, header)
                ws.cell(1, col).font = Font(bold=True)

            for row in range(2, src_ws.max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row, col, src_ws.cell(row, col).value)
            return

        for col, header in enumerate(default_headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)

    def _create_fci(self, wb: Workbook):
        self._create_visual_passthrough_sheet(
            wb,
            "FCI",
            ["FCI"],
            ['Concertación', 'Liquidación', 'Moneda', 'Tipo Operación', 'Cantidad',
             'Tipo de Cambio', 'Precio', 'Bruto', 'Gastos', 'IVA', 'Resultado']
        )

    def _create_opciones(self, wb: Workbook):
        self._create_visual_passthrough_sheet(
            wb,
            "Opciones",
            ["Opciones"],
            ['Instrumento', 'Concertación', 'Liquidación', 'Moneda', 'Tipo Operación', 'Cantidad',
             'Tipo de Cambio', 'Precio', 'Bruto', 'Gastos', 'IVA', 'Resultado']
        )

    def _create_futuros(self, wb: Workbook):
        self._create_visual_passthrough_sheet(
            wb,
            "Futuros",
            ["Futuros", "Resultado de Futuros"],
            ['Moneda', 'Instrumento', 'Tipo de Liquidación', 'Total']
        )

    def _create_pagare_cpd(self, wb: Workbook):
        """Crea hoja Pagare_CPD consolidando Visual y Gallo en un esquema común."""
        ws = wb.create_sheet("Pagare_CPD")

        headers = ['Instrumento', 'Concertación', 'Liquidación', 'Vencimiento', 'Tipo Operación',
                   'Abreviatura', 'Moneda', 'Tipo Cambio', 'Valor Nominal', 'Tasa',
                   'Valor Final', 'Gastos', 'Neto', 'Origen']

        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = Font(bold=True)

        rows_out = []

        def _to_float(value):
            if value in (None, ''):
                return 0.0
            if isinstance(value, (int, float)):
                return float(value)
            text = str(value).strip()
            if not text:
                return 0.0
            text = text.replace('.', '').replace(',', '.')
            try:
                return float(text)
            except Exception:
                return 0.0

        def _moneda_display(value, fallback='Pesos'):
            text = str(value or '').strip()
            if not text:
                return fallback
            lowered = text.lower()
            if 'peso' in lowered or lowered == 'ars':
                return 'Pesos'
            if 'cable' in lowered:
                return 'Dolar Cable'
            if 'mep' in lowered:
                return 'Dolar MEP'
            if any(token in lowered for token in ['dolar', 'dólar', 'usd', 'u$d', 'dol']):
                return 'Dolar'
            return text

        if 'Pagare_CPD' in self.visual_wb.sheetnames:
            visual_ws = self.visual_wb['Pagare_CPD']
            for row in range(2, visual_ws.max_row + 1):
                if not any(visual_ws.cell(row, col).value not in (None, '') for col in range(1, visual_ws.max_column + 1)):
                    continue

                concertacion = visual_ws.cell(row, 2).value
                fecha_dt, year = self._parse_fecha(concertacion)
                if year and year != 2025:
                    continue

                rows_out.append({
                    'instrumento': visual_ws.cell(row, 1).value,
                    'concertacion': fecha_dt if fecha_dt else concertacion,
                    'liquidacion': visual_ws.cell(row, 3).value,
                    'vencimiento': visual_ws.cell(row, 4).value,
                    'tipo_operacion': visual_ws.cell(row, 5).value,
                    'abreviatura': visual_ws.cell(row, 6).value,
                    'moneda': _moneda_display(visual_ws.cell(row, 7).value),
                    'tipo_cambio': visual_ws.cell(row, 8).value,
                    'valor_nominal': visual_ws.cell(row, 9).value,
                    'tasa': visual_ws.cell(row, 10).value,
                    'valor_final': visual_ws.cell(row, 11).value,
                    'gastos': visual_ws.cell(row, 12).value,
                    'neto': visual_ws.cell(row, 13).value,
                    'origen': 'Visual',
                })

        for sheet_name in self.gallo_wb.sheetnames:
            sheet_lower = sheet_name.lower()
            if not any(token in sheet_lower for token in ['cpd', 'pagare', 'pagaré']):
                continue
            if any(skip in sheet_name for skip in ['Posicion', 'Resultado', 'Posición']):
                continue

            gallo_ws = self.gallo_wb[sheet_name]
            moneda = 'Pesos'
            if 'exterior' in sheet_lower:
                moneda = 'Dolar Cable'
            elif 'dolar' in sheet_lower or 'dólar' in sheet_lower:
                moneda = 'Dolar MEP'

            for row in range(2, gallo_ws.max_row + 1):
                operacion = gallo_ws.cell(row, 5).value
                fecha = gallo_ws.cell(row, 4).value
                if not operacion:
                    continue

                fecha_dt, year = self._parse_fecha(fecha)
                if year and year != 2025:
                    continue

                codigo = gallo_ws.cell(row, 2).value
                especie = gallo_ws.cell(row, 3).value
                cantidad = gallo_ws.cell(row, 7).value
                tasa = gallo_ws.cell(row, 8).value
                valor_final = gallo_ws.cell(row, 9).value
                resultado_pesos = gallo_ws.cell(row, 11).value
                resultado_usd = gallo_ws.cell(row, 12).value
                gastos_pesos = gallo_ws.cell(row, 13).value
                gastos_usd = gallo_ws.cell(row, 14).value

                gastos = gastos_pesos if moneda == 'Pesos' else gastos_usd
                neto = resultado_pesos if moneda == 'Pesos' else resultado_usd
                if neto in (None, ''):
                    neto = _to_float(valor_final) - _to_float(gastos)

                rows_out.append({
                    'instrumento': especie or codigo,
                    'concertacion': fecha_dt if fecha_dt else fecha,
                    'liquidacion': '',
                    'vencimiento': '',
                    'tipo_operacion': operacion,
                    'abreviatura': codigo,
                    'moneda': moneda,
                    'tipo_cambio': 1 if moneda == 'Pesos' else None,
                    'valor_nominal': cantidad,
                    'tasa': tasa,
                    'valor_final': valor_final,
                    'gastos': gastos,
                    'neto': neto,
                    'origen': f'Gallo-{sheet_name}',
                })

        def _sort_key(item):
            fecha = item.get('concertacion')
            if isinstance(fecha, datetime):
                return (fecha, str(item.get('instrumento') or ''))
            return (datetime.min, str(item.get('instrumento') or ''))

        rows_out.sort(key=_sort_key)

        for row_out, item in enumerate(rows_out, start=2):
            ws.cell(row_out, 1, item['instrumento'])
            ws.cell(row_out, 2, item['concertacion'])
            ws.cell(row_out, 3, item['liquidacion'])
            ws.cell(row_out, 4, item['vencimiento'])
            ws.cell(row_out, 5, item['tipo_operacion'])
            ws.cell(row_out, 6, item['abreviatura'])
            ws.cell(row_out, 7, item['moneda'])
            ws.cell(row_out, 8, item['tipo_cambio'])
            ws.cell(row_out, 9, item['valor_nominal'])
            ws.cell(row_out, 10, item['tasa'])
            ws.cell(row_out, 11, item['valor_final'])
            ws.cell(row_out, 12, item['gastos'])
            ws.cell(row_out, 13, item['neto'])
            ws.cell(row_out, 14, item['origen'])

    def _create_posicion_titulos(self, wb: Workbook):
        """Crea hoja Posicion Titulos con datos de Visual (fuente principal).
        
        La Posición de Títulos se obtiene de la sección POSICIÓN DE TÍTULOS
        de los PDFs de Visual, NO de Gallo.
        """
        ws = wb.create_sheet("Posicion Titulos")
        
        # Intentar obtener datos de Visual primero (fuente principal)
        visual_sheet_name = None
        for name in ["Posicion Titulos", "Posicion de Titulos", "POSICIÓN DE TÍTULOS"]:
            if name in self.visual_wb.sheetnames:
                visual_sheet_name = name
                break
        
        if visual_sheet_name:
            # Usar datos de Visual (fuente correcta)
            visual_ws = self.visual_wb[visual_sheet_name]
            
            # Copiar encabezados de Visual
            for col in range(1, visual_ws.max_column + 1):
                header = visual_ws.cell(1, col).value
                ws.cell(1, col, header)
                ws.cell(1, col).font = Font(bold=True)
            
            # Copiar datos de Visual
            row_out = 2
            for row in range(2, visual_ws.max_row + 1):
                has_data = False
                for col in range(1, visual_ws.max_column + 1):
                    value = visual_ws.cell(row, col).value
                    if value:
                        has_data = True
                    ws.cell(row_out, col, value)
                
                if has_data:
                    # Agregar columna de origen
                    ws.cell(row_out, visual_ws.max_column + 1, f"Visual-{visual_sheet_name}")
                    row_out += 1
            
            # Agregar header de origen si hay datos
            if row_out > 2:
                ws.cell(1, visual_ws.max_column + 1, "Origen")
                ws.cell(1, visual_ws.max_column + 1).font = Font(bold=True)
        else:
            # Fallback: crear estructura básica si no hay Visual
            headers = ['Instrumento', 'Código', 'Ticker', 'Cantidad', 'Importe', 'Moneda', 'Origen']
            
            for col, header in enumerate(headers, 1):
                ws.cell(1, col, header)
                ws.cell(1, col).font = Font(bold=True)
            
            # Si no hay Visual, intentar generar desde Posicion Final Gallo como fallback
            try:
                pos_final = wb['Posicion Final Gallo']
                
                row_out = 2
                for row in range(2, pos_final.max_row + 1):
                    especie = pos_final.cell(row, 3).value  # especie (col 3)
                    codigo = pos_final.cell(row, 4).value   # Codigo especie (col 4)
                    ticker = pos_final.cell(row, 2).value   # Ticker (col 2)
                    cantidad = pos_final.cell(row, 9).value  # cantidad (col 9)
                    importe = pos_final.cell(row, 17).value  # importe_pesos (col 17)
                    
                    if not especie:
                        continue
                    
                    # Determinar moneda basado en tipo_especie o nombre
                    tipo_especie = pos_final.cell(row, 1).value or ''
                    moneda = 'Pesos'  # Default
                    if 'dolar' in str(especie).lower() or 'usd' in str(especie).lower():
                        moneda = 'Dolar'
                    
                    ws.cell(row_out, 1, especie)
                    ws.cell(row_out, 2, codigo)
                    ws.cell(row_out, 3, ticker)
                    ws.cell(row_out, 4, cantidad)
                    ws.cell(row_out, 5, importe)
                    ws.cell(row_out, 6, moneda)
                    ws.cell(row_out, 7, "Fallback-Gallo (Visual no disponible)")
                    
                    row_out += 1
            except KeyError:
                pass  # No hay datos disponibles
    
    def _add_aux_sheets(self, wb: Workbook):
        """Agrega hojas auxiliares al workbook."""
        aux_files = {
            'EspeciesVisual': self.especies_visual,
            'EspeciesGallo': self.especies_gallo,
            'Cotizacion Dolar Historica': self.cotizacion_dolar,
            'PreciosInicialesEspecies': self.precios_iniciales
        }
        
        for sheet_name, aux_wb in aux_files.items():
            ws_src = aux_wb.active
            ws_dst = wb.create_sheet(sheet_name)
            
            for row in ws_src.iter_rows():
                for cell in row:
                    ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
        
        # Enriquecer PreciosInicialesEspecies con columnas calculadas para fallback
        self._enrich_precios_iniciales(wb)

    def _add_precio_tenencias_sheet(self, wb: Workbook):
        """Agrega la hoja PrecioTenenciasIniciales con columnas Ratio y Precio Ajustado."""
        if not self.precio_tenencias_wb:
            return
        if 'PrecioTenenciasIniciales' in wb.sheetnames:
            return

        ws_src = self.precio_tenencias_wb['PrecioTenenciasIniciales'] if 'PrecioTenenciasIniciales' in self.precio_tenencias_wb.sheetnames else self.precio_tenencias_wb.active
        ws_dst = wb.create_sheet('PrecioTenenciasIniciales')

        # Copy original data
        max_col_src = ws_src.max_column
        for row in ws_src.iter_rows():
            for cell in row:
                ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)

        # Identify columns
        headers = [str(ws_src.cell(1, c).value or '').strip().lower() for c in range(1, max_col_src + 1)]
        def find_col(keyword):
            for idx, h in enumerate(headers, 1):
                if keyword in h:
                    return idx
            return None

        col_codigo = find_col('cod')
        col_ticker = find_col('ticker')
        col_cantidad = find_col('cantidad')
        col_importe = find_col('importe')

        # Add new columns: Ratio CEDEAR, Precio Ajustado
        ratio_col = max_col_src + 1
        adjusted_col = max_col_src + 2
        ws_dst.cell(1, ratio_col, 'Ratio CEDEAR')
        ws_dst.cell(1, ratio_col).font = Font(bold=True)
        ws_dst.cell(1, adjusted_col, 'Precio Ajustado')
        ws_dst.cell(1, adjusted_col).font = Font(bold=True)

        for row in range(2, ws_src.max_row + 1):
            codigo = ws_src.cell(row, col_codigo).value if col_codigo else None
            ticker = ws_src.cell(row, col_ticker).value if col_ticker else None
            cantidad = self._to_float(ws_src.cell(row, col_cantidad).value) if col_cantidad else 0
            importe = self._to_float(ws_src.cell(row, col_importe).value) if col_importe else 0

            # Raw price = importe / cantidad
            raw_price = (importe / cantidad) if cantidad else 0

            # Ratio: real CEDEAR ratio for acciones del exterior, 1 for everything else
            ratio = 1
            cod_clean = self._clean_codigo(str(codigo)) if codigo else ''
            if cod_clean and self._is_accion_exterior(cod_clean):
                especie_name = str(ws_src.cell(row, 3).value or '')
                r = self._get_ratio_for_especie(
                    str(ticker) if ticker else '',
                    especie_name,
                )
                if r:
                    ratio = r

            adjusted_price = raw_price / ratio

            ws_dst.cell(row, ratio_col, ratio)
            ws_dst.cell(row, adjusted_col, adjusted_price)

    def _add_ratios_cedears_sheet(self, wb: Workbook):
        """Agrega la hoja RatiosCedearsAcciones como referencia auxiliar visible."""
        if 'RatiosCedearsAcciones' in wb.sheetnames:
            return
        try:
            aux_path = self.aux_data_dir / 'RatiosCedearsAcciones.xlsx'
            if not aux_path.exists():
                return
            wb_ratios = load_workbook(aux_path)
            ws_src = wb_ratios.active
            ws_dst = wb.create_sheet('RatiosCedearsAcciones')
            for row in ws_src.iter_rows():
                for cell in row:
                    ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            # Bold headers
            for col in range(1, ws_src.max_column + 1):
                ws_dst.cell(1, col).font = Font(bold=True)
        except Exception:
            pass

    def _enrich_precios_iniciales(self, wb: Workbook):
        """
        Agrega columnas calculadas a PreciosInicialesEspecies para usar en VLOOKUPs de fallback.
        
        Columnas existentes: A=Código, B=Nombre, C=Ticker/ORDEN, D-F=otros, G=Precio
        
        Columnas nuevas:
        - H: Tipo Instrumento (VLOOKUP a EspeciesVisual col R)
        - I: Precio Nominal = IF(tipo requiere /100, G/100, G)
        - J: Precio Nominal USD = I / Cotización inicio período
        """
        if 'PreciosInicialesEspecies' not in wb.sheetnames:
            return
        
        ws = wb['PreciosInicialesEspecies']
        
        # Agregar headers
        ws.cell(1, 8, 'Tipo Instrumento')
        ws.cell(1, 9, 'Precio Nominal')
        ws.cell(1, 10, 'Precio Nominal USD')
        
        # Cotización inicio período (para fórmulas)
        cotiz = self._fmt_num_es(self.COTIZACION_INICIO_PERIODO)
        
        # Lista de tipos que requieren división por 100 (para la fórmula)
        # Usamos matching parcial como en _es_tipo_precio_cada_100
        tipos_100 = "obligaciones negociables|obligacion negociable|títulos públicos|titulos publicos|titulo publico|letras del tesoro|letra del tesoro|letras"
        
        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row, 1).value
            precio = ws.cell(row, 7).value
            
            if codigo:
                # Col H: Tipo Instrumento - VLOOKUP a EspeciesVisual
                # EspeciesVisual: Col C=Código, Col R=Tipo Especie (offset 16)
                ws.cell(row, 8, f'=SI(ESERROR(BUSCARV(A{row};EspeciesVisual!C:R;16;FALSO));"";BUSCARV(A{row};EspeciesVisual!C:R;16;FALSO))')
                
                # Col I: Precio Nominal - dividir por 100 si tipo lo requiere
                # Usamos HALLAR para detectar si H contiene alguno de los tipos
                ws.cell(row, 9, f'=SI(O(ESNUMERO(HALLAR("obligacion";MINUSC(H{row})));ESNUMERO(HALLAR("titulo";MINUSC(H{row})));ESNUMERO(HALLAR("letra";MINUSC(H{row}))));G{row}/100;G{row})')
                
                # Col J: Precio Nominal USD = I / cotización
                ws.cell(row, 10, f'=I{row}/{cotiz}')


def merge_gallo_visual(gallo_path: str = None, visual_path: str = None, output_path: str = None, 
                       output_mode: str = "formulas", precio_tenencias_path: str = None) -> str:
    """
    Función principal para ejecutar el merge.
    
    Args:
        gallo_path: Ruta al Excel de Gallo (opcional para casos Visual-only)
        visual_path: Ruta al Excel de Visual
        output_path: Ruta de salida (opcional, genera nombre automático)
        output_mode: "formulas" (default), "values", or "both"
    
    Returns:
        Ruta del archivo generado (o tupla de rutas si output_mode="both")
    """
    merger = GalloVisualMerger(gallo_path, visual_path, precio_tenencias_path=precio_tenencias_path)
    wb_formulas, wb_values = merger.merge(output_mode=output_mode)
    
    if output_path is None:
        # Generar nombre basado en el archivo de entrada
        gallo_name = Path(gallo_path).stem.replace('_Gallo_Generado_OK', '') if gallo_path else Path(visual_path).stem
        output_path = f"{gallo_name}_Merge_Consolidado.xlsx"
    
    if output_mode == "formulas" and wb_formulas:
        wb_formulas.save(output_path)
        return output_path
    elif output_mode == "values" and wb_values:
        wb_values.save(output_path)
        return output_path
    elif output_mode == "both":
        wb_formulas.save(output_path)
        values_path = output_path.replace('.xlsx', '_values.xlsx')
        wb_values.save(values_path)
        return output_path, values_path
    
    return output_path


if __name__ == "__main__":
    # Test con archivos de ejemplo
    import sys
    
    if len(sys.argv) >= 3:
        gallo = sys.argv[1]
        visual = sys.argv[2]
        output = sys.argv[3] if len(sys.argv) > 3 else None
        
        result = merge_gallo_visual(gallo, visual, output)
        print(f"Merge completado: {result}")
    else:
        print("Uso: python merge_gallo_visual.py <gallo.xlsx> <visual.xlsx> [output.xlsx]")
