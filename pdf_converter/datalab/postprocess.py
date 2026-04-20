"""
Post-processing for Datalab-extracted financial data.
Transforms raw extracted data to match the expected Excel format.
"""

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

console = Console()

# Sheet total names mapping
SHEET_TOTAL_NAMES = {
    'tit.privados exentos': 'TOTAL TIT.PRIVADOS EXENTOS',
    'tit. privados exentos': 'TOTAL TIT.PRIVADOS EXENTOS',
    'tit.privados exterior': 'TOTAL TIT.PRIVADOS DEL EXTERIOR',
    'tit. privados exterior': 'TOTAL TIT.PRIVADOS DEL EXTERIOR',
    'renta fija pesos': 'TOTAL RENTA FIJA EN PESOS',
    'renta fija dolares': 'TOTAL RENTA FIJA EN DOLARES',
    'cauciones dolares': 'TOTAL CAUCIONES EN DOLARES',
    'cauciones pesos': 'TOTAL CAUCIONES EN PESOS',
}


def fix_trailing_negative(value: str) -> str:
    """
    Convert trailing negative sign to leading negative.
    Example: "541.39-" -> "-541.39"
    """
    if not value or not isinstance(value, str):
        return value
    
    value = value.strip()
    if value.endswith('-'):
        # Remove trailing minus and add leading minus
        return '-' + value[:-1]
    return value


def parse_numeric(value: str) -> Optional[float]:
    """
    Parse a numeric string, handling trailing negatives and thousand separators.
    """
    if not value or not isinstance(value, str):
        return None
    
    value = fix_trailing_negative(value.strip())
    
    # Remove thousand separators (comma in US format)
    value = value.replace(',', '')
    
    try:
        return float(value)
    except ValueError:
        return None


def is_numeric_column(header: str) -> bool:
    """
    Determine if a column should contain numeric values based on header name.
    """
    if not header:
        return False
    h = header.lower()
    
    # Specific numeric column names (exact or partial match)
    numeric_keywords = [
        'cantidad', 'precio', 'importe', 'total', 'monto', 'valor',
        'pesos', 'dolares', 'usd', 'ars', 'nominal', 'cotizacion',
        'bruto', 'neto', 'gastos', 'iva', 'resultado', 'interes', 'interés',
        'cartera', '%', 'tipo de cambio', 'tipo cambio', 't.cambio',
        # Cauciones-specific
        'contado', 'futuro', 'tasa', 'aranceles', 'derechos', 'costo', 'devengad'
    ]
    
    # Text column keywords - these override numeric if both match
    text_keywords = ['especie', 'tipo operacion', 'tipo instrumento', 'tipo_', 
                     'operacion', 'fecha', 'comprobante', 'numero', 'venc',
                     'detalle', 'custodia', 'moneda', 'instrumento', 'categoria', 
                     'codigo', 'ticker', 'concertacion', 'liquidacion', 'nro']
    
    for kw in text_keywords:
        if kw in h:
            return False
    
    for kw in numeric_keywords:
        if kw in h:
            return True
    
    return False


def parse_parentheses_negative(value: str) -> Optional[float]:
    """
    Parse a numeric string, handling parentheses as negatives.
    Example: "(123.45)" -> -123.45
    Also handles trailing negatives and thousand separators.
    
    For European format:
    - "91.886" (no comma) -> 91886 (thousands separator only)
    - "1.234,56" -> 1234.56 (European decimal)
    - "1,215,0000000" -> 1215.0 (mixed format for exchange rates)
    
    OCR artifact handling:
    - "(300.000,000," -> -300000000 (single paren = negative, trailing comma)
    - "772.000,00," -> 772000000 (garbled thousands: ,00, -> .000)
    - "650.945.200," -> 650945200 (trailing comma stripped)
    """
    if not value or not isinstance(value, str):
        return None
    
    value = value.strip()
    
    # Check for parentheses indicating negative
    is_negative = False
    if value.startswith('(') and value.endswith(')'):
        value = value[1:-1]
        is_negative = True
    elif value.startswith('('):
        # Single opening paren without closing (OCR truncation) -> negative
        value = value[1:]
        is_negative = True
    
    # Handle trailing negative
    value = fix_trailing_negative(value)
    if value.startswith('-'):
        is_negative = True
        value = value[1:]
    
    # --- OCR artifact cleanup ---
    # Fix garbled thousands only for OCR trailing-comma artifacts.
    # Notes:
    # - ",00," is consistently a missing thousands block in OCR (e.g. 772.000,00, -> 772.000.000)
    # - ",000," can be either garbled thousands OR a valid decimal with trailing comma.
    #   We only expand ",000," when clearly large/negative to avoid inflating rows like
    #   2.000,000, (should remain 2000.000, not 2.000.000).
    if re.search(r',00,$', value):
        value = re.sub(r',00,$', '.000', value)
    elif re.search(r',000,$', value):
        integer_part = value[:-5]  # remove trailing ",000,"
        digits = re.sub(r'\D', '', integer_part)
        if is_negative or len(digits) >= 6:
            value = integer_part + '.000'
        else:
            value = value[:-1]
    elif re.match(r'^\d{1,3}(\.\d{3})+,\d{3}$', value):
        integer_part, decimal_group = value.rsplit(',', 1)
        digits = re.sub(r'\D', '', integer_part)
        if re.match(r'^0+$', decimal_group):
            # All-zero decimal (e.g. "580.000,000") → European 580000.000
            # NOT garbled thousands 580.000.000.  OCR options often emit
            # trailing decimal zeros that look like a thousands group.
            value = integer_part.replace('.', '') + '.' + decimal_group
        elif is_negative or len(digits) >= 6:
            value = integer_part + '.' + decimal_group
    # Strip any remaining trailing comma or period (OCR truncation)
    value = value.rstrip(',').rstrip('.')
    
    # Remove thousand separators
    # Detect format based on presence of comma as decimal separator
    if ',' in value:
        # Check for European format: dots are thousands, comma is decimal
        # e.g., "1.234,56" or "1.234.567,89"
        if re.match(r'^[\d.]+,\d+$', value):
            # Standard European: 1.234,56
            value = value.replace('.', '').replace(',', '.')
        elif re.match(r'^\d{1,3}(,\d{3})+,\d+$', value):
            # Mixed format like "1,215,0000000" - multiple commas
            # Take first part as integer, rest as decimal
            parts = value.split(',')
            if len(parts) >= 3:
                # Format: "1,215,0000000" -> integer part = "1215", decimal = "0000000"
                # or "10,000,0000" -> 10000.0
                integer_part = ''.join(parts[:-1])
                decimal_part = parts[-1]
                value = f"{integer_part}.{decimal_part}"
            else:
                value = value.replace(',', '')
        else:
            # US format with comma as thousands: "1,234.56" or just commas "1,234"
            value = value.replace(',', '')
    else:
        # OCR sometimes emits quantities like 125.000.0000 or 1.234.567.90 where the
        # last dot is the decimal separator and previous dots are thousands separators.
        dotted_decimal_tail = re.match(r'^(\d{1,3}(?:\.\d{3})+)\.(\d{2}|\d{4})$', value)
        if dotted_decimal_tail:
            integer_part = dotted_decimal_tail.group(1).replace('.', '')
            decimal_part = dotted_decimal_tail.group(2)
            value = f"{integer_part}.{decimal_part}"
        # No comma - check if dots are thousands separators
        # e.g., "91.886" is 91886 (no decimal), "91.88" could be 91.88
        # If there's a dot followed by exactly 3 digits and no more dots after, it's thousands
        elif re.match(r'^\d{1,3}(\.\d{3})+$', value):
            # Pattern like "91.886" or "1.234.567" - these are thousands separators
            value = value.replace('.', '')
        # Otherwise keep as is (could be a decimal like "1.5" or "123.45")
    
    try:
        result = float(value)
        return -result if is_negative else result
    except ValueError:
        return None


def parse_visual_quantity_value(value) -> Optional[float]:
    """Parse Visual quantity cells, preferring integer quantities over zero decimal tails."""
    if value in (None, ''):
        return None

    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)

    if not isinstance(value, str):
        return None

    raw = value.strip()
    if not raw:
        return None

    is_negative = False
    if raw.startswith('('):
        is_negative = True
        raw = raw[1:]
        if raw.endswith(')'):
            raw = raw[:-1]

    raw = fix_trailing_negative(raw)
    if raw.startswith('-'):
        is_negative = True
        raw = raw[1:]

    raw = raw.rstrip(',').rstrip('.')

    zero_decimal = re.match(r'^(\d{1,3}(?:\.\d{3})+),(0+)$', raw)
    if zero_decimal:
        result = int(zero_decimal.group(1).replace('.', ''))
        return -result if is_negative else result

    thousands_only = re.match(r'^\d{1,3}(?:\.\d{3})+$', raw)
    if thousands_only:
        result = int(raw.replace('.', ''))
        return -result if is_negative else result

    # Truncated thousands: last dot-group has 1-2 digits instead of 3.
    # OCR column-width truncation turns e.g. "1.170.588.235" into "1.170.588.23".
    # Quantities are always integers, so pad with trailing zeros to recover the
    # full integer. This takes priority over interpreting the tail as a decimal.
    truncated_thousands = re.match(r'^(\d{1,3}(?:\.\d{3})+)\.(\d{1,2})$', raw)
    if truncated_thousands:
        full_part = truncated_thousands.group(1).replace('.', '')
        truncated_part = truncated_thousands.group(2)
        result = int(full_part + truncated_part + '0' * (3 - len(truncated_part)))
        return -result if is_negative else result

    parsed = parse_parentheses_negative(value)
    if parsed is None:
        return None
    return int(parsed) if float(parsed).is_integer() else parsed


def _normalize_visual_date_key(value) -> str:
    if value is None:
        return ''
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return f"{value.day}/{value.month}/{value.year}"

    text = str(value).strip()
    if not text:
        return ''

    slash = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', text)
    if slash:
        return f"{int(slash.group(1))}/{int(slash.group(2))}/{slash.group(3)}"

    iso = re.match(r'^(\d{4})-(\d{2})-(\d{2})', text)
    if iso:
        return f"{int(iso.group(3))}/{int(iso.group(2))}/{iso.group(1)}"

    return text


def _normalize_visual_currency_key(value) -> str:
    text = str(value or '').strip().lower()
    if not text:
        return ''
    if 'peso' in text or text == 'ars':
        return 'pesos'
    if 'cable' in text or 'dolar ca' in text:
        return 'dolar cable'
    if 'mep' in text or text.startswith('dolar me'):
        return 'dolar mep'
    if 'dolar' in text or 'dólar' in text or text == 'usd':
        return 'dolar'
    return text


def _normalize_visual_operation_key(value) -> str:
    text = re.sub(r'\s+', ' ', str(value or '').strip().lower())
    text = re.sub(r'\bno$', '', text).strip()
    return text


def _normalize_visual_code(value) -> Optional[int]:
    if isinstance(value, str):
        text = value.strip()
        if re.match(r'^\d{1,3}(?:\.\d{3})+$', text):
            try:
                return int(text.replace('.', ''))
            except Exception:
                pass
    try:
        return int(float(value))
    except Exception:
        return None


def _build_visual_quantity_anchors(wb: Workbook) -> Dict[Tuple[int, str, str, str], int]:
    """Collect trustworthy quantities from Resultado Ventas sheets before Boletos is processed."""
    anchors: Dict[Tuple[int, str, str, str], int] = {}

    for sheet_name in ['Resultado Ventas ARS', 'Resultado Ventas USD']:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]

        code_col = next((i for i, h in enumerate(headers, start=1) if 'cod.instrum' in h or h == 'codigo'), None)
        concert_col = next((i for i, h in enumerate(headers, start=1) if 'concert' in h), None)
        currency_col = next((i for i, h in enumerate(headers, start=1) if h == 'moneda'), None)
        oper_col = next((i for i, h in enumerate(headers, start=1) if 'tipo oper' in h), None)
        qty_col = next((i for i, h in enumerate(headers, start=1) if h == 'cantidad'), None)
        precio_col = next((i for i, h in enumerate(headers, start=1) if h == 'precio'), None)
        bruto_col = next((i for i, h in enumerate(headers, start=1) if h == 'bruto'), None)

        if not all([code_col, concert_col, currency_col, oper_col, qty_col]):
            continue

        for row in range(2, ws.max_row + 1):
            code = _normalize_visual_code(ws.cell(row, code_col).value)
            qty = parse_visual_quantity_value(ws.cell(row, qty_col).value)
            if code is None or qty in (None, 0):
                continue

            key = (
                code,
                _normalize_visual_date_key(ws.cell(row, concert_col).value),
                _normalize_visual_currency_key(ws.cell(row, currency_col).value),
                _normalize_visual_operation_key(ws.cell(row, oper_col).value),
            )

            qty_int = int(qty) if float(qty).is_integer() else None
            if qty_int is None:
                continue

            # Bruto-validated 1000× deflation: OCR may add an extra ".000" group
            # (e.g. "512.974.026.000" → 512974026000 instead of 512974026).
            if precio_col and bruto_col:
                deflated = qty_int // 1000
                if deflated != 0 and (qty_int % 1000 == 0):
                    try:
                        precio_num = abs(float(ws.cell(row, precio_col).value or 0))
                        bruto_num = abs(float(ws.cell(row, bruto_col).value or 0))
                        if precio_num > 0 and bruto_num > 0:
                            err_orig = abs(abs(qty_int) * precio_num - bruto_num) / bruto_num
                            err_defl = abs(abs(deflated) * precio_num - bruto_num) / bruto_num
                            if err_orig > 0.50 and err_defl <= 0.05:
                                qty_int = deflated
                    except (ValueError, TypeError):
                        pass

            current = anchors.get(key)
            if current is None or abs(qty_int) < abs(current):
                anchors[key] = qty_int

    # Cross-currency 1000× sibling deflation: when two anchors for the same
    # code+date differ by ~1000×, the larger is likely an OCR ".000" artifact.
    from collections import defaultdict
    code_date_groups: Dict[Tuple[int, str], List[Tuple]] = defaultdict(list)
    for key, qty_int in anchors.items():
        code_date_groups[(key[0], key[1])].append(key)

    for (code, date), keys in code_date_groups.items():
        if len(keys) < 2:
            continue
        # Pairwise 1000× check: if entry A is ~1000× entry B, deflate A.
        abs_vals = {k: abs(anchors[k]) for k in keys}
        for k_large in keys:
            v_large = abs_vals[k_large]
            if v_large == 0:
                continue
            for k_small in keys:
                if k_small is k_large:
                    continue
                v_small = abs_vals[k_small]
                if v_small == 0:
                    continue
                ratio = v_large / v_small
                if 900 <= ratio <= 1100:
                    sign = 1 if anchors[k_large] > 0 else -1
                    anchors[k_large] = sign * v_small
                    break

    return anchors


def _lookup_visual_quantity_anchor(
    anchors: Dict[Tuple[int, str, str, str], int],
    code,
    concertacion,
    moneda,
    operacion,
) -> Optional[int]:
    code_key = _normalize_visual_code(code)
    if code_key is None:
        return None

    exact_key = (
        code_key,
        _normalize_visual_date_key(concertacion),
        _normalize_visual_currency_key(moneda),
        _normalize_visual_operation_key(operacion),
    )
    if exact_key in anchors:
        return anchors[exact_key]

    date_key = _normalize_visual_date_key(concertacion)
    currency_key = _normalize_visual_currency_key(moneda)
    operation_key = _normalize_visual_operation_key(operacion)

    relaxed = [
        qty
        for (anchor_code, anchor_date, anchor_currency, anchor_operation), qty in anchors.items()
        if anchor_code == code_key
        and anchor_date == date_key
        and anchor_currency == currency_key
        and anchor_operation.startswith(operation_key[:12])
    ]

    if len(relaxed) == 1:
        return relaxed[0]

    return None


def _is_strong_visual_quantity_anomaly(raw_qty) -> bool:
    raw_text = str(raw_qty or '').strip()
    if not raw_text:
        return False

    raw_text = raw_text.strip('()')
    raw_text = fix_trailing_negative(raw_text).lstrip('-')

    if ',' in raw_text:
        integer_part, decimal_part = raw_text.rsplit(',', 1)
        if re.match(r'^\d{1,3}(?:\.\d{3})*$', integer_part) and decimal_part.isdigit():
            # Non-zero decimal in what should be an integer quantity = column-width truncation.
            # e.g. 1.285.714,2 means 1285714285 was truncated to 1285714.2
            if int(decimal_part) != 0:
                return True
            return False
        return True

    if re.match(r'^\d{1,3}(?:\.\d{3})+$', raw_text):
        return False

    if re.match(r'^\d+\.\d$', raw_text):
        return False

    if raw_text.count('.') >= 2:
        return True

    decimal_match = re.match(r'^-?\d+\.(\d+)$', raw_text)
    if decimal_match and len(decimal_match.group(1)) >= 2:
        return True

    return False


def _build_visual_code_anchor_evidence(ws: Worksheet) -> Dict[int, int]:
    """Count strong raw OCR quantity anomalies per code in Boletos."""
    if ws.title.lower() != 'boletos':
        return {}

    headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
    code_col = next((i for i, h in enumerate(headers, start=1) if 'cod.instru' in h or h == 'cod.instrum'), None)
    qty_col = next((i for i, h in enumerate(headers, start=1) if h == 'cantidad'), None)

    if not code_col or not qty_col:
        return {}

    evidence: Dict[int, int] = {}
    for row in range(2, ws.max_row + 1):
        code = _normalize_visual_code(ws.cell(row, code_col).value)
        raw_qty = ws.cell(row, qty_col).value
        if code is None or not _is_strong_visual_quantity_anomaly(raw_qty):
            continue
        evidence[code] = evidence.get(code, 0) + 1

    return evidence


def _build_code_mixed_magnitude(ws: Worksheet) -> Dict[int, bool]:
    """Detect codes whose Boletos quantities span 3+ orders of magnitude (e.g., millions AND billions).

    When a code has mixed magnitudes, some boletos were likely column-width truncated
    (lost a '.000' thousands group).  When all boletos sit at the same scale, the
    quantities are probably correct and differing anchors are OCR artifacts.
    """
    import math
    headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
    cod_col = None
    qty_col = None
    for idx, h in enumerate(headers, start=1):
        if 'cod.instru' in h:
            cod_col = idx
        elif h == 'cantidad':
            qty_col = idx
    if not cod_col or not qty_col:
        return {}

    code_mags: Dict[int, List[int]] = {}
    for row in range(2, ws.max_row + 1):
        code = _normalize_visual_code(ws.cell(row, cod_col).value)
        if code is None:
            continue
        try:
            qty_abs = abs(float(ws.cell(row, qty_col).value or 0))
        except (ValueError, TypeError):
            continue
        if qty_abs < 1:
            continue
        mag = int(math.log10(qty_abs))
        code_mags.setdefault(code, []).append(mag)

    return {
        code: (max(mags) - min(mags)) >= 3
        for code, mags in code_mags.items()
    }


def _should_apply_visual_anchor(raw_qty, parsed_qty, anchor_qty: int, code_anchor_evidence: int = 0, operacion: str = '', code_mixed_magnitude: bool = False) -> bool:
    if anchor_qty in (None, 0):
        return False

    try:
        parsed_num = float(parsed_qty)
    except Exception:
        return True

    if float(anchor_qty) == parsed_num:
        return False

    # Quantities must always be integers. If parsed qty is non-integer and
    # the anchor is integer, accept the anchor when in the same magnitude
    # or a clean 1000× truncation.  Reject wildly different anchors that
    # happen to share the same code+date key but belong to another trade.
    if not float(parsed_num).is_integer() and float(anchor_qty).is_integer():
        _p = abs(parsed_num)
        _a = abs(float(anchor_qty))
        if _p > 0:
            _r = _a / _p
            if 0.5 <= _r <= 2 or 900 <= _r <= 1100:
                return True
        else:
            return True  # parsed is 0, anchor has value → accept

    raw_text = str(raw_qty or '').strip()
    has_strong_anomaly = _is_strong_visual_quantity_anomaly(raw_qty)
    allow_clean_propagation = code_anchor_evidence >= 2 and 'contado continuo' in _normalize_visual_operation_key(operacion)

    parsed_abs = abs(parsed_num)
    anchor_abs = abs(float(anchor_qty))

    if has_strong_anomaly:
        # Strong raw anomaly → trust anchor, but only when the magnitude is
        # compatible (same order or clean 1000× truncation).  This prevents
        # accepting an anchor from a different trade that happens to share
        # the same code+date key.
        if parsed_abs > 0:
            _r = anchor_abs / parsed_abs
            if 0.5 <= _r <= 2 or 900 <= _r <= 1100:
                return True
            # Fall through to more nuanced checks below.
        else:
            return True

    # 1000× ratio indicates exactly one ".000" group was lost to column truncation.
    # Only trust this when the code has mixed-magnitude quantities in Boletos
    # (proving some rows were truncated, vs. all anchors being inflated OCR artifacts).
    if parsed_abs > 0 and code_mixed_magnitude:
        ratio_1000 = anchor_abs / parsed_abs
        if 900 <= ratio_1000 <= 1100:
            return True

    digits = re.sub(r'\D', '', raw_text)
    anchor_digits = str(abs(int(anchor_qty)))
    if digits and anchor_digits.startswith(digits) and len(anchor_digits) > len(digits):
        if not allow_clean_propagation:
            return False
        return True

    if parsed_abs == 0:
        return True

    ratio = max(anchor_abs / parsed_abs, parsed_abs / anchor_abs)
    return allow_clean_propagation and 10 <= ratio <= 1100


def _maybe_rescue_visual_bruto(current_bruto, qty_value, precio_value):
    try:
        qty_num = float(qty_value)
        precio_num = float(precio_value)
    except Exception:
        return current_bruto

    expected = qty_num * precio_num

    try:
        bruto_num = float(current_bruto)
    except Exception:
        return expected

    if bruto_num == 0:
        return expected

    ratio = abs(expected) / abs(bruto_num) if bruto_num else float('inf')
    if ratio >= 10 or ratio <= 0.1:
        return expected

    return current_bruto


def _maybe_rescue_visual_neto(current_neto, qty_value, precio_value, gastos_value):
    try:
        qty_num = float(qty_value)
        precio_num = float(precio_value)
        gastos_num = float(gastos_value or 0)
    except Exception:
        return current_neto

    expected_bruto = qty_num * precio_num
    expected_neto = expected_bruto + gastos_num if qty_num > 0 else expected_bruto - gastos_num

    try:
        neto_num = float(current_neto)
    except Exception:
        return expected_neto

    if neto_num == 0:
        return expected_neto

    ratio = abs(expected_neto) / abs(neto_num) if neto_num else float('inf')
    if ratio >= 10 or ratio <= 0.1:
        return expected_neto

    return current_neto


def _parse_ambiguous_quantity_candidates(value: str) -> List[float]:
    """Genera candidatos para cantidades OCR ambiguas como 150.000,000 o 166.000,0000."""
    if not value or not isinstance(value, str):
        return []

    raw = value.strip()
    is_negative = raw.startswith('(') or raw.endswith('-') or raw.startswith('-')
    cleaned = raw.strip('()')
    cleaned = fix_trailing_negative(cleaned)
    cleaned = cleaned.lstrip('-').rstrip(',').rstrip('.')

    match = re.match(r'^(\d{1,3}(?:\.\d{3})*),(\d{3,4})$', cleaned)
    if not match:
        return []

    decimal_group = match.group(2)
    if set(decimal_group) != {'0'}:
        return []

    integer_digits = match.group(1).replace('.', '')
    sign = -1 if is_negative else 1

    candidates = []
    try:
        candidates.append(sign * float(f"{integer_digits}.0"))
    except Exception:
        pass
    try:
        candidates.append(sign * float(integer_digits + decimal_group))
    except Exception:
        pass

    # De-duplicate preserving order
    deduped = []
    for candidate in candidates:
        if candidate not in deduped:
            deduped.append(candidate)
    return deduped


def _choose_best_boletos_quantity(raw_qty: str, parsed_qty, precio, bruto):
    """Elige la cantidad que mejor cierra con Precio x Cantidad ~= Bruto."""
    candidates = _parse_ambiguous_quantity_candidates(raw_qty)
    if not candidates:
        return parsed_qty

    try:
        precio_num = abs(float(precio or 0))
        bruto_num = abs(float(bruto or 0))
    except Exception:
        return parsed_qty

    if precio_num == 0 or bruto_num == 0:
        return parsed_qty

    best_candidate = parsed_qty
    best_error = None

    parsed_candidates = []
    try:
        parsed_candidates.append(float(parsed_qty))
    except Exception:
        pass

    for candidate in candidates + parsed_candidates:
        expected_bruto = abs(candidate) * precio_num
        error = abs(expected_bruto - bruto_num)
        if best_error is None or error < best_error:
            best_error = error
            best_candidate = candidate

    # Guardrail extra para boletos Visual: si el candidato elegido infla la cantidad en
    # tres órdenes de magnitud frente al patrón con 4 decimales OCR, preferir la versión
    # desescalada cuando también sigue cerrando razonablemente contra el bruto.
    if candidates and len(candidates) >= 2:
        smallest = min(candidates, key=lambda x: abs(x))
        largest = max(candidates, key=lambda x: abs(x))
        if abs(largest) >= abs(smallest) * 1000:
            expected_small = abs(smallest) * precio_num
            rel_small_error = abs(expected_small - bruto_num) / bruto_num if bruto_num else 1
            rel_large_error = abs(abs(largest) * precio_num - bruto_num) / bruto_num if bruto_num else 1
            if rel_small_error <= 0.05:
                best_candidate = smallest
            elif rel_large_error > 0.5 and rel_small_error < rel_large_error:
                best_candidate = smallest

    return int(best_candidate) if float(best_candidate).is_integer() else best_candidate


def replace_null_with_zero(ws: Worksheet) -> None:
    """
    Replace null values with 0 in numeric columns.
    """
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    for col_idx, header in enumerate(headers, 1):
        if is_numeric_column(str(header) if header else ""):
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is None or cell.value == '':
                    cell.value = 0


def process_resultado_totales(ws: Worksheet) -> None:
    """
    Post-process Resultado Totales sheet:
    - Convert trailing negatives to leading negatives
    - Parse numeric values
    - Replace nulls with 0 in numeric columns
    """
    console.print("  [dim]Processing Resultado Totales...[/dim]")
    
    # Find numeric columns (typically 2 and 3 for pesos and usd)
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                numeric = parse_numeric(cell.value)
                if numeric is not None:
                    cell.value = numeric
            elif cell.value is None:
                cell.value = 0


def extract_especie_parts(especie_value: str) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Extract cod_especie, especie name, and tipo_fila from a combined especie string.
    
    Examples:
    - "00007 ALUA ALUAR" -> ("00007", "ALUA ALUAR", None)
    - "Total Renta" -> (None, None, "Total Renta")
    - "Total Enajenacion" -> (None, None, "Total Enajenacion")
    - "." -> (None, None, None)  # separator row
    
    Returns: (cod_especie, especie_name, tipo_fila)
    """
    if not especie_value or not isinstance(especie_value, str):
        return (None, None, None)
    
    value = especie_value.strip()
    
    # Skip separator rows
    if value in ['.', '']:
        return (None, None, None)
    
    # Check if it's a total row (Total Renta, Total Enajenacion, TOTAL X, etc.)
    if value.lower().startswith('total'):
        return (None, None, value)
    
    # Check if it starts with a numeric code
    match = re.match(r'^(\d+)\s+(.+)$', value)
    if match:
        cod_especie = match.group(1)
        especie_name = match.group(2).strip()
        return (cod_especie, especie_name, None)
    
    # No code found, might be a continuation or unknown format
    return (None, value, None)


def get_sheet_total_name(sheet_name: str) -> str:
    """Get the total row name for a given sheet."""
    sheet_lower = sheet_name.lower()
    for key, value in SHEET_TOTAL_NAMES.items():
        if key in sheet_lower:
            return value
    return f"TOTAL {sheet_name.upper()}"


def process_detail_sheet(ws: Worksheet, sheet_name: str) -> Worksheet:
    """
    Post-process detail sheets (Tit.Privados, Renta Fija, Cauciones):
    - Split 'Especie' column into: tipo_fila, cod_especie, especie
    - Propagate especie and cod_especie to blank rows
    - Mark non-total rows as 'transaccion'
    - Fix trailing negatives in numeric columns
    - Replace nulls with 0 in numeric columns
    - Add final TOTAL row for the entire sheet
    
    Returns a new worksheet with the transformed data.
    """
    console.print(f"  [dim]Processing {sheet_name}...[/dim]")
    
    # Read all data first
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    # Find the Especie column index
    especie_col = None
    for i, h in enumerate(headers):
        if h and 'especie' in h.lower():
            especie_col = i
            break
    
    if especie_col is None:
        console.print(f"    [yellow]No 'Especie' column found[/yellow]")
        return ws
    
    # Read all rows
    rows_data = []
    for row in range(2, ws.max_row + 1):
        row_values = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
        rows_data.append(row_values)
    
    # Build new headers
    new_headers = ['tipo_fila', 'cod_especie', 'especie']
    numeric_col_indices = []  # Track which columns in new_headers are numeric
    for i, h in enumerate(headers):
        if i != especie_col:
            new_headers.append(h)
            if is_numeric_column(str(h) if h else ""):
                numeric_col_indices.append(len(new_headers) - 1)
    
    # Process rows: extract especie parts and propagate
    current_cod = None
    current_especie = None
    processed_rows = []
    last_was_sheet_total = False  # Track if last row was a sheet-level total
    
    for row_values in rows_data:
        especie_value = row_values[especie_col] if especie_col < len(row_values) else None
        especie_str = str(especie_value).strip() if especie_value else ""
        cod, name, tipo_fila = extract_especie_parts(especie_str)
        
        # Check if this is a continuation row after a sheet-level TOTAL
        # (especie is empty but row has data in other columns)
        if last_was_sheet_total and not especie_str:
            # Check if there's any data in this row (excluding especie column)
            has_data = any(
                row_values[i] and str(row_values[i]).strip() 
                for i in range(len(row_values)) if i != especie_col
            )
            if has_data and processed_rows:
                # Merge this row's values into the previous TOTAL row
                prev_row = processed_rows[-1]
                for i, val in enumerate(row_values):
                    if i == especie_col:
                        continue
                    if not val or not str(val).strip():
                        continue
                    # Calculate target column in new_row
                    # new_row has: tipo_fila, cod_especie, especie, then original cols minus especie_col
                    # So for i > especie_col: target = i + 2 (3 new cols, minus 1 skipped)
                    # For i < especie_col: target = i + 3 (3 new cols)
                    if i < especie_col:
                        target_idx = i + 3
                    else:
                        target_idx = i + 2
                    
                    numeric = parse_numeric(str(val))
                    final_val = numeric if numeric is not None else val
                    
                    # Extend prev_row if needed
                    while len(prev_row) <= target_idx:
                        prev_row.append(None)
                    
                    # Merge if target is None, 0, or empty string
                    existing = prev_row[target_idx]
                    if existing is None or existing == 0 or existing == '':
                        prev_row[target_idx] = final_val
                # Don't add this row - it's been merged
                last_was_sheet_total = False
                continue
        
        # Update current tracking
        if cod:
            current_cod = cod
        if name and not tipo_fila:
            current_especie = name
        
        # Determine tipo_fila value
        # Check if this is a sheet-level total (TOTAL TIT.PRIVADOS, TOTAL RENTA FIJA, etc.)
        is_sheet_total = tipo_fila and tipo_fila.upper().startswith('TOTAL') and any(
            key in tipo_fila.upper() for key in ['TIT.PRIVADOS', 'RENTA FIJA', 'CAUCIONES']
        )
        last_was_sheet_total = is_sheet_total
        
        if tipo_fila:
            # It's a Total row (Total Renta, Total Enajenacion, TOTAL TIT.PRIVADOS, etc.)
            final_tipo_fila = tipo_fila
        else:
            # Regular transaction row
            final_tipo_fila = 'transaccion'
        
        # Build new row with: tipo_fila, cod_especie, especie, then rest of original columns
        new_row = [final_tipo_fila]
        
        # If it's a sheet-level total, don't include cod_especie or especie
        if is_sheet_total:
            new_row.append(None)
            new_row.append(None)
        elif tipo_fila:
            # It's a subtotal (Total Renta, Total Enajenacion), use current_cod and current_especie
            new_row.append(current_cod)
            new_row.append(current_especie)
        else:
            # Regular transaction row
            new_row.append(current_cod if not cod else cod)
            new_row.append(current_especie if not name else name)
        
        # Add remaining columns (skip the original especie column)
        for i, val in enumerate(row_values):
            if i != especie_col:
                # Fix trailing negatives for numeric values
                if val and isinstance(val, str):
                    numeric = parse_numeric(val)
                    if numeric is not None:
                        val = numeric
                # Replace None with 0 for numeric columns
                col_in_new = len(new_row)
                if val is None and col_in_new in numeric_col_indices:
                    val = 0
                new_row.append(val)
        
        # Skip empty/separator rows (rows without cod and without any data)
        if final_tipo_fila != 'transaccion' or current_cod or any(v for v in new_row[3:] if v and v != 0):
            processed_rows.append(new_row)
    
    # Check if a sheet-level total already exists in the data
    total_name = get_sheet_total_name(sheet_name)
    has_existing_total = any(
        row[0] and str(row[0]).upper().startswith('TOTAL') and any(
            key in str(row[0]).upper() for key in ['TIT.PRIVADOS', 'RENTA FIJA', 'CAUCIONES']
        )
        for row in processed_rows
    )
    
    # Only add final TOTAL row if it doesn't already exist
    if not has_existing_total:
        total_row = [total_name, None, None]  # tipo_fila, no cod_especie, no especie
        
        # Calculate sums for numeric columns
        for col_idx in range(3, len(new_headers)):
            if col_idx in numeric_col_indices:
                col_sum = 0
                for row_data in processed_rows:
                    if col_idx < len(row_data) and row_data[col_idx] is not None:
                        try:
                            col_sum += float(row_data[col_idx])
                        except (ValueError, TypeError):
                            pass
                total_row.append(col_sum)
            else:
                total_row.append(None)
        
        processed_rows.append(total_row)
    
    # Get original dimensions before clearing
    original_max_row = ws.max_row
    original_max_col = ws.max_column
    
    # Clear and rewrite the worksheet
    for row in range(1, original_max_row + 1):
        for col in range(1, original_max_col + 5):
            ws.cell(row, col).value = None
    
    # Write new headers
    for col, header in enumerate(new_headers, 1):
        ws.cell(1, col).value = header
    
    # Write processed data
    for row_idx, row_data in enumerate(processed_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx).value = value
    
    # Delete excess rows if we have fewer processed rows than original
    new_max_row = len(processed_rows) + 1  # +1 for header
    if new_max_row < original_max_row:
        ws.delete_rows(new_max_row + 1, original_max_row - new_max_row)
    
    return ws


def extract_tipo_especie(especie_value: str) -> tuple[Optional[str], Optional[str]]:
    """
    Extract tipo_especie category from especie value in position sheets.
    
    Categories to detect:
    - TITULOS PRIVADOS LOCALES
    - TIT.PRIVADOS DEL EXTERIOR
    - RENTA FIJA EN DOLARES
    - RENTA FIJA EN PESOS
    - CASH
    - TOTAL (for total rows)
    
    Returns: (tipo_especie, cleaned_especie)
    """
    if not especie_value or not isinstance(especie_value, str):
        return (None, None)
    
    value = especie_value.strip()
    
    # Category patterns (in bold tags from markdown)
    categories = [
        ("TITULOS PRIVADOS LOCALES", ["TITULOS PRIVADOS LOCALES", "<b>TITULOS PRIVADOS LOCALES</b>"]),
        ("TIT.PRIVADOS DEL EXTERIOR", ["TIT.PRIVADOS DEL EXTERIOR", "TIT. PRIVADOS DEL EXTERIOR", "<b>TIT.PRIVADOS DEL EXTERIOR</b>"]),
        ("RENTA FIJA EN DOLARES", ["RENTA FIJA EN DOLARES", "<b>RENTA FIJA EN DOLARES</b>"]),
        ("RENTA FIJA EN PESOS", ["RENTA FIJA EN PESOS", "<b>RENTA FIJA EN PESOS</b>"]),
        ("CASH", ["CASH", "<b>CASH</b>"]),
        ("TOTAL", ["TOTAL", "<b>TOTAL</b>"]),
    ]
    
    for cat_name, patterns in categories:
        for pattern in patterns:
            if pattern.upper() in value.upper():
                return (cat_name, None)  # This row is a category header
    
    # Check for "INCREMENTOS/DECREMENTOS" section marker (should be ignored)
    if "INCREMENTOS" in value.upper() or "DECREMENTOS" in value.upper():
        return ("IGNORE", None)
    
    # Regular especie value
    # Clean HTML tags
    clean = re.sub(r'<[^>]+>', '', value).strip()
    return (None, clean if clean else None)


def is_header_row(row_values: List) -> bool:
    """
    Check if a row appears to be a header row (column names repeated).
    This happens when INCREMENTOS/DECREMENTOS section starts.
    """
    text_values = [str(v).lower() if v else "" for v in row_values[:5]]
    header_keywords = ['especie', 'detalle', 'custodia', 'cantidad', 'precio', 'fecha']
    matches = sum(1 for kw in header_keywords if any(kw in tv for tv in text_values))
    return matches >= 2


def process_position_sheet(ws: Worksheet, sheet_name: str, metadata: dict = None) -> Worksheet:
    """
    Post-process Position sheets (Posicion Inicial, Posicion Final):
    - Extract tipo_especie from category rows
    - Add tipo_especie column and propagate to subsequent rows
    - Remove category header rows from data
    - Ignore INCREMENTOS/DECREMENTOS section
    - Fix trailing negatives in numeric columns
    - Replace nulls with 0 in numeric columns
    - For Posicion Inicial: Add calculated TOTAL row at the end
    - For Posicion Final: Keep original TOTAL, TOTAL DE LA INVERSION and RESULTADO rows
    - Add fecha row at the top if metadata contains 'fecha'
    """
    console.print(f"  [dim]Processing {sheet_name}...[/dim]")
    
    if metadata is None:
        metadata = {}
    
    is_posicion_final = "final" in sheet_name.lower()
    
    # Read headers and find especie column
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    especie_col = None
    for i, h in enumerate(headers):
        if h and 'especie' in h.lower():
            especie_col = i
            break
    
    if especie_col is None:
        console.print(f"    [yellow]No 'Especie' column found[/yellow]")
        return ws
    
    # Build new headers with tipo_especie first
    new_headers = ['tipo_especie']
    numeric_col_indices = []
    for i, h in enumerate(headers):
        new_headers.append(h)
        if is_numeric_column(str(h) if h else ""):
            numeric_col_indices.append(len(new_headers) - 1)
    
    # Read all rows
    rows_data = []
    for row in range(2, ws.max_row + 1):
        row_values = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
        rows_data.append(row_values)
    
    # Process rows
    current_tipo = None
    processed_rows = []
    ignore_mode = False  # Flag to ignore INCREMENTOS/DECREMENTOS section
    
    for row_values in rows_data:
        # Check if this is a repeated header row (start of INCREMENTOS/DECREMENTOS section)
        if is_header_row(row_values):
            ignore_mode = True
            continue
        
        # Skip rows while in ignore mode
        if ignore_mode:
            # Check if we've reached a new category that's valid (not INCREMENTOS related)
            especie_value = row_values[especie_col] if especie_col < len(row_values) else None
            if especie_value:
                tipo, _ = extract_tipo_especie(str(especie_value))
                if tipo and tipo != "IGNORE" and tipo in ["TITULOS PRIVADOS LOCALES", "TIT.PRIVADOS DEL EXTERIOR", 
                                                           "RENTA FIJA EN DOLARES", "RENTA FIJA EN PESOS", "CASH", "TOTAL"]:
                    # This is a valid new section, exit ignore mode
                    ignore_mode = False
                    current_tipo = tipo
                    continue
            continue
        
        especie_value = row_values[especie_col] if especie_col < len(row_values) else None
        tipo, clean_especie = extract_tipo_especie(str(especie_value) if especie_value else "")
        
        if tipo == "IGNORE":
            ignore_mode = True
            continue
        
        if tipo:
            # This is a category header row - update current tipo, skip the row
            current_tipo = tipo
            continue
        
        # For Posicion Final: check if TOTAL/RESULTADO is in second column (Detalle)
        # These rows have empty first column but contain totals in the second column
        if is_posicion_final and not clean_especie:
            second_col_value = row_values[especie_col + 1] if (especie_col + 1) < len(row_values) else None
            if second_col_value:
                second_val = str(second_col_value).upper().strip()
                if 'TOTAL' in second_val or second_val == 'RESULTADO':
                    # Keep these summary rows - they have the label in second column
                    new_row = [None]  # No tipo_especie for summary rows
                    for i, val in enumerate(row_values):
                        if val and isinstance(val, str):
                            numeric = parse_numeric(val)
                            if numeric is not None:
                                val = numeric
                        new_row.append(val)
                    processed_rows.append(new_row)
                    continue
        
        if not clean_especie:
            # Empty row, skip
            continue
        
        # For Posicion Final: keep TOTAL, TOTAL DE LA INVERSION, RESULTADO rows
        # For Posicion Inicial: skip TOTAL lines (they'll be recalculated)
        first_val = str(clean_especie).upper() if clean_especie else ""
        if first_val.startswith('TOTAL') or first_val == 'RESULTADO':
            if is_posicion_final:
                # Keep these rows in Posicion Final - they come from the PDF
                new_row = [None]  # No tipo_especie for summary rows
                for i, val in enumerate(row_values):
                    if i == especie_col:
                        new_row.append(clean_especie)
                    else:
                        if val and isinstance(val, str):
                            numeric = parse_numeric(val)
                            if numeric is not None:
                                val = numeric
                        new_row.append(val)
                processed_rows.append(new_row)
            continue
        
        # Build new row with tipo_especie as first column
        new_row = [current_tipo]
        
        for i, val in enumerate(row_values):
            if i == especie_col:
                new_row.append(clean_especie)
            else:
                # Fix trailing negatives
                if val and isinstance(val, str):
                    numeric = parse_numeric(val)
                    if numeric is not None:
                        val = numeric
                # Replace None with 0 for numeric columns
                col_in_new = len(new_row)
                if val is None and col_in_new in numeric_col_indices:
                    val = 0
                new_row.append(val)
        
        processed_rows.append(new_row)
    
    # Add calculated TOTAL row only for Posicion Inicial
    # Posicion Final already has TOTAL rows from the PDF
    if not is_posicion_final:
        total_label = "TOTAL POSICION INICIAL"
        total_row = [total_label]
        
        # For each column, calculate sum if numeric
        for col_idx in range(1, len(new_headers)):
            if col_idx in numeric_col_indices:
                col_sum = 0
                for row_data in processed_rows:
                    if col_idx < len(row_data) and row_data[col_idx] is not None:
                        try:
                            col_sum += float(row_data[col_idx])
                        except (ValueError, TypeError):
                            pass
                total_row.append(col_sum)
            else:
                total_row.append(None)
        
        processed_rows.append(total_row)
    
    # Get original dimensions before clearing
    original_max_row = ws.max_row
    original_max_col = ws.max_column
    
    # Clear and rewrite
    for row in range(1, original_max_row + 1):
        for col in range(1, original_max_col + 5):
            ws.cell(row, col).value = None
    
    # Write headers
    for col, header in enumerate(new_headers, 1):
        ws.cell(1, col).value = header
    
    # Write data
    for row_idx, row_data in enumerate(processed_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx).value = value
    
    # Delete excess rows if we have fewer processed rows than original
    new_max_row = len(processed_rows) + 1  # +1 for header
    if new_max_row < original_max_row:
        ws.delete_rows(new_max_row + 1, original_max_row - new_max_row)
    
    # Add fecha in the first empty column if metadata contains it
    if metadata and 'fecha' in metadata:
        fecha = metadata['fecha']
        
        # Find first empty column (check if all cells in that column are empty)
        fecha_col = None
        for col_idx in range(1, len(new_headers) + 10):  # Check up to 10 columns beyond headers
            # Check if this column is empty (all data rows have None/empty)
            is_empty = True
            for row_idx in range(2, new_max_row + 1):  # Start from row 2 (after header)
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value is not None and str(cell_value).strip() != '':
                    is_empty = False
                    break
            
            if is_empty:
                fecha_col = col_idx
                break
        
        # If we found an empty column, add fecha to all data rows
        if fecha_col:
            # Add header
            ws.cell(1, fecha_col).value = "fecha"
            
            # Add fecha value to all data rows
            for row_idx in range(2, new_max_row + 1):
                ws.cell(row_idx, fecha_col).value = fecha
    
    return ws


def postprocess_gallo_workbook(wb: Workbook, tables: dict = None) -> Workbook:
    """
    Apply all Gallo format post-processing to a workbook.
    
    Args:
        wb: The workbook to process
        tables: Optional dict of TableData objects with metadata
    """
    console.print("\n[cyan]📐 Post-processing Gallo format...[/cyan]")
    
    # Process each sheet type
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lower = sheet_name.lower()
        
        # Get metadata for this sheet if available
        metadata = None
        if tables and sheet_name in tables:
            metadata = tables[sheet_name].metadata
        
        if sheet_lower == 'preciotenenciasiniciales':
            process_precio_tenencias_sheet(ws)
        elif 'resultado' in sheet_lower and 'total' in sheet_lower:
            process_resultado_totales(ws)
        
        elif any(x in sheet_lower for x in ['privados', 'renta fija', 'cauciones']):
            process_detail_sheet(ws, sheet_name)
        
        elif 'posicion' in sheet_lower:
            process_position_sheet(ws, sheet_name, metadata)
    
    console.print("[green]✓ Post-processing complete[/green]")
    return wb


def is_integer_column(header: str) -> bool:
    """
    Determine if a column should contain integer values (no decimals).
    """
    if not header:
        return False
    h = header.lower()
    integer_keywords = ['boleto', 'cod.instrum', 'codigo', 'nro', 'ndc', 'plaz']
    for kw in integer_keywords:
        if kw in h:
            return True
    return False


def process_visual_sheet(
    ws: Worksheet,
    sheet_name: str,
    quantity_anchors: Optional[Dict[Tuple[int, str, str, str], int]] = None,
    code_anchor_evidence: Optional[Dict[int, int]] = None,
    code_mixed_magnitude: Optional[Dict[int, bool]] = None,
) -> None:
    """
    Process a Visual format sheet:
    - Convert parentheses to negative numbers
    - Convert numeric strings to actual numbers
    - Replace None/empty with 0 in numeric columns
    - Convert Nro. Boleto and Cod.Instrum to integers
    """
    console.print(f"  [dim]Processing {sheet_name}...[/dim]")
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers_lower = [str(h or '').strip().lower() for h in headers]

    cantidad_col = None
    precio_col = None
    bruto_col = None
    boleto_col = None
    moneda_col = None
    codigo_col = None
    gastos_col = None
    neto_col = None
    # Find cantidad column for ALL sheets (integer enforcement is universal)
    for idx, header_str in enumerate(headers_lower, start=1):
        if header_str == 'cantidad':
            cantidad_col = idx
        elif header_str == 'precio':
            precio_col = idx
        elif header_str == 'bruto':
            bruto_col = idx
    if sheet_name.lower() == 'boletos':
        for idx, header_str in enumerate(headers_lower, start=1):
            if 'boleto' in header_str:
                boleto_col = idx
            elif header_str == 'moneda':
                moneda_col = idx
            elif 'cod.instru' in header_str or header_str == 'cod.instrum':
                codigo_col = idx
            elif 'gasto' in header_str:
                gastos_col = idx
            elif 'neto' in header_str:
                neto_col = idx

    if sheet_name.lower() == 'boletos':
        tipo_col = None
        oper_col = None
        for idx, header in enumerate(headers, start=1):
            header_str = str(header or '').strip().lower()
            if header_str == 'tipo de instrumento':
                tipo_col = idx
            elif header_str == 'tipo operación' or header_str == 'tipo operacion':
                oper_col = idx

        if tipo_col and oper_col:
            for row in range(2, ws.max_row + 1):
                tipo_val = str(ws.cell(row, tipo_col).value or '').strip()
                oper_val = str(ws.cell(row, oper_col).value or '').strip().lower()
                needs_reclass = (not tipo_val) or ('sin datos' in tipo_val.lower())
                if not needs_reclass or not oper_val:
                    continue

                if 'futuro' in oper_val:
                    ws.cell(row, tipo_col).value = 'Futuros'
                elif 'opción' in oper_val or 'opcion' in oper_val:
                    ws.cell(row, tipo_col).value = 'Opciones'
                elif 'ejercicio' in oper_val:
                    ws.cell(row, tipo_col).value = 'Acciones'
    
    for row in range(2, ws.max_row + 1):
        raw_row_values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            header = headers[col - 1] if col <= len(headers) else None
            header_str = str(header).lower() if header else ""
            
            if isinstance(cell.value, str):
                val = cell.value.strip()
                
                if val:  # Non-empty string
                    # Quantities in Visual often need integer-aware parsing.
                    if header_str == 'cantidad':
                        numeric = parse_visual_quantity_value(val)
                    else:
                        numeric = parse_parentheses_negative(val)
                    if numeric is not None:
                        # Check if it should be an integer
                        if is_integer_column(header_str):
                            cell.value = int(numeric)
                        else:
                            cell.value = numeric
                else:  # Empty string
                    # Replace empty with 0 in numeric columns
                    if is_numeric_column(header_str) or is_integer_column(header_str):
                        cell.value = 0
                    else:
                        cell.value = None  # Keep as None for non-numeric
            elif cell.value is None:
                # Replace None with 0 in numeric columns
                if is_numeric_column(header_str) or is_integer_column(header_str):
                    cell.value = 0

        # Zero out gastos when it is nearly equal to bruto (column-width OCR artifact).
        if gastos_col and bruto_col:
            try:
                _g = float(ws.cell(row, gastos_col).value or 0)
                _b = float(ws.cell(row, bruto_col).value or 0)
                if _b != 0 and _g != 0 and abs(_g / _b) > 0.9:
                    ws.cell(row, gastos_col).value = 0
            except (ValueError, TypeError):
                pass

        if cantidad_col and precio_col and bruto_col:
            raw_qty = raw_row_values[cantidad_col - 1]
            qty_value = ws.cell(row, cantidad_col).value
            precio_value = ws.cell(row, precio_col).value
            bruto_value = ws.cell(row, bruto_col).value

            if isinstance(raw_qty, str):
                best_qty = _choose_best_boletos_quantity(raw_qty, qty_value, precio_value, bruto_value)
                ws.cell(row, cantidad_col).value = best_qty
                qty_value = best_qty

            # When qty was corrected (e.g. truncated-thousands → integer),
            # bruto and neto may also be truncated by the same column width.
            # Rescue them using qty × precio.
            if isinstance(raw_qty, str) and sheet_name.lower() == 'boletos':
                rescued_bruto = _maybe_rescue_visual_bruto(bruto_value, qty_value, precio_value)
                if rescued_bruto != bruto_value:
                    ws.cell(row, bruto_col).value = rescued_bruto
                    bruto_value = rescued_bruto
                    if neto_col and gastos_col:
                        ws.cell(row, neto_col).value = _maybe_rescue_visual_neto(
                            ws.cell(row, neto_col).value,
                            qty_value,
                            precio_value,
                            ws.cell(row, gastos_col).value,
                        )

            if sheet_name.lower() == 'boletos' and quantity_anchors and codigo_col and moneda_col:
                anchor_qty = _lookup_visual_quantity_anchor(
                    quantity_anchors,
                    ws.cell(row, codigo_col).value,
                    ws.cell(row, 2).value,
                    ws.cell(row, moneda_col).value,
                    ws.cell(row, 6).value,
                )
                if _should_apply_visual_anchor(
                    raw_qty,
                    qty_value,
                    anchor_qty,
                    code_anchor_evidence=(code_anchor_evidence or {}).get(_normalize_visual_code(ws.cell(row, codigo_col).value) or -1, 0),
                    operacion=ws.cell(row, 6).value,
                    code_mixed_magnitude=(code_mixed_magnitude or {}).get(_normalize_visual_code(ws.cell(row, codigo_col).value) or -1, False),
                ):
                    ws.cell(row, cantidad_col).value = anchor_qty
                    qty_value = anchor_qty
                    ws.cell(row, bruto_col).value = _maybe_rescue_visual_bruto(bruto_value, qty_value, precio_value)
                    if neto_col and gastos_col:
                        ws.cell(row, neto_col).value = _maybe_rescue_visual_neto(
                            ws.cell(row, neto_col).value,
                            qty_value,
                            precio_value,
                            ws.cell(row, gastos_col).value,
                        )

        # Universal integer enforcement: quantities are always integers.
        # Runs for ALL sheets with a cantidad column, after all rescue attempts.
        if cantidad_col:
            qty_value = ws.cell(row, cantidad_col).value
            try:
                qty_float = float(qty_value or 0)
            except (ValueError, TypeError):
                qty_float = 0.0
            if qty_float != 0 and not qty_float.is_integer():
                precio_for_derive = ws.cell(row, precio_col).value if precio_col else None
                bruto_for_derive = ws.cell(row, bruto_col).value if bruto_col else None
                derived = _derive_integer_quantity_from_bruto(qty_float, precio_for_derive, bruto_for_derive)
                if derived is not None:
                    ws.cell(row, cantidad_col).value = derived
                else:
                    ws.cell(row, cantidad_col).value = round(qty_float)


def _derive_integer_quantity_from_bruto(qty, precio, bruto) -> Optional[int]:
    """Derive integer quantity from bruto/precio when OCR truncated the cantidad."""
    try:
        precio_num = abs(float(precio or 0))
        bruto_num = abs(float(bruto or 0))
    except (ValueError, TypeError):
        return None
    if precio_num == 0 or bruto_num == 0:
        return None
    sign = -1 if float(qty) < 0 else 1
    qty_int = round(bruto_num / precio_num)
    if qty_int == 0:
        return None
    # Validate: derived qty × precio must be within 5% of bruto
    expected = qty_int * precio_num
    if bruto_num > 0 and abs(expected - bruto_num) / bruto_num <= 0.05:
        return sign * qty_int
    return None


def process_precio_tenencias_sheet(ws: Worksheet) -> None:
    """
    Procesa la hoja PrecioTenenciasIniciales:
    - Divide la columna Especie en: Cod.Especie, Ticker, Especie
    - Calcula Precio Tenencia Inicial = Importe invertido / Cantidad
    """
    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    headers_lower = [h.lower() for h in headers]

    def find_col(keyword: str) -> Optional[int]:
        for idx, h in enumerate(headers_lower, start=1):
            if keyword in h:
                return idx
        return None

    especie_col = find_col('especie')
    cantidad_col = find_col('cantidad')
    importe_col = find_col('importe')
    resultado_col = find_col('resultado')
    cod_col = find_col('cod')
    ticker_col = find_col('ticker')
    precio_col = find_col('precio tenencia')

    if not especie_col or not cantidad_col or not importe_col:
        return

    new_headers = [
        'Cod.Especie',
        'Ticker',
        'Especie',
        'Cantidad tenencia',
        'Importe invertido',
        'Resultado',
        'Precio tenencia inicial'
    ]

    def to_float(val) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            numeric = parse_parentheses_negative(str(val).strip())
            if numeric is not None:
                return float(numeric)
        except Exception:
            pass
        try:
            numeric = parse_numeric(str(val).strip())
            if numeric is not None:
                return float(numeric)
        except Exception:
            pass
        try:
            return float(str(val).replace('.', '').replace(',', '.'))
        except Exception:
            return 0.0

    def parse_cantidad_tenencia(val) -> float:
        """Corrige cantidades con tres ceros decimales (ej: 844.000 -> 844)."""
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        raw = str(val).strip()
        # Detectar formato con .000 o ,000 al final
        if re.match(r'^\d+[\.,]000$', raw):
            raw = re.sub(r'[\.,]000$', '', raw)
            try:
                return float(raw)
            except Exception:
                return 0.0
        return to_float(raw)

    def _clean_codigo(val: str) -> str:
        if val is None:
            return ""
        s = str(val).strip()
        if s.endswith('.0'):
            s = s[:-2]
        s = s.replace('.', '').replace(',', '')
        try:
            return str(int(float(s)))
        except Exception:
            return s

    def _normalize_ratio_key(val: str) -> str:
        if not val:
            return ""
        return re.sub(r"[^A-Z0-9]", "", str(val).strip().upper())

    def _load_ratio_cache() -> dict:
        try:
            aux_path = Path(__file__).parent / 'aux_data' / 'RatiosCedearsAcciones.xlsx'
            if not aux_path.exists():
                return {}
            wb_ratios = load_workbook(aux_path)
            ws_ratios = wb_ratios.active
            cache = {}
            for r in range(2, ws_ratios.max_row + 1):
                nombre = ws_ratios.cell(r, 1).value
                ratio_val = ws_ratios.cell(r, 2).value
                key = ws_ratios.cell(r, 3).value
                if ratio_val is None:
                    continue
                try:
                    ratio_num = float(ratio_val)
                except Exception:
                    continue
                if key:
                    normalized_key = _normalize_ratio_key(key)
                    if normalized_key:
                        cache[normalized_key] = ratio_num
                if nombre:
                    nombre_str = str(nombre).strip()
                    nombre_key = _normalize_ratio_key(nombre_str.split()[0])
                    if nombre_key:
                        cache.setdefault(nombre_key, ratio_num)
                    # Extract stock ticker from Nombre (format: "Company Name TICKER EXCHANGE")
                    tokens = nombre_str.split()
                    if len(tokens) >= 2:
                        ticker_candidate = tokens[-2]
                        ticker_key = _normalize_ratio_key(ticker_candidate)
                        if ticker_key and len(ticker_key) <= 6:
                            cache.setdefault(ticker_key, ratio_num)
            return cache
        except Exception:
            return {}

    def _load_acciones_exterior_codigos() -> set:
        try:
            aux_path = Path(__file__).parent / 'aux_data' / 'EspeciesVisual.xlsx'
            if not aux_path.exists():
                return set()
            wb_especies = load_workbook(aux_path)
            ws_especies = wb_especies.active
            cods = set()
            for r in range(2, ws_especies.max_row + 1):
                codigo = ws_especies.cell(r, 3).value  # Col C
                moneda_emision = ws_especies.cell(r, 7).value  # Col G
                tipo_especie = ws_especies.cell(r, 18).value  # Col R
                if not codigo:
                    continue
                if str(moneda_emision).strip() == "Dolar Cable (exterior)" and str(tipo_especie).strip() == "Acciones":
                    cods.add(_clean_codigo(codigo))
            return cods
        except Exception:
            return set()

    ratio_cache = _load_ratio_cache()
    acciones_exterior_codigos = _load_acciones_exterior_codigos()

    # Si ya está estructurada (Cod/Ticker/Precio Tenencia), solo recalcular y ajustar ratio
    if cod_col and ticker_col and precio_col:
        for row in range(2, ws.max_row + 1):
            cod = ws.cell(row, cod_col).value or ""
            ticker = ws.cell(row, ticker_col).value or ""
            nombre = ws.cell(row, especie_col).value or ""

            cantidad_val = ws.cell(row, cantidad_col).value
            importe_val = ws.cell(row, importe_col).value
            resultado_val = ws.cell(row, resultado_col).value if resultado_col else 0

            cantidad_num = parse_cantidad_tenencia(cantidad_val)
            importe_num = to_float(importe_val)
            resultado_num = to_float(resultado_val)

            # Fix invalid rows: cantidad > 0 but importe <= 0
            if cantidad_num > 0 and importe_num <= 0:
                if importe_num == 0:
                    # Use resultado / cantidad as fallback price
                    precio_tenencia = abs(resultado_num / cantidad_num) if cantidad_num else 0
                    # Also fix the importe cell to resultado so the sheet is consistent
                    ws.cell(row, importe_col, value=abs(resultado_num))
                else:
                    # Negative importe: flip sign
                    importe_num = abs(importe_num)
                    ws.cell(row, importe_col, value=importe_num)
                    precio_tenencia = importe_num / cantidad_num
            else:
                precio_tenencia = (importe_num / cantidad_num) if cantidad_num else 0

            cod_clean = _clean_codigo(cod)
            if cod_clean in acciones_exterior_codigos:
                search_text = f"{ticker} {nombre}".strip().upper()
                key = _normalize_ratio_key(search_text.split()[0]) if search_text else ""
                ratio = ratio_cache.get(key)
                # Try ticker without -US/-D suffix
                if not ratio and ticker:
                    base_ticker = str(ticker).strip().upper().split('-')[0]
                    ratio = ratio_cache.get(_normalize_ratio_key(base_ticker))
                if ratio and cantidad_num:
                    precio_tenencia = (importe_num / cantidad_num) / ratio

            ws.cell(row, precio_col, value=precio_tenencia)

        return

    rows = []
    for row in range(2, ws.max_row + 1):
        especie_val = ws.cell(row, especie_col).value or ""
        tokens = str(especie_val).strip().split()
        cod = tokens[0] if len(tokens) > 0 else ""
        ticker = tokens[1] if len(tokens) > 1 else ""
        nombre = " ".join(tokens[2:]) if len(tokens) > 2 else ""

        cantidad_val = ws.cell(row, cantidad_col).value
        importe_val = ws.cell(row, importe_col).value
        resultado_val = ws.cell(row, resultado_col).value if resultado_col else 0

        cantidad_num = parse_cantidad_tenencia(cantidad_val)
        importe_num = to_float(importe_val)
        resultado_num = to_float(resultado_val)

        # Fix invalid rows: cantidad > 0 but importe <= 0
        if cantidad_num > 0 and importe_num <= 0:
            if importe_num == 0:
                # Use resultado / cantidad as fallback price
                precio_tenencia = abs(resultado_num / cantidad_num) if cantidad_num else 0
                importe_num = abs(resultado_num)
            else:
                # Negative importe: flip sign
                importe_num = abs(importe_num)
                precio_tenencia = importe_num / cantidad_num
        else:
            precio_tenencia = (importe_num / cantidad_num) if cantidad_num else 0

        # Ajuste por ratio para Acciones del Exterior (CEDEAR)
        cod_clean = _clean_codigo(cod)
        if cod_clean in acciones_exterior_codigos:
            search_text = f"{ticker} {nombre}".strip().upper()
            key = _normalize_ratio_key(search_text.split()[0]) if search_text else ""
            ratio = ratio_cache.get(key)
            # Try ticker without -US/-D suffix
            if not ratio and ticker:
                base_ticker = str(ticker).strip().upper().split('-')[0]
                ratio = ratio_cache.get(_normalize_ratio_key(base_ticker))
            if ratio and cantidad_num:
                precio_tenencia = (importe_num / cantidad_num) / ratio

        rows.append([
            cod,
            ticker,
            nombre,
            cantidad_num,
            importe_num,
            to_float(resultado_val),
            precio_tenencia,
        ])

    # Limpiar hoja y escribir nueva estructura
    ws.delete_rows(1, ws.max_row)

    for col_idx, header in enumerate(new_headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def postprocess_visual_workbook(wb: Workbook) -> Workbook:
    """
    Apply Visual format post-processing to a workbook.
    """
    console.print("\n[cyan]📐 Post-processing Visual format...[/cyan]")

    # First normalize all non-Boletos sheets so Resultado Ventas carries stable integer anchors.
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == 'boletos':
            continue

        ws = wb[sheet_name]
        process_visual_sheet(ws, sheet_name)
        if sheet_name.lower() == 'preciotenenciasiniciales':
            process_precio_tenencias_sheet(ws)

    quantity_anchors = _build_visual_quantity_anchors(wb)
    code_anchor_evidence = _build_visual_code_anchor_evidence(wb['Boletos']) if 'Boletos' in wb.sheetnames else {}
    code_mixed_magnitude = _build_code_mixed_magnitude(wb['Boletos']) if 'Boletos' in wb.sheetnames else {}

    if 'Boletos' in wb.sheetnames:
        process_visual_sheet(
            wb['Boletos'],
            'Boletos',
            quantity_anchors=quantity_anchors,
            code_anchor_evidence=code_anchor_evidence,
            code_mixed_magnitude=code_mixed_magnitude,
        )
    
    console.print("[green]✓ Post-processing complete[/green]")
    return wb
