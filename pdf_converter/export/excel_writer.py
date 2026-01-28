"""
Excel Writer for generating structured XLSX files.
Supports multiple sheets with proper formatting.
"""

from typing import List, Dict, Any, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from rich.console import Console

console = Console()


class ExcelWriter:
    """
    Excel Writer for creating multi-sheet workbooks with formatting.
    """
    
    # Default styles
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Number format for currency
    NUMBER_FORMAT = '#,##0.00'
    PERCENT_FORMAT = '0.00%'
    
    def __init__(self):
        self.wb = Workbook()
        # Remove default sheet
        if self.wb.active:
            self.wb.remove(self.wb.active)
        
        self.sheets_added = []
    
    def add_sheet(
        self,
        name: str,
        data: List[Dict],
        schema: List[str],
        numeric_fields: Optional[List[str]] = None
    ):
        """
        Add a sheet with data and formatting.
        
        Args:
            name: Sheet name
            data: List of row dictionaries
            schema: List of column names in order
            numeric_fields: Fields that should be formatted as numbers
        """
        if not data:
            console.print(f"[dim]  Skipping empty sheet: {name}[/dim]")
            return
        
        # Clean sheet name (Excel limits to 31 chars)
        clean_name = name[:31]
        
        # Create sheet
        ws = self.wb.create_sheet(clean_name)
        self.sheets_added.append(clean_name)
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Reorder columns to match schema
        available_cols = [c for c in schema if c in df.columns]
        missing_cols = [c for c in schema if c not in df.columns]
        
        if missing_cols:
            console.print(f"[dim]  Missing columns in {name}: {missing_cols}[/dim]")
        
        # Add missing columns with default values
        for col in missing_cols:
            df[col] = 0 if numeric_fields and col in numeric_fields else ""
        
        df = df[schema]
        
        # Write headers
        for col_idx, col_name in enumerate(schema, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, value) in enumerate(zip(schema, row), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                
                # Format based on column type
                if numeric_fields and col_name in numeric_fields:
                    cell.alignment = self.NUMBER_ALIGNMENT
                    if isinstance(value, (int, float)):
                        cell.number_format = self.NUMBER_FORMAT
                else:
                    cell.alignment = self.DATA_ALIGNMENT
        
        # Auto-fit columns
        self._auto_fit_columns(ws, schema, df)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        console.print(f"  [green]✓ Sheet '{clean_name}' created with {len(data)} rows[/green]")
    
    def _auto_fit_columns(self, ws, schema: List[str], df: pd.DataFrame):
        """Auto-fit column widths based on content."""
        for col_idx, col_name in enumerate(schema, 1):
            # Calculate max width
            max_length = len(col_name)
            
            if col_name in df.columns:
                col_data = df[col_name].astype(str)
                if len(col_data) > 0:
                    max_data_length = col_data.str.len().max()
                    if pd.notna(max_data_length):
                        max_length = max(max_length, int(max_data_length))
            
            # Set width with some padding, max 50 chars
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    def add_validation_sheet(self, validation_results: List[Dict]):
        """
        Add a validation summary sheet.
        
        Args:
            validation_results: List of validation result dictionaries
        """
        if not validation_results:
            return
        
        ws = self.wb.create_sheet("Validación")
        self.sheets_added.append("Validación")
        
        # Headers
        headers = ["Campo", "Calculado", "Esperado", "Diferencia", "Match"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # Data
        for row_idx, result in enumerate(validation_results, 2):
            ws.cell(row=row_idx, column=1, value=result.get("field", ""))
            ws.cell(row=row_idx, column=2, value=result.get("calculated", 0)).number_format = self.NUMBER_FORMAT
            ws.cell(row=row_idx, column=3, value=result.get("expected", 0)).number_format = self.NUMBER_FORMAT
            ws.cell(row=row_idx, column=4, value=result.get("difference", 0)).number_format = self.NUMBER_FORMAT
            
            match_cell = ws.cell(row=row_idx, column=5, value="✓" if result.get("match") else "✗")
            if not result.get("match"):
                match_cell.font = Font(color="FF0000", bold=True)
            else:
                match_cell.font = Font(color="00AA00", bold=True)
            
            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).border = self.THIN_BORDER
        
        # Auto-fit
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 10
        
        ws.freeze_panes = "A2"
    
    def save(self, path: str):
        """
        Save the workbook to a file.
        
        Args:
            path: Output file path
        """
        output_path = Path(path)
        
        # Ensure directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Add at least one sheet if empty
        if not self.sheets_added:
            ws = self.wb.create_sheet("Info")
            ws.cell(row=1, column=1, value="No data extracted")
        
        self.wb.save(str(output_path))
        console.print(f"\n[bold green]📥 Excel saved: {output_path}[/bold green]")
    
    def get_sheets_count(self) -> int:
        """Return number of sheets added."""
        return len(self.sheets_added)


def create_excel_from_data(
    data: Dict[str, List[Dict]],
    output_path: str,
    report_type: str,
    schemas: Dict[str, List[str]],
    sheet_names: Dict[str, str],
    numeric_fields: Dict[str, List[str]],
    validation_results: Optional[List[Dict]] = None
) -> str:
    """
    Create an Excel file from extracted data.
    
    Args:
        data: Dictionary with section_key -> list of rows
        output_path: Output file path
        report_type: "gallo" or "visual"
        schemas: Dictionary with section_key -> list of column names
        sheet_names: Dictionary with section_key -> sheet name
        numeric_fields: Dictionary with section_key -> list of numeric field names
        validation_results: Optional validation results to add as a sheet
    
    Returns:
        Path to the created Excel file
    """
    writer = ExcelWriter()
    
    console.print(f"\n[cyan]Creating Excel file...[/cyan]")
    
    # Define section order
    if report_type == "gallo":
        section_order = [
            "resultado_totales",
            "tit_privados_exentos",
            "tit_privados_exterior",
            "renta_fija_pesos",
            "renta_fija_dolares",
            "fci",
            "opciones",
            "futuros",
            "cauciones_pesos",
            "cauciones_dolares",
            "posicion_inicial",
            "posicion_final",
        ]
    else:
        section_order = [
            "resumen",
            "boletos",
            "resultado_ventas_ars",
            "resultado_ventas_usd",
            "rentas_dividendos_ars",
            "rentas_dividendos_usd",
            "posicion_titulos",
        ]
    
    # Add sheets in order
    for section_key in section_order:
        if section_key in data and data[section_key]:
            sheet_name = sheet_names.get(section_key, section_key)
            schema = schemas.get(section_key, list(data[section_key][0].keys()) if data[section_key] else [])
            numeric = numeric_fields.get(section_key, [])
            
            writer.add_sheet(sheet_name, data[section_key], schema, numeric)
    
    # Add validation sheet
    if validation_results:
        writer.add_validation_sheet(validation_results)
    
    # Save
    writer.save(output_path)
    
    return output_path
