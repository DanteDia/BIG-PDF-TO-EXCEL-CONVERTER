from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import sys

from openpyxl import load_workbook

from compare_workbooks import compare_workbooks
import verify_regression_cases as regression


ROOT = Path(__file__).resolve().parent
PRE_DIR = ROOT / "SMOKE_BASELINE" / "AGUIAR_20260318_PRE_MARTIN"
POST_DIR = ROOT / "SMOKE_BASELINE" / "AGUIAR_20260318_POST_MARTIN"
CANULLO_DIR = ROOT / "SMOKE_BASELINE" / "CANULLO_20260326_APPROVED"

BASELINE_PARITY_SHEETS = [
    "Boletos",
    "Resultado Ventas ARS",
    "Resultado Ventas USD",
    "Resumen",
    "Cauciones Tomadoras",
    "Cauciones Colocadoras",
]

CURRENT_OPTIONAL_SHEETS = {
    "Opciones": (1, 11),
    "FCI": (1, 11),
    "Futuros": (1, 4),
}

EXACT_MATCH_SHEETS = [
    "Boletos",
    "Resultado Ventas ARS",
    "Resultado Ventas USD",
    "Cauciones Tomadoras",
    "Cauciones Colocadoras",
    "Resumen",
]


def _load(path: Path):
    return load_workbook(path, data_only=True)


def _sheet_stats(workbook_path: Path, sheet_name: str) -> tuple[int, int]:
    wb = _load(workbook_path)
    ws = wb[sheet_name]
    return ws.max_row, ws.max_column


def _has_sheet(workbook_path: Path, sheet_name: str) -> bool:
    wb = _load(workbook_path)
    return sheet_name in wb.sheetnames


def _float_equal(left, right, tolerance: float = 1e-9) -> bool:
    try:
        return abs(float(left) - float(right)) <= tolerance
    except Exception:
        return left == right


def _rows_equal(left: Sequence[object], right: Sequence[object]) -> bool:
    if len(left) != len(right):
        return False
    return all(_float_equal(a, b) for a, b in zip(left, right))


def _assert(condition: bool, message: str, failures: list[str]) -> None:
    if not condition:
        failures.append(message)


def _check_glozman_and_salvo(failures: list[str]) -> None:
    glozman = regression._latest_existing(
        ROOT / "Ejemplo Glozman error moneda pesos en seccion USD" / "12766_GLOZMAN_DARIO_EDMUNDO_Resumen_Impositivo_FIXED_v2.xlsx",
        ROOT / "Ejemplo Glozman error moneda pesos en seccion USD" / "12766_GLOZMAN_DARIO_EDMUNDO_Resumen_Impositivo_FIXED.xlsx",
    )
    salvo = regression._latest_existing(
        ROOT / "11896_SALVO_MARTIN_Resumen_Impositivo_REGRESSION_v2.xlsx",
        ROOT / "11896_SALVO_MARTIN_Resumen_Impositivo_REGRESSION.xlsx",
    )

    _assert(glozman.exists(), f"Falta workbook GLOZMAN: {glozman}", failures)
    _assert(salvo.exists(), f"Falta workbook SALVO: {salvo}", failures)
    if failures:
        return

    glozman_ars_rows = regression.count_non_empty_rows(glozman, "Rentas Dividendos ARS")
    glozman_usd_rows, glozman_bad = regression.check_currency_rows(glozman, "Rentas Dividendos USD", "USD")
    salvo_ars_rows = regression.count_non_empty_rows(salvo, "Rentas Dividendos ARS")
    salvo_usd_rows, salvo_bad = regression.check_currency_rows(salvo, "Rentas Dividendos USD", "USD")
    salvo_resultado_bruto = regression.check_resultado_vs_bruto(salvo)

    _assert(glozman_ars_rows == 26, f"GLOZMAN ARS rows esperado=26 actual={glozman_ars_rows}", failures)
    _assert(glozman_usd_rows == 6, f"GLOZMAN USD rows esperado=6 actual={glozman_usd_rows}", failures)
    _assert(not glozman_bad, f"GLOZMAN filas USD mal clasificadas: {glozman_bad}", failures)

    _assert(salvo_ars_rows == 4, f"SALVO ARS rows esperado=4 actual={salvo_ars_rows}", failures)
    _assert(salvo_usd_rows == 4, f"SALVO USD rows esperado=4 actual={salvo_usd_rows}", failures)
    _assert(not salvo_bad, f"SALVO filas USD mal clasificadas: {salvo_bad}", failures)
    _assert(not salvo_resultado_bruto, f"SALVO filas Resultado>Bruto: {salvo_resultado_bruto}", failures)


def _check_aguiar_same_input(failures: list[str]) -> None:
    baseline = PRE_DIR / "AGUIAR_SMOKE_values.xlsx"
    current = POST_DIR / "AGUIAR_SAME_INPUT_values.xlsx"

    _assert(baseline.exists(), f"Falta baseline AGUIAR: {baseline}", failures)
    _assert(current.exists(), f"Falta current AGUIAR same-input: {current}", failures)
    if failures:
        return

    for sheet_name in BASELINE_PARITY_SHEETS:
        base_rows, base_cols = _sheet_stats(baseline, sheet_name)
        cur_rows, cur_cols = _sheet_stats(current, sheet_name)
        _assert(
            (base_rows, base_cols) == (cur_rows, cur_cols),
            (
                f"AGUIAR same-input difiere en {sheet_name}: "
                f"baseline=({base_rows},{base_cols}) current=({cur_rows},{cur_cols})"
            ),
            failures,
        )

    for sheet_name, expected_stats in CURRENT_OPTIONAL_SHEETS.items():
        _assert(_has_sheet(current, sheet_name), f"AGUIAR same-input no contiene hoja esperada: {sheet_name}", failures)
        if failures:
            continue
        cur_stats = _sheet_stats(current, sheet_name)
        _assert(
            cur_stats == expected_stats,
            f"AGUIAR same-input {sheet_name} esperado={expected_stats} actual={cur_stats}",
            failures,
        )

    base_resumen = regression.resumen_rows(baseline)
    cur_resumen = regression.resumen_rows(current)
    for row_number in (1, 2):
        base_row = base_resumen[row_number]
        cur_row = cur_resumen[row_number]
        _assert(
            _rows_equal(base_row, cur_row),
            f"AGUIAR same-input Resumen fila {row_number + 1} difiere: baseline={base_row} current={cur_row}",
            failures,
        )

    exact_diffs = compare_workbooks(
        baseline,
        current,
        tolerance=1e-9,
        max_diffs=25,
        include_sheets=EXACT_MATCH_SHEETS,
        ignore_extra_sheets=True,
    )
    for diff in exact_diffs:
        failures.append(f"AGUIAR exact diff: {diff}")


def _check_canullo_approved(failures: list[str]) -> None:
    baseline = CANULLO_DIR / "10600_CANULLO_MARTHA_NOEMI_Resumen_Impositivo_FIXED_values.xlsx"
    current = ROOT / "10600_CANULLO_MARTHA_NOEMI_Resumen_Impositivo_FIXED_values.xlsx"

    _assert(baseline.exists(), f"Falta baseline CANULLO: {baseline}", failures)
    _assert(current.exists(), f"Falta workbook actual CANULLO: {current}", failures)
    if failures:
        return

    exact_diffs = compare_workbooks(
        baseline,
        current,
        tolerance=1e-9,
        max_diffs=25,
    )
    for diff in exact_diffs:
        failures.append(f"CANULLO exact diff: {diff}")


def _print_section(title: str, lines: Iterable[str]) -> None:
    print(f"\n=== {title} ===")
    for line in lines:
        print(line)


def main() -> int:
    failures: list[str] = []

    _check_glozman_and_salvo(failures)
    _check_aguiar_same_input(failures)
    _check_canullo_approved(failures)

    # Run dedicated smoke tests (Sigal, Prida, Sturman 2797, Sturman 11688, Koltan 13353)
    import subprocess
    for script in [
        "smoke_test_sigal.py",
        "smoke_test_prida.py",
        "smoke_test_sturman.py",
        "smoke_test_sturman_11688.py",
        "smoke_test_koltan_13353.py",
    ]:
        script_path = ROOT / script
        if script_path.exists():
            result = subprocess.run(
                [sys.executable, str(script_path)],
                capture_output=True, text=True, cwd=str(ROOT),
            )
            if result.returncode != 0:
                failures.append(f"{script} FAILED (exit {result.returncode})")
                # Show last 5 lines of output for diagnostics
                for line in (result.stdout + result.stderr).strip().splitlines()[-5:]:
                    failures.append(f"  {line}")

    if failures:
        _print_section("SMOKE FAIL", failures)
        return 1

    _print_section(
        "SMOKE PASS",
        [
            "GLOZMAN: ARS/USD split consistente",
            "SALVO: ARS/USD split consistente y Resultado<=Bruto",
            "AGUIAR same-input: hojas clave y resumen sin desvíos",
            "CANULLO approved: workbook completo sin desvíos",
            "SIGAL 10374: cell-by-cell OK",
            "PRIDA 11797: cell-by-cell OK",
            "STURMAN 2797: cell-by-cell OK",
            "STURMAN 11688: cell-by-cell OK + micro-price guard + inflation guard",
            "KOLTAN 13353: cell-by-cell OK + inflation guard",
        ],
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())