from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable, Sequence

from openpyxl import load_workbook


DEFAULT_AUX_SHEETS = {
    "EspeciesVisual",
    "EspeciesGallo",
    "Cotizacion Dolar Historica",
    "PreciosInicialesEspecies",
    "PrecioTenenciasIniciales",
    "RatiosCedearsAcciones",
}

WorkbookGuard = Callable[[object], Sequence[str]]
SummaryBuilder = Callable[[object], Sequence[str]]
ExtraSaveAction = Callable[[], Sequence[str] | None]
PipelineRunner = Callable[[], object]


def _floats_equal(a, b, *, float_rtol: float, float_atol: float) -> bool:
    if a == 0 and b == 0:
        return True
    if a == 0 or b == 0:
        return abs(a - b) <= float_atol
    rel = abs(a - b) / max(abs(a), abs(b))
    return rel <= float_rtol or abs(a - b) <= float_atol


def values_equal(a, b, *, float_rtol: float, float_atol: float) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        if (a is None and b == 0) or (b is None and a == 0):
            return True
        if (a is None and b == "") or (b is None and a == ""):
            return True
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return _floats_equal(float(a), float(b), float_rtol=float_rtol, float_atol=float_atol)
    if isinstance(a, (datetime, date)) and isinstance(b, (datetime, date)):
        return a == b
    return str(a) == str(b)


def serialise_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        return round(value, 10)
    return value


def _normalize_audit_value(value):
    if not isinstance(value, str):
        return value
    marker = " | ALERTA: "
    if marker not in value:
        return value
    return value.split(marker, 1)[0].strip()


def snapshot_workbook_sheet(workbook, name: str) -> dict:
    worksheet = workbook[name]
    sheet = {
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "headers": [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)],
        "rows": {},
    }
    for row in range(2, worksheet.max_row + 1):
        sheet["rows"][str(row)] = [
            serialise_value(worksheet.cell(row, col).value)
            for col in range(1, worksheet.max_column + 1)
        ]
    return sheet


def build_snapshot(workbook, *, aux_sheets: set[str]) -> dict:
    snapshot = {}
    for name in workbook.sheetnames:
        if name in aux_sheets:
            worksheet = workbook[name]
            snapshot[name] = {
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "skipped": True,
            }
        else:
            snapshot[name] = snapshot_workbook_sheet(workbook, name)
    return snapshot


def compare_workbooks(
    baseline_workbook,
    current_workbook,
    *,
    aux_sheets: set[str],
    float_rtol: float,
    float_atol: float,
) -> list[dict]:
    diffs: list[dict] = []

    baseline_sheets = set(baseline_workbook.sheetnames)
    current_sheets = set(current_workbook.sheetnames)

    for sheet_name in sorted(baseline_sheets - current_sheets):
        diffs.append({"type": "SHEET_MISSING", "sheet": sheet_name})
    for sheet_name in sorted(current_sheets - baseline_sheets):
        diffs.append({"type": "SHEET_ADDED", "sheet": sheet_name})

    for name in sorted(baseline_sheets & current_sheets):
        baseline_sheet = baseline_workbook[name]
        current_sheet = current_workbook[name]

        if name in aux_sheets:
            if (
                baseline_sheet.max_row != current_sheet.max_row
                or baseline_sheet.max_column != current_sheet.max_column
            ):
                diffs.append(
                    {
                        "type": "AUX_SIZE_CHANGE",
                        "sheet": name,
                        "baseline": f"{baseline_sheet.max_row}x{baseline_sheet.max_column}",
                        "current": f"{current_sheet.max_row}x{current_sheet.max_column}",
                    }
                )
            print(f"  {name}: dims OK (skipped cell-by-cell)", flush=True)
            continue

        max_row = max(baseline_sheet.max_row, current_sheet.max_row)
        max_col = max(baseline_sheet.max_column, current_sheet.max_column)

        if baseline_sheet.max_row != current_sheet.max_row:
            diffs.append(
                {
                    "type": "ROW_COUNT",
                    "sheet": name,
                    "baseline": baseline_sheet.max_row,
                    "current": current_sheet.max_row,
                }
            )

        headers = [
            baseline_sheet.cell(1, col).value or current_sheet.cell(1, col).value
            for col in range(1, max_col + 1)
        ]

        sheet_diffs = 0
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                baseline_value = (
                    baseline_sheet.cell(row, col).value
                    if row <= baseline_sheet.max_row and col <= baseline_sheet.max_column
                    else None
                )
                current_value = (
                    current_sheet.cell(row, col).value
                    if row <= current_sheet.max_row and col <= current_sheet.max_column
                    else None
                )
                if not values_equal(
                    baseline_value,
                    current_value,
                    float_rtol=float_rtol,
                    float_atol=float_atol,
                ):
                    header = headers[col - 1] if col <= len(headers) else f"Col{col}"
                    if header == "Auditoría" and values_equal(
                        _normalize_audit_value(baseline_value),
                        _normalize_audit_value(current_value),
                        float_rtol=float_rtol,
                        float_atol=float_atol,
                    ):
                        continue
                    diffs.append(
                        {
                            "type": "CELL_DIFF",
                            "sheet": name,
                            "row": row,
                            "col": col,
                            "header": str(header),
                            "baseline": repr(baseline_value),
                            "current": repr(current_value),
                        }
                    )
                    sheet_diffs += 1

        cells = max_row * max_col
        print(f"  {name}: {cells:,} cells, {sheet_diffs} diff(s)", flush=True)

    return diffs


def print_diff_report(diffs: Sequence[dict], *, total_cells: int) -> None:
    print()
    print("=" * 70)
    print(f"  FAIL - {len(diffs)} difference(s)")
    print("=" * 70)

    by_sheet: dict[str, list[dict]] = {}
    for diff in diffs:
        by_sheet.setdefault(diff.get("sheet", "N/A"), []).append(diff)

    for sheet_name, sheet_diffs in sorted(by_sheet.items()):
        print(f"\n--- {sheet_name} ({len(sheet_diffs)} diff(s)) ---")
        for diff in sheet_diffs[:50]:
            if diff["type"] == "CELL_DIFF":
                print(
                    f"  Row {diff['row']}, Col {diff['col']} ({diff['header']}):  "
                    f"{diff['baseline']}  ->  {diff['current']}"
                )
            elif diff["type"] == "ROW_COUNT":
                print(f"  Row count: {diff['baseline']}  ->  {diff['current']}")
            else:
                print(f"  {diff['type']}")
        if len(sheet_diffs) > 50:
            print(f"  ... and {len(sheet_diffs) - 50} more")

    print(f"\nTotal diffs: {len(diffs)}   Cells checked: {total_cells:,}")


def run_guard_suite(guards: Sequence[tuple[str, WorkbookGuard]], workbook) -> tuple[bool, list[str]]:
    messages: list[str] = []
    for title, guard in guards:
        issues = list(guard(workbook))
        if issues:
            messages.append(title)
            messages.extend(issues)
    return (len(messages) == 0, messages)


@dataclass
class SmokeTestConfig:
    title: str
    description: str
    baseline_dir: Path
    baseline_values: Path
    baseline_json: Path
    input_labels: Sequence[tuple[str, Path]]
    run_pipeline: PipelineRunner
    summary_builder: SummaryBuilder
    baseline_load_data_only: bool = True
    float_rtol: float = 1e-9
    float_atol: float = 0.005
    aux_sheets: set[str] = field(default_factory=lambda: set(DEFAULT_AUX_SHEETS))
    save_guards: Sequence[tuple[str, WorkbookGuard]] = ()
    run_guards: Sequence[tuple[str, WorkbookGuard]] = ()
    extra_save_actions: Sequence[ExtraSaveAction] = ()


def save_baseline(config: SmokeTestConfig) -> int:
    print("Running pipeline to generate baseline ...")
    workbook = config.run_pipeline()

    guards_ok, guard_messages = run_guard_suite(config.save_guards, workbook)
    if not guards_ok:
        print("ERROR: refusing to save baseline:")
        for message in guard_messages:
            print(message)
        return 1

    config.baseline_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(str(config.baseline_values))
    print(f"  Saved workbook : {config.baseline_values}")

    for action in config.extra_save_actions:
        for message in action() or ():
            print(f"  {message}")

    snapshot = build_snapshot(workbook, aux_sheets=config.aux_sheets)
    with open(config.baseline_json, "w", encoding="utf-8") as handle:
        json.dump(snapshot, handle, indent=2, ensure_ascii=False, default=str)
    print(f"  Saved snapshot : {config.baseline_json}")

    lines = [
        f"{config.title} approved baseline - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for label, path in config.input_labels:
        lines.append(f"{label}: {path.name}")
    lines.extend(["", "Sheets:"])
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        lines.append(f"  {name}: {worksheet.max_row} rows x {worksheet.max_column} cols")
    lines.extend(config.summary_builder(workbook))

    summary = config.baseline_dir / "baseline_summary.txt"
    summary.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Saved summary  : {summary}")

    total = sum(workbook[name].max_row * workbook[name].max_column for name in workbook.sheetnames)
    print(f"\nBaseline saved.  Sheets: {len(workbook.sheetnames)}  Cells: {total:,}")
    return 0


def run_smoke(config: SmokeTestConfig) -> int:
    if not config.baseline_values.exists():
        print(f"ERROR: baseline not found at {config.baseline_values}")
        print(f"Run  python {Path(sys.argv[0]).name} --save  first.")
        return 1

    print("=" * 70)
    print(f"SMOKE TEST - {config.title}")
    print("=" * 70)
    print(f"Baseline : {config.baseline_values.name}")
    for label, path in config.input_labels:
        print(f"{label:<9}: {path.name}")
    print()

    print("Loading baseline ...", flush=True)
    baseline = load_workbook(str(config.baseline_values), data_only=config.baseline_load_data_only)

    print("Running pipeline ...", flush=True)
    current = config.run_pipeline()

    guards_ok, guard_messages = run_guard_suite(config.run_guards, current)
    if not guards_ok:
        print("\nFAIL - runtime guards triggered:")
        for message in guard_messages:
            print(message)
        return 1

    print("Comparing every cell ...", flush=True)
    diffs = compare_workbooks(
        baseline,
        current,
        aux_sheets=config.aux_sheets,
        float_rtol=config.float_rtol,
        float_atol=config.float_atol,
    )

    total = sum(baseline[name].max_row * baseline[name].max_column for name in baseline.sheetnames)
    if not diffs:
        print()
        print("=" * 70)
        print("  PASS - 0 differences")
        print(f"  Sheets : {len(baseline.sheetnames)}")
        print(f"  Cells  : {total:,}")
        print("=" * 70)
        return 0

    print_diff_report(diffs, total_cells=total)
    return 1


def run_cli(config: SmokeTestConfig) -> int:
    parser = argparse.ArgumentParser(description=config.description)
    parser.add_argument(
        "--save",
        action="store_true",
        help="Save current pipeline output as approved baseline",
    )
    args = parser.parse_args()

    if args.save:
        return save_baseline(config)
    return run_smoke(config)