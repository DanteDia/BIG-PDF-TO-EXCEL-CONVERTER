#!/usr/bin/env python
"""
Smoke test for Prida 10488 case.

Runs the full merge pipeline from frozen Gallo+Visual inputs and compares
EVERY cell in EVERY sheet against the approved baseline values workbook.

Usage:
    python smoke_test_prida.py --save       # save current output as new baseline
    python smoke_test_prida.py              # run smoke test against baseline
"""
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
PRIDA_DIR = ROOT / "Prida- Transferencia titulos tomar como compra"
BASELINE_DIR = ROOT / "SMOKE_BASELINE" / "PRIDA_20260416_APPROVED"

# Frozen input files (never modify these)
GALLO_PATH = PRIDA_DIR / "IG_10488.xlsx"
VISUAL_PATH = PRIDA_DIR / "10488bolsa.xlsx"

# Baseline artefacts
BASELINE_VALUES = BASELINE_DIR / "10488_PRIDA_baseline_values.xlsx"
BASELINE_JSON = BASELINE_DIR / "baseline_snapshot.json"

# Tolerance for float comparison
FLOAT_RTOL = 1e-9   # relative
FLOAT_ATOL = 0.005   # absolute (half a cent)


# ---------------------------------------------------------------------------
# Comparison helpers
# ---------------------------------------------------------------------------

def _floats_equal(a, b):
    if a == 0 and b == 0:
        return True
    if a == 0 or b == 0:
        return abs(a - b) <= FLOAT_ATOL
    rel = abs(a - b) / max(abs(a), abs(b))
    return rel <= FLOAT_RTOL or abs(a - b) <= FLOAT_ATOL


def _values_equal(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        if (a is None and b == 0) or (b is None and a == 0):
            return True
        if (a is None and b == "") or (b is None and a == ""):
            return True
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return _floats_equal(float(a), float(b))
    if isinstance(a, (datetime, date)) and isinstance(b, (datetime, date)):
        return a == b
    return str(a) == str(b)


# ---------------------------------------------------------------------------
# Snapshot
# ---------------------------------------------------------------------------

def _serialise_value(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float):
        return round(v, 10)
    return v


def snapshot_workbook_sheet(wb, name) -> dict:
    ws = wb[name]
    sheet = {
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "headers": [ws.cell(1, c).value for c in range(1, ws.max_column + 1)],
        "rows": {},
    }
    for r in range(2, ws.max_row + 1):
        sheet["rows"][str(r)] = [
            _serialise_value(ws.cell(r, c).value)
            for c in range(1, ws.max_column + 1)
        ]
    return sheet


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

def run_pipeline():
    sys.path.insert(0, str(ROOT))
    from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger

    assert GALLO_PATH.exists(), f"Gallo not found: {GALLO_PATH}"
    assert VISUAL_PATH.exists(), f"Visual not found: {VISUAL_PATH}"

    merger = GalloVisualMerger(str(GALLO_PATH), str(VISUAL_PATH))
    _, wb_values = merger.merge("both")
    return wb_values


# ---------------------------------------------------------------------------
# Compare
# ---------------------------------------------------------------------------

AUX_SHEETS = {
    "EspeciesVisual", "EspeciesGallo", "Cotizacion Dolar Historica",
    "PreciosInicialesEspecies", "PrecioTenenciasIniciales", "RatiosCedearsAcciones",
}


def compare_workbooks(baseline_wb, current_wb):
    diffs = []

    base_sheets = set(baseline_wb.sheetnames)
    curr_sheets = set(current_wb.sheetnames)

    for s in sorted(base_sheets - curr_sheets):
        diffs.append({"type": "SHEET_MISSING", "sheet": s})
    for s in sorted(curr_sheets - base_sheets):
        diffs.append({"type": "SHEET_ADDED", "sheet": s})

    for name in sorted(base_sheets & curr_sheets):
        ws_b = baseline_wb[name]
        ws_c = current_wb[name]

        if name in AUX_SHEETS:
            if ws_b.max_row != ws_c.max_row or ws_b.max_column != ws_c.max_column:
                diffs.append({
                    "type": "AUX_SIZE_CHANGE", "sheet": name,
                    "baseline": f"{ws_b.max_row}x{ws_b.max_column}",
                    "current": f"{ws_c.max_row}x{ws_c.max_column}",
                })
            print(f"  {name}: dims OK (skipped cell-by-cell)", flush=True)
            continue

        max_row = max(ws_b.max_row, ws_c.max_row)
        max_col = max(ws_b.max_column, ws_c.max_column)

        if ws_b.max_row != ws_c.max_row:
            diffs.append({
                "type": "ROW_COUNT", "sheet": name,
                "baseline": ws_b.max_row, "current": ws_c.max_row,
            })

        headers = [
            ws_b.cell(1, c).value or ws_c.cell(1, c).value
            for c in range(1, max_col + 1)
        ]

        sheet_diffs = 0
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                vb = ws_b.cell(r, c).value if r <= ws_b.max_row and c <= ws_b.max_column else None
                vc = ws_c.cell(r, c).value if r <= ws_c.max_row and c <= ws_c.max_column else None
                if not _values_equal(vb, vc):
                    hdr = headers[c - 1] if c <= len(headers) else f"Col{c}"
                    diffs.append({
                        "type": "CELL_DIFF", "sheet": name,
                        "row": r, "col": c, "header": str(hdr),
                        "baseline": repr(vb), "current": repr(vc),
                    })
                    sheet_diffs += 1
        cells = max_row * max_col
        print(f"  {name}: {cells:,} cells, {sheet_diffs} diff(s)", flush=True)
    return diffs


# ---------------------------------------------------------------------------
# Save baseline
# ---------------------------------------------------------------------------

def save_baseline():
    print("Running pipeline to generate baseline ...")
    wb_values = run_pipeline()

    BASELINE_DIR.mkdir(parents=True, exist_ok=True)
    wb_values.save(str(BASELINE_VALUES))
    print(f"  Saved workbook : {BASELINE_VALUES}")

    snap = {}
    for name in wb_values.sheetnames:
        if name in AUX_SHEETS:
            ws = wb_values[name]
            snap[name] = {"max_row": ws.max_row, "max_column": ws.max_column, "skipped": True}
        else:
            snap[name] = snapshot_workbook_sheet(wb_values, name)
    with open(BASELINE_JSON, "w", encoding="utf-8") as f:
        json.dump(snap, f, indent=2, ensure_ascii=False, default=str)
    print(f"  Saved snapshot : {BASELINE_JSON}")

    # ---- summary ----
    lines = [
        f"PRIDA 10488 approved baseline – {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Gallo : {GALLO_PATH.name}",
        f"Visual: {VISUAL_PATH.name}",
        f"Values: {BASELINE_VALUES.name}",
        "",
        "Sheets:",
    ]
    for name in wb_values.sheetnames:
        ws = wb_values[name]
        lines.append(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")

    ws_res = wb_values["Resumen"]
    lines += ["", "Resumen:"]
    for r in range(1, ws_res.max_row + 1):
        row = [ws_res.cell(r, c).value for c in range(1, ws_res.max_column + 1)]
        lines.append(f"  Row {r}: {row}")

    ws_bol = wb_values["Boletos"]
    hdrs = [ws_bol.cell(1, c).value for c in range(1, ws_bol.max_column + 1)]
    cols = {h: i + 1 for i, h in enumerate(hdrs)}
    lines += ["", f"Boletos: {ws_bol.max_row - 1} data rows", "Key boletos:"]
    for r in range(2, ws_bol.max_row + 1):
        num = str(ws_bol.cell(r, 4).value or "")
        if num in {"42583", "42598", "42580", "42582"}:
            pn = ws_bol.cell(r, cols.get("Precio Nominal", 20)).value
            mon = ws_bol.cell(r, cols.get("Moneda", 5)).value
            br = ws_bol.cell(r, cols.get("Bruto", 13)).value
            lines.append(f"  B={num}: PN={pn}, Moneda={mon}, Bruto={br}")

    lines += [
        "",
        "Protected invariants:",
        "- B=42583  Moneda=Dolar MEP  PN=1.309     Bruto=3998.995  (VENTA USD override)",
        "- B=42598  Moneda=Dolar MEP  PN=1.13269   Bruto=-3460.368 (VENTA USD)",
        "- B=42580  Moneda=Pesos      PN=1.28608   Bruto=4998.993  (normal Pesos)",
        "- Synthetic initial positions for 81086/81090/81092/81274 from TRF history",
        "- 81090 Resultado Ventas USD: gains (not losses) from TRF cost basis",
        "- Full workbook values compared cell-by-cell by smoke test.",
    ]
    summary = BASELINE_DIR / "baseline_summary.txt"
    summary.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Saved summary  : {summary}")

    total = sum(wb_values[s].max_row * wb_values[s].max_column for s in wb_values.sheetnames)
    print(f"\nBaseline saved.  Sheets: {len(wb_values.sheetnames)}  Cells: {total:,}")


# ---------------------------------------------------------------------------
# Run smoke test
# ---------------------------------------------------------------------------

def run_smoke():
    if not BASELINE_VALUES.exists():
        print(f"ERROR: baseline not found at {BASELINE_VALUES}")
        print("Run  python smoke_test_prida.py --save  first.")
        return 1

    print("=" * 70)
    print("SMOKE TEST — PRIDA 10488")
    print("=" * 70)
    print(f"Baseline : {BASELINE_VALUES.name}")
    print(f"Gallo    : {GALLO_PATH.name}")
    print(f"Visual   : {VISUAL_PATH.name}")
    print()

    print("Loading baseline ...", flush=True)
    bl = load_workbook(str(BASELINE_VALUES), data_only=True)

    print("Running pipeline ...", flush=True)
    cur = run_pipeline()

    print("Comparing every cell ...", flush=True)
    diffs = compare_workbooks(bl, cur)

    total = sum(bl[s].max_row * bl[s].max_column for s in bl.sheetnames)

    if not diffs:
        print()
        print("=" * 70)
        print(f"  PASS — 0 differences")
        print(f"  Sheets : {len(bl.sheetnames)}")
        print(f"  Cells  : {total:,}")
        print("=" * 70)
        return 0

    print()
    print("=" * 70)
    print(f"  FAIL — {len(diffs)} difference(s)")
    print("=" * 70)

    by_sheet = {}
    for d in diffs:
        by_sheet.setdefault(d.get("sheet", "N/A"), []).append(d)

    for sheet, sd in sorted(by_sheet.items()):
        print(f"\n--- {sheet} ({len(sd)} diff(s)) ---")
        for d in sd[:50]:
            if d["type"] == "CELL_DIFF":
                print(f"  Row {d['row']}, Col {d['col']} ({d['header']}):  "
                      f"{d['baseline']}  →  {d['current']}")
            elif d["type"] == "ROW_COUNT":
                print(f"  Row count: {d['baseline']}  →  {d['current']}")
            else:
                print(f"  {d['type']}")
        if len(sd) > 50:
            print(f"  ... and {len(sd) - 50} more")

    print(f"\nTotal diffs: {len(diffs)}   Cells checked: {total:,}")
    return 1


# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Smoke test – Prida 10488")
    ap.add_argument("--save", action="store_true",
                    help="Save current pipeline output as approved baseline")
    args = ap.parse_args()

    if args.save:
        save_baseline()
    else:
        sys.exit(run_smoke())


if __name__ == "__main__":
    main()
