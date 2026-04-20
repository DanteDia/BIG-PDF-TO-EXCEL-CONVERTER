#!/usr/bin/env python
"""
Smoke test for Sturman 11688 case.

Protects against regressions in:
- USD micro-price bruto/neto corruption (STU-MICRO-PRICE-RESCUE-BUG-001)
- Dolar Cable 1000x quantity inflation (STU-ANCHOR-1000X-DEFLATION-001)
- Futuros boletos bruto=0 rule (STU-FUTUROS-BOLETOS-001)
- Full workbook cell-by-cell comparison

Frozen inputs from OCR run 2026-04-20 of Sturman 11688 PDFs.
Reference: web output 20260420_0242 (pre-fix) and 20260420_1624 (post-fix).

Usage:
    python smoke_test_sturman_11688.py --save   # save current output as new baseline
    python smoke_test_sturman_11688.py           # run smoke test against baseline
"""
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
BASELINE_DIR = ROOT / "SMOKE_BASELINE" / "STURMAN_11688_20260420_APPROVED"

# Frozen input files (never modify these — from OCR run 2026-04-20)
GALLO_PATH = BASELINE_DIR / "11688_gallo_frozen.xlsx"
VISUAL_PATH = BASELINE_DIR / "11688_visual_frozen.xlsx"
PRECIO_PATH = BASELINE_DIR / "11688_precio_tenencias_frozen.xlsx"

# Baseline artefacts
BASELINE_VALUES = BASELINE_DIR / "11688_STURMAN_JAVIER_baseline_values.xlsx"
BASELINE_JSON = BASELINE_DIR / "baseline_snapshot.json"

# Tolerance for float comparison
FLOAT_RTOL = 1e-9   # relative
FLOAT_ATOL = 0.005   # absolute (half a cent)

# Max allowed quantity — anything above is a 1000x inflation artifact
MAX_SANE_QTY = 10_000_000_000


# ---------------------------------------------------------------------------
# Comparison helpers
# ---------------------------------------------------------------------------

def _floats_equal(a, b):
    if a == 0 and b == 0:
        return True
    if a == 0 or b == 0:
        return abs(a - b) <= FLOAT_ATOL
    rel = abs(a - b) / max(abs(a), abs(b))
    return rel <= FLOAT_RTOL or abs(a - b) <= FLOAT_ATOL


def _values_equal(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        if (a is None and b == 0) or (b is None and a == 0):
            return True
        if (a is None and b == "") or (b is None and a == ""):
            return True
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return _floats_equal(float(a), float(b))
    if isinstance(a, (datetime, date)) and isinstance(b, (datetime, date)):
        return a == b
    return str(a) == str(b)


# ---------------------------------------------------------------------------
# Snapshot
# ---------------------------------------------------------------------------

def _serialise_value(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float):
        return round(v, 10)
    return v


def snapshot_workbook_sheet(wb, name) -> dict:
    ws = wb[name]
    sheet = {
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "headers": [ws.cell(1, c).value for c in range(1, ws.max_column + 1)],
        "rows": {},
    }
    for r in range(2, ws.max_row + 1):
        sheet["rows"][str(r)] = [
            _serialise_value(ws.cell(r, c).value)
            for c in range(1, ws.max_column + 1)
        ]
    return sheet


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

def run_pipeline():
    sys.path.insert(0, str(ROOT))
    from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger

    assert GALLO_PATH.exists(), f"Gallo not found: {GALLO_PATH}"
    assert VISUAL_PATH.exists(), f"Visual not found: {VISUAL_PATH}"

    precio = str(PRECIO_PATH) if PRECIO_PATH.exists() else None
    merger = GalloVisualMerger(str(GALLO_PATH), str(VISUAL_PATH),
                               precio_tenencias_path=precio)
    _, wb_values = merger.merge("both")
    return wb_values


# ---------------------------------------------------------------------------
# Inflation guard
# ---------------------------------------------------------------------------

def check_inflation(wb):
    """Return list of cells with |value| > MAX_SANE_QTY."""
    bad = []
    for sn in wb.sheetnames:
        if sn in AUX_SHEETS:
            continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    continue
                try:
                    if abs(float(v)) > MAX_SANE_QTY:
                        hdr = ws.cell(1, c).value or f"Col{c}"
                        bad.append(f"  {sn} Row {r} Col {c} ({hdr}): {v}")
                except (TypeError, ValueError):
                    pass
    return bad


# ---------------------------------------------------------------------------
# Micro-price guard
# ---------------------------------------------------------------------------

# Boletos that must have correct neto after STU-MICRO-PRICE-RESCUE-BUG-001
MICROPRICE_EXPECTED = {
    "71174": 995000,
    "71873": 994975,
    "72719": 999000,
    "122342": 1000000,
}


def check_microprice(wb):
    """Verify micro-price boletos have correct neto values."""
    ws = wb["Boletos"]
    bad = []
    found = {}
    for r in range(2, ws.max_row + 1):
        bol = str(ws.cell(r, 4).value or "").strip()
        if bol in MICROPRICE_EXPECTED:
            neto = ws.cell(r, 16).value
            expected = MICROPRICE_EXPECTED[bol]
            try:
                if abs(float(neto or 0) - expected) > 1.0:
                    bad.append(f"  Bol {bol}: neto={neto}, expected={expected}")
            except (TypeError, ValueError):
                bad.append(f"  Bol {bol}: neto={neto!r} (non-numeric), expected={expected}")
            found[bol] = neto
    missing = set(MICROPRICE_EXPECTED) - set(found)
    for bol in sorted(missing):
        bad.append(f"  Bol {bol}: NOT FOUND in Boletos sheet")
    return bad


# ---------------------------------------------------------------------------
# Compare
# ---------------------------------------------------------------------------

AUX_SHEETS = {
    "EspeciesVisual", "EspeciesGallo", "Cotizacion Dolar Historica",
    "PreciosInicialesEspecies", "PrecioTenenciasIniciales", "RatiosCedearsAcciones",
}


def compare_workbooks(baseline_wb, current_wb):
    diffs = []

    base_sheets = set(baseline_wb.sheetnames)
    curr_sheets = set(current_wb.sheetnames)

    for s in sorted(base_sheets - curr_sheets):
        diffs.append({"type": "SHEET_MISSING", "sheet": s})
    for s in sorted(curr_sheets - base_sheets):
        diffs.append({"type": "SHEET_ADDED", "sheet": s})

    for name in sorted(base_sheets & curr_sheets):
        ws_b = baseline_wb[name]
        ws_c = current_wb[name]

        if name in AUX_SHEETS:
            if ws_b.max_row != ws_c.max_row or ws_b.max_column != ws_c.max_column:
                diffs.append({
                    "type": "AUX_SIZE_CHANGE", "sheet": name,
                    "baseline": f"{ws_b.max_row}x{ws_b.max_column}",
                    "current": f"{ws_c.max_row}x{ws_c.max_column}",
                })
            print(f"  {name}: dims OK (skipped cell-by-cell)", flush=True)
            continue

        max_row = max(ws_b.max_row, ws_c.max_row)
        max_col = max(ws_b.max_column, ws_c.max_column)

        if ws_b.max_row != ws_c.max_row:
            diffs.append({
                "type": "ROW_COUNT", "sheet": name,
                "baseline": ws_b.max_row, "current": ws_c.max_row,
            })

        headers = [
            ws_b.cell(1, c).value or ws_c.cell(1, c).value
            for c in range(1, max_col + 1)
        ]

        sheet_diffs = 0
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                vb = ws_b.cell(r, c).value if r <= ws_b.max_row and c <= ws_b.max_column else None
                vc = ws_c.cell(r, c).value if r <= ws_c.max_row and c <= ws_c.max_column else None
                if not _values_equal(vb, vc):
                    hdr = headers[c - 1] if c <= len(headers) else f"Col{c}"
                    diffs.append({
                        "type": "CELL_DIFF", "sheet": name,
                        "row": r, "col": c, "header": str(hdr),
                        "baseline": repr(vb), "current": repr(vc),
                    })
                    sheet_diffs += 1
        cells = max_row * max_col
        print(f"  {name}: {cells:,} cells, {sheet_diffs} diff(s)", flush=True)
    return diffs


# ---------------------------------------------------------------------------
# Save baseline
# ---------------------------------------------------------------------------

def save_baseline():
    print("Running pipeline to generate baseline ...")
    wb_values = run_pipeline()

    # Inflation guard
    bad = check_inflation(wb_values)
    if bad:
        print("ERROR: inflated quantities found — refusing to save baseline:")
        for b in bad:
            print(b)
        sys.exit(1)

    # Micro-price guard
    bad = check_microprice(wb_values)
    if bad:
        print("ERROR: micro-price boletos have wrong values — refusing to save baseline:")
        for b in bad:
            print(b)
        sys.exit(1)

    BASELINE_DIR.mkdir(parents=True, exist_ok=True)
    wb_values.save(str(BASELINE_VALUES))
    print(f"  Saved workbook : {BASELINE_VALUES}")

    snap = {}
    for name in wb_values.sheetnames:
        if name in AUX_SHEETS:
            ws = wb_values[name]
            snap[name] = {"max_row": ws.max_row, "max_column": ws.max_column, "skipped": True}
        else:
            snap[name] = snapshot_workbook_sheet(wb_values, name)
    with open(BASELINE_JSON, "w", encoding="utf-8") as f:
        json.dump(snap, f, indent=2, ensure_ascii=False, default=str)
    print(f"  Saved snapshot : {BASELINE_JSON}")

    # ---- summary ----
    ws_bol = wb_values["Boletos"]
    lines = [
        f"STURMAN 11688 approved baseline – {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Gallo : {GALLO_PATH.name}",
        f"Visual: {VISUAL_PATH.name}",
        f"Precio: {PRECIO_PATH.name}",
        f"Values: {BASELINE_VALUES.name}",
        "",
        "Sheets:",
    ]
    for name in wb_values.sheetnames:
        ws = wb_values[name]
        lines.append(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")

    lines += [
        "",
        "Micro-price boletos (STU-MICRO-PRICE-RESCUE-BUG-001):",
    ]
    for bol_num, exp_neto in sorted(MICROPRICE_EXPECTED.items()):
        for r in range(2, ws_bol.max_row + 1):
            if str(ws_bol.cell(r, 4).value or "").strip() == bol_num:
                actual = ws_bol.cell(r, 16).value
                lines.append(f"  Bol {bol_num}: neto={actual} (expected={exp_neto})")
                break

    lines += [
        "",
        "Protected invariants:",
        "- No quantity > 10,000,000,000 in any sheet (1000x inflation guard)",
        "- Bol 71174: neto=995,000 (not 1,053,529 — micro-price fix)",
        "- Bol 71873: neto=994,975 (not 1,053,503 — micro-price fix)",
        "- Bol 72719: neto=999,000 (not 1,057,765 — micro-price fix)",
        "- Bol 122342: neto=1,000,000 (not 1,058,824 — micro-price fix)",
        "- Futuros boletos: bruto=0, neto=gastos (no qty*precio)",
        "- Boletos Dolar Cable code 9249 qty=757575758 (not 757575758000)",
        "- Boletos Dolar Cable code 9323 qty=512974026 (not 512974026000)",
        "- Full workbook values compared cell-by-cell by smoke test",
        "",
        "Reference artifacts in this directory:",
        "- 11688_web_reference_20260420_0242.xlsx: web output BEFORE micro-price fix",
        "- 11688_web_reference_20260420_1624.pdf: web output AFTER micro-price fix",
    ]
    summary = BASELINE_DIR / "baseline_summary.txt"
    summary.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Saved summary  : {summary}")

    total = sum(wb_values[s].max_row * wb_values[s].max_column for s in wb_values.sheetnames)
    print(f"\nBaseline saved.  Sheets: {len(wb_values.sheetnames)}  Cells: {total:,}")


# ---------------------------------------------------------------------------
# Run smoke test
# ---------------------------------------------------------------------------

def run_smoke():
    if not BASELINE_VALUES.exists():
        print(f"ERROR: baseline not found at {BASELINE_VALUES}")
        print("Run  python smoke_test_sturman_11688.py --save  first.")
        return 1

    print("=" * 70)
    print("SMOKE TEST — STURMAN 11688")
    print("=" * 70)
    print(f"Baseline : {BASELINE_VALUES.name}")
    print(f"Gallo    : {GALLO_PATH.name}")
    print(f"Visual   : {VISUAL_PATH.name}")
    print(f"Precio   : {PRECIO_PATH.name}")
    print()

    print("Loading baseline ...", flush=True)
    bl = load_workbook(str(BASELINE_VALUES))

    print("Running pipeline ...", flush=True)
    cur = run_pipeline()

    # Inflation guard on current run
    bad = check_inflation(cur)
    if bad:
        print("\nFAIL — inflated quantities detected:")
        for b in bad:
            print(b)
        return 1

    # Micro-price guard on current run
    bad = check_microprice(cur)
    if bad:
        print("\nFAIL — micro-price boletos corrupted:")
        for b in bad:
            print(b)
        return 1

    print("Comparing every cell ...", flush=True)
    diffs = compare_workbooks(bl, cur)

    total = sum(bl[s].max_row * bl[s].max_column for s in bl.sheetnames)

    if not diffs:
        print()
        print("=" * 70)
        print(f"  PASS \u00f9 0 differences")
        print(f"  Sheets : {len(bl.sheetnames)}")
        print(f"  Cells  : {total:,}")
        print("=" * 70)
        return 0

    print()
    print("=" * 70)
    print(f"  FAIL — {len(diffs)} difference(s)")
    print("=" * 70)

    by_sheet = {}
    for d in diffs:
        by_sheet.setdefault(d.get("sheet", "N/A"), []).append(d)

    for sheet, sd in sorted(by_sheet.items()):
        print(f"\n--- {sheet} ({len(sd)} diff(s)) ---")
        for d in sd[:50]:
            if d["type"] == "CELL_DIFF":
                print(f"  Row {d['row']}, Col {d['col']} ({d['header']}):  "
                      f"{d['baseline']}  \u2192  {d['current']}")
            elif d["type"] == "ROW_COUNT":
                print(f"  Row count: {d['baseline']}  \u2192  {d['current']}")
            else:
                print(f"  {d['type']}")
        if len(sd) > 50:
            print(f"  ... and {len(sd) - 50} more")

    print(f"\nTotal diffs: {len(diffs)}   Cells checked: {total:,}")
    return 1


# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Smoke test \u2013 Sturman 11688")
    ap.add_argument("--save", action="store_true",
                    help="Save current pipeline output as approved baseline")
    args = ap.parse_args()

    if args.save:
        save_baseline()
    else:
        sys.exit(run_smoke())


if __name__ == "__main__":
    main()
