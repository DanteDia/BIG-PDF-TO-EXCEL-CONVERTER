from pathlib import Path
from pdf_converter.convert_with_datalab import convert_pdf_to_excel
from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger
from openpyxl import load_workbook

root = Path(r"c:/Users/xarodan/Downloads/Resumen Impositivo- Branch dots.OCR")
visual_pdf = root / "Aguiar_2025_Visual.pdf"
gallo_pdf = root / "Aguiar_2025_Gallo.PDF"
precio_pdf = root / "juanaguiar preciostenencias.pdf"
visual_excel = root / "AGUIAR_Visual_from_PDF.xlsx"
gallo_excel = root / "AGUIAR_Gallo_from_PDF.xlsx"
precio_excel = root / "AGUIAR_PrecioTenencias_from_PDF.xlsx"

print("Converting Visual...")
convert_pdf_to_excel(str(visual_pdf), str(visual_excel))
print("Converting Gallo...")
convert_pdf_to_excel(str(gallo_pdf), str(gallo_excel))
print("Converting Precio Tenencias...")
convert_pdf_to_excel(str(precio_pdf), str(precio_excel))

aux_dir = root / "pdf_converter" / "datalab" / "aux_data"
merger = GalloVisualMerger(str(gallo_excel), str(visual_excel), str(aux_dir), precio_tenencias_path=str(precio_excel))
wb_formulas, wb_values = merger.merge(output_mode="both")

out_formulas = root / "AGUIAR_MERGE_FULL_FLOW_formulas.xlsx"
out_values = root / "AGUIAR_MERGE_FULL_FLOW_values.xlsx"
wb_formulas.save(out_formulas)
wb_values.save(out_values)
print("Saved:", out_formulas.name, out_values.name)

ws = wb_values["Resultado Ventas ARS"]
cod_col = 4
qty_ini_col = 17
qty_fin_col = 22
rows = [r for r in range(2, ws.max_row + 1) if str(ws.cell(r, cod_col).value).strip() == "839"]
print("Rows for Cod 839:", rows)
for r in rows:
    print("Row", r, "Concertacion", ws.cell(r, 5).value, "Cantidad", ws.cell(r, 9).value, "Q", ws.cell(r, qty_ini_col).value, "V", ws.cell(r, qty_fin_col).value)
