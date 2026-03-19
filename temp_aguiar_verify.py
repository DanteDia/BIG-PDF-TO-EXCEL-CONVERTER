"""Regenerate Aguiar full flow and verify PrecioTenenciasIniciales fixes + Posicion Inicial Gallo."""
import os, sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

os.environ["DATALAB_API_KEY"] = "Z1JLPnKJIAcNosYvpyF-GZrjVizmsf5MsBlNGL7Szuk"

from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger
from openpyxl import load_workbook

root = Path(__file__).parent

# Reuse already-OCR'd intermediate Excels (avoid re-OCR cost)
gallo_excel = root / "AGUIAR_Gallo_from_PDF.xlsx"
visual_excel = root / "AGUIAR_Visual_from_PDF.xlsx"
precio_excel = root / "AGUIAR_PrecioTenencias_from_PDF.xlsx"

# But we need to re-postprocess PrecioTenencias because postprocess was updated
from pdf_converter.datalab.postprocess import process_precio_tenencias_sheet
wb_precio = load_workbook(str(precio_excel))
ws_precio = wb_precio.active
print("=== PrecioTenenciasIniciales BEFORE fix ===")
headers = [ws_precio.cell(1, c).value for c in range(1, ws_precio.max_column + 1)]
print("Headers:", headers)
# Show the problematic rows (24, 25, 30 in original; look for cantidad>0 and importe<=0)
for r in range(2, ws_precio.max_row + 1):
    cantidad = ws_precio.cell(r, 4).value  # col D
    importe = ws_precio.cell(r, 5).value   # col E
    resultado = ws_precio.cell(r, 6).value # col F
    try:
        c = float(cantidad) if cantidad else 0
        i = float(importe) if importe else 0
    except:
        c, i = 0, 0
    if c > 0 and i <= 0:
        print(f"  Row {r}: Cod={ws_precio.cell(r,1).value} Ticker={ws_precio.cell(r,2).value} "
              f"Cantidad={cantidad} Importe={importe} Resultado={resultado} Precio={ws_precio.cell(r,7).value}")

# Re-run postprocess on the precio sheet
process_precio_tenencias_sheet(ws_precio)

print("\n=== PrecioTenenciasIniciales AFTER fix ===")
for r in range(2, ws_precio.max_row + 1):
    cantidad = ws_precio.cell(r, 4).value
    importe = ws_precio.cell(r, 5).value
    resultado = ws_precio.cell(r, 6).value
    precio = ws_precio.cell(r, 7).value
    try:
        c = float(cantidad) if cantidad else 0
        i = float(importe) if importe else 0
    except:
        c, i = 0, 0
    if c > 0 and (precio and float(precio) > 0):
        # Show rows that previously had issues
        cod = ws_precio.cell(r, 1).value
        if cod in ['2923', '2916', '1167']:  # expected problematic codes
            print(f"  Row {r}: Cod={cod} Ticker={ws_precio.cell(r,2).value} "
                  f"Cantidad={cantidad} Importe={importe} Resultado={resultado} Precio={precio}")

# Save fixed precio
fixed_precio = root / "AGUIAR_PrecioTenencias_FIXED.xlsx"
wb_precio.save(str(fixed_precio))
print(f"\nSaved fixed PrecioTenencias: {fixed_precio.name}")

# Now run the merge
aux_dir = root / "pdf_converter" / "datalab" / "aux_data"
merger = GalloVisualMerger(str(gallo_excel), str(visual_excel), str(aux_dir),
                           precio_tenencias_path=str(fixed_precio))
wb_formulas, wb_values = merger.merge(output_mode="both")

out_formulas = root / "AGUIAR_MERGE_FIXED_formulas.xlsx"
out_values = root / "AGUIAR_MERGE_FIXED_values.xlsx"
wb_formulas.save(str(out_formulas))
wb_values.save(str(out_values))
print(f"\nSaved merged: {out_formulas.name}, {out_values.name}")

# Verify Posicion Inicial Gallo: check col N (Origen Precio Costo)
ws_pos = wb_values['Posicion Inicial Gallo']
print("\n=== Posicion Inicial Gallo: Origen Precio Costo ===")
headers_pos = [ws_pos.cell(1, c).value for c in range(1, ws_pos.max_column + 1)]
# Find col N index
try:
    col_n = headers_pos.index('Origen Precio Costo') + 1
except:
    col_n = 14  # fallback
try:
    col_p = headers_pos.index('Precio a Utilizar') + 1
except:
    col_p = 16
try:
    col_v = headers_pos.index('Precio Nominal') + 1
except:
    col_v = 22

print(f"Col N({col_n}) = Origen Precio Costo, Col P({col_p}) = Precio a Utilizar, Col V({col_v}) = Precio Nominal")
problematic_found = False
for r in range(2, ws_pos.max_row + 1):
    origen = ws_pos.cell(r, col_n).value or ""
    if "PreciosInicialesEspecies" in str(origen):
        problematic_found = True
        print(f"  ⚠️  Row {r}: Cod={ws_pos.cell(r,4).value} Origen='{origen}' "
              f"Precio={ws_pos.cell(r, col_p).value} PrecioNom={ws_pos.cell(r, col_v).value}")
    else:
        print(f"  ✅ Row {r}: Cod={ws_pos.cell(r,4).value} Origen='{origen}' "
              f"Precio={ws_pos.cell(r, col_p).value}")

if not problematic_found:
    print("\n✅ ALL rows in Posicion Inicial Gallo source from PrecioTenenciasIniciales!")
else:
    print("\n⚠️  Some rows still fallback to PreciosInicialesEspecies")

# Also verify Resultado Ventas ARS cod 839 stock carryover still works
ws_ars = wb_values['Resultado Ventas ARS']
rows_839 = [r for r in range(2, ws_ars.max_row + 1) if str(ws_ars.cell(r, 4).value).strip() == '839']
print(f"\n=== Resultado Ventas ARS - Cod 839 stock carryover ===")
for r in rows_839:
    print(f"  Row {r}: Concert={ws_ars.cell(r,5).value} Qty={ws_ars.cell(r,9).value} "
          f"Q={ws_ars.cell(r,17).value} V={ws_ars.cell(r,22).value}")
