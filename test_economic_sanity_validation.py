from openpyxl import Workbook

from pdf_converter.datalab.economic_sanity import add_validation_sheet, validate_workbook


def test_economic_sanity_flags_human_review_triggers_and_writes_sheet():
    wb = Workbook()
    boletos = wb.active
    boletos.title = "Boletos"
    boletos.append(["Moneda", "Tipo Operación", "Cantidad", "Precio", "Bruto", "Neto"])
    boletos.append(["Dolar Cable", "Compra", 1, 4_000_000, 4_000_000, 4_000_000])
    boletos.append(["Pesos", "Compra", 1, 100, 100, 100])

    ars = wb.create_sheet("Resultado Ventas ARS")
    ars.append(["Tipo Operación", "Bruto", "Costo", "Resultado Calculado(final)"])
    ars.append(["Venta", 1000, 100, 1200])

    usd = wb.create_sheet("Resultado Ventas USD")
    usd.append(["Tipo Operación", "Bruto en USD", "Costo", "Resultado Calculado(final)"])
    usd.append(["Venta", 1000, 100, 120])

    report = validate_workbook(wb)
    rule_ids = {issue.rule_id for issue in report.issues}

    assert "BOL-USD-LARGE-001" in rule_ids
    assert "BOL-TINY-001" in rule_ids
    assert "RES-ARS-RATIO-001" in rule_ids
    assert "RES-USD-RATIO-001" in rule_ids

    add_validation_sheet(wb, report)
    assert wb.sheetnames[0] == "Validacion"
    assert wb["Validacion"].cell(2, 2).value == report.issue_count


def test_economic_sanity_uses_usd_cost_denominator_before_bruto():
    wb = Workbook()
    wb.active.title = "Boletos"
    wb["Boletos"].append(["Moneda", "Bruto"])

    usd = wb.create_sheet("Resultado Ventas USD")
    usd.append(["Tipo Operación", "Bruto en USD", "Costo", "Resultado Calculado(final)"])
    usd.append(["Venta", 10, 1000, 9])

    report = validate_workbook(wb)

    assert not [issue for issue in report.issues if issue.rule_id == "RES-USD-RATIO-001"]


def test_economic_sanity_allows_ars_full_gain_with_zero_cost_initial_stock():
    wb = Workbook()
    wb.active.title = "Boletos"
    wb["Boletos"].append(["Moneda", "Bruto"])

    pos = wb.create_sheet("Posicion Inicial Gallo")
    pos.append(["Codigo especie", "Origen precio costo"])
    pos.append([30043, "PrecioTenenciasCostoRecuperado"])

    ars = wb.create_sheet("Resultado Ventas ARS")
    ars.append(["Tipo Operación", "Cod.Instrum", "Bruto", "Costo", "Resultado Calculado(final)", "Cantidad Stock Inicial"])
    ars.append(["Venta", 30043, -1000, 0, 1000, 10])
    ars.append(["Venta", 30043, -1000, -50, 950, 10])
    ars.append(["Venta", 99999, -1000, 0, 1000, 10])
    ars.append(["Venta", 30043, -1000, 0, 1200, 10])

    report = validate_workbook(wb)
    ars_ratio_rows = [issue.row for issue in report.issues if issue.rule_id == "RES-ARS-RATIO-001"]

    assert 2 not in ars_ratio_rows
    assert 3 not in ars_ratio_rows
    assert 4 in ars_ratio_rows
    assert 5 in ars_ratio_rows
