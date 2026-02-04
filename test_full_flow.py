#!/usr/bin/env python
"""
Test completo del flujo PDF -> Excel -> Merge -> Verificación

PDFs de test:
- Visual: VeroLandro2025.pdf
- Gallo: Vero_2025_gallo.PDF
"""

import os
import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from pdf_converter.convert_with_datalab import convert_pdf_to_excel
from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger
from openpyxl import load_workbook


def test_full_flow(skip_pdf_conversion=False):
    """Ejecutar el flujo completo y verificar resultados."""
    
    # PDFs de test (anotar para futuros tests)
    visual_pdf = 'VeroLandro2025.pdf'
    gallo_pdf = 'Vero_2025_gallo.PDF'
    
    # Output Excels intermedios
    visual_excel = 'TEST_Visual_from_PDF.xlsx'
    gallo_excel = 'TEST_Gallo_from_PDF.xlsx'
    
    # Output final del merge
    merge_output_formulas = 'TEST_MERGE_FULL_FLOW_formulas.xlsx'
    merge_output_values = 'TEST_MERGE_FULL_FLOW_values.xlsx'
    
    # Directorio de archivos auxiliares
    aux_dir = Path(__file__).parent / 'pdf_converter' / 'datalab' / 'aux_data'
    
    print('=' * 60)
    print('FLUJO COMPLETO: PDF -> Excel -> Merge')
    print('=' * 60)
    
    if not skip_pdf_conversion:
        # PASO 1: Convertir Visual PDF
        print('\n=== PASO 1: Convirtiendo Visual PDF a Excel ===')
        print(f'Input: {visual_pdf}')
        try:
            convert_pdf_to_excel(visual_pdf, visual_excel)
            print(f'✓ Visual Excel generado: {visual_excel}')
        except Exception as e:
            print(f'✗ Error: {e}')
            return False
        
        # PASO 2: Convertir Gallo PDF
        print('\n=== PASO 2: Convirtiendo Gallo PDF a Excel ===')
        print(f'Input: {gallo_pdf}')
        try:
            convert_pdf_to_excel(gallo_pdf, gallo_excel)
            print(f'✓ Gallo Excel generado: {gallo_excel}')
        except Exception as e:
            print(f'✗ Error: {e}')
            return False
    else:
        print('\n=== PASO 1-2: SALTANDO conversión PDF (usando Excel existentes) ===')
        # Usar los Excel ya generados
        visual_excel = '12128_LANDRO_VERONICA_INES_Visual_Generado_OK.xlsx'
        gallo_excel = '12128_LANDRO_VERONICA_INES_Gallo_Generado_OK.xlsx'
        print(f'  Visual Excel: {visual_excel}')
        print(f'  Gallo Excel: {gallo_excel}')
    
    # PASO 3: Merge
    print('\n=== PASO 3: Ejecutando Merge Gallo + Visual ===')
    try:
        merger = GalloVisualMerger(gallo_excel, visual_excel, str(aux_dir))
        wb_formulas, wb_values = merger.merge(output_mode="both")
        
        # Guardar ambos archivos
        wb_formulas.save(merge_output_formulas)
        print(f'✓ Merge con fórmulas generado: {merge_output_formulas}')
        
        wb_values.save(merge_output_values)
        print(f'✓ Merge con valores generado: {merge_output_values}')
    except Exception as e:
        print(f'✗ Error en merge: {e}')
        import traceback
        traceback.print_exc()
        return False
    
    # PASO 4: Verificar columna Precio Nominal
    print('\n=== PASO 4: Verificando columna "Precio Nominal" ===')
    try:
        # Verificar en ambos archivos
        for output_file in [merge_output_formulas, merge_output_values]:
            print(f'\n--- Verificando: {output_file} ---')
            wb = load_workbook(output_file)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1]]
                
                if 'Precio Nominal' in headers:
                    col_idx = headers.index('Precio Nominal') + 1
                    print(f'✓ Hoja "{sheet_name}": Columna "Precio Nominal" encontrada en posición {col_idx}')
                    
                    # Mostrar algunos valores de ejemplo
                    print(f'  Primeras filas:')
                    for row in range(2, min(6, ws.max_row + 1)):
                        tipo = ws.cell(row, 1).value
                        precio = ws.cell(row, 11).value if ws.max_column >= 11 else None  # Precio Unitario
                        precio_nominal = ws.cell(row, col_idx).value
                        tipo_str = str(tipo)[:30] if tipo else "N/A"
                        print(f'    Fila {row}: Tipo="{tipo_str}..." | Precio={precio} | Precio Nominal={precio_nominal}')
                else:
                    if sheet_name in ['Boletos', 'Resultado Ventas ARS', 'Resultado Ventas USD']:
                        print(f'✗ Hoja "{sheet_name}": NO tiene columna "Precio Nominal"!')
                        print(f'  Headers: {headers}')
                    else:
                        print(f'  Hoja "{sheet_name}": {len(headers)} columnas (sin Precio Nominal - OK)')
            
            wb.close()
        
        print('\n✓ VERIFICACIÓN COMPLETADA')
        return True
        
    except Exception as e:
        print(f'✗ Error verificando: {e}')
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    # Si pasamos --skip-pdf, saltamos la conversión de PDFs
    skip_pdf = '--skip-pdf' in sys.argv
    success = test_full_flow(skip_pdf_conversion=skip_pdf)
    sys.exit(0 if success else 1)
