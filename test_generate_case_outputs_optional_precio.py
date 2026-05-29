import sys

from openpyxl import Workbook

import generate_case_outputs


class _FakeReport:
    def to_dict(self):
        return {"issue_count": 0, "counts_by_severity": {}, "issues": []}


def test_generate_case_outputs_allows_missing_precio_tenencias(tmp_path, monkeypatch):
    visual = tmp_path / "visual.xlsx"
    gallo = tmp_path / "gallo.xlsx"
    Workbook().save(visual)
    Workbook().save(gallo)

    seen = {}

    class FakeMerger:
        def __init__(self, gallo_path, visual_path, aux_data_dir, precio_tenencias_path=None):
            seen["precio_tenencias_path"] = precio_tenencias_path

        def merge(self, output_mode="both"):
            return Workbook(), Workbook()

    class FakeExporter:
        def __init__(self, workbook_path, client):
            self.periodo_inicio = None
            self.periodo_fin = None
            self.anio = None

        def export_to_pdf(self, output_path):
            with open(output_path, "wb") as handle:
                handle.write(b"%PDF-1.4\n")

    monkeypatch.setattr(generate_case_outputs, "GalloVisualMerger", FakeMerger)
    monkeypatch.setattr(generate_case_outputs, "validate_workbook", lambda wb: _FakeReport())
    monkeypatch.setattr(generate_case_outputs, "add_validation_sheet", lambda wb, report: None)
    monkeypatch.setattr(generate_case_outputs, "ExcelToPdfExporter", FakeExporter)
    monkeypatch.setattr(sys, "argv", [
        "generate_case_outputs.py",
        "--root", str(tmp_path),
        "--visual-xlsx", str(visual),
        "--gallo-xlsx", str(gallo),
        "--case-prefix", "nested/CASE_NO_PRECIO",
        "--client-number", "13988",
        "--client-name", "Cliente Sin Tenencias",
    ])

    assert generate_case_outputs.main() == 0
    assert seen["precio_tenencias_path"] is None
    assert (tmp_path / "nested" / "CASE_NO_PRECIO_Resumen_Impositivo_FIXED_values.xlsx").exists()
    assert (tmp_path / "nested" / "CASE_NO_PRECIO_Resumen_Impositivo_VALIDATION.json").exists()