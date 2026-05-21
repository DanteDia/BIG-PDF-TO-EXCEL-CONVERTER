import textwrap

from openpyxl import load_workbook

from pdf_converter.datalab.md_to_excel import MarkdownTableParser
from pdf_converter.datalab.md_to_excel import convert_markdown_to_excel


def test_gallo_inline_initial_position_row_is_not_replaced_by_later_position_section():
    markdown = textwrap.dedent(
        """
        Industrial Valores S.A.

        ## Analisis de la Inversion

        | POSICION AL 01/01/25 | | | | | | | | |
        |---|---|---|---|---|---|---|---|---|
        | Especie | Detalle | Custodia | Cantidad | Precio | Importe en Pesos | % de Cartera | Importe en Dolares | % de Cartera |
        | SUPV GRUPO SUPERVIELLE S.A. | | CAJA VALORES | 4,500.00 | 3575.000 | 16,087,500.00 | 0.92 | 13,775.84 | 0.92 |
        | LOMA LOMA NEGRA | | CAJA VALORES | 78,000.00 | 2880.000 | 224,640,000.00 | 12.86 | 192,360.80 | 12.86 |

        ### POSICION AL 31/05/25

        | Especie | Detalle | Custodia | Cantidad | Precio | Importe en Pesos | % de Cartera | Importe en Dolares | % de Cartera |
        |---|---|---|---|---|---|---|---|---|
        | SUPV GRUPO SUPERVIELLE S.A. | | CAJA VALORES | 7,000.00 | 3135.000 | 21,945,000.00 | 1.41 | 18,390.25 | 1.41 |
        | LOMA LOMA NEGRA | | CAJA VALORES | 50,000.00 | 3045.000 | 152,250,000.00 | 9.80 | 127,587.87 | 9.80 |
        """
    )

    parser = MarkdownTableParser(markdown)
    tables = parser.parse()

    assert 'Posicion Inicial' in tables
    assert 'Posicion Final' in tables

    inicial = tables['Posicion Inicial']
    final = tables['Posicion Final']

    assert inicial.metadata.get('fecha') == '01/01/25'
    assert final.metadata.get('fecha') == '31/05/25'
    assert inicial.rows[0][3] == '4,500.00'
    assert inicial.rows[1][3] == '78,000.00'
    assert final.rows[0][3] == '7,000.00'
    assert final.rows[1][3] == '50,000.00'


def test_gallo_position_continuation_category_stays_in_initial_position(tmp_path):
    markdown_path = tmp_path / "gallo.md"
    output_path = tmp_path / "gallo.xlsx"
    markdown_path.write_text(
        textwrap.dedent(
            """
            Industrial Valores S.A.

            | POSICION AL 01/01/25 | | | | | | | | |
            |---|---|---|---|---|---|---|---|---|
            | Especie | Detalle | Custodia | Cantidad | Precio | Importe en Pesos | % de Cartera | Importe en Dolares | % de Cartera |
            | <b>TITULOS PRIVADOS LOCALES</b> | | | | | | | | |
            | GGAL GGAL GRUPO FINANCIERO GALICIA | | CAJA VALORES | 1,200.00 | 7450.000 | 8,940,000.00 | 0.51 | 7,655.38 | 0.51 |

            ### TIT.PRIVADOS DEL EXTERIOR

            | Especie | Fecha | Operacion | Numero | Cantidad | Precio | Importe | Costo | Resultado en Pesos | Resultado en USD | Gastos en Pesos | Gastos en USD |
            |---|---|---|---|---|---|---|---|---|---|---|---|
            | VTRS-US VIATRIS INC | | CAJA VALORES | | 7.00 | 12.450 | 101,774.25 | 0.01 | 87.15 | 0.01 | | |

            ### POSICION AL 31/05/25

            | Especie | Detalle | Custodia | Cantidad | Precio | Importe en Pesos | % de Cartera | Importe en Dolares | % de Cartera |
            |---|---|---|---|---|---|---|---|---|
            | VTRS-US VIATRIS INC | | CAJA VALORES | 7.00 | 8.790 | 73,423.46 | | 61.53 | |
            """
        ),
        encoding="utf-8",
    )

    convert_markdown_to_excel(markdown_path, output_path)
    wb = load_workbook(output_path, data_only=True)

    inicial = wb["Posicion Inicial"]
    vtrs_row = None
    for row in inicial.iter_rows(min_row=2, values_only=True):
        if row[1] == "VTRS-US VIATRIS INC":
            vtrs_row = row
            break

    assert vtrs_row is not None
    assert vtrs_row[0] == "TIT.PRIVADOS DEL EXTERIOR"
    assert vtrs_row[4] == 7
    assert vtrs_row[5] == 12.45
    assert vtrs_row[6] == 101774.25
    assert vtrs_row[8] == 87.15