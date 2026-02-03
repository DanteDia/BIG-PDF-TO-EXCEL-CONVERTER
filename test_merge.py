"""Test script para merge_gallo_visual"""

import sys
sys.path.insert(0, '.')

from pathlib import Path
from openpyxl import Workbook, load_workbook

# Importar directamente el módulo
import importlib.util
spec = importlib.util.spec_from_file_location("merge_gallo_visual", 
    "pdf_converter/datalab/merge_gallo_visual.py")
merge_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(merge_module)

GalloVisualMerger = merge_module.GalloVisualMerger

# Config
aux_dir = Path('pdf_converter/datalab/aux_data')
gallo = '12128_LANDRO_VERONICA_INES_Gallo_Generado_OK.xlsx'
visual = '12128_LANDRO_VERONICA_INES_Visual_Generado_OK.xlsx'
output = 'TEST_Merge_v2.xlsx'

print("=== Test Merge Gallo-Visual ===")
print(f"Gallo: {gallo}")
print(f"Visual: {visual}")
print(f"Aux: {aux_dir}")
print()

# Verificar archivos auxiliares
print("Archivos auxiliares disponibles:")
for f in aux_dir.glob('*.xlsx'):
    wb = load_workbook(f)
    ws = wb.active
    print(f"  - {f.name}: {ws.max_row} rows x {ws.max_column} cols")

print()

# Inicializar merger
print("Inicializando merger...")
m = GalloVisualMerger(gallo, visual, str(aux_dir))

print(f"  Especies Visual cache: {len(m._especies_visual_cache)} entries")
print(f"  Especies Gallo cache: {len(m._especies_gallo_cache)} entries") 
print(f"  Cotización cache: {len(m._cotizacion_cache)} entries")
print(f"  Precios iniciales cache: {len(m._precios_iniciales_cache)} entries")

# Muestra algunas entradas del cache
print("\nEjemplo Especies Visual (primeras 5):")
for i, (k, v) in enumerate(list(m._especies_visual_cache.items())[:5]):
    print(f"  {k}: {v.get('nombre_con_moneda', 'N/A')}")

print("\nEjemplo Precios Iniciales (primeras 5):")
for i, (k, v) in enumerate(list(m._precios_iniciales_cache.items())[:5]):
    print(f"  {k}: {v}")

# Ejecutar merge
print("\n\nEjecutando merge...")
wb_formulas, wb_values = m.merge(output_mode="both")

print(f"\nHojas generadas ({len(wb_formulas.sheetnames)}):")
for name in wb_formulas.sheetnames:
    ws = wb_formulas[name]
    print(f"  - {name}: {ws.max_row} rows x {ws.max_column} cols")

# Guardar ambas versiones
output = "TEST_Merge_v8.xlsx"
wb_formulas.save(output)
print(f"\nOK Guardado (formulas): {output}")

output_values = "TEST_Merge_v8_values.xlsx"
wb_values.save(output_values)
print(f"OK Guardado (values): {output_values}")
