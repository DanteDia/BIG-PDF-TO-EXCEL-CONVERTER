import math
import sys
from openpyxl import Workbook

sys.path.insert(0, '.')

from pdf_converter.datalab.postprocess import postprocess_visual_workbook


def test_precio_tenencias_embedded_quantity_rescue():
    wb = Workbook()
    ws = wb.active
    ws.title = 'PrecioTenenciasIniciales'
    ws.append([
        'Cod.Especie',
        'Ticker',
        'Especie',
        'Cantidad tenencia',
        'Importe invertido',
        'Resultado',
        'Precio tenencia inicial',
    ])
    ws.append([
        '09296',
        'LT',
        'REP ARGENTINA CAP V29/280,676,000.000',
        350000165.24,
        15580324.76,
        15580324.76,
        0.04451519258374165,
    ])
    ws.append([
        '46774',
        'LEBAC',
        'VTO 06/07/16 I06L6',
        0,
        -8854.84,
        8854.84,
        0,
    ])

    postprocess_visual_workbook(wb)

    assert ws.cell(2, 3).value == 'REP ARGENTINA CAP V29'
    assert ws.cell(2, 4).value == 280676000
    assert math.isclose(ws.cell(2, 5).value, 350000165.24, rel_tol=0, abs_tol=0.01)
    assert math.isclose(ws.cell(2, 7).value, 350000165.24 / 280676000, rel_tol=0, abs_tol=1e-12)

    # Control: barras dentro del nombre no deben disparar el rescate.
    assert ws.cell(3, 3).value == 'VTO 06/07/16 I06L6'
    assert ws.cell(3, 4).value == 0
    assert math.isclose(ws.cell(3, 5).value, -8854.84, rel_tol=0, abs_tol=0.01)


if __name__ == '__main__':
    test_precio_tenencias_embedded_quantity_rescue()
    print('ok')