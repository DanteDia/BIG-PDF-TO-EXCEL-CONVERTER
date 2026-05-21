from openpyxl import Workbook

from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger
from pdf_converter.datalab.postprocess import process_precio_tenencias_sheet


def test_postprocess_preserves_negative_invested_amount_as_zero_cost_basis():
    wb = Workbook()
    ws = wb.active
    ws.title = "PrecioTenenciasIniciales"
    ws.append(["Especie", "Cantidad tenencia", "Importe invertido", "Importe tenencia", "Resultado"])
    ws.append(["00534 GGAL GRUPO FINANCIERO GA", "1,200.000", "38,131,810.89-", "8,940,000.00", "47,071,810.89"])

    process_precio_tenencias_sheet(ws)

    assert ws.cell(2, 1).value == "00534"
    assert ws.cell(2, 2).value == "GGAL"
    assert ws.cell(2, 4).value == 1200
    assert ws.cell(2, 5).value == -38131810.89
    assert ws.cell(2, 7).value == 0


def test_negative_invested_amount_with_positive_quantity_is_zero_cost_basis():
    merger = GalloVisualMerger.__new__(GalloVisualMerger)
    merger._precio_tenencias_by_codigo = {}
    merger._precio_tenencias_by_ticker = {}
    merger._precio_tenencias_qty_by_codigo = {}
    merger._precio_tenencias_qty_by_ticker = {}
    merger._precio_tenencias_zero_cost_codes = set()
    merger._precio_tenencias_zero_cost_tickers = set()

    wb = Workbook()
    ws = wb.active
    ws.title = "PrecioTenenciasIniciales"
    ws.append([
        "Cod.Especie",
        "Ticker",
        "Especie",
        "Cantidad tenencia",
        "Importe invertido",
        "Resultado",
        "Precio tenencia inicial",
    ])
    ws.append(["00534", "GGAL", "GRUPO FINANCIERO GA", 1200, -38131810.89, 47071810.89, 31776.509075])

    merger._build_precio_tenencias_cache(ws)

    assert merger._precio_tenencias_by_codigo["534"] == 0
    assert merger._precio_tenencias_by_ticker["GGAL"] == 0
    assert merger._precio_tenencias_qty_by_codigo["534"] == 1200
    assert merger._has_zero_cost_precio_tenencia("00534", "GGAL")
