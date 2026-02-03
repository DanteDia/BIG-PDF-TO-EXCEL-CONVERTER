"""
Test script para verificar cálculos Python contra Excel de referencia.
Compara las columnas de running stock (Q-W para ARS, T-Z para USD).
"""

from pathlib import Path
from openpyxl import load_workbook
from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger

# Archivos
GALLO_EXCEL = "12128_LANDRO_VERONICA_INES_Gallo_Generado_OK.xlsx"
VISUAL_EXCEL = "12128_LANDRO_VERONICA_INES_Visual_Generado_OK.xlsx"
REFERENCE_EXCEL = "12128_LANDRO_VERONICA_INES_Resumen_Impositivo_20260203_0218.xlsx"
OUTPUT_FORMULAS = "TEST_PYTHON_CALC_formulas.xlsx"
OUTPUT_VALUES = "TEST_PYTHON_CALC_values.xlsx"

def main():
    base_dir = Path(__file__).parent
    
    print("=" * 60)
    print("TEST: Cálculos Python vs Excel de Referencia")
    print("=" * 60)
    
    # 1. Ejecutar merge con output dual
    print("\n1. Ejecutando merge con cálculos Python...")
    merger = GalloVisualMerger(
        gallo_path=str(base_dir / GALLO_EXCEL),
        visual_path=str(base_dir / VISUAL_EXCEL),
    )
    
    wb_formulas, wb_values = merger.merge(output_mode="both")
    
    # Guardar ambos archivos
    wb_formulas.save(str(base_dir / OUTPUT_FORMULAS))
    wb_values.save(str(base_dir / OUTPUT_VALUES))
    print(f"   OK Guardado: {OUTPUT_FORMULAS}")
    print(f"   OK Guardado: {OUTPUT_VALUES}")
    
    # 2. Cargar Excel de referencia
    print("\n2. Cargando Excel de referencia...")
    ref_wb = load_workbook(str(base_dir / REFERENCE_EXCEL), data_only=True)
    
    # 3. Comparar Resultado Ventas ARS (columnas Q-W)
    print("\n3. Comparando 'Resultado Ventas ARS' columnas Q-W...")
    compare_sheet(
        wb_values, ref_wb, 
        sheet_name="Resultado Ventas ARS",
        columns_to_compare=[17, 18, 19, 20, 21, 22, 23],  # Q-W
        column_names=["Q:StockIniQty", "R:StockIniPrice", "S:Costo", "T:Neto", "U:Resultado", "V:StockFinQty", "W:StockFinPrice"]
    )
    
    # 4. Comparar Resultado Ventas USD (columnas T-Z)
    print("\n4. Comparando 'Resultado Ventas USD' columnas T-Z...")
    compare_sheet(
        wb_values, ref_wb,
        sheet_name="Resultado Ventas USD",
        columns_to_compare=[20, 21, 22, 23, 24, 25, 26],  # T-Z
        column_names=["T:StockIniQty", "U:StockIniPrice", "V:Costo", "W:Neto", "X:Resultado", "Y:StockFinQty", "Z:StockFinPrice"]
    )
    
    # 5. Comparar Boletos (columnas A, I, L, M, P, R)
    print("\n5. Comparando 'Boletos' columnas con fórmulas...")
    compare_sheet(
        wb_values, ref_wb,
        sheet_name="Boletos",
        columns_to_compare=[1, 9, 12, 13, 16, 18],  # A, I, L, M, P, R
        column_names=["A:TipoInstr", "I:InstrConMoneda", "L:TipoCambio", "M:Bruto", "P:Neto", "R:MonedaEmision"]
    )
    
    print("\n" + "=" * 60)
    print("TEST COMPLETADO")
    print("=" * 60)


def compare_sheet(wb_test, wb_ref, sheet_name: str, columns_to_compare: list, column_names: list):
    """Compara columnas específicas entre dos workbooks."""
    
    if sheet_name not in wb_test.sheetnames:
        print(f"   WARN Hoja '{sheet_name}' no existe en test workbook")
        return
    
    if sheet_name not in wb_ref.sheetnames:
        print(f"   WARN Hoja '{sheet_name}' no existe en referencia")
        return
    
    ws_test = wb_test[sheet_name]
    ws_ref = wb_ref[sheet_name]
    
    max_rows = min(ws_test.max_row, ws_ref.max_row)
    print(f"   Comparando {max_rows - 1} filas de datos...")
    
    errors = []
    matches = 0
    total = 0
    
    for row in range(2, max_rows + 1):
        for col_idx, col_num in enumerate(columns_to_compare):
            val_test = ws_test.cell(row, col_num).value
            val_ref = ws_ref.cell(row, col_num).value
            
            total += 1
            
            # Comparar valores
            if compare_values(val_test, val_ref):
                matches += 1
            else:
                if len(errors) < 20:  # Limitar errores mostrados
                    errors.append({
                        'row': row,
                        'col': column_names[col_idx],
                        'test': val_test,
                        'ref': val_ref
                    })
    
    # Mostrar resultados
    pct = (matches / total * 100) if total > 0 else 0
    print(f"   Coincidencias: {matches}/{total} ({pct:.1f}%)")
    
    if errors:
        print(f"   WARN Primeras {len(errors)} diferencias:")
        for err in errors[:10]:
            print(f"      Fila {err['row']}, {err['col']}: Test={format_val(err['test'])} vs Ref={format_val(err['ref'])}")
    else:
        print(f"   OK Todas las celdas coinciden!")


def compare_values(val1, val2, tolerance=0.01) -> bool:
    """Compara dos valores con tolerancia para floats."""
    # Ambos None o vacíos
    if val1 is None and val2 is None:
        return True
    if val1 == "" and val2 is None:
        return True
    if val1 is None and val2 == "":
        return True
    if val1 == "" and val2 == "":
        return True
    
    # Uno es None/vacío y otro no
    if val1 is None or val1 == "":
        return False
    if val2 is None or val2 == "":
        return False
    
    # Strings
    if isinstance(val1, str) and isinstance(val2, str):
        return val1.strip() == val2.strip()
    
    # Números
    try:
        num1 = float(val1)
        num2 = float(val2)
        
        # Tolerancia relativa para valores grandes
        if abs(num2) > 1:
            return abs(num1 - num2) / abs(num2) < tolerance
        else:
            return abs(num1 - num2) < tolerance
    except (ValueError, TypeError):
        return str(val1) == str(val2)


def format_val(val):
    """Formatea valor para mostrar."""
    if val is None:
        return "None"
    if isinstance(val, float):
        return f"{val:.2f}"
    return str(val)[:30]


if __name__ == "__main__":
    main()
