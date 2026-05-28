from openpyxl import Workbook

from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger


def test_usd_fixed_income_position_uses_direct_usd_unit_price_over_tenencias():
    merger = GalloVisualMerger.__new__(GalloVisualMerger)
    merger.gallo_wb = Workbook()
    ws = merger.gallo_wb.active
    ws.title = "Posicion Inicial"
    ws.append([
        "tipo_especie", "Especie", "Detalle", "Custodia", "Cantidad", "Precio",
        "Importe en Pesos", "% de Cartera", "Importe en Dolares", "% de Cartera",
    ])
    ws.append([
        "RENTA FIJA EN DOLARES",
        "GD30 BONOS REP. ARG. U$S STEP UP V.09/07/30",
        None,
        "CAJA VALORES",
        97814,
        88870,
        86927301.8,
        29.88,
        74436.45,
        29.88,
    ])

    merger._precios_iniciales_cache = {"GD30": {"codigo": 81086, "precio": 89390}}
    merger._precio_tenencias_by_codigo = {"81086": 107.1877240476823}
    merger._precio_tenencias_by_ticker = {}
    merger._precio_tenencias_zero_cost_codes = set()
    merger._precio_tenencias_zero_cost_tickers = set()
    merger._especies_visual_cache = {"81086": {"tipo_especie": "Titulos Publicos"}}
    merger._compute_synthetic_initial_positions = lambda: {}

    output = Workbook()
    output.remove(output.active)
    merger._create_posicion_inicial(output)

    result = output["Posicion Inicial Gallo"]
    expected_price = 74436.45 / 97814
    assert result.cell(2, 4).value == 81086
    assert result.cell(2, 14).value == "PosicionInicialUSD"
    assert result.cell(2, 16).value == expected_price
    assert result.cell(2, 22).value == expected_price


def test_usd_fixed_income_unit_price_is_not_converted_again():
    merger = GalloVisualMerger.__new__(GalloVisualMerger)
    merger._especies_visual_cache = {}

    assert not merger._usd_stock_price_needs_fx_conversion(0.7609999591060584, "Titulos Publicos", "81086")
    assert merger._usd_stock_price_needs_fx_conversion(107.1877240476823, "Titulos Publicos", "81086")
