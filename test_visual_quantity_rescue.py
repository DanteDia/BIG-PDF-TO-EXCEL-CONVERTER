import math
import sys
from openpyxl import Workbook

sys.path.insert(0, '.')

from pdf_converter.datalab.postprocess import postprocess_visual_workbook


BOLETOS_HEADERS = [
    'Tipo de Instrumento', 'Concertación', 'Liquidación', 'Nro. Boleto', 'Moneda',
    'Tipo Operación', 'Cod.Instru', 'Instrumento', 'Cantidad', 'Precio',
    'Tipo Cambio', 'Bruto', 'Interés', 'Gastos', 'Neto'
]

RESULTADO_HEADERS = [
    'Tipo de Instrumento', 'Instrumento', 'Cod.Instrum', 'Concertación', 'Liquidación',
    'Moneda', 'Tipo Operación', 'Cantidad', 'Precio', 'Bruto', 'Interés',
    'Tipo de Cambio', 'Gastos', 'IVA', 'Resultado'
]


def _find_boleto_row(ws, boleto):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 4).value == boleto:
            return row
    raise AssertionError(f'Boleto not found: {boleto}')


def _find_resultado_row(ws, code, concertacion, moneda, operacion):
    for row in range(2, ws.max_row + 1):
        if (
            ws.cell(row, 3).value == code
            and ws.cell(row, 4).value == concertacion
            and ws.cell(row, 6).value == moneda
            and ws.cell(row, 7).value == operacion
        ):
            return row
    raise AssertionError(f'Resultado row not found: {(code, concertacion, moneda, operacion)}')


def test_visual_quantity_rescue_from_resultado_anchors():
    wb = Workbook()
    ws_boletos = wb.active
    ws_boletos.title = 'Boletos'
    ws_boletos.append(BOLETOS_HEADERS)
    ws_boletos.append([
        'Letras del Tesoro nac', '10/7/2025', '10/7/2025', 80261, 'Pesos',
        'Venta Contado Continuo No', 9305, 'LETRAS DEL TESO', -2173913, 1.4568,
        1, -3166999.9, 0, 0, -3166999.9
    ])
    ws_boletos.append([
        'Letras del Tesoro nac', '10/7/2025', '10/7/2025', 80260, 'Dolar Ca',
        'Compra Contado Continuo', 9305, 'LETRAS DEL TESO', 2173913, 0.0012,
        1255, 2500000, 0, 0, 2500000
    ])
    ws_boletos.append([
        'Letras del Tesoro nac', '24/7/2025', '24/7/2025', 88186, 'Pesos',
        'Venta Contado Continuo No', 9305, 'LETRAS DEL TESO', '(2.217.391.3', 1.4544,
        1, '(3.224.973.9', 0, 0, '(3.224.973.9'
    ])
    ws_boletos.append([
        'Letras del Tesoro nac', '24/7/2025', '24/7/2025', 88185, 'Dolar Ca',
        'Compra Contado Continuo', 9305, 'LETRAS DEL TESO', '2.217.391.3', 0.0012,
        1258, 2550000, 0, 0, 2550000
    ])
    ws_boletos.append([
        'Letras del Tesoro nac', '5/9/2025', '5/9/2025', 111965, 'Pesos',
        'Venta Contado Continuo No', 9301, 'LETRAS DEL TESO', -352670.14, 1.574,
        1, -555102.8, 0, 0, -555102.8
    ])
    ws_boletos.append([
        'Letras del Tesoro nac', '29/8/2025', '29/8/2025', 107310, 'Pesos',
        'Compra Contado Continuo', 9301, 'LETRAS DEL TESO', 352670140, 1.5627,
        1, 551117627.78, 0, 0, 551117627.78
    ])
    ws_boletos.append([
        'Obligaciones Negociables', '12/9/2025', '12/9/2025', 116455, 'Pesos',
        'Venta Contado Continuo No', 58449, 'ON BCO GALICIA C', -1285714.2, 1.0197,
        1, -1311029.9, 0, 0, -1311029.9
    ])
    ws_boletos.append([
        'Obligaciones Negociables', '12/9/2025', '12/9/2025', 116454, 'Dolar Ca',
        'Compra Contado Continuo', 58449, 'ON BCO GALICIA C', 1285714.2, 0.0007,
        1432, 900000, 0, 0, 900000
    ])
    ws_boletos.append([
        'Letras del Tesoro nac', '17/6/2025', '17/6/2025', 68292, 'Pesos',
        'Venta Contado Continuo No', 9295, 'LT REP ARGENTIN', -1632072.1, 1.4439,
        1, -2356499.9, 0, 0, -2356499.9
    ])
    ws_boletos.append([
        'Letras del Tesoro nac', '17/6/2025', '17/6/2025', 68291, 'Dolar Ca',
        'Compra Contado Continuo', 9295, 'LT REP ARGENTIN', 1632072.1, 0.0012,
        1182, 1991128, 0, 0, 1991128
    ])
    ws_boletos.append([
        'Títulos Públicos', '29/7/2025', '29/7/2025', 90061, 'Dolar ME',
        'Venta Contado Continuo No', 5921, 'BONO REP. ARGE', -2080213, 0.6009,
        1295, -1249999.95, 0, 0, -1249999.95
    ])
    ws_boletos.append([
        'Títulos Públicos', '29/7/2025', '29/7/2025', 90060, 'Dolar Ca',
        'Compra Contado Continuo', 5921, 'BONO REP. ARGE', 2080213, 0.6,
        1295, 1248127.8, 0, 0, 1248127.8
    ])
    ws_boletos.append([
        'Títulos Públicos', '21/8/2025', '21/8/2025', 102074, 'Dolar ME',
        'Venta Contado Continuo No', 5921, 'BONO REP. ARGE', -4500000, 0.6006,
        1301, -2702700.09, 0, 0, -2702700.09
    ])
    ws_boletos.append([
        'Títulos Públicos', '21/8/2025', '21/8/2025', 102073, 'Dolar Ca',
        'Compra Contado Continuo', 5921, 'BONO REP. ARGE', 4500000, 0.6,
        1301, 2700000, 0, 0, 2700000
    ])
    ws_boletos.append([
        'Títulos Públicos', '3/10/2025', '3/10/2025', 128850, 'Dolar ME',
        'Venta Contado Continuo No', 5921, 'BONO REP. ARGE', -462963, 0.5508,
        1450, -255000.02, 0, 0, -255000.02
    ])
    ws_boletos.append([
        'Títulos Públicos', '3/10/2025', '3/10/2025', 128849, 'Dolar Ca',
        'Compra Contado Continuo', 5921, 'BONO REP. ARGE', 462963, 0.54,
        1424.5, 250000.02, 0, 0, 250000.02
    ])

    ws_result = wb.create_sheet('Resultado Ventas ARS')
    ws_result.append(RESULTADO_HEADERS)
    ws_result.append([
        'Letras del Tesoro nac', 'LETRAS DEL TESORO CAP $ V31/07/25', 9305, '10/7/2025',
        '10/7/2025', 'Pesos', 'Venta Contado Continuo', '(2.173.913.043,)', 1.4568,
        '3.166.999.999,3', 0, 1, 0, 0, '29.499.978,25'
    ])
    ws_result.append([
        'Letras del Tesoro nac', 'LETRAS DEL TESORO CAP $ V31/07/25', 9305, '10/7/2025',
        '10/7/2025', 'Dolar Cable', 'Compra Contado Continuo', '2.173.913.043,0', 0.0012,
        '(3.137.500.000,)', 0, '1.255,00000000', 0, 0, 0
    ])
    ws_result.append([
        'Letras del Tesoro nac', 'LETRAS DEL TESORO CAP $ V31/07/25', 9305, '24/7/2025',
        '24/7/2025', 'Pesos', 'Venta Contado Continuo', '(2.217.391.304,)', 1.4544,
        '3.224.973.912,5', 0, 1, 0, 0, '18.480.703,04'
    ])
    ws_result.append([
        'Letras del Tesoro nac', 'LETRAS DEL TESORO CAP $ V31/07/25', 9305, '24/7/2025',
        '24/7/2025', 'Dolar Cable', 'Compra Contado Continuo', '2.217.391.304,0', 0.0012,
        '(3.207.900.000,)', 0, '1.258,00000000', 0, 0, 0
    ])
    ws_result.append([
        'Letras del Tesoro nac', 'LETRAS DEL TESORO CAP $ V12/09/25', 9301, '29/8/2025',
        '29/8/2025', 'Pesos', 'Compra Contado Continuo', '352.670.140.000,0', 1.5627,
        '(551.117.627,78)', 0, 1, 0, 0, 0
    ])
    ws_result.append([
        'Letras del Tesoro nac', 'LETRAS DEL TESORO CAP $ V12/09/25', 9301, '5/9/2025',
        '5/9/2025', 'Pesos', 'Venta Contado Continuo', '(352.670.140,00)', 1.574,
        '555.102.800,36', 0, 1, 0, 0, '3.985.172,58'
    ])
    ws_result.append([
        'Obligaciones Negociables', 'ON BCO GALICIA CL.21 V10/02/26 $ CG', 58449, '12/9/2025',
        '12/9/2025', 'Pesos', 'Venta Contado Continuo', '(1.285.714.285,0)', 1.0197,
        '1.311.029.999,2', 0, 1, 0, 0, '22.229.987,13'
    ])
    ws_result.append([
        'Obligaciones Negociables', 'ON BCO GALICIA CL.21 V10/02/26 $ CG', 58449, '12/9/2025',
        '12/9/2025', 'Dolar Cable', 'Compra Contado Continuo', '1.285.714.285,0', 0.0007,
        '(1.288.800.000,)', 0, '1.432,00000000', 0, 0, 0
    ])

    ws_result_usd = wb.create_sheet('Resultado Ventas USD')
    ws_result_usd.append(RESULTADO_HEADERS)
    ws_result_usd.append([
        'Títulos Públicos', 'BONO REP. ARGENTINA USD STEP UP 2030', 5921, '29/7/2025',
        '29/7/2025', 'Dolar MEP (I)', 'Venta Contado Continuo', -2080213000, 0.6009,
        1249999.95, 0, 1, 0, 0, 1872.11
    ])
    ws_result_usd.append([
        'Títulos Públicos', 'BONO REP. ARGENTINA USD STEP UP 2030', 5921, '29/7/2025',
        '29/7/2025', 'Dolar Cable', 'Compra Contado Continuo', 2080213000, 0.6,
        -1248127.84, 0, 1, 0, 0, 0
    ])
    ws_result_usd.append([
        'Títulos Públicos', 'BONO REP. ARGENTINA USD STEP UP 2030', 5921, '21/8/2025',
        '21/8/2025', 'Dolar MEP (I)', 'Venta Contado Continuo', -4500000000, 0.6006,
        2702700.09, 0, 1, 0, 0, 2700.14
    ])
    ws_result_usd.append([
        'Títulos Públicos', 'BONO REP. ARGENTINA USD STEP UP 2030', 5921, '21/8/2025',
        '21/8/2025', 'Dolar Cable', 'Compra Contado Continuo', 4500000000, 0.6,
        -2699999.97, 0, 1, 0, 0, 0
    ])
    ws_result_usd.append([
        'Títulos Públicos', 'BONO REP. ARGENTINA USD STEP UP 2030', 5921, '3/10/2025',
        '3/10/2025', 'Dolar MEP (I)', 'Venta Contado Continuo', -462963000, 0.5508,
        255000.03, 0, 1, 0, 0, 5000.01
    ])
    ws_result_usd.append([
        'Títulos Públicos', 'BONO REP. ARGENTINA USD STEP UP 2030', 5921, '3/10/2025',
        '3/10/2025', 'Dolar Cable', 'Compra Contado Continuo', 462963000, 0.54,
        -250000.02, 0, 1, 0, 0, 0
    ])
    ws_result.append([
        'Letras del Tesoro nac', 'LT REP ARGENTINA CAP V30/06/25 $ CG', 9295, '17/6/2025',
        '17/6/2025', 'Pesos', 'Venta Contado Continuo', -1632072131, 1.4439,
        2356499987.7, 0, 1, 0, 0, 2986692
    ])
    ws_result.append([
        'Letras del Tesoro nac', 'LT REP ARGENTINA CAP V30/06/25 $ CG', 9295, '17/6/2025',
        '17/6/2025', 'Dolar Cable', 'Compra Contado Continuo', 1632072131, 0.0012,
        -2353513296, 0, 1182, 0, 0, 0
    ])

    postprocess_visual_workbook(wb)

    row_80261 = _find_boleto_row(ws_boletos, 80261)
    row_80260 = _find_boleto_row(ws_boletos, 80260)
    row_88186 = _find_boleto_row(ws_boletos, 88186)
    row_88185 = _find_boleto_row(ws_boletos, 88185)
    row_111965 = _find_boleto_row(ws_boletos, 111965)
    row_107310 = _find_boleto_row(ws_boletos, 107310)
    row_116455 = _find_boleto_row(ws_boletos, 116455)
    row_116454 = _find_boleto_row(ws_boletos, 116454)
    row_68292 = _find_boleto_row(ws_boletos, 68292)
    row_68291 = _find_boleto_row(ws_boletos, 68291)
    row_90060 = _find_boleto_row(ws_boletos, 90060)
    row_102073 = _find_boleto_row(ws_boletos, 102073)
    row_128850 = _find_boleto_row(ws_boletos, 128850)
    row_128849 = _find_boleto_row(ws_boletos, 128849)

    assert ws_boletos.cell(row_80261, 9).value == -2173913043
    assert ws_boletos.cell(row_80260, 9).value == 2173913043
    assert ws_boletos.cell(row_88186, 9).value == -2217391304
    assert ws_boletos.cell(row_88185, 9).value == 2217391304
    assert ws_boletos.cell(row_111965, 9).value == -352670140
    assert ws_boletos.cell(row_107310, 9).value == 352670140
    assert ws_boletos.cell(row_116455, 9).value == -1285714285
    assert ws_boletos.cell(row_116454, 9).value == 1285714285
    assert ws_boletos.cell(row_68292, 9).value == -1632072131
    assert ws_boletos.cell(row_68291, 9).value == 1632072131
    assert ws_boletos.cell(row_90060, 9).value == 2080213
    assert ws_boletos.cell(row_102073, 9).value == 4500000
    assert ws_boletos.cell(row_128850, 9).value == -462963
    assert ws_boletos.cell(row_128849, 9).value == 462963

    assert math.isclose(
        ws_boletos.cell(row_111965, 12).value,
        ws_boletos.cell(row_111965, 9).value * ws_boletos.cell(row_111965, 10).value,
        rel_tol=0,
        abs_tol=0.01,
    )
    assert math.isclose(ws_boletos.cell(row_111965, 15).value, ws_boletos.cell(row_111965, 12).value, rel_tol=0, abs_tol=0.01)
    assert math.isclose(ws_boletos.cell(row_88186, 12).value, -3224973906.72, rel_tol=0, abs_tol=0.05)
    assert math.isclose(ws_boletos.cell(row_88186, 15).value, -3224973906.72, rel_tol=0, abs_tol=0.05)

    # The anchor sheets should keep integer quantities instead of inflating zero decimal tails.
    row_result_9305 = _find_resultado_row(ws_result, 9305, '10/7/2025', 'Pesos', 'Venta Contado Continuo')
    row_result_9301 = _find_resultado_row(ws_result, 9301, '5/9/2025', 'Pesos', 'Venta Contado Continuo')
    row_result_58449 = _find_resultado_row(ws_result, 58449, '12/9/2025', 'Pesos', 'Venta Contado Continuo')
    row_result_9295 = _find_resultado_row(ws_result, 9295, '17/6/2025', 'Pesos', 'Venta Contado Continuo')
    assert ws_result.cell(row_result_9305, 8).value == -2173913043
    assert ws_result.cell(row_result_9301, 8).value == -352670140
    assert ws_result.cell(row_result_58449, 8).value == -1285714285
    assert ws_result.cell(row_result_9295, 8).value == -1632072131


def test_visual_quantity_rescue_exact_1000x_match_without_punctuation_anomaly():
    wb = Workbook()
    ws_boletos = wb.active
    ws_boletos.title = 'Boletos'
    ws_boletos.append(BOLETOS_HEADERS)
    ws_boletos.append([
        'Títulos Públicos', '30/6/2025', '1/7/2025', 73831, 'Pesos',
        'Venta Contado', 9323, 'BONO NACION TA', -130000, 1.065,
        1, -138450, 0, -83070, -138366.93
    ])
    ws_boletos.append([
        'Títulos Públicos', '30/7/2025', '30/7/2025', 90535, 'Pesos',
        'Compra Contado', 9314, 'BONO TESORO NA', 8300000, 1.168,
        1, 9694400, 0, 10663.84, 9705063.8
    ])
    ws_boletos.append([
        'Títulos Públicos', '19/9/2025', '22/9/2025', 120514, 'Pesos',
        'Venta Contado', 9314, 'BONO TESORO NA', -108300, 1.1525,
        1, -124815.75, 0, -12481.58, -124803.26
    ])
    ws_boletos.append([
        'Títulos Públicos', '19/9/2025', '22/9/2025', 120515, 'Pesos',
        'Venta Contado', 9318, 'BONO TESORO NA', -100000, 0.9725,
        1, -97250, 0, -9725, -97240275
    ])

    ws_result = wb.create_sheet('Resultado Ventas ARS')
    ws_result.append(RESULTADO_HEADERS)
    ws_result.append([
        'Títulos Públicos', 'BONO TESORO NACI CAP V.13/02/26 $ CG - Pesos', 9314, '30/7/2025',
        '30/7/2025', 'Pesos', 'Compra Contado', 8300000, 1.168,
        -9694400, 0, 1, 10663.84, 0, 0
    ])
    ws_result.append([
        'Títulos Públicos', 'BONO TESORO NACI CAP V.13/02/26 $ CG - Pesos', 9314, '19/9/2025',
        '22/9/2025', 'Pesos', 'Venta Contado', -108300000, 1.1525,
        124815750, 0, 1, -12481.58, 0, -140270.5
    ])
    ws_result.append([
        'Títulos Públicos', 'BONO NACIÓN TASA DUAL 15/12/26 $ CG - Pesos', 9323, '30/6/2025',
        '1/7/2025', 'Pesos', 'Venta Contado', -130000000, 1.065,
        138450000, 0, 1, -83070, 0, 0
    ])
    ws_result.append([
        'Títulos Públicos', 'BONO TESORO NACI CAP V.30/06/26 $ CG - Pesos', 9318, '19/9/2025',
        '22/9/2025', 'Pesos', 'Venta Contado', -100000000, 0.9725,
        97250000, 0, 1, -9725, 0, -22389
    ])

    postprocess_visual_workbook(wb)

    row_73831 = _find_boleto_row(ws_boletos, 73831)
    row_120514 = _find_boleto_row(ws_boletos, 120514)
    row_120515 = _find_boleto_row(ws_boletos, 120515)
    row_90535 = _find_boleto_row(ws_boletos, 90535)

    assert ws_boletos.cell(row_73831, 9).value == -130000000
    assert math.isclose(ws_boletos.cell(row_73831, 12).value, -138450000, rel_tol=0, abs_tol=0.01)
    assert ws_boletos.cell(row_120514, 9).value == -108300000
    assert math.isclose(ws_boletos.cell(row_120514, 12).value, -124815750, rel_tol=0, abs_tol=0.01)
    assert ws_boletos.cell(row_120515, 9).value == -100000000
    assert math.isclose(ws_boletos.cell(row_120515, 12).value, -97250000, rel_tol=0, abs_tol=0.01)
    assert math.isclose(ws_boletos.cell(row_120515, 15).value, -97240275, rel_tol=0, abs_tol=0.01)

    # Control sano: una fila ya correcta no debe cambiar.
    assert ws_boletos.cell(row_90535, 9).value == 8300000
    assert math.isclose(ws_boletos.cell(row_90535, 12).value, 9694400, rel_tol=0, abs_tol=0.01)


def test_visual_boletos_prefers_descaled_trailing_dot_zero_group_when_bruto_matches():
    wb = Workbook()
    ws_boletos = wb.active
    ws_boletos.title = 'Boletos'
    ws_boletos.append(BOLETOS_HEADERS)
    ws_boletos.append([
        'Acciones', '4/11/2025', '4/11/2025', 146258, 'Pesos',
        'Venta Contado', 94, 'BBVA ARG ESC S', '(157.000,000000)', 9020,
        1, -1416140, 0, -2913, -1413227
    ])
    ws_boletos.append([
        'Acciones', '4/11/2025', '4/11/2025', 146257, 'Pesos',
        'Block Venta', 94, 'BBVA ARG ESC S', '(843.000,000000)', 9020,
        1, -7603860, 0, -6440.47, -7597419.5
    ])
    ws_boletos.append([
        'Acciones', '28/10/2025', '28/10/2025', 141673, 'Pesos',
        'Venta Contado', 710, 'YPF S.A. ESCRIT.', '(350.000,000000)', 49400,
        1, -17290000, 0, -35565.53, -17254434.47
    ])
    ws_boletos.append([
        'Acciones', '14/11/2025', '14/11/2025', 152732, 'Pesos',
        'Block Venta', 457, 'PAMPA ENERGIA S', '(7.000,0000000)', 5145,
        1, -36015, 0, -30.5, -35984.5
    ])
    ws_boletos.append([
        'Acciones', '25/6/2025', '26/6/2025', 72214, 'Pesos',
        'Block Compra', 457, 'PAMPA ENERGIA S', '1.000,00000', 3365,
        1, 3365000, 0, 2850.16, 3367850.16
    ])

    postprocess_visual_workbook(wb)

    assert ws_boletos.cell(_find_boleto_row(ws_boletos, 146258), 9).value == -157
    assert ws_boletos.cell(_find_boleto_row(ws_boletos, 146257), 9).value == -843
    assert ws_boletos.cell(_find_boleto_row(ws_boletos, 141673), 9).value == -350
    assert ws_boletos.cell(_find_boleto_row(ws_boletos, 152732), 9).value == -7
    assert ws_boletos.cell(_find_boleto_row(ws_boletos, 72214), 9).value == 1000


def test_visual_repeated_comma_zero_tail_quantities_stay_aligned_between_boletos_and_resultado():
    wb = Workbook()
    ws_boletos = wb.active
    ws_boletos.title = 'Boletos'
    ws_boletos.append(BOLETOS_HEADERS)
    ws_boletos.append([
        'Acciones', '3/11/2025', '4/11/2025', 146258, 'Pesos',
        'Venta Contado', 94, 'BBVA ARG ESC S', '(157,000,000)', 9020,
        1, '(1.416.140,0)', 0, '(2.913,00)', '(1.413.227,0)'
    ])
    ws_boletos.append([
        'Acciones', '14/11/2025', '14/11/2025', 152732, 'Pesos',
        'Block Venta', 457, 'PAMPA ENERGIA S', '(7,000,000,00)', 5145,
        1, '(36.015,00)', 0, '(30,50)', '(35.984,50)'
    ])
    ws_boletos.append([
        'Acciones', '7/11/2025', '10/11/2025', 149217, 'Pesos',
        'Block Compra', 94, 'BBVA ARG ESC S', '1.000,000,00', 7795,
        1, '7.795.000,0', 0, '6.602,37', '7.801.602,3'
    ])

    ws_result = wb.create_sheet('Resultado Ventas ARS')
    ws_result.append(RESULTADO_HEADERS)
    ws_result.append([
        'Acciones', 'BBVA ARG ESC S 1 V. - Pesos', 94, '3/11/2025',
        '4/11/2025', 'Pesos', 'Venta Contado', '(157,000,000)', 9020,
        '(1.416.140,0)', 0, 1, '(2.913,00)', 0, '552.981,35'
    ])
    ws_result.append([
        'Acciones', 'PAMPA ENERGIA S.A. ESCRIT. 1 VOTO - Pesos', 457, '14/11/2025',
        '14/11/2025', 'Pesos', 'Block Venta', '(7,000,000,00)', 5145,
        '(36.015,00)', 0, 1, '(30,50)', 0, '29.226,55'
    ])

    postprocess_visual_workbook(wb)

    row_146258 = _find_boleto_row(ws_boletos, 146258)
    row_152732 = _find_boleto_row(ws_boletos, 152732)
    row_149217 = _find_boleto_row(ws_boletos, 149217)
    result_94 = _find_resultado_row(ws_result, 94, '3/11/2025', 'Pesos', 'Venta Contado')
    result_457 = _find_resultado_row(ws_result, 457, '14/11/2025', 'Pesos', 'Block Venta')

    assert ws_boletos.cell(row_146258, 9).value == -157
    assert ws_boletos.cell(row_152732, 9).value == -7
    assert ws_boletos.cell(row_149217, 9).value == 1000
    assert ws_result.cell(result_94, 8).value == -157
    assert ws_result.cell(result_457, 8).value == -7


if __name__ == '__main__':
    test_visual_quantity_rescue_from_resultado_anchors()
    test_visual_quantity_rescue_exact_1000x_match_without_punctuation_anomaly()
    test_visual_boletos_prefers_descaled_trailing_dot_zero_group_when_bruto_matches()
    test_visual_repeated_comma_zero_tail_quantities_stay_aligned_between_boletos_and_resultado()
    print('ok')
