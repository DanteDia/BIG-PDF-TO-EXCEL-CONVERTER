from datetime import datetime

from openpyxl import Workbook

from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger


def test_visual_rentas_usd_importe_is_not_reduced_by_gastos_again():
    merger = GalloVisualMerger.__new__(GalloVisualMerger)
    merger._especies_visual_cache = {
        "5921": {
            "moneda_emision": "Dolar MEP (local)",
            "tipo_especie": "Títulos Públicos",
        }
    }

    wb = Workbook()
    wb.remove(wb.active)
    source = wb.create_sheet("Rentas y Dividendos Gallo")
    headers = [
        "Tipo de Instrumento", "Concertación", "Liquidación", "Nro. Boleto", "Moneda",
        "Tipo Operación", "Cod.Instrum", "Instrumento Crudo", "InstrumentoConMoneda",
        "Cantidad", "Precio", "Tipo Cambio", "Bruto", "Interés", "Gastos", "Costo",
        "Neto Calculado", "Origen", "moneda emision", "Auditoría",
    ]
    for col, header in enumerate(headers, 1):
        source.cell(1, col, header)

    source.append([
        "Títulos Públicos", datetime(2025, 7, 10), "10/7/2025", 10709, "Dolar MEP",
        "RENTA", 5921, "BONO REP. ARGENTINA USD STEP UP 2030", "", 0, 1, 1,
        25.77, 0, 0.04, 0, 25.73, "Visual-Rentas Dividendos USD",
        "Dolar MEP (local)", "Origen: Visual-Rentas Dividendos USD",
    ])
    source.append([
        "Títulos Públicos", datetime(2025, 1, 9), None, 1149, "Dolar MEP",
        "RENTA", 5921, "BONO USD 2030 LA", "", 0, 1, 1,
        28.16, 0, 0.51, 0, 27.65, "Gallo-Renta Fija Dolares",
        "Dolar MEP (local)", "Origen: Gallo-Renta Fija Dolares",
    ])

    merger._create_rentas_dividendos_usd(wb)
    result = wb["Rentas Dividendos USD"]

    by_origin = {result.cell(row, 14).value: result.cell(row, 13).value for row in range(2, result.max_row + 1)}
    assert by_origin["Visual-Rentas Dividendos USD"] == 25.77
    assert by_origin["Gallo-Renta Fija Dolares"] == 27.65