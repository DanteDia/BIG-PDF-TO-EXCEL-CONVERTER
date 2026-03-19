from openpyxl import load_workbook
from pathlib import Path

p = Path(r"c:/Users/xarodan/Downloads/Resumen Impositivo- Branch dots.OCR/LOCAL_VERIFY_20260303_162032/AGUIAR_MERGED_values.xlsx")
wb = load_workbook(p, data_only=True)

out=[]
for sheet in ["Cauciones Tomadoras","Cauciones Colocadoras"]:
    ws = wb[sheet]
    total=miss_boleto=miss_tasa=neg_cf=pos_cf=0
    by_origin={}
    for r in range(2, ws.max_row+1):
        if ws.cell(r,4).value is None:
            continue
        total += 1
        orig = str(ws.cell(r,16).value)
        by_origin[orig] = by_origin.get(orig,0)+1
        b=ws.cell(r,5).value
        t=ws.cell(r,9).value
        cf=ws.cell(r,14).value
        if b in (None,''):
            miss_boleto += 1
        if t in (None,''):
            miss_tasa += 1
        if isinstance(cf,(int,float)):
            if cf<0: neg_cf +=1
            elif cf>0: pos_cf +=1
    out.append(f"{sheet}: total={total} miss_boleto={miss_boleto} miss_tasa={miss_tasa} neg_cf={neg_cf} pos_cf={pos_cf}")
    for k,v in sorted(by_origin.items(), key=lambda x:x[1], reverse=True)[:5]:
        out.append(f"  {k}: {v}")

for sheet in ["Cauciones Tomadoras","Cauciones Colocadoras"]:
    ws=wb[sheet]
    out.append(f"First Visual rows in {sheet}:")
    c=0
    for r in range(2, ws.max_row+1):
        o=str(ws.cell(r,16).value or '')
        if 'Visual' in o:
            vals=[ws.cell(r,i).value for i in range(1,15)]
            out.append(f"  r{r}: {vals}")
            c+=1
            if c>=3: break

print('\n'.join(out))
