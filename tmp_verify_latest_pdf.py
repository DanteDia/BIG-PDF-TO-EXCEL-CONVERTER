from pathlib import Path
from openpyxl import load_workbook
from pdf_converter.datalab.excel_to_pdf import ExcelToPdfExporter

root = Path(r"c:/Users/xarodan/Downloads/Resumen Impositivo- Branch dots.OCR")
merged = root / "LOCAL_VERIFY_CAU_20260303_200308" / "AGUIAR_MERGED_values.xlsx"
wb = load_workbook(merged, data_only=True)

for sheet in ["Cauciones Tomadoras", "Cauciones Colocadoras"]:
    ws = wb[sheet]
    miss = 0
    total = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 4).value:
            total += 1
            if ws.cell(r, 5).value in (None, ''):
                miss += 1
    print(f"{sheet}: missing boleto in Excel = {miss}/{total}")

out_pdf = root / "LOCAL_VERIFY_CAU_20260303_200308" / "AGUIAR_Resumen_Impositivo_VERIFY_CAU_v2.pdf"
exp = ExcelToPdfExporter(str(merged), {'numero': '13056', 'nombre': 'AGUIAR JUAN MARTIN'})
out_pdf.write_bytes(exp.export_to_pdf())
print("PDF generated:", out_pdf)
