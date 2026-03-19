#!/usr/bin/env python
"""Verify TEST_output_v3 against desired output."""
from openpyxl import load_workbook

wb_test = load_workbook('TEST_output_v3_values.xlsx')
wb_desired = load_workbook('12128_LANDRO_VERONICA_INES_Resumen_Impositivo_Valores_OK.xlsx')

ws_test = wb_test['Posicion Inicial Gallo']
ws_desired = wb_desired['Posicion Inicial Gallo']

print('=== Posicion Inicial Gallo Comparison ===')
print(f'Test rows: {ws_test.max_row - 1}, Desired rows: {ws_desired.max_row - 1}')
print()

header = f'{"Ticker":<15} {"P.a.Util(test)":>18} {"P.a.Util(desired)":>18} {"P.Nom(test)":>18} {"P.Nom(desired)":>18} {"Match":>6}'
print(header)
print('-' * len(header))

def approx_eq(a, b, tol=0.01):
    try:
        fa = float(a or 0)
    except:
        # Handle formulas like "=359659.414/24"
        if isinstance(a, str) and a.startswith('='):
            try:
                fa = eval(a[1:])
            except:
                return str(a) == str(b)
        else:
            return str(a) == str(b)
    try:
        fb = float(b or 0)
    except:
        if isinstance(b, str) and b.startswith('='):
            try:
                fb = eval(b[1:])
            except:
                return str(a) == str(b)
        else:
            return str(a) == str(b)
    return abs(fa - fb) < tol

mismatches = 0
for row in range(2, max(ws_test.max_row, ws_desired.max_row) + 1):
    t_ticker = ws_test.cell(row, 2).value if row <= ws_test.max_row else ''
    d_ticker = ws_desired.cell(row, 2).value if row <= ws_desired.max_row else ''

    t_pu = ws_test.cell(row, 16).value if row <= ws_test.max_row else ''
    d_pu = ws_desired.cell(row, 16).value if row <= ws_desired.max_row else ''

    t_pn = ws_test.cell(row, 22).value if row <= ws_test.max_row else ''
    d_pn = ws_desired.cell(row, 22).value if row <= ws_desired.max_row else ''

    match_u = approx_eq(t_pu, d_pu)
    match_n = approx_eq(t_pn, d_pn)
    status = 'OK' if (match_u and match_n) else 'FAIL'
    if status == 'FAIL':
        mismatches += 1

    ticker = t_ticker or d_ticker or ''
    print(f'{str(ticker):<15} {str(t_pu):>18} {str(d_pu):>18} {str(t_pn):>18} {str(d_pn):>18} {status:>6}')

print()
print(f'Total mismatches: {mismatches}')

# Also check the new sheets exist
print()
print('=== New sheets check ===')
test_sheets = wb_test.sheetnames
print(f'RatiosCedearsAcciones present: {"RatiosCedearsAcciones" in test_sheets}')
print(f'PrecioTenenciasIniciales present: {"PrecioTenenciasIniciales" in test_sheets}')

# Check PrecioTenenciasIniciales has new columns
if 'PrecioTenenciasIniciales' in test_sheets:
    ws_pt = wb_test['PrecioTenenciasIniciales']
    headers = [ws_pt.cell(1, c).value for c in range(1, ws_pt.max_column + 1)]
    print(f'PrecioTenenciasIniciales headers: {headers}')
    # Show NVIDIA and TESLA rows
    for row in range(2, ws_pt.max_row + 1):
        ticker = ws_pt.cell(row, 2).value
        if ticker and str(ticker).upper() in ['NVIDIA', 'TESLA']:
            vals = [ws_pt.cell(row, c).value for c in range(1, ws_pt.max_column + 1)]
            print(f'  {ticker}: {vals}')
