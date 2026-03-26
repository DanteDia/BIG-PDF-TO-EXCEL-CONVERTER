from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(r"c:/Users/xarodan/Downloads/Resumen Impositivo- Branch dots.OCR")


def _latest_existing(*paths: Path):
    for path in paths:
        if path.exists():
            return path
    return paths[0]


def _load(path: Path):
    return load_workbook(path, data_only=True)


def _sheet_rows(ws):
    return max(ws.max_row - 1, 0)


def check_currency_rows(workbook_path: Path, sheet_name: str, expected: str):
    wb = _load(workbook_path)
    ws = wb[sheet_name]
    bad_rows = []
    for row in range(2, ws.max_row + 1):
        moneda = str(ws.cell(row, 10).value or '').lower()
        if expected == 'USD':
            ok = 'dolar' in moneda or moneda == 'usd'
        else:
            ok = 'peso' in moneda or moneda == 'ars'
        if not ok:
            bad_rows.append(row)
    return _sheet_rows(ws), bad_rows


def count_non_empty_rows(workbook_path: Path, sheet_name: str):
    wb = _load(workbook_path)
    ws = wb[sheet_name]
    return sum(1 for row in range(2, ws.max_row + 1) if any(ws.cell(row, col).value not in (None, '') for col in range(1, ws.max_column + 1)))


def check_resultado_vs_bruto(workbook_path: Path, sheet_name: str = 'Resultado Ventas ARS'):
    wb = _load(workbook_path)
    ws = wb[sheet_name]
    bad_rows = []
    for row in range(2, ws.max_row + 1):
        bruto = ws.cell(row, 11).value or 0
        resultado = ws.cell(row, 21).value or 0
        try:
            if abs(float(resultado)) > abs(float(bruto)) and abs(float(bruto)) > 0:
                bad_rows.append((row, ws.cell(row, 4).value, bruto, resultado))
        except Exception:
            continue
    return bad_rows


def resumen_rows(workbook_path: Path):
    wb = _load(workbook_path)
    ws = wb['Resumen']
    return [
        [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        for row in range(1, 4)
    ]


def main():
    glozman = _latest_existing(
        ROOT / 'Ejemplo Glozman error moneda pesos en seccion USD' / '12766_GLOZMAN_DARIO_EDMUNDO_Resumen_Impositivo_FIXED_v2.xlsx',
        ROOT / 'Ejemplo Glozman error moneda pesos en seccion USD' / '12766_GLOZMAN_DARIO_EDMUNDO_Resumen_Impositivo_FIXED.xlsx',
    )
    salvo = _latest_existing(
        ROOT / '11896_SALVO_MARTIN_Resumen_Impositivo_REGRESSION_v2.xlsx',
        ROOT / '11896_SALVO_MARTIN_Resumen_Impositivo_REGRESSION.xlsx',
    )

    print('=== GLOZMAN ===')
    if glozman.exists():
        ars_rows = count_non_empty_rows(glozman, 'Rentas Dividendos ARS')
        usd_rows, bad = check_currency_rows(glozman, 'Rentas Dividendos USD', 'USD')
        print(f'Rentas Dividendos ARS rows: {ars_rows}')
        print(f'Rentas Dividendos USD rows: {usd_rows}')
        print(f'Bad USD currency rows: {bad}')
        for row in resumen_rows(glozman):
            print(row)
    else:
        print('Missing file:', glozman)

    print('\n=== SALVO ===')
    if salvo.exists():
        ars_rows = count_non_empty_rows(salvo, 'Rentas Dividendos ARS')
        usd_rows, bad = check_currency_rows(salvo, 'Rentas Dividendos USD', 'USD')
        print(f'Rentas Dividendos ARS rows: {ars_rows}')
        print(f'Rentas Dividendos USD rows: {usd_rows}')
        print(f'Bad USD currency rows: {bad}')
        print(f'Resultado>Bruto rows: {check_resultado_vs_bruto(salvo)}')
        for row in resumen_rows(salvo):
            print(row)
    else:
        print('Missing file:', salvo)


if __name__ == '__main__':
    main()
