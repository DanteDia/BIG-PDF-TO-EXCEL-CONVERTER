#!/usr/bin/env python
"""Generate and verify output v4."""
import sys
sys.path.insert(0, '.')
from pathlib import Path
from pdf_converter.datalab.merge_gallo_visual import GalloVisualMerger
from openpyxl import load_workbook

# Generate
merger = GalloVisualMerger(
    '12128_LANDRO_VERONICA_INES_Gallo_Generado_OK.xlsx',
    '12128_LANDRO_VERONICA_INES_Visual_Generado_OK.xlsx',
    str(Path('pdf_converter/datalab/aux_data')),
    precio_tenencias_path='precio_estructurado.xlsx'
)
wb_f, wb_v = merger.merge(output_mode='both')
wb_v.save('TEST_output_v4_values.xlsx')
wb_f.save('TEST_output_v4_formulas.xlsx')
print("Generated TEST_output_v4")

# Verify
wb = load_workbook('TEST_output_v4_values.xlsx')

# Check PrecioTenenciasIniciales
ws_pt = wb['PrecioTenenciasIniciales']
print("\n=== PrecioTenenciasIniciales ===")
headers = [ws_pt.cell(1, c).value for c in range(1, ws_pt.max_column + 1)]
print(f"Headers: {headers}")
for row in range(2, min(ws_pt.max_row + 1, 40)):
    cod = ws_pt.cell(row, 1).value
    ticker = ws_pt.cell(row, 2).value
    ratio = ws_pt.cell(row, 8).value  # Ratio CEDEAR
    adj = ws_pt.cell(row, 9).value    # Precio Ajustado
    if ticker:
        print(f"  {str(cod):>8} {str(ticker):<12} ratio={ratio}  adj={adj}")

# Check Posicion Inicial Gallo
ws_pi = wb['Posicion Inicial Gallo']
print("\n=== Posicion Inicial Gallo ===")
for row in range(2, ws_pi.max_row + 1):
    ticker = ws_pi.cell(row, 2).value
    pu = ws_pi.cell(row, 16).value
    pn = ws_pi.cell(row, 22).value
    same = "SAME" if pu == pn else f"DIFF!"
    print(f"  {str(ticker):<15} P.Util={str(pu):>18}  P.Nom={str(pn):>18}  {same}")

# Compare with desired
wb_d = load_workbook('12128_LANDRO_VERONICA_INES_Resumen_Impositivo_Valores_OK.xlsx')
ws_d = wb_d['Posicion Inicial Gallo']

def eval_val(v):
    if isinstance(v, str) and v.startswith('='):
        try:
            return eval(v[1:])
        except:
            return v
    return v

print("\n=== Comparison with desired ===")
mismatches = 0
for row in range(2, max(ws_pi.max_row, ws_d.max_row) + 1):
    t = ws_pi.cell(row, 2).value or ''
    pu_t = ws_pi.cell(row, 16).value
    pn_t = ws_pi.cell(row, 22).value
    pu_d = eval_val(ws_d.cell(row, 16).value)
    pn_d = eval_val(ws_d.cell(row, 22).value)
    
    try:
        pu_ok = abs(float(pu_t or 0) - float(pu_d or 0)) < 0.01
    except:
        pu_ok = str(pu_t) == str(pu_d)
    try:
        pn_ok = abs(float(pn_t or 0) - float(pn_d or 0)) < 0.01
    except:
        pn_ok = str(pn_t) == str(pn_d)
    
    status = "OK" if pu_ok and pn_ok else "FAIL"
    if status == "FAIL":
        mismatches += 1
        print(f"  {str(t):<15} PU: {pu_t} vs {pu_d} ({pu_ok})  PN: {pn_t} vs {pn_d} ({pn_ok})")
    else:
        print(f"  {str(t):<15} OK")

print(f"\nTotal mismatches: {mismatches}")

# Check sheet presence
print(f"\nRatiosCedearsAcciones sheet: {'RatiosCedearsAcciones' in wb.sheetnames}")
if 'RatiosCedearsAcciones' in wb.sheetnames:
    ws_r = wb['RatiosCedearsAcciones']
    print(f"  {ws_r.max_row - 1} rows, headers: {[ws_r.cell(1,c).value for c in range(1, ws_r.max_column+1)]}")
